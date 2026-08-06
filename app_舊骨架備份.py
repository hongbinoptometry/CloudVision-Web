from __future__ import annotations

import csv
import io
import json
import os
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("CLOUDVISION_DATA_DIR", BASE_DIR / "data"))
DATA_DIR.mkdir(parents=True, exist_ok=True)
JSONL_PATH = DATA_DIR / "cloudvision_results.jsonl"
CSV_PATH = DATA_DIR / "cloudvision_results.csv"

app = Flask(__name__)

FIELDS = [
    "id", "date", "time", "name", "phone", "visual_acuity_right", "visual_acuity_left",
    "near_reading", "astigmatism_right", "astigmatism_left", "amsler_right", "amsler_left",
    "appointment", "appointment_date", "appointment_time", "note", "device", "visitor_id",
]


def clean(value: Any, limit: int = 300) -> str:
    return str(value or "").strip()[:limit]


def read_rows() -> list[dict[str, str]]:
    if not JSONL_PATH.exists():
        return []
    rows: list[dict[str, str]] = []
    for line in JSONL_PATH.read_text(encoding="utf-8").splitlines():
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return rows


def save_row(payload: dict[str, Any]) -> dict[str, str]:
    now = datetime.now()
    row = {key: clean(payload.get(key)) for key in FIELDS}
    row.update({"id": uuid.uuid4().hex[:12], "date": now.strftime("%Y-%m-%d"), "time": now.strftime("%H:%M:%S")})
    with JSONL_PATH.open("a", encoding="utf-8") as f:
        f.write(json.dumps(row, ensure_ascii=False) + "\n")
    exists = CSV_PATH.exists()
    with CSV_PATH.open("a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS)
        if not exists:
            writer.writeheader()
        writer.writerow(row)
    return row


def build_excel(rows: list[dict[str, str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "完整資料"
    headers = [
        "日期", "時間", "姓名", "電話", "右眼視力", "左眼視力", "30cm近距閱讀",
        "右眼散光鐘", "左眼散光鐘", "右眼Amsler", "左眼Amsler", "是否預約",
        "預約日期", "預約時段", "備註", "裝置", "資料編號",
    ]
    keys = ["date", "time", "name", "phone", "visual_acuity_right", "visual_acuity_left", "near_reading",
            "astigmatism_right", "astigmatism_left", "amsler_right", "amsler_left", "appointment",
            "appointment_date", "appointment_time", "note", "device", "id"]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(k, "") for k in keys])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCEBFF")
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        width = min(max(len(str(c.value or "")) for c in col) + 2, 35)
        ws.column_dimensions[col[0].column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@app.get("/")
def root():
    return redirect(url_for("cloud_home"))


@app.get("/health")
def health():
    return jsonify(ok=True, service="CloudVision-Web")


@app.get("/cloud")
def cloud_home():
    return render_template("index.html")


@app.get("/cloud/calibration")
def calibration():
    return render_template("calibration.html")


@app.get("/cloud/assessment")
def assessment():
    return render_template("assessment.html")


@app.post("/cloud/result")
def result():
    payload = request.get_json(silent=True) or request.form.to_dict()
    row = save_row(payload)
    return jsonify(ok=True, id=row["id"])


@app.get("/cloud/today")
def today():
    date_key = datetime.now().strftime("%Y-%m-%d")
    rows = [r for r in read_rows() if r.get("date") == date_key]
    rows.reverse()
    return render_template("today.html", rows=rows, today=date_key)


@app.get("/cloud/today.xlsx")
def today_excel():
    date_key = datetime.now().strftime("%Y-%m-%d")
    rows = [r for r in read_rows() if r.get("date") == date_key]
    return send_file(io.BytesIO(build_excel(rows)), as_attachment=True,
                     download_name=f"CloudVision_{date_key}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=False)
