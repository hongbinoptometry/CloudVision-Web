# Cloud Vision V3.0：教學／研究版；一般使用者匿名、專業人員分類登錄、使用統計與問卷自動寫入背景 Excel。
# Cloud Vision V11.1-1：後台統計卡可點擊、今日名單可篩選、個人詳細資料。
# Cloud Vision V10.7：只修改第一個確認視標頁，改為可捲動並加入 Safari 底部安全留白。
# Cloud Vision V6.3：恢復完整校正流程，完整視力表頁可上下捲動到底部。
# V4.24：修正受試者瀏覽器收到題目後未顯示題目；中央同步顯示 active_targets。
# -*- coding: utf-8 -*-
"""
V8.8：阿姆斯勒方格受試者端改用內嵌 SVG 20×20 格線，修正 iPad Safari 只顯示中央黑點。
字母型視力表｜第四功能 V4.21 瀏覽器第一步測試版
全畫面、觀看距離連動、Bailey–Lovie 風格排列

用途：
- 作為「視覺功能研究平台」之外的獨立第四支程式。
- 每列固定 5 個 Sloan 字母。
- 相鄰列為 0.1 logMAR 級距。
- 輸入觀看距離後，依 5 arcmin 完整字母視角自動換算實體字高。
- 提供 5 cm 校正線，正式測量前必須用尺校正。

快捷鍵：
- F11：切換全畫面
- Esc：離開全畫面
- F2：顯示／隱藏設定列
- F5：重新隨機排列
- Home：回到最上方
- F10：Bagolini 單點全黑模式（再按一次返回）
- 滑鼠滾輪：上下瀏覽
"""

# V4.21 第一階段：沿用 V4.20 全部功能。
# Cloud Vision V5.2：受試者端改為「校正 → 等待 → 作答」三個獨立畫面。
# 本版僅建立可辨識的測試版本，不重寫視標、尺寸、連線或作答邏輯。

from __future__ import annotations

import shutil
import sys
import math
import json
import os
import random
import tkinter as tk
import socket
import threading
import queue
import time
import subprocess
import re
import secrets
import base64
import io
import html
import csv
import webbrowser
import smtplib
import ssl
from email.message import EmailMessage
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs, quote
from tkinter import messagebox, ttk, simpledialog
from PIL import Image, ImageDraw, ImageFont, ImageOps, ImageTk

APP_TITLE = "CloudVision Web V4.0｜結果頁顯示右眼左眼雙眼視力"
SLOAN_LETTERS = "CDHKNORSVZ"
CARDINAL_DIRECTIONS = ("up", "right", "down", "left")
LANDOLT_DIRECTIONS = ("up", "up_right", "right", "down_right", "down", "down_left", "left", "up_left")
OPTOTYPE_DIRECTIONS = CARDINAL_DIRECTIONS
DIRECTION_SYMBOLS = {
    "up": "↑", "up_right": "↗", "right": "→", "down_right": "↘",
    "down": "↓", "down_left": "↙", "left": "←", "up_left": "↖",
}
REWARD_STICKERS = ["🍎", "⭐", "🐶", "🚗", "🦖", "🐱", "🧸", "🐼", "🚀", "👑"]

# 0.1 logMAR 級距，十進位視力由 0.10 至 1.58。
DECIMAL_LEVELS = [
    0.100, 0.126, 0.158, 0.200, 0.251, 0.316, 0.398,
    0.501, 0.631, 0.794, 1.000, 1.259, 1.585,
]


class FullscreenAcuityChart:
    def __init__(self, root: tk.Tk, cloud_mode: bool = False) -> None:
        self.cloud_mode = cloud_mode
        self.root = root
        self.root.title(APP_TITLE)
        self.root.configure(background="white")

        self.distance_var = tk.StringVar(value="300")
        self.calibration_var = tk.StringVar(value="1.190476")  # 本機實測：5 cm 線僅 4.2 cm，故 5/4.2
        self.show_labels_var = tk.BooleanVar(value=True)
        self.optotype_mode_var = tk.StringVar(value="letter")
        self.fullscreen_var = tk.BooleanVar(value=True)
        self.status_var = tk.StringVar(value="已套用本機實測校正：原 5 cm 線為 4.2 cm，倍率 1.190476；請再用尺確認。")

        self.photo_images: list[ImageTk.PhotoImage] = []
        self.row_letters: list[list[str]] = []
        self.row_directions: list[list[str]] = []
        self.row_widgets: list[tk.Widget] = []
        # 手機選擇視力值後，只顯示該列；None 代表完整視力表。
        self.remote_single_row_index: int | None = None
        # 起始刺激位置：0～4 對應該列五個字母。
        self.remote_stimulus_index: int | None = None
        # 顯示刺激數量：1～5；5 代表完整一排。
        self.remote_stimulus_count = 1
        # 手機操作模式：full=完整視力表；single=單一刺激。
        self.remote_mode = "full"
        self.controls_visible = True
        self._resize_after_id: str | None = None
        self.current_view = "chart"
        self.dial_canvas: tk.Canvas | None = None
        self.amsler_canvas: tk.Canvas | None = None
        self.worth_canvas: tk.Canvas | None = None
        self.worth_mode = "near"
        self.worth_focus_mode = False
        self._worth_previous_fullscreen = False
        self.bagolini_canvas: tk.Canvas | None = None
        self.bagolini_mode = "near"
        self.bagolini_dot_px = 3
        self.bagolini_focus_mode = False
        self._bagolini_previous_fullscreen = False
        self.thorington_canvas: tk.Canvas | None = None
        self.thorington_focus_mode = False
        self.thorington_distance_cm = 100.0
        self.thorington_dot_px = 4
        self._thorington_previous_fullscreen = False
        self.remote_server = None
        self.remote_thread = None
        self.remote_port = int(os.environ.get("CLOUDVISION_BACKEND_PORT", "8765"))
        # 連線方式：wifi=電腦與手機連同一個 Wi-Fi；hotspot=手機連電腦行動熱點。
        # 預設使用共用 Wi-Fi，避免手機已連家中 Wi-Fi 時，QR Code 卻誤用 192.168.137.1。
        self.connection_mode = "wifi"
        self.connection_ip = ""
        self.remote_url = ""
        self.control_url = ""
        self.answer_url = ""
        # 每次啟動程式都產生新的連線識別碼，避免舊手機頁面誤報連線。
        self.connection_session_token = secrets.token_urlsafe(18)
        # 電腦行動熱點 Wi-Fi QR Code：由使用者輸入熱點名稱與密碼，程式會記住。
        self.hotspot_ssid_var = tk.StringVar(value="JohnVA")
        self.hotspot_password_var = tk.StringVar(value="12345678")
        self.hotspot_security_var = tk.StringVar(value="WPA2（相容模式）")
        # 測驗者後台密碼可在電腦端修改；公開頁面不顯示後台入口。
        self.examiner_password_var = tk.StringVar(value="123456")
        self.examiner_sessions: set[str] = set()
        self.examiner_sessions_lock = threading.RLock()
        self.hotspot_config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "字母手機版_熱點設定.json")
        self.remote_command_queue: queue.Queue[tuple[str, str]] = queue.Queue()
        self.test_mode = "child"
        self.question_status = "waiting_send"
        self.question_started_at = time.monotonic()
        self.active_targets: list[str] = []
        self.answer_sequence: list[str] = []
        self.answer_results: list[bool] = []
        self.last_answer = ""
        self.last_answer_correct: bool | None = None
        self.last_elapsed: float | None = None
        self.question_id = 0
        # 兒童獎勵貼圖由測驗者手動送出，並持續累積。
        self.sticker_collection: list[str] = []
        self.pending_stickers: list[str] = []
        self.sticker_version = 0
        self.sticker_collection_label: tk.Label | None = None
        self.pending_sticker_var = tk.StringVar(value="已選：尚未選擇貼圖")
        self.qr_window: tk.Toplevel | None = None
        self.qr_photo: ImageTk.PhotoImage | None = None
        self.qr_photo_control: ImageTk.PhotoImage | None = None
        self.qr_photo_answer: ImageTk.PhotoImage | None = None
        self.qr_photo_hotspot: ImageTk.PhotoImage | None = None
        self.qr_photo_current_wifi: ImageTk.PhotoImage | None = None
        self.current_wifi_ssid = ""
        self.current_wifi_password = ""
        # 正式版固定資料位置：不再把 Excel 與程式版本混放在 Downloads。
        # Windows 會建立在 C:\Users\<使用者>\CloudVision；其他系統則建立在家目錄。
        self.app_base_dir = os.path.dirname(os.path.abspath(__file__))
        self.cloudvision_home_dir = os.path.join(os.path.expanduser("~"), "CloudVision")
        self.public_excel_dir = os.path.join(self.cloudvision_home_dir, "Excel")
        self.system_data_dir = os.path.join(self.cloudvision_home_dir, ".System")
        self.backup_dir = os.path.join(self.cloudvision_home_dir, "Backup")
        for folder in (self.cloudvision_home_dir, self.public_excel_dir, self.system_data_dir, self.backup_dir):
            os.makedirs(folder, exist_ok=True)
        self._migrate_legacy_excel_folder()
        self.public_results_path = os.path.join(self.system_data_dir, "公開測驗資料.jsonl")
        self.public_results_csv_path = os.path.join(self.system_data_dir, "公開測驗資料.csv")
        self.public_events_path = os.path.join(self.system_data_dir, "公開瀏覽事件.jsonl")
        self.appointment_read_path = os.path.join(self.system_data_dir, "預約通知已讀.json")
        self.email_settings_path = os.path.join(self.system_data_dir, "Email設定.json")
        self.email_log_path = os.path.join(self.system_data_dir, "Email寄送紀錄.jsonl")
        self.appointment_status_path = os.path.join(self.system_data_dir, "預約管理狀態.json")
        self._migrate_legacy_public_data()
        self.public_results_lock = threading.RLock()

        # 手機／iPad 連線狀態。網頁開啟後會由 HTTP 伺服器回報到主程式。
        self.device_last_seen = {"control": 0.0, "participant": 0.0}
        self.device_ip = {"control": "", "participant": ""}
        self.device_name = {"control": "", "participant": ""}
        self.remote_estimated_va = None
        self.remote_estimated_va_version = 0
        self.remote_estimated_va_seen_version = 0
        self.remote_estimated_va_last_submit = "尚未送出"
        self.remote_estimated_va_last_control_read = "尚未讀取"
        self.remote_estimated_va_last_submit_ip = ""
        self.remote_estimated_va_last_control_ip = ""
        self.device_event_queue: queue.Queue[tuple[str, str]] = queue.Queue()
        self._device_notified_online = {"control": False, "participant": False}
        self.connection_control_var = tk.StringVar(value="測驗者：🔴 等待連線")
        self.connection_participant_var = tk.StringVar(value="受試者：🔴 等待連線")
        self.connection_count_var = tk.StringVar(value="目前已連線：0 / 2")

        self.result_question_var = tk.StringVar(value="目前題目：尚未選題")
        self.result_answer_var = tk.StringVar(value="受試者答案：等待作答")
        self.result_judgement_var = tk.StringVar(value="")
        self.result_time_var = tk.StringVar(value="")

        self._load_hotspot_settings()
        self._build_ui()
        self._bind_keys()
        self.randomize_letters(refresh=False)

        self.root.after(100, self._poll_remote_commands)
        if self.cloud_mode:
            # Railway 後端：隱藏 Tkinter 視窗，只啟動原本的網頁伺服器。
            self.root.withdraw()
            self.connection_ip = "127.0.0.1"
            self.start_remote_server()
        else:
            # 本機教學版維持原本的全畫面與連線選擇流程。
            self.root.after(50, lambda: self.set_fullscreen(True))
            self.root.after(150, self.refresh_chart)
            self.root.after(350, self.choose_connection_mode)


    def _load_hotspot_settings(self) -> None:
        try:
            with open(self.hotspot_config_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            self.hotspot_ssid_var.set(str(data.get("ssid", "")))
            self.hotspot_password_var.set(str(data.get("password", "")))
            saved_examiner_password = str(data.get("examiner_password", "123456")) or "123456"
            self.examiner_password_var.set(saved_examiner_password if len(saved_examiner_password) == 6 else "123456")
            security = str(data.get("security", "WPA2（相容模式）"))
            allowed = {"WPA2（相容模式）", "WPA", "WEP", "NOPASS"}
            self.hotspot_security_var.set(security if security in allowed else "WPA2（相容模式）")
        except (OSError, ValueError, TypeError):
            pass

    def _save_hotspot_settings(self) -> None:
        data = {
            "ssid": self.hotspot_ssid_var.get().strip(),
            "password": self.hotspot_password_var.get(),
            "security": self.hotspot_security_var.get(),
            "examiner_password": (self.examiner_password_var.get().strip() if len(self.examiner_password_var.get().strip()) == 6 else "123456"),
        }
        try:
            with open(self.hotspot_config_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except OSError as exc:
            self.status_var.set(f"熱點設定無法儲存：{exc}")

    @staticmethod
    def _wifi_qr_escape(value: str) -> str:
        return value.replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,").replace(":", "\\:").replace('"', '\\"')

    def _hotspot_wifi_payload(self) -> str:
        ssid = self._wifi_qr_escape(self.hotspot_ssid_var.get().strip())
        password = self._wifi_qr_escape(self.hotspot_password_var.get())
        security_label = self.hotspot_security_var.get().strip().upper()
        if security_label == "NOPASS":
            return f"WIFI:T:nopass;S:{ssid};H:false;;"
        # Wi-Fi QR 標準沒有 WPA2 專用代號；Windows WPA2 熱點需使用 T:WPA。
        security = "WEP" if security_label == "WEP" else "WPA"
        return f"WIFI:T:{security};S:{ssid};P:{password};H:false;;"


    def _detect_current_wifi_credentials(self) -> tuple[str, str]:
        """Windows：自動取得目前連線中的 Wi-Fi SSID 與已儲存密碼。

        若系統不是 Windows、未連線、或無法讀取密碼，會回傳空字串；
        程式仍可正常使用，使用者可改用 JohnVA 行動熱點。
        """
        if os.name != "nt":
            return "", ""
        try:
            active = subprocess.run(
                ["netsh", "wlan", "show", "interfaces"],
                capture_output=True, text=True, encoding="utf-8", errors="ignore",
                creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
                timeout=8,
            ).stdout
            ssid = ""
            for line in active.splitlines():
                # 排除 BSSID，只抓 SSID 欄。中英文 Windows 均以冒號分隔。
                left, sep, right = line.partition(":")
                if sep and left.strip().upper() == "SSID":
                    ssid = right.strip()
                    break
            if not ssid:
                return "", ""

            profile = subprocess.run(
                ["netsh", "wlan", "show", "profile", f"name={ssid}", "key=clear"],
                capture_output=True, text=True, encoding="utf-8", errors="ignore",
                creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
                timeout=8,
            ).stdout
            password = ""
            for line in profile.splitlines():
                left, sep, right = line.partition(":")
                key_name = left.strip().lower()
                if sep and key_name in {"key content", "金鑰內容", "密鑰內容"}:
                    password = right.strip()
                    break
            return ssid, password
        except Exception:
            return "", ""

    def _current_wifi_payload(self) -> str:
        ssid = self._wifi_qr_escape(self.current_wifi_ssid)
        password = self._wifi_qr_escape(self.current_wifi_password)
        if not password:
            return f"WIFI:T:nopass;S:{ssid};H:false;;"
        return f"WIFI:T:WPA;S:{ssid};P:{password};H:false;;"

    # ---------- UI ----------
    def _build_ui(self) -> None:
        # 上方控制區改為兩列，避免 15 吋筆電畫面寬度不足時按鈕超出視窗。
        self.control_frame = ttk.Frame(self.root, padding=(12, 7))
        self.control_frame.pack(fill=tk.X)

        settings_row = ttk.Frame(self.control_frame)
        settings_row.pack(fill=tk.X)

        ttk.Label(
            settings_row,
            text="三視標視力表",
            font=("Microsoft JhengHei", 17, "bold"),
        ).pack(side=tk.LEFT, padx=(0, 18))

        ttk.Label(settings_row, text="視標：").pack(side=tk.LEFT)
        mode_box = ttk.Combobox(
            settings_row, textvariable=self.optotype_mode_var, state="readonly", width=12,
            values=("letter", "landolt_c", "tumbling_e"),
        )
        mode_box.pack(side=tk.LEFT, padx=(3, 10))
        mode_box.bind("<<ComboboxSelected>>", lambda _e: self.change_optotype_mode())

        ttk.Label(settings_row, text="觀看距離：").pack(side=tk.LEFT)
        distance_entry = ttk.Entry(
            settings_row, textvariable=self.distance_var, width=7, justify="center"
        )
        distance_entry.pack(side=tk.LEFT, padx=(3, 2))
        ttk.Label(settings_row, text="cm").pack(side=tk.LEFT, padx=(0, 12))

        ttk.Label(settings_row, text="校正倍率：").pack(side=tk.LEFT)
        calibration_entry = ttk.Entry(
            settings_row, textvariable=self.calibration_var, width=8, justify="center"
        )
        calibration_entry.pack(side=tk.LEFT, padx=(3, 4))
        ttk.Button(
            settings_row, text="實尺校正 5 cm", command=self.calibrate_from_ruler
        ).pack(side=tk.LEFT, padx=(0, 10))

        ttk.Checkbutton(
            settings_row,
            text="顯示左右視力標示",
            variable=self.show_labels_var,
            command=self.refresh_chart,
        ).pack(side=tk.LEFT, padx=(0, 8))

        ttk.Button(settings_row, text="套用距離", command=self.apply_current_view).pack(side=tk.LEFT, padx=3)
        ttk.Button(settings_row, text="連線模式／QR Code", command=self.choose_connection_mode).pack(side=tk.LEFT, padx=3)
        ttk.Button(settings_row, text="測驗者後台", command=self.open_examiner_dashboard).pack(side=tk.LEFT, padx=3)
        ttk.Button(settings_row, text="離開", command=self.root.destroy).pack(side=tk.RIGHT, padx=3)

        function_row = ttk.Frame(self.control_frame)
        function_row.pack(fill=tk.X, pady=(7, 0))

        buttons = [
            ("Sloan Letter", lambda: self.set_optotype_mode("letter")),
            ("Landolt C", lambda: self.set_optotype_mode("landolt_c")),
            ("Tumbling E", lambda: self.set_optotype_mode("tumbling_e")),
            ("重新排列視標", self.randomize_letters),
            ("散光鐘", self.show_astigmatic_dial),
            ("阿姆斯勒方格", self.show_amsler_grid),
            ("Worth 四點 F7", self.show_worth_four_dot),
            ("純四點 F9", self.toggle_worth_focus_mode),
            ("Bagolini", self.show_bagolini_test),
            ("Thorington", self.show_thorington_test),
            ("單點全黑 F10", self.toggle_test_focus_mode),
            ("回視力表", self.show_acuity_chart),
            ("全畫面 F11", self.toggle_fullscreen),
            ("隱藏設定 F2", self.toggle_controls),
        ]
        for text, command in buttons:
            ttk.Button(function_row, text=text, command=command).pack(side=tk.LEFT, padx=3)

        distance_entry.bind("<Return>", lambda _e: self.apply_current_view())
        calibration_entry.bind("<Return>", lambda _e: self.apply_current_view())

        # 校正與說明列。
        self.info_frame = tk.Frame(self.root, background="white")
        self.info_frame.pack(fill=tk.X)

        self.instruction_label = tk.Label(
            self.info_frame,
            text=(
                "Bailey–Lovie 風格：每列固定 5 個視標，可切換 Sloan Letter、Landolt C、Tumbling E。"
                "三種視標在相同視力等級使用完全相同的外框尺寸；正式使用前請先完成螢幕校正。"
            ),
            background="white",
            foreground="#333333",
            font=("Microsoft JhengHei", 11, "bold"),
            anchor="w",
        )
        self.instruction_label.pack(fill=tk.X, padx=18, pady=(5, 2))

        self.calibration_canvas = tk.Canvas(
            self.info_frame,
            height=48,
            background="white",
            highlightthickness=0,
        )
        self.calibration_canvas.pack(fill=tk.X, padx=18)

        # 可捲動視力表。
        body = tk.Frame(self.root, background="white")
        body.pack(fill=tk.BOTH, expand=True)

        self.canvas = tk.Canvas(
            body,
            background="white",
            highlightthickness=0,
            borderwidth=0,
        )
        self.v_scroll = ttk.Scrollbar(body, orient=tk.VERTICAL, command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.v_scroll.set)

        self.v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.chart_frame = tk.Frame(self.canvas, background="white")
        self.chart_window_id = self.canvas.create_window(
            (0, 0), window=self.chart_frame, anchor="n"
        )

        self.chart_frame.bind(
            "<Configure>",
            lambda _e: self.canvas.configure(scrollregion=self.canvas.bbox("all")),
        )
        self.canvas.bind("<Configure>", self._on_canvas_resize)

        # 測驗者專用即時判分區（受試者不會看到）。
        self.result_frame = tk.Frame(self.root, background="#eef5ff", bd=1, relief="solid")
        self.result_frame.pack(fill=tk.X, padx=12, pady=(4, 2))
        tk.Label(
            self.result_frame, text="受試者即時作答結果", background="#eef5ff",
            foreground="#174ea6", font=("Microsoft JhengHei", 12, "bold")
        ).pack(side=tk.LEFT, padx=(12, 18), pady=7)
        tk.Label(
            self.result_frame, textvariable=self.result_question_var, background="#eef5ff",
            font=("Microsoft JhengHei", 11, "bold")
        ).pack(side=tk.LEFT, padx=8)
        tk.Label(
            self.result_frame, textvariable=self.result_answer_var, background="#eef5ff",
            font=("Microsoft JhengHei", 11)
        ).pack(side=tk.LEFT, padx=8)
        self.result_judgement_label = tk.Label(
            self.result_frame, textvariable=self.result_judgement_var, background="#eef5ff",
            font=("Microsoft JhengHei", 12, "bold")
        )
        self.result_judgement_label.pack(side=tk.LEFT, padx=10)
        tk.Label(
            self.result_frame, textvariable=self.result_time_var, background="#eef5ff",
            font=("Microsoft JhengHei", 11)
        ).pack(side=tk.LEFT, padx=8)
        ttk.Button(
            self.result_frame, text="◀ 上一個視力值", command=self.remote_previous_va_row
        ).pack(side=tk.RIGHT, padx=(4, 8), pady=4)
        ttk.Button(
            self.result_frame, text="下一個視力值 ▶", command=self.remote_next_va_row
        ).pack(side=tk.RIGHT, padx=4, pady=4)
        ttk.Button(
            self.result_frame, text="確認並清除", command=self.clear_current_question
        ).pack(side=tk.RIGHT, padx=4, pady=4)
        ttk.Button(
            self.result_frame, text="發送目前題目", command=self.send_current_question
        ).pack(side=tk.RIGHT, padx=4, pady=4)

        # 測驗者先選擇一個或多個貼圖，再按「發送貼圖」一次送出。
        self.sticker_control_frame = tk.Frame(self.root, background="#fff8df", bd=1, relief="solid")
        self.sticker_control_frame.pack(fill=tk.X, padx=12, pady=(0, 3))
        tk.Label(
            self.sticker_control_frame, text="選擇貼圖：", background="#fff8df",
            foreground="#9a4b00", font=("Microsoft JhengHei", 11, "bold")
        ).pack(side=tk.LEFT, padx=(12, 8), pady=5)
        for sticker in REWARD_STICKERS:
            tk.Button(
                self.sticker_control_frame, text=sticker, command=lambda x=sticker: self.select_reward_sticker(x),
                font=("Segoe UI Emoji", 18), width=2, relief="raised", cursor="hand2"
            ).pack(side=tk.LEFT, padx=2, pady=3)
        tk.Label(
            self.sticker_control_frame, textvariable=self.pending_sticker_var,
            background="#fff8df", foreground="#7a3d00",
            font=("Segoe UI Emoji", 11, "bold")
        ).pack(side=tk.LEFT, padx=(10, 6))
        ttk.Button(
            self.sticker_control_frame, text="清除選擇", command=self.clear_pending_stickers
        ).pack(side=tk.RIGHT, padx=3, pady=4)
        ttk.Button(
            self.sticker_control_frame, text="發送貼圖", command=self.send_selected_stickers
        ).pack(side=tk.RIGHT, padx=3, pady=4)
        ttk.Button(
            self.sticker_control_frame, text="清空收藏", command=self.clear_reward_stickers
        ).pack(side=tk.RIGHT, padx=(6, 3), pady=4)

        self.status_label = tk.Label(
            self.root,
            textvariable=self.status_var,
            background="#f3f5f7",
            foreground="#444444",
            anchor="w",
            font=("Microsoft JhengHei", 9),
            padx=12,
            pady=4,
        )
        self.status_label.pack(fill=tk.X)

    def _bind_keys(self) -> None:
        self.root.bind("<F11>", lambda _e: self.toggle_fullscreen())
        self.root.bind("<Escape>", lambda _e: self.set_fullscreen(False))
        self.root.bind("<F2>", lambda _e: self.toggle_controls())
        self.root.bind("<F5>", lambda _e: self.randomize_letters())
        self.root.bind("<Home>", lambda _e: self.canvas.yview_moveto(0.0))
        self.root.bind("<F3>", lambda _e: self.show_astigmatic_dial())
        self.root.bind("<F4>", lambda _e: self.show_acuity_chart())
        self.root.bind("<F6>", lambda _e: self.show_amsler_grid())
        self.root.bind("<F7>", lambda _e: self.show_worth_four_dot())
        self.root.bind("<F8>", lambda _e: self.show_bagolini_test())
        self.root.bind("<F9>", lambda _e: self.toggle_worth_focus_mode())
        self.root.bind("<F10>", lambda _e: self.toggle_test_focus_mode())
        self.root.bind("<F12>", lambda _e: self.show_thorington_test())
        self.root.bind("<Key-1>", lambda _e: self.set_thorington_distance(40.0))
        self.root.bind("<Key-2>", lambda _e: self.set_thorington_distance(100.0))
        self.root.bind("<Key-3>", lambda _e: self.set_thorington_distance(300.0))
        self.root.bind("<plus>", lambda _e: self.adjust_bagolini_dot(1))
        self.root.bind("<KP_Add>", lambda _e: self.adjust_bagolini_dot(1))
        self.root.bind("<minus>", lambda _e: self.adjust_bagolini_dot(-1))
        self.root.bind("<KP_Subtract>", lambda _e: self.adjust_bagolini_dot(-1))
        self.root.bind_all("<MouseWheel>", self._on_mousewheel)
        self.root.bind_all("<Button-4>", lambda _e: self.canvas.yview_scroll(-3, "units"))
        self.root.bind_all("<Button-5>", lambda _e: self.canvas.yview_scroll(3, "units"))

    # ---------- 區域網路遙控 ----------
    @staticmethod
    def _candidate_local_ips() -> list[str]:
        """取得本機可供手機連線的 IPv4 位址。

        Windows 行動熱點通常使用 192.168.137.1；一般 Wi-Fi／區網則常見
        192.168.x.x、10.x.x.x 或 172.16～31.x.x。
        """
        ips: list[str] = []
        try:
            for info in socket.getaddrinfo(socket.gethostname(), None, socket.AF_INET):
                ip = info[4][0]
                if ip and not ip.startswith("127.") and ip not in ips:
                    ips.append(ip)
        except OSError:
            pass

        # UDP 方法通常可以找到目前預設網路介面的位址。
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            sock.connect(("8.8.8.8", 80))
            ip = sock.getsockname()[0]
            sock.close()
            if ip and not ip.startswith("127.") and ip not in ips:
                ips.insert(0, ip)
        except OSError:
            pass
        return ips

    def _select_connection_ip(self, mode: str) -> str:
        ips = self._candidate_local_ips()
        if not ips:
            return "127.0.0.1"
        if mode == "hotspot":
            # Windows 行動熱點預設閘道固定為 192.168.137.1。
            # 即使 Python 暫時沒有列出該介面，也直接使用此網址產生 QR Code。
            for ip in ips:
                if ip == "192.168.137.1":
                    return ip
            for ip in ips:
                if ip.startswith("192.168.137."):
                    return ip
            return "192.168.137.1"
        else:
            # 共用 Wi-Fi 模式避開典型的 Windows 行動熱點介面。
            for ip in ips:
                if not ip.startswith("192.168.137."):
                    return ip
        return ips[0]

    def choose_connection_mode(self) -> None:
        """選擇連線方式，並在同一視窗下方即時顯示可掃描的 QR Code。"""
        # 先啟動伺服器，讓測驗者與受試者 QR Code 一開啟就能顯示。
        self.connection_ip = self._select_connection_ip(self.connection_mode)
        self.start_remote_server()
        self._update_remote_urls()

        try:
            import qrcode
        except ImportError:
            try:
                import subprocess, sys
                subprocess.check_call(
                    [sys.executable, "-m", "pip", "install", "qrcode[pil]"],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )
                import qrcode
            except Exception:
                qrcode = None

        win = tk.Toplevel(self.root)
        self.qr_window = win
        win.title("選擇手機連線方式／QR Code")
        win.configure(background="white")
        win.attributes("-topmost", True)
        win.transient(self.root)
        win.grab_set()
        win.geometry("1280x760")
        win.minsize(1100, 700)

        mode_var = tk.StringVar(value=self.connection_mode)
        tk.Label(win, text="請選擇手機連線方式", bg="white",
                 font=("Microsoft JhengHei", 18, "bold")).pack(padx=24, pady=(14, 8))

        modes = tk.Frame(win, bg="white")
        modes.pack(fill=tk.X, padx=22)
        wifi_box = tk.Frame(modes, bg="#eef5ff", bd=2, relief="groove")
        wifi_box.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        tk.Radiobutton(
            wifi_box, variable=mode_var, value="wifi", bg="#eef5ff",
            activebackground="#eef5ff", anchor="w", justify="left",
            text="① 共用 Wi-Fi 模式\n電腦與兩支手機連接同一個 Wi-Fi",
            font=("Microsoft JhengHei", 12, "bold")
        ).pack(fill=tk.X, padx=12, pady=10)

        hotspot_box = tk.Frame(modes, bg="#fff5eb", bd=2, relief="groove")
        hotspot_box.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 0))
        tk.Radiobutton(
            hotspot_box, variable=mode_var, value="hotspot", bg="#fff5eb",
            activebackground="#fff5eb", anchor="w", justify="left",
            text="② 電腦行動熱點模式\n手機連接電腦提供的熱點，不需使用家裡 Wi-Fi",
            font=("Microsoft JhengHei", 12, "bold")
        ).pack(fill=tk.X, padx=12, pady=10)

        settings = tk.LabelFrame(
            win, text="電腦行動熱點設定", bg="white",
            font=("Microsoft JhengHei", 11, "bold"), padx=10, pady=5
        )
        settings.pack(fill=tk.X, padx=22, pady=(8, 4))
        tk.Label(settings, text="熱點名稱（SSID）：", bg="white").grid(row=0, column=0, sticky="e", padx=4, pady=3)
        ssid_entry = ttk.Entry(settings, textvariable=self.hotspot_ssid_var, width=26)
        ssid_entry.grid(row=0, column=1, sticky="w", padx=4, pady=3)
        tk.Label(settings, text="熱點密碼：", bg="white").grid(row=0, column=2, sticky="e", padx=(20, 4), pady=3)
        password_entry = ttk.Entry(settings, textvariable=self.hotspot_password_var, width=22)
        password_entry.grid(row=0, column=3, sticky="w", padx=4, pady=3)
        tk.Label(settings, text="加密方式：", bg="white").grid(row=0, column=4, sticky="e", padx=(20, 4), pady=3)
        security_combo = ttk.Combobox(
            settings, textvariable=self.hotspot_security_var,
            values=("WPA2（相容模式）", "WPA", "WEP", "NOPASS"), state="readonly", width=18
        )
        security_combo.grid(row=0, column=5, sticky="w", padx=4, pady=3)
        tk.Label(settings, text="測驗者密碼：", bg="white").grid(row=1, column=0, sticky="e", padx=4, pady=3)
        examiner_password_entry = ttk.Entry(settings, textvariable=self.examiner_password_var, width=22, show="•")
        examiner_password_entry.grid(row=1, column=1, sticky="w", padx=4, pady=3)
        ttk.Label(settings, text="（固定 6 個字元；測驗者／今日資料專用）").grid(row=1, column=2, columnspan=3, sticky="w", padx=4, pady=3)

        info_var = tk.StringVar(value="")
        tk.Label(win, textvariable=info_var, bg="white", fg="#555555",
                 font=("Microsoft JhengHei", 10, "bold")).pack(pady=(2, 2))

        connection_bar = tk.Frame(win, bg="#f7f7f7", bd=1, relief="solid")
        connection_bar.pack(fill=tk.X, padx=22, pady=(3, 4))
        tk.Label(connection_bar, textvariable=self.connection_control_var, bg="#f7f7f7",
                 font=("Microsoft JhengHei", 11, "bold")).pack(side=tk.LEFT, padx=14, pady=6)
        tk.Label(connection_bar, textvariable=self.connection_participant_var, bg="#f7f7f7",
                 font=("Microsoft JhengHei", 11, "bold")).pack(side=tk.LEFT, padx=14, pady=6)
        tk.Label(connection_bar, textvariable=self.connection_count_var, bg="#f7f7f7", fg="#174ea6",
                 font=("Microsoft JhengHei", 11, "bold")).pack(side=tk.RIGHT, padx=14, pady=6)

        qr_row = tk.Frame(win, bg="white")
        qr_row.pack(fill=tk.BOTH, expand=True, padx=18, pady=4)
        qr_refs: list[ImageTk.PhotoImage] = []

        def make_card(parent, title, subtitle, bg, fg):
            card = tk.Frame(parent, bg=bg, bd=2, relief="groove")
            card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=6, pady=3)
            tk.Label(card, text=title, bg=bg, fg=fg,
                     font=("Microsoft JhengHei", 13, "bold")).pack(pady=(8, 1))
            tk.Label(card, text=subtitle, bg=bg,
                     font=("Microsoft JhengHei", 10)).pack()
            image_label = tk.Label(card, bg=bg)
            image_label.pack(pady=4)
            text_label = tk.Label(card, bg=bg, fg=fg,
                                  font=("Arial", 9, "bold"), wraplength=280)
            text_label.pack(padx=6, pady=(0, 7))
            return card, image_label, text_label

        wifi_card, wifi_image, wifi_text = make_card(
            qr_row, "① 加入目前共用 Wi-Fi", "自動抓取電腦目前的 Wi-Fi", "#f2efff", "#5435a5"
        )
        hotspot_card, hotspot_image, hotspot_text = make_card(
            qr_row, "② 加入 JohnVA 行動熱點", "未使用共用 Wi-Fi 時掃這張", "#eaf8ea", "#176b2c"
        )
        control_card, control_image, control_text = make_card(
            qr_row, "③ 測驗者手機／iPad", "遙控出題", "#eef5ff", "#174ea6"
        )
        answer_card, answer_image, answer_text = make_card(
            qr_row, "④ 受試者手機", "只用來按答案", "#fff5eb", "#9a4b00"
        )

        def set_qr(label: tk.Label, payload: str) -> None:
            if qrcode is None:
                label.configure(text="QR Code 元件無法載入", image="")
                return
            qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=8, border=4)
            qr.add_data(payload)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white").convert("RGB").resize((165, 165))
            photo = ImageTk.PhotoImage(img)
            qr_refs.append(photo)
            label.configure(image=photo, text="")
            label.image = photo

        def refresh_qr(*_args) -> None:
            qr_refs.clear()
            new_mode = mode_var.get()
            if new_mode != self.connection_mode:
                self.connection_mode = new_mode
                self._reset_device_connections(renew_session=True)
            else:
                self.connection_mode = new_mode
            self.connection_ip = self._select_connection_ip(self.connection_mode)
            self._update_remote_urls()

            # 自動抓取電腦目前連線中的共用 Wi-Fi，直接產生加入網路 QR Code。
            self.current_wifi_ssid, self.current_wifi_password = self._detect_current_wifi_credentials()
            if self.current_wifi_ssid:
                set_qr(wifi_image, self._current_wifi_payload())
                password_text = self.current_wifi_password if self.current_wifi_password else "無密碼／系統未提供密碼"
                wifi_text.configure(text=f"Wi-Fi：{self.current_wifi_ssid}\n密碼：{password_text}")
            else:
                wifi_image.configure(image="", text="目前未偵測到共用 Wi-Fi")
                wifi_text.configure(text="請先讓電腦連上 Wi-Fi，再按下方『重新偵測』")

            # 測驗者與受試者網址會依目前選取的網路介面立即更新。
            set_qr(control_image, self.control_url)
            control_text.configure(text=self.control_url)
            set_qr(answer_image, self.answer_url)
            answer_text.configure(text=self.answer_url)

            # 三張 QR Code 固定顯示：即使目前選擇共用 Wi-Fi，仍保留 JohnVA 熱點條碼，
            # 方便現場直接讓手機掃描加入電腦行動熱點，不必切換模式後才看得到。
            if not hotspot_card.winfo_ismapped():
                hotspot_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=6, pady=3, before=control_card)

            ssid = self.hotspot_ssid_var.get().strip()
            if ssid:
                set_qr(hotspot_image, self._hotspot_wifi_payload())
                hotspot_text.configure(text=f"熱點：{ssid}\n密碼：{self.hotspot_password_var.get()}")
            else:
                hotspot_image.configure(image="", text="請先輸入熱點名稱")
                hotspot_text.configure(text="")

            if self.connection_mode == "hotspot":
                info_var.set(f"行動熱點模式｜目前內網 IP：{self.connection_ip}｜先掃綠色，再掃藍色或橘色")
            else:
                info_var.set(f"共用 Wi-Fi 模式｜目前內網 IP：{self.connection_ip}｜綠色 JohnVA 條碼仍固定顯示")

        def show_wifi_qr_details() -> None:
            payload = self._hotspot_wifi_payload()
            ssid = self.hotspot_ssid_var.get().strip()
            password = self.hotspot_password_var.get()
            messagebox.showinfo(
                "Wi-Fi QR Code 測驗",
                f"熱點名稱：{ssid}\n密碼：{password}\n\nQR Code 內容：\n{payload}\n\n"
                "請先確認 Windows 行動熱點已開啟。若舊款 iPad 仍無法加入，請在 Windows 將網路頻帶改為 2.4 GHz。",
                parent=win,
            )

        def apply_and_close() -> None:
            self.connection_mode = mode_var.get()
            if self.connection_mode == "hotspot" and not self.hotspot_ssid_var.get().strip():
                messagebox.showwarning("尚未輸入熱點名稱", "請輸入 Windows 行動熱點名稱（SSID）。", parent=win)
                return
            self._save_hotspot_settings()
            refresh_qr()
            try:
                win.grab_release()
            except tk.TclError:
                pass
            if self.qr_window is win:
                self.qr_window = None
            win.destroy()

        # 選模式、修改名稱或密碼時，最下方 QR Code 立即重畫，不必再按產生。
        mode_var.trace_add("write", refresh_qr)
        self.hotspot_ssid_var.trace_add("write", refresh_qr)
        self.hotspot_password_var.trace_add("write", refresh_qr)
        self.hotspot_security_var.trace_add("write", refresh_qr)

        # 公開瀏覽器入口：給開發者在這台電腦直接查看 Cloud Vision 公開版。
        def open_cloud_browser() -> None:
            try:
                self.start_remote_server()
                self._update_remote_urls()
                # 使用 localhost，避免電腦切換 Wi-Fi 後舊 IP 造成無法開啟。
                browser_url = f"http://127.0.0.1:{self.remote_port}/cloud"
                opened = webbrowser.open_new_tab(browser_url)
                if opened:
                    self.status_var.set(f"已開啟 Cloud Vision 瀏覽器：{browser_url}")
                else:
                    messagebox.showinfo(
                        "Cloud Vision 網址",
                        f"瀏覽器未自動開啟，請複製以下網址：\n\n{browser_url}",
                        parent=win,
                    )
            except Exception as exc:
                messagebox.showerror(
                    "無法開啟 Cloud Vision",
                    f"瀏覽器頁面啟動失敗：\n{exc}",
                    parent=win,
                )

        browser_btns = ttk.Frame(win)
        browser_btns.pack(pady=(6, 4))
        ttk.Button(
            browser_btns,
            text="🌐 開啟 Cloud Vision 瀏覽器",
            command=open_cloud_browser,
        ).pack(ipadx=24, ipady=7)

        btns = ttk.Frame(win)
        btns.pack(pady=(4, 12))
        ttk.Button(btns, text="重新偵測目前 Wi-Fi", command=refresh_qr).pack(side=tk.LEFT, padx=6)
        def clear_old_connections() -> None:
            self._reset_device_connections(renew_session=True)
            refresh_qr()
            self.status_var.set("已清除舊連線；舊手機頁面已失效，請重新掃描最新 QR Code。")

        ttk.Button(btns, text="清除舊連線狀態", command=clear_old_connections).pack(side=tk.LEFT, padx=6)
        ttk.Button(btns, text="測驗 JohnVA QR 內容", command=show_wifi_qr_details).pack(side=tk.LEFT, padx=6)
        ttk.Button(btns, text="儲存設定並關閉", command=apply_and_close).pack(side=tk.LEFT, padx=6)
        ttk.Button(btns, text="另外開啟大型 QR Code", command=self.show_remote_qr).pack(side=tk.LEFT, padx=6)
        ttk.Button(btns, text="取消", command=win.destroy).pack(side=tk.LEFT, padx=6)

        def refresh_connection_labels() -> None:
            if not win.winfo_exists():
                return
            self._update_device_connection_vars()
            win.after(500, refresh_connection_labels)

        refresh_qr()
        refresh_connection_labels()
        def close_connection_window() -> None:
            try:
                if win.grab_current() is win:
                    win.grab_release()
            except tk.TclError:
                pass
            if self.qr_window is win:
                self.qr_window = None
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", close_connection_window)

    def _update_remote_urls(self) -> None:
        public_base = os.environ.get("PUBLIC_BASE_URL", "").strip().rstrip("/")
        if public_base:
            self.remote_url = public_base
        else:
            if not self.connection_ip:
                self.connection_ip = self._select_connection_ip(self.connection_mode)
            self.remote_url = f"http://{self.connection_ip}:{self.remote_port}"
        token = self.connection_session_token
        self.control_url = self.remote_url + f"/control?session={token}"
        self.answer_url = self.remote_url + f"/participant?session={token}"
        mode_text = "電腦行動熱點" if self.connection_mode == "hotspot" else "共用 Wi-Fi"
        self.status_var.set(
            f"{mode_text}已啟動｜IP：{self.connection_ip}｜測驗者：{self.control_url}｜受試者：{self.answer_url}"
        )

    def _reset_device_connections(self, renew_session: bool = False) -> None:
        """清除連線狀態；需要時同步作廢所有舊手機頁面。"""
        self.device_last_seen = {"control": 0.0, "participant": 0.0}
        self.device_ip = {"control": "", "participant": ""}
        self.device_name = {"control": "", "participant": ""}
        self._device_notified_online = {"control": False, "participant": False}
        if renew_session:
            self.connection_session_token = secrets.token_urlsafe(18)
            self._update_remote_urls()
        self._update_device_connection_vars()

    def _valid_connection_session(self, query: dict[str, list[str]]) -> bool:
        supplied = query.get("session", [""])[0]
        return bool(supplied) and secrets.compare_digest(supplied, self.connection_session_token)

    def _ip_matches_current_mode(self, ip: str) -> bool:
        """只接受目前連線模式所在網段的裝置，避免舊 Wi-Fi 頁面被誤算。"""
        if not ip:
            return False
        if self.connection_mode == "hotspot":
            return ip.startswith("192.168.137.")
        # 共用 Wi-Fi：接受與電腦目前網址相同的前三段網段，並排除行動熱點。
        if ip.startswith("192.168.137."):
            return False
        parts = self.connection_ip.split(".")
        if len(parts) == 4:
            return ip.startswith(".".join(parts[:3]) + ".")
        return True

    @staticmethod
    def _device_name_from_user_agent(user_agent: str) -> str:
        ua = (user_agent or "").lower()
        if "ipad" in ua:
            return "iPad"
        if "iphone" in ua:
            return "iPhone"
        if "android" in ua:
            return "Android 手機／平板"
        if "windows" in ua:
            return "Windows 裝置"
        return "手機／平板"

    def _register_device_connection(self, role: str, ip: str, notify: bool = True, user_agent: str = "") -> None:
        if role not in ("control", "participant"):
            return
        # 只要裝置能實際向本程式回報，就視為連線成功。
        was_offline = (time.monotonic() - self.device_last_seen.get(role, 0.0)) > 8.0
        self.device_last_seen[role] = time.monotonic()
        self.device_ip[role] = ip
        if user_agent:
            self.device_name[role] = self._device_name_from_user_agent(user_agent)
        if notify and was_offline:
            self.device_event_queue.put((role, ip))

    def _update_device_connection_vars(self) -> None:
        now = time.monotonic()
        control_online = now - self.device_last_seen.get("control", 0.0) <= 8.0
        participant_online = now - self.device_last_seen.get("participant", 0.0) <= 8.0
        control_detail = ""
        if control_online and self.device_ip["control"]:
            control_detail = f"（{self.device_name.get('control') or '手機／平板'}｜{self.device_ip['control']}）"
        participant_detail = ""
        if participant_online and self.device_ip["participant"]:
            participant_detail = f"（{self.device_name.get('participant') or '手機／平板'}｜{self.device_ip['participant']}）"
        self.connection_control_var.set(
            f"測驗者：{'🟢 已連線' if control_online else '🔴 等待連線'}" + control_detail
        )
        self.connection_participant_var.set(
            f"受試者：{'🟢 已連線' if participant_online else '🔴 等待連線'}" + participant_detail
        )
        self.connection_count_var.set(f"目前已連線：{int(control_online) + int(participant_online)} / 2")
        if not control_online:
            self._device_notified_online["control"] = False
        if not participant_online:
            self._device_notified_online["participant"] = False

    def _public_cloud_url(self) -> str:
        base = (self.remote_url or "").rstrip("/")
        return (base + "/cloud") if base else "/cloud"

    def _public_qr_data_uri(self) -> str:
        """直接把公開首頁網址做成內嵌 QR Code，不依賴舊版 QR 圖片。"""
        try:
            import qrcode
            qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=8, border=3)
            qr.add_data(self._public_cloud_url())
            qr.make(fit=True)
            image = qr.make_image(fill_color="black", back_color="white").convert("RGB")
            buf = io.BytesIO()
            image.save(buf, format="PNG")
            return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("ascii")
        except Exception:
            return ""

    def _migrate_legacy_excel_folder(self) -> None:
        """將舊版放在程式旁或 Downloads 的每日 Excel 複製到固定 Excel 資料夾。"""
        candidates = [
            os.path.join(self.app_base_dir, "Cloud Vision Excel"),
            os.path.join(self.app_base_dir, "Cloud_Vision_每日Excel"),
            os.path.join(os.path.expanduser("~"), "Downloads", "Cloud Vision Excel"),
            os.path.join(os.path.expanduser("~"), "Downloads", "Cloud_Vision_每日Excel"),
        ]
        for old_dir in candidates:
            if not os.path.isdir(old_dir) or os.path.abspath(old_dir) == os.path.abspath(self.public_excel_dir):
                continue
            try:
                for name in os.listdir(old_dir):
                    if not (name.startswith("CloudVision_") and name.lower().endswith(".xlsx")):
                        continue
                    source = os.path.join(old_dir, name)
                    target = os.path.join(self.public_excel_dir, name)
                    if not os.path.exists(target):
                        shutil.copy2(source, target)
            except Exception as exc:
                print(f"[Cloud Vision] 舊 Excel 搬移略過：{exc}", flush=True)

    def _open_on_server(self, path: str) -> None:
        """在執行 Cloud Vision 的電腦開啟檔案或資料夾。"""
        if not os.path.exists(path):
            raise FileNotFoundError(path)
        if os.name == "nt":
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])

    def _migrate_legacy_public_data(self) -> None:
        """將舊版散落在程式旁的資料移入系統資料夾；舊檔保留，不做刪除。"""
        legacy_map = {
            "Cloud_Vision_公開測驗資料.jsonl": self.public_results_path,
            "Cloud_Vision_公開測驗資料.csv": self.public_results_csv_path,
            "Cloud_Vision_公開瀏覽事件.jsonl": self.public_events_path,
            "Cloud_Vision_預約通知已讀.json": self.appointment_read_path,
        }
        for old_name, new_path in legacy_map.items():
            old_path = os.path.join(self.app_base_dir, old_name)
            if os.path.exists(old_path) and not os.path.exists(new_path):
                try:
                    import shutil
                    shutil.copy2(old_path, new_path)
                except OSError:
                    pass

    def _load_email_settings(self) -> dict:
        defaults = {
            "enabled": False,
            "recipient": "tojohn123456@yahoo.com.tw",
            "sender": "tojohn123456@yahoo.com.tw",
            "smtp_server": "smtp.mail.yahoo.com",
            "smtp_port": 465,
            "app_password": "",
        }
        try:
            with open(self.email_settings_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                defaults.update(data)
        except (OSError, ValueError, TypeError):
            pass
        defaults["smtp_port"] = int(defaults.get("smtp_port", 465) or 465)
        return defaults

    def _save_email_settings(self, settings: dict) -> None:
        os.makedirs(self.system_data_dir, exist_ok=True)
        temp = self.email_settings_path + ".tmp"
        with open(temp, "w", encoding="utf-8") as f:
            json.dump(settings, f, ensure_ascii=False, indent=2)
        os.replace(temp, self.email_settings_path)
        try:
            os.chmod(self.email_settings_path, 0o600)
        except OSError:
            pass

    def _write_email_log(self, ok: bool, message: str, record: dict | None = None) -> None:
        row = {
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "ok": bool(ok),
            "message": str(message)[:500],
            "record_id": str((record or {}).get("id", "")),
        }
        with open(self.email_log_path, "a", encoding="utf-8") as f:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")

    def _send_email_message(self, subject: str, body: str) -> None:
        settings = self._load_email_settings()
        sender = str(settings.get("sender", "")).strip()
        recipient = str(settings.get("recipient", "")).strip()
        app_password = str(settings.get("app_password", "")).replace(" ", "").strip()
        if not sender or not recipient or not app_password:
            raise RuntimeError("Email 尚未完成設定，請先在測驗者後台儲存 Yahoo 應用程式密碼。")
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = sender
        msg["To"] = recipient
        msg.set_content(body)
        context = ssl.create_default_context()
        server = str(settings.get("smtp_server", "smtp.mail.yahoo.com")).strip() or "smtp.mail.yahoo.com"
        port = int(settings.get("smtp_port", 465) or 465)
        with smtplib.SMTP_SSL(server, port, context=context, timeout=20) as smtp:
            smtp.login(sender, app_password)
            smtp.send_message(msg)

    def _send_appointment_email(self, record: dict) -> None:
        settings = self._load_email_settings()
        if not bool(settings.get("enabled")):
            return
        body = "\n".join([
            "Cloud Vision 收到一筆新的驗光預約", "",
            f"送出時間：{record.get('date','')} {record.get('time','')}",
            f"姓名：{record.get('name','') or '未填寫'}",
            f"電話：{record.get('phone','') or '未填寫'}",
            f"預約日期：{record.get('appointment_date','') or '未選擇'}",
            f"預約時段：{record.get('appointment_time','') or '未選擇'}",
            f"備註：{record.get('note','') or '無'}", "",
            f"右眼散光鐘：{record.get('astigmatism_right','')}",
            f"左眼散光鐘：{record.get('astigmatism_left','')}",
            f"右眼黃斑部：{record.get('amsler_right','')}",
            f"左眼黃斑部：{record.get('amsler_left','')}", "",
            "請登入 Cloud Vision 測驗者後台查看完整資料。",
        ])
        try:
            self._send_email_message("Cloud Vision 新預約通知", body)
            self._write_email_log(True, "預約通知寄送成功", record)
        except Exception as exc:
            self._write_email_log(False, str(exc), record)
            print(f"[Cloud Vision] Email 通知寄送失敗：{exc}", flush=True)

    def _send_test_email(self) -> None:
        self._send_email_message(
            "Cloud Vision 測試通知",
            "這是一封 Cloud Vision 測試信。\n\n若您收到此信，代表 Yahoo Email 自動通知已設定成功。",
        )
        self._write_email_log(True, "測試信寄送成功")

    def _save_public_result(self, payload: dict, client_ip: str, user_agent: str) -> dict:
        now = datetime.now()
        # 本研究／教學版關閉一般使用者個資與預約功能。
        # 只保存視覺功能測驗結果、匿名裝置識別碼與使用回饋。
        allowed = {
            "consent", "visual_acuity_right", "visual_acuity_left", "visual_acuity_both",
            "astigmatism_right", "astigmatism_left", "amsler_right", "amsler_left",
            "near_reading", "note", "visitor_id", "session_id", "user_type",
            "professional_name", "professional_role", "professional_purpose",
            "satisfaction", "needs_improvement", "improvement_suggestion"
        }
        clean = {k: str(payload.get(k, "")).strip()[:500] for k in allowed}
        clean.update({
            "name": "", "phone": "", "appointment": "",
            "appointment_date": "", "appointment_time": "",
        })
        clean["user_type"] = "專業人員" if clean.get("user_type") == "專業人員" else "一般使用者"
        clean["anonymous_id"] = clean.get("visitor_id", "") if clean["user_type"] == "一般使用者" else ""
        if clean["user_type"] != "專業人員":
            clean["professional_name"] = ""
            clean["professional_role"] = ""
            clean["professional_purpose"] = ""
        # 以本次登錄的 session_id 配對進站與完成時間；同一人重複登錄也各算一次。
        session_id = clean.get("session_id", "")
        timing = self._session_timing(session_id)
        started_at = timing.get("started_at", "")
        duration_seconds = timing.get("duration_seconds", "")
        clean.update({
            "session_id": session_id,
            "test_id": timing.get("test_id", "") or ("CV-" + now.strftime("%Y%m%d-%H%M%S-") + secrets.token_hex(2).upper()),
            "login_at": started_at,
            "completed_at": now.isoformat(timespec="seconds"),
            "duration_seconds": duration_seconds,
            "duration_text": self._format_duration(duration_seconds),
            "completion_status": "已完成",
            "id": now.strftime("%Y%m%d%H%M%S") + "-" + secrets.token_hex(3),
            "date": now.strftime("%Y-%m-%d"),
            "time": now.strftime("%H:%M:%S"),
            "created_at": now.isoformat(timespec="seconds"),
            "client_ip": client_ip,
            "device": user_agent[:300],
        })
        csv_fields = [
            "id", "date", "time", "created_at", "name", "phone", "consent",
            "visual_acuity_right", "visual_acuity_left", "visual_acuity_both",
            "astigmatism_right", "astigmatism_left",
            "amsler_right", "amsler_left", "near_reading",
            "appointment", "appointment_date", "appointment_time", "note",
            "user_type", "anonymous_id", "professional_name", "professional_role", "professional_purpose",
            "satisfaction", "needs_improvement", "improvement_suggestion",
            "visitor_id", "session_id", "test_id", "login_at", "completed_at", "duration_seconds", "duration_text", "completion_status", "client_ip", "device",
        ]
        with self.public_results_lock:
            # 同一個 session 只保留一筆完成紀錄。進入結果頁先建立；送出問卷時再更新同一筆，避免重複計數。
            existing_rows = []
            try:
                with open(self.public_results_path, "r", encoding="utf-8") as f:
                    for line in f:
                        try:
                            row = json.loads(line)
                            if isinstance(row, dict):
                                existing_rows.append(row)
                        except (ValueError, TypeError):
                            continue
            except OSError:
                pass

            replace_index = None
            if session_id:
                for index in range(len(existing_rows) - 1, -1, -1):
                    if str(existing_rows[index].get("session_id", "")) == session_id:
                        replace_index = index
                        break
            if replace_index is not None:
                previous = existing_rows[replace_index]
                merged = previous.copy()
                for key, value in clean.items():
                    # 問卷尚未填寫時，不用空字串覆蓋先前已填內容。
                    if value != "" or key not in {"satisfaction", "needs_improvement", "improvement_suggestion"}:
                        merged[key] = value
                merged["id"] = previous.get("id", clean["id"])
                merged["login_at"] = previous.get("login_at", "") or clean.get("login_at", "")
                merged["completed_at"] = previous.get("completed_at", "") or clean.get("completed_at", "")
                merged["duration_seconds"] = previous.get("duration_seconds", "") or clean.get("duration_seconds", "")
                merged["duration_text"] = previous.get("duration_text", "") or clean.get("duration_text", "")
                existing_rows[replace_index] = merged
                clean = merged
            else:
                existing_rows.append(clean)

            temp_jsonl = self.public_results_path + ".tmp"
            with open(temp_jsonl, "w", encoding="utf-8") as f:
                for row in existing_rows:
                    f.write(json.dumps(row, ensure_ascii=False) + "\n")
            os.replace(temp_jsonl, self.public_results_path)

            # CSV 由 JSONL 全量重建，確保同一 session 不會因問卷更新而出現重複列。
            temp_csv = self.public_results_csv_path + ".tmp"
            with open(temp_csv, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=csv_fields, extrasaction="ignore")
                writer.writeheader()
                for row in existing_rows:
                    writer.writerow({key: row.get(key, "") for key in csv_fields})
            os.replace(temp_csv, self.public_results_csv_path)

            # 資料寫入 JSONL／CSV 後，立即同步更新今天的 Excel。
            try:
                self._save_today_excel_to_disk()
            except Exception as exc:
                print(f"[Cloud Vision] 每日 Excel 自動保存失敗：{exc}", flush=True)
        # 預約與 Email 通知在本版本停用。
        return clean

    def _load_appointment_read_ids(self) -> set[str]:
        try:
            with open(self.appointment_read_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            values = data.get("read_ids", []) if isinstance(data, dict) else []
            return {str(x) for x in values if str(x).strip()}
        except (OSError, ValueError, TypeError):
            return set()

    def _save_appointment_read_ids(self, read_ids: set[str]) -> None:
        folder = os.path.dirname(self.appointment_read_path)
        if folder:
            os.makedirs(folder, exist_ok=True)
        temp_path = self.appointment_read_path + ".tmp"
        with open(temp_path, "w", encoding="utf-8") as f:
            json.dump({"read_ids": sorted(read_ids)}, f, ensure_ascii=False, indent=2)
        os.replace(temp_path, self.appointment_read_path)

    def _today_appointment_notifications(self) -> tuple[list[dict], list[dict]]:
        rows = [
            r for r in self._today_public_results()
            if r.get("appointment") == "希望安排預約"
        ]
        rows.sort(key=lambda r: (r.get("date", ""), r.get("time", "")), reverse=True)
        read_ids = self._load_appointment_read_ids()
        unread = [r for r in rows if r.get("id", "") not in read_ids]
        return rows, unread

    def _mark_today_appointments_read(self) -> int:
        rows, unread = self._today_appointment_notifications()
        read_ids = self._load_appointment_read_ids()
        read_ids.update(str(r.get("id", "")) for r in rows if r.get("id"))
        self._save_appointment_read_ids(read_ids)
        return len(unread)

    def _load_appointment_statuses(self) -> dict[str, dict]:
        try:
            with open(self.appointment_status_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data if isinstance(data, dict) else {}
        except (OSError, ValueError, TypeError):
            return {}

    def _save_appointment_statuses(self, statuses: dict[str, dict]) -> None:
        os.makedirs(os.path.dirname(self.appointment_status_path), exist_ok=True)
        temp_path = self.appointment_status_path + ".tmp"
        with open(temp_path, "w", encoding="utf-8") as f:
            json.dump(statuses, f, ensure_ascii=False, indent=2)
        os.replace(temp_path, self.appointment_status_path)

    def _appointment_status_for(self, record_id: str) -> str:
        item = self._load_appointment_statuses().get(str(record_id), {})
        status = str(item.get("status", "未聯絡")).strip()
        return status if status in {"未聯絡", "已聯絡", "已完成", "已取消"} else "未聯絡"

    def _set_appointment_status(self, record_id: str, status: str) -> None:
        if status not in {"未聯絡", "已聯絡", "已完成", "已取消"}:
            raise ValueError("預約狀態不正確")
        statuses = self._load_appointment_statuses()
        statuses[str(record_id)] = {
            "status": status,
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }
        self._save_appointment_statuses(statuses)
        self._save_today_excel_to_disk()

    def _save_public_event(self, payload: dict, client_ip: str, user_agent: str) -> dict:
        now = datetime.now()
        event = str(payload.get("event", "")).strip()[:40]
        if event not in {"view", "general_entry", "professional_entry", "start", "complete", "appointment"}:
            raise ValueError("invalid_event")
        user_type = str(payload.get("user_type", "")).strip()
        if user_type not in {"一般使用者", "專業人員"}:
            user_type = "專業人員" if event == "professional_entry" else "一般使用者"
        # 同一個 session 的同一入口事件只保留一筆，避免首頁與一般入口頁重複送出造成重複計數。
        incoming_session = str(payload.get("session_id", "")).strip()[:120]
        if incoming_session and event in {"general_entry", "professional_entry"}:
            try:
                with self.public_results_lock:
                    with open(self.public_events_path, "r", encoding="utf-8") as f:
                        for line in f:
                            try:
                                existing = json.loads(line)
                            except (ValueError, TypeError):
                                continue
                            if existing.get("session_id") == incoming_session and existing.get("event") == event:
                                return existing
            except OSError:
                pass
        row = {
            "id": now.strftime("%Y%m%d%H%M%S") + "-" + secrets.token_hex(3),
            "date": now.strftime("%Y-%m-%d"),
            "time": now.strftime("%H:%M:%S"),
            "created_at": now.isoformat(timespec="seconds"),
            "event": event,
            "user_type": user_type,
            "visitor_id": str(payload.get("visitor_id", "")).strip()[:120],
            "session_id": str(payload.get("session_id", "")).strip()[:120] or (now.strftime("%Y%m%d%H%M%S") + "-" + secrets.token_hex(4)),
            "test_id": "CV-" + now.strftime("%Y%m%d-%H%M%S-") + secrets.token_hex(2).upper(),
            "anonymous_id": str(payload.get("visitor_id", "")).strip()[:120] if user_type == "一般使用者" else "",
            "professional_name": str(payload.get("professional_name", "")).strip()[:120] if user_type == "專業人員" else "",
            "professional_role": str(payload.get("professional_role", "")).strip()[:120] if user_type == "專業人員" else "",
            "professional_purpose": str(payload.get("professional_purpose", "")).strip()[:160] if user_type == "專業人員" else "",
            "client_ip": client_ip,
            "device": user_agent[:300],
        }
        with self.public_results_lock:
            with open(self.public_events_path, "a", encoding="utf-8") as f:
                f.write(json.dumps(row, ensure_ascii=False) + "\n")
        try:
            self._save_today_excel_to_disk()
        except Exception as exc:
            print(f"[Cloud Vision] 使用者紀錄同步 Excel 失敗：{exc}", flush=True)
        return row

    @staticmethod
    def _format_duration(seconds) -> str:
        try:
            total = max(0, int(float(seconds)))
        except (TypeError, ValueError):
            return "—"
        minutes, secs = divmod(total, 60)
        hours, minutes = divmod(minutes, 60)
        if hours:
            return f"{hours} 小時 {minutes} 分 {secs} 秒"
        if minutes:
            return f"{minutes} 分 {secs} 秒"
        return f"{secs} 秒"

    def _session_timing(self, session_id: str) -> dict:
        if not session_id:
            return {"started_at": "", "duration_seconds": "", "test_id": ""}
        entry = None
        candidates = []
        try:
            with self.public_results_lock:
                with open(self.public_events_path, "r", encoding="utf-8") as f:
                    for line in f:
                        try:
                            row = json.loads(line)
                        except (ValueError, TypeError):
                            continue
                        if row.get("session_id") == session_id and row.get("event") in {"general_entry", "professional_entry", "start"}:
                            candidates.append(row)
            if candidates:
                entry = min(candidates, key=lambda r: str(r.get("created_at", "")))
        except OSError:
            pass
        if not entry:
            return {"started_at": "", "duration_seconds": "", "test_id": ""}
        started_at = str(entry.get("created_at", ""))
        try:
            duration = int((datetime.now() - datetime.fromisoformat(started_at)).total_seconds())
        except (ValueError, TypeError):
            duration = ""
        return {"started_at": started_at, "duration_seconds": duration, "test_id": str(entry.get("test_id", ""))}

    @staticmethod
    def _display_datetime(value: str) -> str:
        text = str(value or "").strip()
        if not text:
            return "未完成"
        try:
            dt = datetime.fromisoformat(text)
            return dt.strftime("%Y/%m/%d %H:%M:%S")
        except (ValueError, TypeError):
            return text.replace("T", " ")

    def _today_test_records(self) -> list[dict]:
        """每次登錄各為一筆；將完成結果與登錄事件依 session_id 合併。"""
        entries = [e.copy() for e in self._today_public_events() if e.get("event") in {"general_entry", "professional_entry"}]
        result_by_session = {str(r.get("session_id", "")): r for r in self._today_public_results() if r.get("session_id")}
        records = []
        for e in entries:
            sid = str(e.get("session_id", ""))
            result = result_by_session.get(sid)
            row = e.copy()
            row["login_at"] = e.get("created_at", "")
            row["test_id"] = e.get("test_id", "") or sid
            row["completion_status"] = "未完成"
            row["completed_at"] = ""
            row["duration_seconds"] = ""
            row["duration_text"] = "—"
            if result:
                row.update(result)
                row["test_id"] = result.get("test_id", "") or e.get("test_id", "") or sid
                row["login_at"] = result.get("login_at", "") or e.get("created_at", "")
                row["completion_status"] = "已完成"
            records.append(row)
        # 舊資料可能沒有入口事件，仍保留在清單中。
        known = {str(r.get("session_id", "")) for r in records}
        for r in self._today_public_results():
            if str(r.get("session_id", "")) not in known:
                records.append(r.copy())
        records.sort(key=lambda r: str(r.get("login_at") or r.get("created_at") or ""), reverse=True)
        return records

    def _today_public_events(self) -> list[dict]:
        today = datetime.now().strftime("%Y-%m-%d")
        rows = []
        try:
            with self.public_results_lock:
                with open(self.public_events_path, "r", encoding="utf-8") as f:
                    for line in f:
                        try:
                            row = json.loads(line)
                            if row.get("date") == today:
                                rows.append(row)
                        except (ValueError, TypeError):
                            continue
        except OSError:
            pass
        return rows

    def _today_dashboard_stats(self) -> dict:
        events = self._today_public_events()
        records = self._today_test_records()
        completed = [r for r in records if r.get("completion_status") == "已完成"]

        def durations(user_type: str) -> list[int]:
            values = []
            for r in completed:
                if r.get("user_type") != user_type:
                    continue
                try:
                    values.append(max(0, int(float(r.get("duration_seconds", "")))))
                except (TypeError, ValueError):
                    continue
            return values

        general_durations = durations("一般使用者")
        professional_durations = durations("專業人員")
        all_durations = general_durations + professional_durations
        surveys = sum(1 for r in completed if r.get("satisfaction") or r.get("needs_improvement") or r.get("improvement_suggestion"))
        return {
            "views": sum(1 for e in events if e.get("event") == "view"),
            "general_users": sum(1 for r in records if r.get("user_type") == "一般使用者"),
            "professional_users": sum(1 for r in records if r.get("user_type") == "專業人員"),
            "starts": len(records),
            "completes": len(completed),
            "incomplete": len(records) - len(completed),
            "general_avg_seconds": round(sum(general_durations) / len(general_durations)) if general_durations else 0,
            "professional_avg_seconds": round(sum(professional_durations) / len(professional_durations)) if professional_durations else 0,
            "overall_avg_seconds": round(sum(all_durations) / len(all_durations)) if all_durations else 0,
            "surveys": surveys, "contacts": 0, "appointments": 0,
        }

    @staticmethod
    def _screening_attention_reasons(record: dict) -> list[str]:
        """整理需優先查看的自我測驗結果；僅依已回傳欄位判斷。"""
        reasons: list[str] = []

        for eye_label, key in (("右眼", "astigmatism_right"), ("左眼", "astigmatism_left")):
            value = str(record.get(key, "")).strip()
            if value and value != "所有線條一樣深":
                reasons.append(f"{eye_label}散光鐘：{value}")

        for eye_label, key in (("右眼", "amsler_right"), ("左眼", "amsler_left")):
            value = str(record.get(key, "")).strip()
            if value and value != "線條筆直":
                reasons.append(f"{eye_label}黃斑部：{value}")

        near_value = str(record.get("near_reading", "")).strip()
        if near_value:
            normal_near_values = {"閱讀清楚", "可以閱讀", "可清楚閱讀", "正常"}
            if near_value not in normal_near_values:
                reasons.append(f"近距閱讀：{near_value}")

        return reasons

    def _build_today_excel(self) -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError as exc:
            raise RuntimeError("此電腦尚未安裝 openpyxl，請先執行 pip install openpyxl") from exc

        rows = self._today_test_records()
        completed_rows = [r for r in rows if r.get("completion_status") == "已完成"]
        stats = self._today_dashboard_stats()
        wb = Workbook()

        full_headers = ["日期", "時間", "Test ID", "登入時間", "完成時間", "測驗耗時", "完成狀態", "使用者類別", "匿名編號", "專業人員姓名／機構", "專業身分", "使用目的", "姓名", "電話", "右眼視力", "左眼視力", "雙眼視力", "右眼散光鐘", "左眼散光鐘", "右眼黃斑部", "左眼黃斑部", "近距閱讀", "是否預約", "預約日期", "預約時段", "備註", "滿意度", "是否需要改善", "改善建議"]

        def full_row(r: dict) -> list:
            return [
                r.get("date", ""), r.get("time", ""), r.get("test_id", "") or r.get("session_id", ""), self._display_datetime(r.get("login_at", "")), self._display_datetime(r.get("completed_at", "")) if r.get("completed_at") else "未完成", r.get("duration_text", ""), r.get("completion_status", "已完成"), r.get("user_type", "一般使用者"),
                r.get("anonymous_id", r.get("visitor_id", "")), r.get("professional_name", ""),
                r.get("professional_role", ""), r.get("professional_purpose", ""),
                r.get("name", ""), r.get("phone", ""),
                r.get("visual_acuity_right", ""), r.get("visual_acuity_left", ""), r.get("visual_acuity_both", ""),
                r.get("astigmatism_right", ""), r.get("astigmatism_left", ""),
                r.get("amsler_right", ""), r.get("amsler_left", ""), r.get("near_reading", ""),
                r.get("appointment", ""), r.get("appointment_date", ""),
                r.get("appointment_time", ""), r.get("note", ""),
                r.get("satisfaction", ""), r.get("needs_improvement", ""),
                r.get("improvement_suggestion", ""),
            ]

        # 1. 今日統計：放在第一頁，開啟 Excel 時即可先看到整體數量。
        summary = wb.active
        summary.title = "今日統計"
        summary.append(["Cloud Vision 今日統計", datetime.now().strftime("%Y-%m-%d")])
        summary.append([])
        for label, key in [
            ("今日一般使用者", "general_users"), ("今日專業人員", "professional_users"),
            ("今日登錄／測驗次數", "starts"), ("完成測驗次數", "completes"),
            ("未完成次數", "incomplete"), ("今日收到問卷", "surveys"),
        ]:
            summary.append([label, stats[key]])
        summary["A1"].font = Font(bold=True, size=16)
        summary["B1"].font = Font(bold=True, size=16)
        summary.column_dimensions["A"].width = 24
        summary.column_dimensions["B"].width = 18

        # 2. 每次登錄紀錄：包含已完成及未完成，手機後台與 Excel 使用同一份資料。
        all_tests = wb.create_sheet("每次登錄紀錄")
        all_tests.append(full_headers)
        for r in rows:
            all_tests.append(full_row(r))

        # 3. 全部完成資料。
        data = wb.create_sheet("全部完成資料")
        data.append(full_headers)
        for r in completed_rows:
            data.append(full_row(r))

        # 3. 需要注意的測驗結果：自動列出散光鐘、Amsler 或近距閱讀異常原因。
        attention = wb.create_sheet("需要注意的測驗結果")
        attention.append(["注意原因"] + full_headers)
        for r in completed_rows:
            reasons = self._screening_attention_reasons(r)
            if reasons:
                attention.append(["；".join(reasons)] + full_row(r))

        # 4. 預約驗光名單。
        appt = wb.create_sheet("預約驗光名單")
        appt.append(["預約日期", "預約時段", "姓名", "電話", "備註", "送出日期", "送出時間", "預約狀態"])
        for r in rows:
            if r.get("appointment") == "希望安排預約":
                appt.append([
                    r.get("appointment_date", ""), r.get("appointment_time", ""),
                    r.get("name", ""), r.get("phone", ""), r.get("note", ""),
                    r.get("date", ""), r.get("time", ""),
                    self._appointment_status_for(r.get("id", "")),
                ])

        # 5. 留下聯絡資料。
        contacts = wb.create_sheet("留下聯絡資料")
        contacts.append(["姓名", "電話", "建立日期", "建立時間", "是否預約", "預約日期", "預約時段", "備註"])
        for r in rows:
            if r.get("name") or r.get("phone"):
                contacts.append([
                    r.get("name", ""), r.get("phone", ""), r.get("date", ""), r.get("time", ""),
                    r.get("appointment", ""), r.get("appointment_date", ""),
                    r.get("appointment_time", ""), r.get("note", ""),
                ])

        # 6. 完成測驗但沒有留下姓名與電話者。
        anonymous = wb.create_sheet("僅完成未留資料")
        anonymous.append(full_headers)
        for r in rows:
            if not r.get("name") and not r.get("phone"):
                anonymous.append(full_row(r))

        # 7. 使用問卷。
        survey = wb.create_sheet("使用問卷")
        survey.append(["日期", "時間", "Test ID", "登入時間", "完成時間", "測驗耗時", "使用者類別", "匿名編號／專業人員", "滿意度", "是否需要改善", "改善建議"])
        for r in completed_rows:
            if r.get("satisfaction") or r.get("needs_improvement") or r.get("improvement_suggestion"):
                survey.append([
                    r.get("date", ""), r.get("time", ""), r.get("test_id", "") or r.get("session_id", ""),
                    self._display_datetime(r.get("login_at", "")), self._display_datetime(r.get("completed_at", "")),
                    r.get("duration_text", ""), r.get("user_type", "一般使用者"),
                    r.get("professional_name", "") or r.get("anonymous_id", r.get("visitor_id", "")),
                    r.get("satisfaction", ""), r.get("needs_improvement", ""), r.get("improvement_suggestion", ""),
                ])

        # 8. 所有進入平台的使用者紀錄（一般使用者匿名、專業人員保留其自填資料）。
        usage = wb.create_sheet("使用者進入紀錄")
        usage.append(["日期", "時間", "Test ID", "Session ID", "使用者類別", "事件", "匿名編號", "專業人員姓名／機構", "專業身分", "使用目的", "裝置"])
        event_names = {"view": "瀏覽首頁", "general_entry": "進入一般使用者", "professional_entry": "專業人員登錄", "start": "開始測驗", "complete": "完成測驗", "appointment": "其他"}
        for e in reversed(self._today_public_events()):
            usage.append([
                e.get("date", ""), e.get("time", ""), e.get("test_id", ""), e.get("session_id", ""), e.get("user_type", ""),
                event_names.get(e.get("event", ""), e.get("event", "")),
                e.get("anonymous_id", ""), e.get("professional_name", ""),
                e.get("professional_role", ""), e.get("professional_purpose", ""), e.get("device", ""),
            ])

        # 9. 專業人員資料：每次由專業入口送出的自我聲明均會留下紀錄。
        professionals = wb.create_sheet("專業人員資料")
        professionals.append(["日期", "時間", "姓名／機構", "專業身分", "使用目的", "裝置"])
        for e in reversed(self._today_public_events()):
            if e.get("event") == "professional_entry":
                professionals.append([e.get("date", ""), e.get("time", ""), e.get("professional_name", ""), e.get("professional_role", ""), e.get("professional_purpose", ""), e.get("device", "")])

        table_sheets = (data, attention, appt, contacts, anonymous, survey, usage, professionals)
        for sheet in table_sheets:
            for c in sheet[1]:
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="DCEBFF")
                c.alignment = Alignment(horizontal="center", vertical="center")
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for col in sheet.columns:
                width = min(max(len(str(cell.value or "")) for cell in col) + 2, 42)
                sheet.column_dimensions[col[0].column_letter].width = width
            for row_cells in sheet.iter_rows(min_row=2):
                for cell in row_cells:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        # 需要注意的資料以淡黃色醒目標示。
        for row_cells in attention.iter_rows(min_row=2):
            for cell in row_cells:
                cell.fill = PatternFill("solid", fgColor="FFF4CC")

        for row_cells in summary.iter_rows(min_row=3, max_row=7, min_col=1, max_col=2):
            row_cells[0].font = Font(bold=True)
            row_cells[0].fill = PatternFill("solid", fgColor="EAF3FF")

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    def _today_excel_disk_path(self) -> str:
        os.makedirs(self.public_excel_dir, exist_ok=True)
        return os.path.join(
            self.public_excel_dir,
            "CloudVision_" + datetime.now().strftime("%Y-%m-%d") + ".xlsx",
        )

    def _save_today_excel_to_disk(self) -> str:
        """將今日完整資料寫成實體 Excel；每次回傳都覆寫更新同一天檔案。"""
        path = self._today_excel_disk_path()
        data = self._build_today_excel()
        tmp = path + ".tmp"
        with open(tmp, "wb") as f:
            f.write(data)
        os.replace(tmp, path)
        return path

    def open_examiner_dashboard(self) -> None:
        """在電腦預設瀏覽器開啟受密碼保護的測驗者後台。"""
        if not self.remote_url:
            self.start_remote_server()
        url = (self.remote_url or "").rstrip("/") + "/examiner"
        if not url.startswith("http"):
            messagebox.showinfo("測驗者後台", "請先啟動連線模式／QR Code，再開啟測驗者後台。")
            return
        try:
            webbrowser.open(url, new=2)
            self.status_var.set(f"已開啟測驗者後台：{url}")
        except Exception as exc:
            messagebox.showerror("測驗者後台", f"無法開啟瀏覽器：{exc}")

    def open_today_public_results(self) -> None:
        """相容舊按鈕：改為先進入受密碼保護的測驗者後台。"""
        self.open_examiner_dashboard()

    def _today_public_results(self) -> list[dict]:
        today = datetime.now().strftime("%Y-%m-%d")
        rows = []
        try:
            with self.public_results_lock:
                with open(self.public_results_path, "r", encoding="utf-8") as f:
                    for line in f:
                        try:
                            row = json.loads(line)
                            if row.get("date") == today:
                                rows.append(row)
                        except (ValueError, TypeError):
                            continue
        except OSError:
            pass
        return list(reversed(rows))

    def public_home_html(self) -> str:
        """公開首頁僅介紹平台與研究資訊，不在首頁啟動校正或測驗。"""
        return """<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1,viewport-fit=cover'><title>Cloud Vision｜視覺功能平台</title><style>
*{box-sizing:border-box}body{margin:0;background:linear-gradient(180deg,#eef5ff,#f7f9fc);color:#172033;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif}.wrap{max-width:900px;margin:auto;padding:34px 18px 70px}.hero{text-align:center;padding:22px 10px}.logo{font-size:48px}h1{font-size:38px;margin:5px 0}.subtitle{font-size:20px;color:#4d6078;margin:8px 0 24px}.card{background:#fff;border:1px solid #dce5f0;border-radius:22px;padding:26px;box-shadow:0 8px 28px rgba(29,51,84,.08);margin-bottom:18px}.intro{font-size:18px;line-height:1.8;text-align:center}.research{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-top:20px}.info{background:#f7faff;border:1px solid #dce7f6;border-radius:15px;padding:16px;text-align:center}.label{color:#64748b;font-size:14px}.value{font-size:20px;font-weight:900;margin-top:5px}.actions{display:grid;grid-template-columns:1fr 1fr;gap:14px}.examiner-entry{display:block;margin-top:14px;text-decoration:none;color:#fff;text-align:center;border-radius:16px;padding:18px 12px;font-size:20px;font-weight:900;background:#27364a;border:2px solid #172033}.examiner-entry .small{font-size:14px}.examiner-entry:hover{filter:brightness(1.06)}.btn{display:block;text-decoration:none;color:#fff;text-align:center;border-radius:16px;padding:20px 12px;font-size:21px;font-weight:900}.general{background:#1769e0}.professional{background:#14823b}.small{display:block;font-size:14px;font-weight:600;opacity:.92;margin-top:6px}.notice{font-size:14px;line-height:1.7;color:#64748b;text-align:center}.link{color:#244f91;font-weight:800}.device-badge{position:fixed;top:14px;right:14px;z-index:9999;padding:10px 14px;border-radius:999px;background:#172033;color:#fff;font-weight:900;font-size:15px;box-shadow:0 6px 18px rgba(0,0,0,.18)}
/* 三種裝置明顯不同，方便現場立即驗證 */
html.device-desktop .wrap{max-width:980px;padding-top:42px}
html.device-desktop .actions{grid-template-columns:1fr 1fr}
html.device-desktop .btn{min-height:86px}
html.device-tablet body{background:linear-gradient(180deg,#fff8e8,#f7f9fc)}
html.device-tablet .wrap{max-width:880px;padding:34px 28px 72px}
html.device-tablet .hero{padding-top:34px}
html.device-tablet .card{padding:30px;border-radius:26px}
html.device-tablet .intro{font-size:20px}
html.device-tablet .btn{padding:26px 16px;font-size:25px;min-height:94px}
html.device-tablet .device-badge{background:#a15c00}
html.device-phone body{background:linear-gradient(180deg,#eefcf3,#f7f9fc)}
html.device-phone .wrap{max-width:100%;padding:58px 12px 42px}
html.device-phone .hero{padding:6px 6px 14px}
html.device-phone .logo{font-size:36px}
html.device-phone h1{font-size:29px}
html.device-phone .subtitle{font-size:16px;margin-bottom:14px}
html.device-phone .card{padding:17px;border-radius:17px}
html.device-phone .intro{font-size:16px;line-height:1.65}
html.device-phone .research,html.device-phone .actions{grid-template-columns:1fr}html.device-phone .examiner-entry{padding:20px 12px;font-size:22px;min-height:82px}
html.device-phone .research{gap:9px}
html.device-phone .info{padding:13px}
html.device-phone .btn{padding:22px 12px;font-size:23px;min-height:88px;border-radius:18px}
html.device-phone .device-badge{top:8px;right:8px;background:#0b7a3b;font-size:14px;padding:9px 12px}
@media(max-width:650px){h1{font-size:29px}.research,.actions{grid-template-columns:1fr}.wrap{padding-top:58px}.card{padding:17px}}
</style></head><body><div id='deviceBadge' class='device-badge'>正在判斷裝置…</div><main class='wrap'><section class='hero'><div class='logo'>☁️👁️</div><h1>Cloud Vision</h1><div class='subtitle'>視覺功能測驗與教學平台</div></section><section class='card'><div class='intro'>本平台為教學與研究使用之 Beta 版本，提供視力、散光鐘、黃斑部 Amsler 方格、螢幕尺度校正，以及視覺功能相關教學工具。請依使用身分選擇入口，使用後可提供回饋協助改善。</div><div class='research'><div class='info'><div class='label'>開發者／學生</div><div class='value'>黃昭維</div><div>大葉大學研究所二年級</div></div><div class='info'><div class='label'>指導老師</div><div class='value'>黃敬堯</div><div>研究指導</div></div><div class='info'><div class='label'>聯絡方式</div><div class='value'>LINE ID</div><div>a0937587396</div></div><div class='info'><div class='label'>平台用途</div><div class='value'>研究・測驗・教學</div><div>瀏覽器直接使用</div></div></div></section><section class='actions'><a id='generalEntry' class='btn general' href='/cloud/general'>一般使用者<span class='small'>開始視覺功能自我測驗</span></a><a id='professionalEntry' class='btn professional' href='/cloud/professional'>專業使用者<span class='small'>填寫資料後進入測驗或教學</span></a></section><a id='examinerEntry' class='examiner-entry' href='/examiner'>🔒 檢查者後台<span class='small'>密碼登入・查看今日資料・進入手機／iPad控制台</span></a><section class='card notice'>本平台結果僅供研究、教育與初步自我觀察，不作為醫療診斷依據。<br><a class='link' href='/cloud/disclaimer'>查看平台說明與完整免責聲明</a><br><br>© 2026 Cloud Vision｜開發者：黃昭維｜指導老師：黃敬堯</section></main><script>
(function detectCloudVisionDevice(){
  const ua=navigator.userAgent||'';
  const uaData=navigator.userAgentData;
  const touchPoints=navigator.maxTouchPoints||0;
  const shortest=Math.min(window.innerWidth||0,window.innerHeight||0);
  const longest=Math.max(window.innerWidth||0,window.innerHeight||0);
  const isIPad=/iPad/i.test(ua)||(/Macintosh/i.test(ua)&&touchPoints>1);
  const isAndroid=/Android/i.test(ua);
  const isMobileUA=uaData&&typeof uaData.mobile==='boolean'?uaData.mobile:/Mobi|iPhone|iPod|Windows Phone/i.test(ua);
  let device='desktop';
  if(isIPad||(isAndroid&&!/Mobile/i.test(ua))||(touchPoints>1&&shortest>=600&&longest<=1400)) device='tablet';
  else if(isMobileUA||shortest<600) device='phone';
  const root=document.documentElement;
  root.classList.remove('device-phone','device-tablet','device-desktop');
  root.classList.add('device-'+device);
  root.dataset.device=device;
  const badge=document.getElementById('deviceBadge');
  if(badge){badge.textContent=device==='phone'?'📱 手機版':device==='tablet'?'▣ 平板版':'💻 電腦版';}
  try{localStorage.setItem('cloudVisionDeviceType',device)}catch(e){}
  for(const id of ['generalEntry','professionalEntry','examinerEntry']){
    const link=document.getElementById(id);
    if(link){
      const u=new URL(link.href,location.href);
      u.searchParams.set('device',device);
      link.href=u.pathname+u.search;
    }
  }
  window.cloudVisionDeviceType=device;
})();
const vk='cloudVisionVisitorV1019';
let vid=localStorage.getItem(vk);
if(!vid){vid=(crypto.randomUUID?crypto.randomUUID():Date.now()+'-'+Math.random());localStorage.setItem(vk,vid)}
const makeSession=()=>((crypto.randomUUID?crypto.randomUUID():Date.now()+'-'+Math.random()));
function eventPayload(event,user_type,session_id=''){
  return JSON.stringify({event,visitor_id:vid,user_type,session_id});
}
function sendEventReliable(event,user_type,session_id=''){
  const body=eventPayload(event,user_type,session_id);
  if(navigator.sendBeacon){
    try{
      const ok=navigator.sendBeacon('/cloud/event',new Blob([body],{type:'application/json'}));
      if(ok)return Promise.resolve(true);
    }catch(e){}
  }
  return fetch('/cloud/event',{method:'POST',headers:{'Content-Type':'application/json'},body,keepalive:true,cache:'no-store'})
    .then(()=>true).catch(()=>false);
}
sendEventReliable('view','一般使用者');
document.getElementById('generalEntry').addEventListener('click',async(e)=>{
  e.preventDefault();
  const sid=makeSession();
  sessionStorage.setItem('cloudVisionCurrentSession',sid);
  sessionStorage.setItem('cloudVisionCurrentUserType','一般使用者');
  await Promise.race([sendEventReliable('general_entry','一般使用者',sid),new Promise(r=>setTimeout(r,700))]);
  location.href=e.currentTarget.href;
});
</script></body></html>"""

    def public_general_entry_html(self) -> str:
        """一般使用者按下入口後，才測驗是否完成 5 cm 校正。"""
        token = self.connection_session_token
        return f"""<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>一般使用者｜Cloud Vision</title><style>body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif;background:#f5f8fc;margin:0;color:#172033}}.wrap{{max-width:720px;margin:auto;padding:40px 18px}}.card{{background:#fff;border:1px solid #dce5f0;border-radius:20px;padding:28px;text-align:center;box-shadow:0 6px 22px rgba(29,51,84,.07)}}h1{{margin-top:0}}p{{line-height:1.8;color:#52647b}}.btn{{display:block;padding:15px;border-radius:13px;background:#1769e0;color:#fff;text-decoration:none;font-weight:900;margin-top:16px}}</style></head><body><main class='wrap'><section class='card'><h1>一般使用者測驗流程</h1><p id='msg'>正在建立本次登入與使用時間紀錄……</p><a class='btn' href='/cloud'>返回首頁</a></section></main><script>
const calibrationKey='cloudVisionCalibrationV10';
const visitorKey='cloudVisionVisitorV1019';
let sid=sessionStorage.getItem('cloudVisionCurrentSession')||'';
if(!sid){{sid=(crypto.randomUUID?crypto.randomUUID():Date.now()+'-'+Math.random());sessionStorage.setItem('cloudVisionCurrentSession',sid)}}
sessionStorage.setItem('cloudVisionCurrentUserType','一般使用者');
let visitor='';
try{{visitor=localStorage.getItem(visitorKey)||'';if(!visitor){{visitor=(crypto.randomUUID?crypto.randomUUID():Date.now()+'-'+Math.random());localStorage.setItem(visitorKey,visitor)}}}}catch(e){{}}
async function ensureGeneralEntry(){{
  const body=JSON.stringify({{event:'general_entry',user_type:'一般使用者',visitor_id:visitor,session_id:sid}});
  try{{
    const r=await fetch('/cloud/event',{{method:'POST',headers:{{'Content-Type':'application/json'}},body,cache:'no-store'}});
    if(!r.ok)throw new Error('record failed');
    sessionStorage.setItem('cloudVisionGeneralEntrySaved',sid);
    return true;
  }}catch(e){{
    document.getElementById('msg').textContent='正在重新建立登入紀錄，請稍候……';
    await new Promise(r=>setTimeout(r,500));
    try{{
      const r=await fetch('/cloud/event',{{method:'POST',headers:{{'Content-Type':'application/json'}},body,cache:'no-store'}});
      return r.ok;
    }}catch(err){{return false}}
  }}
}}
(async()=>{{
  await ensureGeneralEntry();
  let d=null;try{{d=JSON.parse(localStorage.getItem(calibrationKey)||'null')}}catch(e){{}}
  if(d&&d.factor)location.replace('/participant?session={token}');
  else location.replace('/cloud/calibration?next=general');
}})();
</script></body></html>"""

    def public_disclaimer_html(self) -> str:
        """平台說明與免責聲明；不要求指導老師簽名，也不顯示系統使用者姓名。"""
        return """<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1,viewport-fit=cover'><title>平台說明與免責聲明｜Cloud Vision</title><style>
        *{box-sizing:border-box}body{margin:0;background:#f5f8fc;color:#172033;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif}.wrap{max-width:820px;margin:auto;padding:24px 16px 70px}.card{background:#fff;border:1px solid #dce5f0;border-radius:20px;padding:24px;box-shadow:0 6px 22px rgba(29,51,84,.07);margin-bottom:16px}h1{margin:0 0 18px;font-size:30px}h2{font-size:21px;color:#244f91;margin:22px 0 8px}p,li{line-height:1.75}.name{font-weight:900;font-size:19px}.btn{display:block;text-align:center;text-decoration:none;background:#1769e0;color:#fff;border-radius:13px;padding:14px;font-weight:900;margin-top:18px}.muted{color:#64748b;font-size:14px}.notice{background:#fff7df;border:1px solid #f1d58a;border-radius:14px;padding:14px;font-weight:750}</style></head><body><main class='wrap'><section class='card'><h1>Cloud Vision 平台說明</h1><p>本平台提供瀏覽器式視覺功能工具，包含螢幕尺度校正、視力測驗、散光鐘與黃斑部 Amsler 方格。可使用手機、平板或電腦操作，不需下載應用程式。</p><h2>研究資訊</h2><p class='name'>開發者：黃昭維</p><p>學校：大葉大學<br>身分：研究所二年級<br>指導老師：黃敬堯</p><h2>聯絡資訊</h2><p class='name'>LINE ID：a0937587396</p><p class='muted'>如需平台操作、研究合作或使用回饋，可透過 LINE 聯絡。</p><h2>免責聲明</h2><div class='notice'>本工具提供的是初步視覺功能測驗與自我觀察結果，不構成醫療診斷，也不能取代眼科醫師診察、完整驗光或其他專業醫療評估。</div><p>測驗結果可能受螢幕尺寸、亮度、觀看距離、環境光線、裝置校正、眼鏡配戴狀況與操作方式影響。若出現視力突然下降、影像扭曲、黑影、閃光、視野缺損、眼痛或其他不適，應儘速尋求眼科專業評估。</p><p>使用者應依畫面指示完成 5 cm 尺度校正並維持指定觀看距離。本版本不提供線上驗光或預約服務；一般使用者不需填寫姓名、電話、Email、生日、年齡或性別。平台僅保存匿名測驗結果與使用回饋。</p><a class='btn' href='/cloud'>我已閱讀，返回平台</a></section></main></body></html>"""

    def public_professional_html(self) -> str:
        """專業使用者先填寫基本資料，再進入免責聲明。"""
        return """<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1,viewport-fit=cover'><title>專業使用者入口｜Cloud Vision</title><style>
*{box-sizing:border-box}body{margin:0;background:#f5f8fc;color:#172033;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif}.wrap{max-width:760px;margin:auto;padding:28px 16px 70px}.card{background:#fff;border:1px solid #dce5f0;border-radius:20px;padding:26px;box-shadow:0 6px 22px rgba(29,51,84,.07)}h1{margin:0 0 8px}.sub{color:#64748b;line-height:1.7}label{display:block;font-weight:850;margin:16px 0 6px}input,select{width:100%;padding:14px;border:1px solid #b8c6d8;border-radius:11px;font-size:17px;background:#fff}button{width:100%;border:0;border-radius:13px;background:#14823b;color:#fff;font-size:19px;font-weight:900;padding:15px;margin-top:22px}.note{background:#eef8f1;border:1px solid #b8dfc4;border-radius:12px;padding:13px;margin-top:17px;line-height:1.65}.back{display:block;text-align:center;margin-top:18px;color:#244f91;text-decoration:none;font-weight:800}</style></head><body><main class='wrap'><section class='card'><h1>專業使用者資料</h1><p class='sub'>請填寫姓名、專業身分與使用目的。送出後會先閱讀免責聲明，再進入「測驗」或「教學」。</p><form id='verify'><label>姓名或機構名稱</label><input id='professionalName' required placeholder='請輸入姓名或機構'><label>專業身分</label><select id='professionalRole' required><option value=''>請選擇</option><option>驗光人員</option><option>眼科相關人員</option><option>視光或醫療教師</option><option>研究人員／研究生</option><option>其他視覺功能相關專業人員</option></select><label>使用目的</label><select id='professionalPurpose' required><option value=''>請選擇</option><option>專業測驗</option><option>教學展示</option><option>研究測試</option><option>視標大小與距離換算</option></select><div class='note'>第一階段採自我聲明。送出資料不代表醫療資格認證，平台工具亦不取代完整驗光或眼科診斷。</div><button type='submit'>下一步：閱讀免責聲明</button></form><a class='back' href='/cloud'>返回首頁</a></section></main><script>const key='cloudVisionProfessionalV2';let saved=null;try{saved=JSON.parse(localStorage.getItem(key)||'null')}catch(e){}if(saved&&saved.name){document.getElementById('professionalName').value=saved.name||'';document.getElementById('professionalRole').value=saved.role||'';document.getElementById('professionalPurpose').value=saved.purpose||''}document.getElementById('verify').addEventListener('submit',async e=>{e.preventDefault();const sid=(crypto.randomUUID?crypto.randomUUID():Date.now()+'-'+Math.random());sessionStorage.setItem('cloudVisionCurrentSession',sid);sessionStorage.setItem('cloudVisionCurrentUserType','專業人員');const data={name:document.getElementById('professionalName').value.trim(),role:document.getElementById('professionalRole').value,purpose:document.getElementById('professionalPurpose').value,openedAt:new Date().toISOString(),session_id:sid};localStorage.setItem(key,JSON.stringify(data));let visitor='';try{visitor=localStorage.getItem('cloudVisionVisitorV1019')||''}catch(err){}try{await fetch('/cloud/event',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({event:'professional_entry',user_type:'專業人員',visitor_id:visitor,professional_name:data.name,professional_role:data.role,professional_purpose:data.purpose,session_id:sid})})}catch(err){}location.href='/cloud/professional/disclaimer'});</script></body></html>"""

    def public_professional_disclaimer_html(self) -> str:
        return """<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>專業使用免責聲明｜Cloud Vision</title><style>*{box-sizing:border-box}body{margin:0;background:#f5f8fc;color:#172033;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif}.wrap{max-width:780px;margin:auto;padding:28px 16px 70px}.card{background:#fff;border:1px solid #dce5f0;border-radius:20px;padding:26px;box-shadow:0 6px 22px rgba(29,51,84,.07)}h1{margin-top:0}.notice{background:#fff7df;border:1px solid #f1d58a;border-radius:14px;padding:17px;line-height:1.8}p{line-height:1.8}.agree{display:flex;gap:10px;align-items:flex-start;background:#f7faff;padding:14px;border-radius:12px;margin-top:17px}.agree input{width:22px;height:22px}.btn{width:100%;border:0;border-radius:13px;background:#14823b;color:#fff;font-size:19px;font-weight:900;padding:15px;margin-top:20px}.btn:disabled{background:#9aaba0}.back{display:block;text-align:center;margin-top:17px;color:#244f91;text-decoration:none;font-weight:800}</style></head><body><main class='wrap'><section class='card'><h1>專業使用免責聲明</h1><div class='notice'>Cloud Vision 提供視覺功能測驗、教學及研究輔助工具，不構成醫療診斷，也不能取代眼科醫師診察、完整驗光或其他專業評估。</div><p>測驗結果可能受到螢幕尺寸、校正、亮度、觀看距離、環境光線及操作方式影響。使用者應依專業判斷解讀結果；如有視力突然下降、影像扭曲、黑影、閃光、視野缺損或眼痛，應儘速安排眼科評估。</p><div class='agree'><input id='agree' type='checkbox'><label for='agree'>我已閱讀並了解以上內容，同意依專業與研究用途使用本平台。</label></div><button id='continue' class='btn' disabled>同意並進入專業工具</button><a class='back' href='/cloud/professional'>返回修改資料</a></section></main><script>const data=localStorage.getItem('cloudVisionProfessionalV2');if(!data)location.replace('/cloud/professional');const c=document.getElementById('agree'),b=document.getElementById('continue');c.addEventListener('change',()=>b.disabled=!c.checked);b.addEventListener('click',()=>{localStorage.setItem('cloudVisionProfessionalDisclaimerV2',new Date().toISOString());location.href='/cloud/professional/hub'});</script></body></html>"""

    def public_professional_hub_html(self) -> str:
        """專業人員工作台：畫面只保留測驗與教學，背景資料與 Excel 功能維持不變。"""
        return """<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>專業人員工作台｜Cloud Vision</title><style>
*{box-sizing:border-box}body{margin:0;background:#f5f8fc;color:#172033;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif}.wrap{max-width:980px;margin:auto;padding:28px 16px 70px}.head{text-align:center;margin-bottom:24px}.head h1{margin:0}.who{color:#64748b;margin-top:8px}.grid{display:grid;grid-template-columns:repeat(2,1fr);gap:14px}.choice{display:block;text-decoration:none;color:#172033;background:#fff;border:1px solid #dce5f0;border-radius:20px;padding:28px;box-shadow:0 6px 20px rgba(29,51,84,.06)}.choice:hover{border-color:#1769e0}.icon{font-size:39px}.title{font-size:24px;font-weight:950;margin:8px 0}.desc{line-height:1.65;color:#607086}.blue{border-top:7px solid #1769e0}.green{border-top:7px solid #14823b}.back{display:block;text-align:center;margin-top:22px;color:#244f91;text-decoration:none;font-weight:800}@media(max-width:620px){.grid{grid-template-columns:1fr}}
</style></head><body><main class='wrap'><header class='head'><h1>Cloud Vision 專業人員工作台</h1><div id='who' class='who'></div></header>
<section class='grid'><a id='professionalTestEntry' class='choice blue' href='/cloud/professional/test'><div class='icon'>🧪</div><div class='title'>測驗</div><div class='desc'>沿用原本的 5 cm 校正、視力、散光鐘、黃斑部與視標大小工具。</div></a><a class='choice green' href='/cloud/professional/teaching'><div class='icon'>📚</div><div class='title'>教學</div><div class='desc'>沿用原本完整的視力幾何、視標大小、散光鐘與操作教學。</div></a></section><a class='back' href='/cloud'>返回 Cloud Vision 首頁</a></main><script>let d=null;try{d=JSON.parse(localStorage.getItem('cloudVisionProfessionalV2')||'null')}catch(e){}if(!d||!localStorage.getItem('cloudVisionProfessionalDisclaimerV2'))location.replace('/cloud/professional');else{document.getElementById('who').textContent=(d.name||'專業使用者')+'｜'+(d.role||'');const a=document.getElementById('professionalTestEntry');if(a)a.addEventListener('click',()=>{let vid='';try{vid=localStorage.getItem('cloudVisionVisitorV1019')||''}catch(e){}fetch('/cloud/event',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({event:'start',user_type:'專業人員',visitor_id:vid,professional_name:d.name||'',professional_role:d.role||'',professional_purpose:d.purpose||'',session_id:sessionStorage.getItem('cloudVisionCurrentSession')||d.session_id||''}),keepalive:true}).catch(()=>{})})}</script></body></html>"""

    def public_professional_data_html(self, query_text: str = "") -> str:
        """專業人員唯讀資料頁；不修改原測驗、Excel 或管理後台。"""
        rows = self._today_public_results()
        q = (query_text or "").strip().lower()
        if q:
            rows = [r for r in rows if q in " ".join(str(r.get(k, "")) for k in ("name", "phone", "time", "visual_acuity_right", "visual_acuity_left", "note")).lower()]
        body = []
        for r in rows:
            reasons = self._screening_attention_reasons(r)
            status = "；".join(reasons) if reasons else "完成，未見系統標記項目"
            name = html.escape(str(r.get("name") or "匿名受試者"))
            phone = html.escape(str(r.get("phone") or "—"))
            body.append(f"""<tr><td>{html.escape(str(r.get('time','')))}</td><td><b>{name}</b><br><small>{phone}</small></td><td>右：{html.escape(str(r.get('visual_acuity_right') or '—'))}<br>左：{html.escape(str(r.get('visual_acuity_left') or '—'))}<br>雙眼：{html.escape(str(r.get('visual_acuity_both') or '—'))}</td><td>右：{html.escape(str(r.get('astigmatism_right') or '—'))}<br>左：{html.escape(str(r.get('astigmatism_left') or '—'))}</td><td>右：{html.escape(str(r.get('amsler_right') or '—'))}<br>左：{html.escape(str(r.get('amsler_left') or '—'))}</td><td>{html.escape(status)}</td><td>{html.escape(str(r.get('appointment') or '—'))}<br><small>{html.escape(str(r.get('appointment_date') or ''))} {html.escape(str(r.get('appointment_time') or ''))}</small></td></tr>""")
        rows_html = "".join(body) or "<tr><td colspan='7' class='empty'>找不到符合條件的今日資料。</td></tr>"
        q_value = html.escape(query_text or "")
        return f"""<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>今日測驗資料｜Cloud Vision</title><style>*{{box-sizing:border-box}}body{{margin:0;background:#f5f8fc;color:#172033;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif}}.wrap{{max-width:1200px;margin:auto;padding:24px 14px 70px}}.card{{background:#fff;border:1px solid #dce5f0;border-radius:18px;padding:20px;box-shadow:0 6px 22px rgba(29,51,84,.06)}}.top{{display:flex;justify-content:space-between;gap:12px;align-items:center;flex-wrap:wrap}}h1{{margin:0}}.actions{{display:flex;gap:8px;flex-wrap:wrap}}.btn{{display:inline-block;text-decoration:none;border:0;border-radius:11px;background:#1769e0;color:white;padding:11px 14px;font-weight:850}}.green{{background:#14823b}}form{{display:flex;gap:8px;margin:18px 0;flex-wrap:wrap}}input{{flex:1;min-width:220px;padding:12px;border:1px solid #b8c6d8;border-radius:10px;font-size:16px}}button{{border:0;border-radius:10px;background:#1769e0;color:white;padding:12px 18px;font-weight:850}}.tableWrap{{overflow:auto}}table{{width:100%;border-collapse:collapse;min-width:1000px}}th,td{{padding:11px;border-bottom:1px solid #e2e9f1;text-align:left;vertical-align:top;line-height:1.5}}th{{background:#eef5ff;position:sticky;top:0}}small{{color:#64748b}}.empty{{text-align:center;color:#64748b;padding:30px}}.note{{color:#64748b;font-size:14px;margin-top:14px;line-height:1.6}}</style></head><body><main class='wrap'><section class='card'><div class='top'><div><h1>今日測驗資料</h1><div>共 {len(rows)} 筆符合條件的完成資料</div></div><div class='actions'><a class='btn green' href='/cloud/professional/today.xlsx'>下載今日 Excel</a><a class='btn' href='/cloud/professional/hub'>返回工作台</a></div></div><form method='get' action='/cloud/professional/data'><input name='q' value='{q_value}' placeholder='搜尋姓名、電話、時間、視力或備註'><button type='submit'>搜尋</button></form><div class='tableWrap'><table><thead><tr><th>時間</th><th>受試者</th><th>視力</th><th>散光鐘</th><th>黃斑部</th><th>系統整理</th><th>預約</th></tr></thead><tbody>{rows_html}</tbody></table></div><p class='note'>此頁只新增於專業人員工作台，原本一般使用者、測驗、教學、黃斑部、散光鐘與 Excel 自動整理功能均未更動。結果僅供專業判讀與研究管理，不作為醫療診斷。</p></section></main><script>if(!localStorage.getItem('cloudVisionProfessionalDisclaimerV2'))location.replace('/cloud/professional');</script></body></html>"""

    def public_professional_test_html(self) -> str:
        """專業測驗入口：尚未校正時，先進入 5 cm 校正。"""
        token = self.connection_session_token
        return f"""<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>專業測驗｜Cloud Vision</title><style>*{{box-sizing:border-box}}body{{margin:0;background:#f5f8fc;color:#172033;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif}}.wrap{{max-width:860px;margin:auto;padding:28px 16px 70px}}.card{{background:#fff;border:1px solid #dce5f0;border-radius:20px;padding:25px;box-shadow:0 6px 22px rgba(29,51,84,.07)}}.status{{padding:13px;border-radius:12px;background:#fff7df;border:1px solid #f1d58a;font-weight:800;margin-bottom:18px}}.grid{{display:grid;grid-template-columns:1fr 1fr;gap:13px}}.tool{{display:block;text-decoration:none;color:#172033;border:1px solid #dce5f0;border-radius:15px;padding:20px;background:#f9fbfd;font-weight:900;font-size:18px}}.back{{display:block;text-align:center;margin-top:20px;color:#244f91;text-decoration:none;font-weight:800}}@media(max-width:620px){{.grid{{grid-template-columns:1fr}}}}</style></head><body><main class='wrap'><section class='card'><h1>專業測驗</h1><div id='status' class='status'>正在確認 5 cm 校正狀態……</div><div id='tools' class='grid' style='display:none'><a class='tool' href='/cloud/calibration?next=professional-test'>📏 重新進行 5 cm 校正</a><a class='tool' href='/cloud/tools/size'>📐 視標大小與距離換算</a><a class='tool' href='/cloud/tools/geometry'>📊 視力幾何教學中心</a><a class='tool' href='/cloud/general?source=professional'>🔤 完整視覺功能測驗</a><a class='tool' href='/cloud/general?source=professional'>✳️ 散光鐘與黃斑部完整雙眼流程</a></div><a class='back' href='/cloud/professional/hub'>返回測驗／教學選擇</a></section></main><script>const k='cloudVisionCalibrationV10';let d=null;try{{d=JSON.parse(localStorage.getItem(k)||'null')}}catch(e){{}}if(!(d&&d.factor)){{document.getElementById('status').textContent='第一次使用測驗工具，請先完成 5 cm 螢幕校正。';setTimeout(()=>location.replace('/cloud/calibration?next=professional-test'),900)}}else{{document.getElementById('status').textContent='✅ 本裝置已完成 5 cm 校正。請使用「完整視覺功能測驗」，系統會建立全新測驗流程。';document.getElementById('status').style.background='#eaf8ef';document.getElementById('status').style.borderColor='#b7e4c7';document.getElementById('tools').style.display='grid'}}</script></body></html>"""

    def public_professional_teaching_html(self) -> str:
        return """<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>專業教學｜Cloud Vision</title><style>*{box-sizing:border-box}body{margin:0;background:#f5f8fc;color:#172033;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif}.wrap{max-width:900px;margin:auto;padding:28px 16px 70px}.card{background:#fff;border:1px solid #dce5f0;border-radius:20px;padding:25px;box-shadow:0 6px 22px rgba(29,51,84,.07)}.grid{display:grid;grid-template-columns:1fr 1fr;gap:13px}.tool{display:block;text-decoration:none;color:#172033;border:1px solid #dce5f0;border-radius:15px;padding:20px;background:#f9fbfd;font-weight:900}.tool span{display:block;font-weight:500;color:#64748b;margin-top:7px;line-height:1.55}.back{display:block;text-align:center;margin-top:20px;color:#244f91;text-decoration:none;font-weight:800}@media(max-width:620px){.grid{grid-template-columns:1fr}}</style></head><body><main class='wrap'><section class='card'><h1>專業教學</h1><p>教學區只說明原理、流程與注意事項；需要實際操作時，再由教學頁前往測驗工具。</p><div class='grid'><a class='tool' href='/cloud/teaching/calibration'>📏 螢幕尺度校正教學<span>說明為何使用 5 cm 實體尺校正，以及校正後如何套用。</span></a><a class='tool' href='/cloud/teaching/size'>📐 視標大小與距離換算<span>介紹觀看距離、字體大小與視角之間的關係。</span></a><a class='tool' href='/cloud/tools/geometry'>📊 視力幾何教學中心<span>完整 A／B／C 計算、MAR、CPD、cycles/letter 與互動判讀。</span></a><a class='tool' href='/cloud/teaching/acuity'>🔤 視力測驗教學<span>說明遮眼、觀看距離、視標呈現與記錄原則。</span></a><a class='tool' href='/cloud/teaching/astigmatism'>✳️ 散光鐘教學<span>說明散光鐘觀察方式與「線條是否同樣深」的紀錄。</span></a><a class='tool' href='/cloud/teaching/amsler'>▦ 黃斑部方格教學<span>說明注視中央點及格線扭曲、缺損的觀察方式。</span></a><a class='tool' href='/cloud/disclaimer'>📄 平台說明與免責聲明<span>查看研究資訊、適用範圍與使用注意事項。</span></a></div><a class='back' href='/cloud/professional/hub'>返回測驗／教學選擇</a></section></main></body></html>"""

    def public_size_tool_html(self) -> str:
        """實際可操作的視標大小與距離換算工具，歸在「測驗」區。"""
        return r"""<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>視標大小與距離換算｜Cloud Vision</title><style>
*{box-sizing:border-box}body{margin:0;background:#f5f8fc;color:#172033;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif}.wrap{max-width:980px;margin:auto;padding:24px 16px 70px}.hero,.card{background:#fff;border:1px solid #dce5f0;border-radius:20px;padding:24px;box-shadow:0 6px 22px rgba(29,51,84,.07);margin-bottom:16px}h1{margin:0 0 8px;font-size:31px}h2{margin:0 0 14px;font-size:23px}.sub{color:#64748b;line-height:1.7}.grid{display:grid;grid-template-columns:1fr 1fr;gap:16px}.fields{display:grid;grid-template-columns:1fr 1fr;gap:12px}label{display:block;font-weight:850;margin-bottom:6px}input,select{width:100%;padding:13px;border:1px solid #b8c6d8;border-radius:11px;font-size:18px;background:#fff}.btn{border:0;border-radius:12px;background:#1769e0;color:#fff;padding:14px 16px;font-size:18px;font-weight:900;cursor:pointer;width:100%;margin-top:14px}.secondary{background:#14823b}.result{margin-top:15px;background:#eef6ff;border:1px solid #bed8ff;border-radius:14px;padding:17px}.big{font-size:30px;font-weight:950;color:#124f9f}.formula{font-family:ui-monospace,SFMono-Regular,Consolas,monospace;background:#f7f9fc;border-radius:10px;padding:10px;margin-top:10px;overflow:auto}.presets{display:flex;flex-wrap:wrap;gap:8px;margin:10px 0}.chip{border:1px solid #aac2df;background:#fff;border-radius:999px;padding:8px 12px;font-weight:800;cursor:pointer}.previewWrap{text-align:center}.preview{display:flex;align-items:center;justify-content:center;min-height:210px;border:1px dashed #9fb3c9;border-radius:15px;background:#fff;overflow:auto}.letter{font-family:Arial,sans-serif;font-weight:900;line-height:1;color:#111}.status{margin-top:10px;color:#64748b;line-height:1.6}.table{width:100%;border-collapse:collapse;margin-top:12px}.table th,.table td{padding:10px;border-bottom:1px solid #e0e7ef;text-align:center}.back{display:block;text-align:center;margin-top:18px;color:#244f91;text-decoration:none;font-weight:850}.warn{background:#fff7df;border:1px solid #f1d58a;border-radius:12px;padding:12px;line-height:1.65;margin-top:14px}@media(max-width:760px){.grid,.fields{grid-template-columns:1fr}.wrap{padding-top:14px}.hero,.card{padding:19px}.big{font-size:26px}}
</style></head><body><main class='wrap'><section class='hero'><h1>視標大小與距離換算</h1><p class='sub'>這一頁不是只有文字說明。可直接換算不同觀看距離下的視標實體大小，也可依小數視力計算標準 5 分弧視標高度。</p></section>
<section class='grid'><div class='card'><h2>① 已知尺寸，換算另一個距離</h2><div class='fields'><div><label>原始觀看距離（cm）</label><input id='d1' type='number' step='0.01' value='57'></div><div><label>原始視標高度（cm）</label><input id='h1' type='number' step='0.001' value='0.50'></div><div><label>目標觀看距離（cm）</label><input id='d2' type='number' step='0.01' value='100'></div><div><label>保留小數位</label><select id='digits'><option>2</option><option selected>3</option><option>4</option></select></div></div><div class='presets'><button class='chip' onclick='setD2(30)'>30 cm</button><button class='chip' onclick='setD2(40)'>40 cm</button><button class='chip' onclick='setD2(57)'>57 cm</button><button class='chip' onclick='setD2(100)'>1 m</button><button class='chip' onclick='setD2(300)'>3 m</button><button class='chip' onclick='setD2(600)'>6 m</button></div><button class='btn' onclick='convertRatio()'>計算目標視標大小</button><div class='result'><div>目標視標高度</div><div id='ratioResult' class='big'>— cm</div><div class='formula'>目標高度 = 原始高度 × 目標距離 ÷ 原始距離</div></div></div>
<div class='card'><h2>② 依視力等級計算視標高度</h2><div class='fields'><div><label>小數視力（VA）</label><input id='va' type='number' step='0.01' value='1.0'></div><div><label>觀看距離（cm）</label><input id='distance' type='number' step='0.01' value='57'></div></div><div class='presets'><button class='chip' onclick='setVA(0.1)'>0.1</button><button class='chip' onclick='setVA(0.2)'>0.2</button><button class='chip' onclick='setVA(0.5)'>0.5</button><button class='chip' onclick='setVA(1.0)'>1.0</button><button class='chip' onclick='setVA(1.5)'>1.5</button></div><button class='btn secondary' onclick='calcVA()'>計算標準視標高度</button><div class='result'><div>5 分弧視標外框高度</div><div id='vaResult' class='big'>— cm</div><div id='vaDetail' class='status'></div></div><div class='warn'>這裡以標準視標整體高度約為 5 × MAR 分弧計算。實際 Sloan、Landolt C、Tumbling E 的內部幾何仍需依各視標規格製作。</div></div></section>
<section class='card previewWrap'><h2>③ 校正後螢幕預覽</h2><p class='sub'>先在本裝置完成 5 cm 螢幕校正，預覽才會依校正係數顯示接近實際尺寸。</p><div class='fields' style='max-width:620px;margin:0 auto 12px'><div><label>預覽高度（cm）</label><input id='previewCm' type='number' step='0.001' value='0.50'></div><div><label>預覽字母</label><select id='previewChar'><option>C</option><option>E</option><option selected>S</option><option>K</option><option>Z</option></select></div></div><button class='btn' style='max-width:620px' onclick='renderPreview()'>顯示實際尺寸預覽</button><div class='preview'><div id='letter' class='letter'>S</div></div><div id='calStatus' class='status'></div><a class='back' href='/cloud/calibration?next=professional-test'>前往測驗區的 5 cm 螢幕校正</a></section>
<section class='card'><h2>常用距離快速表</h2><table class='table'><thead><tr><th>觀看距離</th><th>VA 1.0 高度</th><th>VA 0.5 高度</th><th>VA 0.1 高度</th></tr></thead><tbody id='quickTable'></tbody></table><a class='back' href='/cloud/professional/test'>返回專業測驗</a></section></main><script>
const $=id=>document.getElementById(id);function n(id){return parseFloat($(id).value)}function valid(...v){return v.every(x=>Number.isFinite(x)&&x>0)}function setD2(v){$('d2').value=v;convertRatio()}function setVA(v){$('va').value=v;calcVA()}function convertRatio(){const a=n('d1'),b=n('h1'),c=n('d2'),dg=parseInt($('digits').value||3);if(!valid(a,b,c)){$('ratioResult').textContent='請輸入大於 0 的數值';return}const r=b*c/a;$('ratioResult').textContent=r.toFixed(dg)+' cm';$('previewCm').value=r.toFixed(dg)}function optotypeCm(va,dist){const arcmin=5/va;const rad=(arcmin/60)*Math.PI/180;return 2*dist*Math.tan(rad/2)}function calcVA(){const va=n('va'),d=n('distance');if(!valid(va,d)){$('vaResult').textContent='請輸入大於 0 的數值';return}const cm=optotypeCm(va,d);$('vaResult').textContent=cm.toFixed(4)+' cm';$('vaDetail').textContent='約 '+(cm*10).toFixed(3)+' mm；視標角度約 '+(5/va).toFixed(3)+' 分弧。';$('previewCm').value=cm.toFixed(4)}function renderPreview(){const cm=n('previewCm'),ch=$('previewChar').value;if(!valid(cm))return;let cal=null;try{cal=JSON.parse(localStorage.getItem('cloudVisionCalibrationV10')||'null')}catch(e){}const basePxPerCm=96/2.54;const factor=(cal&&Number(cal.factor)>0)?Number(cal.factor):1;const px=cm*basePxPerCm*factor;$('letter').textContent=ch;$('letter').style.fontSize=px+'px';$('letter').style.width=px+'px';$('letter').style.height=px+'px';$('calStatus').textContent=(cal&&cal.factor)?'✅ 已套用本裝置校正係數：'+factor.toFixed(5)+'；預覽外框約 '+cm+' cm。':'⚠️ 尚未找到本裝置校正資料，目前使用瀏覽器預設 96 DPI 估算，實際尺寸可能不準。'}function buildTable(){const ds=[30,40,57,100,300,600];$('quickTable').innerHTML=ds.map(d=>'<tr><td>'+d+' cm</td><td>'+optotypeCm(1,d).toFixed(3)+' cm</td><td>'+optotypeCm(.5,d).toFixed(3)+' cm</td><td>'+optotypeCm(.1,d).toFixed(3)+' cm</td></tr>').join('')}convertRatio();calcVA();renderPreview();buildTable();</script></body></html>"""

    def public_teaching_topic_html(self, topic: str) -> str:
        """教學區只負責原理、步驟與注意事項，並連往測驗區的實際工具。"""
        topics = {
            'calibration': {
                'icon': '📏',
                'title': '螢幕尺度校正教學',
                'intro': '不同裝置的像素密度與瀏覽器縮放比例不同，因此同一條線在不同螢幕上不一定具有相同的實際長度。',
                'steps': [
                    '準備一把實體直尺，將尺靠近螢幕但不要壓迫面板。',
                    '測量畫面上的黑線目前實際長度。',
                    '輸入實測值，系統會計算校正係數並重新調整。',
                    '再次量測，確認黑線實際長度為 5.00 cm 後儲存。',
                    '更換裝置、瀏覽器縮放比例或螢幕設定後，應重新校正。'
                ],
                'note': '校正只影響這個瀏覽器與裝置。教學頁本身不需要校正；只有實際呈現或測量視標時才需要。',
                'tool': '/cloud/calibration?next=professional-test',
                'tool_label': '前往測驗區：5 cm 螢幕校正'
            },
            'size': {
                'icon': '📐',
                'title': '視標大小與距離換算教學',
                'intro': '當觀看距離改變時，為維持相同視角，視標的實際大小也必須依距離同比例改變。',
                'steps': [
                    '先確認原始觀看距離與原始視標高度。',
                    '新的視標高度可依「原始高度 × 新距離 ÷ 原始距離」換算。',
                    '小數視力愈低，對應視標所需的視角與外框高度愈大。',
                    '實際顯示前仍須完成 5 cm 螢幕校正，否則螢幕上的公分尺寸可能不準。',
                    'Sloan、Landolt C 與 Tumbling E 雖可具有相同外框視角，但內部幾何規格不同。'
                ],
                'note': '教學區說明原理；數字輸入、計算、快速表與實際尺寸預覽集中放在測驗工具中。',
                'tool': '/cloud/tools/size',
                'tool_label': '前往測驗區：視標大小與距離換算工具'
            },
            'acuity': {
                'icon': '🔤',
                'title': '視力測驗教學',
                'intro': '視力測驗需要固定觀看距離、正確遮眼與一致的視標呈現，才能讓結果具有可比較性。',
                'steps': [
                    '先完成螢幕校正，並依畫面要求保持指定觀看距離。',
                    '左右眼分開測量；遮眼時不要壓迫眼球，也不要從遮眼縫隙偷看。',
                    '由較大的視標開始，逐步進入較小視標。',
                    '記錄使用的視標類型、觀看距離、裝置與校正狀態。',
                    '自我測量只能作為初步觀察，不能取代完整驗光或眼科測驗。'
                ],
                'note': '測量環境光、螢幕亮度、眼鏡配戴狀況與疲勞都可能影響結果。',
                'tool': '/cloud/professional/test',
                'tool_label': '前往測驗區：視力測驗'
            },
            'astigmatism': {
                'icon': '✳️',
                'title': '散光鐘教學',
                'intro': '散光鐘用來觀察不同方向線條是否呈現相同深淺，屬於初步觀察工具。',
                'steps': [
                    '左右眼分開觀察，另一眼需適當遮蔽。',
                    '注視線條共同中心，不要沿著某一條線追看。',
                    '比較各方向線條是否同樣清楚、同樣深。',
                    '簡化紀錄為「全部一樣深」或「深淺不一樣」。',
                    '深淺不同不等於可直接推算散光軸度或度數。'
                ],
                'note': '結果異常或視覺不適時，仍需由眼科醫師或驗光人員進一步評估。',
                'tool': '/cloud/test/astigmatism',
                'tool_label': '前往測驗區：散光鐘'
            },
            'amsler': {
                'icon': '▦',
                'title': '黃斑部 Amsler 方格教學',
                'intro': 'Amsler 方格用來觀察中央視野中的格線是否筆直、完整，並留意扭曲、模糊或缺損。',
                'steps': [
                    '配戴平常閱讀用眼鏡，保持建議的近距離。',
                    '左右眼分開測量，始終注視中央黑點。',
                    '不要移動視線逐格搜尋，只用周邊感覺觀察整體格線。',
                    '紀錄格線是否筆直，以及有無彎曲、缺格、模糊或黑影。',
                    '若突然出現異常，應儘速尋求眼科評估。'
                ],
                'note': 'Amsler 方格不能排除所有黃斑部疾病，正常結果也不代表眼底一定正常。',
                'tool': '/cloud/test/amsler',
                'tool_label': '前往測驗區：黃斑部 Amsler 方格'
            }
        }
        d = topics.get(topic)
        if not d:
            d = {'icon':'📚','title':'教學內容','intro':'此教學內容正在整理。','steps':['請返回專業教學選擇其他主題。'],'note':'','tool':'/cloud/professional/test','tool_label':'前往專業測驗'}
        items = ''.join(f"<li>{html.escape(x)}</li>" for x in d['steps'])
        note = f"<div class='note'>{html.escape(d['note'])}</div>" if d.get('note') else ''
        return f"""<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>{html.escape(d['title'])}｜Cloud Vision</title><style>
*{{box-sizing:border-box}}body{{margin:0;background:#f5f8fc;color:#172033;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif}}.wrap{{max-width:820px;margin:auto;padding:28px 16px 70px}}.card{{background:#fff;border:1px solid #dce5f0;border-radius:20px;padding:28px;box-shadow:0 6px 22px rgba(29,51,84,.07)}}.icon{{font-size:44px}}h1{{margin:8px 0 10px}}.intro{{font-size:18px;line-height:1.85;color:#52647b}}h2{{font-size:21px;margin-top:24px}}li{{font-size:17px;line-height:1.8;margin:7px 0}}.note{{background:#fff7df;border:1px solid #f1d58a;border-radius:13px;padding:14px;line-height:1.7;margin-top:20px}}.tool{{display:block;text-align:center;text-decoration:none;background:#1769e0;color:#fff;border-radius:14px;padding:15px;font-size:18px;font-weight:900;margin-top:22px}}.back{{display:block;text-align:center;margin-top:18px;color:#244f91;text-decoration:none;font-weight:850}}
</style></head><body><main class='wrap'><section class='card'><div class='icon'>{d['icon']}</div><h1>{html.escape(d['title'])}</h1><p class='intro'>{html.escape(d['intro'])}</p><h2>教學重點</h2><ol>{items}</ol>{note}<a class='tool' href='{d['tool']}'>{html.escape(d['tool_label'])}</a><a class='back' href='/cloud/professional/teaching'>返回專業教學</a></section></main></body></html>"""

    def public_astigmatism_test_html(self) -> str:
        """專業／教學共用散光鐘：完整 19 條線、10°～180°標示與倍率調整。"""
        return r"""<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>散光鐘測驗｜Cloud Vision</title><style>
*{box-sizing:border-box}body{margin:0;background:#f4f7fb;color:#172033;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif}.wrap{max-width:1120px;margin:auto;padding:22px 14px 70px}.card{background:#fff;border:1px solid #dbe5f0;border-radius:20px;padding:24px;box-shadow:0 7px 24px rgba(29,51,84,.08)}h1{margin:0 0 10px}.notice{background:#fff7df;border:1px solid #efd58c;border-radius:13px;padding:13px;line-height:1.65;margin:14px 0}.controls,.answers,.nav{display:flex;gap:10px;flex-wrap:wrap;justify-content:center;align-items:center}.scaleBtn,.answer,.btn{border:0;border-radius:13px;padding:12px 17px;font-size:17px;font-weight:900;cursor:pointer}.scaleBtn{background:#e8eef7;color:#244f91}.scaleBtn.active{background:#1769e0;color:#fff}.readout{font-weight:900;min-width:145px;text-align:center;color:#244f91}.dialWrap{display:flex;justify-content:center;align-items:center;min-height:570px;overflow:auto}.dialStage{width:900px;height:560px;transform-origin:center bottom;overflow:visible}.dialSvg{width:100%;height:100%;display:block;overflow:visible}.answer{flex:1;min-width:260px;background:#f4f7fb;border:2px solid #cfdaea}.answer.selected{background:#e8f6ed;border-color:#159447;color:#0e6c34}.btn{background:#1769e0;color:#fff;text-decoration:none}.btn.green{background:#14823b}.nav{margin-top:20px}.back{display:block;text-align:center;margin-top:18px;color:#244f91;text-decoration:none;font-weight:850}@media(max-width:760px){.dialStage{width:720px;height:470px}.dialWrap{min-height:490px}}@media(max-width:560px){.dialStage{width:580px;height:390px}.dialWrap{min-height:410px}.card{padding:17px}}
</style></head><body><main class='wrap'><section class='card'><h1>散光鐘自我觀察</h1><div class='notice'>請戴平常使用的眼鏡，在約 <b>30 公分</b>距離，遮住一眼並注視下方中央點。完整散光鐘共有 <b>19 條線</b>，每隔 10° 一條，標示 10°～180°。比較各方向線條是否同樣深、同樣清楚。</div><div class='controls'><button class='scaleBtn' onclick='setScale(.75)'>75%</button><button class='scaleBtn active' onclick='setScale(1)'>100%</button><button class='scaleBtn' onclick='setScale(1.25)'>125%</button><button class='scaleBtn' onclick='setScale(1.5)'>150%</button><button class='scaleBtn' onclick='adjust(-.025)'>－微縮</button><div id='readout' class='readout'>倍率 100%</div><button class='scaleBtn' onclick='adjust(.025)'>＋微放</button></div><div class='dialWrap'><div id='stage' class='dialStage'><svg id='dial' class='dialSvg' viewBox='0 0 900 560' aria-label='完整 19 線散光鐘'></svg></div></div><div class='answers'><button class='answer' onclick='pick(this)'>所有線條看起來一樣</button><button class='answer' onclick='pick(this)'>有些方向較深或較清楚</button></div><div class='nav'><a class='btn green' href='/cloud/test/amsler'>下一個：黃斑部 Amsler 方格</a></div><a class='back' href='/cloud/professional/test'>返回專業測驗</a></section></main><script>
const KEY='cloudVisionDialScaleV20';let scale=parseFloat(localStorage.getItem(KEY)||'1')||1;scale=Math.max(.6,Math.min(1.8,scale));
function build(){const svg=document.getElementById('dial'),NS='http://www.w3.org/2000/svg';svg.innerHTML='';const cx=450,cy=500,r=330,labelR=376;for(let theta=180;theta>=0;theta-=10){const rad=theta*Math.PI/180;const x=cx+r*Math.cos(rad),y=cy-r*Math.sin(rad);const ln=document.createElementNS(NS,'line');ln.setAttribute('x1',cx);ln.setAttribute('y1',cy);ln.setAttribute('x2',x);ln.setAttribute('y2',y);ln.setAttribute('stroke','#111');ln.setAttribute('stroke-width','2.3');ln.setAttribute('stroke-linecap','round');svg.appendChild(ln);let label;if(theta===90)label=180;else if(theta>90)label=270-theta;else label=90-theta;const tx=cx+labelR*Math.cos(rad),ty=cy-labelR*Math.sin(rad)+8;const t=document.createElementNS(NS,'text');t.setAttribute('x',tx);t.setAttribute('y',ty);t.setAttribute('text-anchor','middle');t.setAttribute('dominant-baseline','middle');t.setAttribute('fill','#111');t.setAttribute('font-family','Arial, Microsoft JhengHei, sans-serif');t.setAttribute('font-size','24');t.setAttribute('font-weight','900');t.textContent=String(label);svg.appendChild(t)}const dot=document.createElementNS(NS,'circle');dot.setAttribute('cx',cx);dot.setAttribute('cy',cy);dot.setAttribute('r','9');dot.setAttribute('fill','#111');svg.appendChild(dot)}
function render(){document.getElementById('stage').style.transform='scale('+scale+')';document.getElementById('readout').textContent='倍率 '+Math.round(scale*100)+'%';document.querySelectorAll('.scaleBtn').forEach(b=>b.classList.toggle('active',b.textContent===Math.round(scale*100)+'%'));localStorage.setItem(KEY,String(scale))}function setScale(v){scale=v;render()}function adjust(v){scale=Math.max(.6,Math.min(1.8,scale+v));render()}function pick(b){document.querySelectorAll('.answer').forEach(x=>x.classList.remove('selected'));b.classList.add('selected')}build();render();
</script></body></html>"""

    def public_amsler_test_html(self) -> str:
        """專業／教學共用黃斑部 Amsler 方格；不再誤連到散光鐘。"""
        return r"""<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>黃斑部 Amsler 方格｜Cloud Vision</title><style>
*{box-sizing:border-box}body{margin:0;background:#f4f7fb;color:#172033;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif}.wrap{max-width:980px;margin:auto;padding:22px 14px 70px}.card{background:#fff;border:1px solid #dbe5f0;border-radius:20px;padding:24px;box-shadow:0 7px 24px rgba(29,51,84,.08)}h1{margin:0 0 10px}.notice{background:#fff7df;border:1px solid #efd58c;border-radius:13px;padding:14px;line-height:1.7;margin:14px 0}.controls,.answers,.nav{display:flex;gap:10px;flex-wrap:wrap;justify-content:center;align-items:center}.control,.answer,.btn{border:0;border-radius:13px;padding:12px 17px;font-size:17px;font-weight:900;cursor:pointer}.control{background:#e8eef7;color:#244f91}.readout{font-weight:900;min-width:240px;text-align:center;color:#244f91}.gridWrap{display:flex;justify-content:center;padding:18px;overflow:auto}.grid{width:378px;height:378px;flex:0 0 auto;background-color:#fff;background-image:linear-gradient(#111 1px,transparent 1px),linear-gradient(90deg,#111 1px,transparent 1px);background-size:5% 5%;border:2px solid #111;position:relative}.grid:after{content:'';position:absolute;left:50%;top:50%;width:11px;height:11px;border-radius:50%;background:#111;transform:translate(-50%,-50%)}.answer{flex:1;min-width:260px;background:#f4f7fb;border:2px solid #cfdaea}.answer.selected{background:#e8f6ed;border-color:#159447;color:#0e6c34}.btn{background:#1769e0;color:#fff;text-decoration:none}.btn.gray{background:#64748b}.nav{margin-top:20px}.back{display:block;text-align:center;margin-top:18px;color:#244f91;text-decoration:none;font-weight:850}
</style></head><body><main class='wrap'><section class='card'><h1>黃斑部 Amsler 方格自我觀察</h1><div class='notice'><b>測試距離：30 公分。</b><br><b>標準方格：寬度 10 公分、高度 10 公分。</b><br>請先以實體尺確認外框寬度與高度。若不是 10 公分，使用下方按鈕微調。遮住一眼後，持續注視中央黑點，觀察格線是否筆直、完整。</div><div class='controls'><button class='control' onclick='adjust(-.02)'>－縮小</button><div id='readout' class='readout'></div><button class='control' onclick='adjust(.02)'>＋放大</button><button class='control' onclick='resetGrid()'>恢復 10 公分</button></div><div class='gridWrap'><div id='grid' class='grid'></div></div><div class='answers'><button class='answer' onclick='pick(this)'>格線筆直且完整</button><button class='answer' onclick='pick(this)'>格線彎曲、模糊或缺損</button></div><div class='nav'><a class='btn gray' href='/cloud/test/astigmatism'>返回散光鐘</a></div><a class='back' href='/cloud/professional/test'>返回專業測驗</a></section></main><script>
const KEY='cloudVisionAmslerScaleV20',CAL='cloudVisionCalibrationV10';let scale=parseFloat(localStorage.getItem(KEY)||'1')||1,cal=1;try{const d=JSON.parse(localStorage.getItem(CAL)||'null');if(d&&d.factor)cal=Number(d.factor)||1}catch(e){}function render(){const px=378*cal*scale;const g=document.getElementById('grid');g.style.width=px+'px';g.style.height=px+'px';document.getElementById('readout').textContent='目標 10.0 × 10.0 cm｜目前倍率 '+Math.round(scale*100)+'%';localStorage.setItem(KEY,String(scale))}function adjust(v){scale=Math.max(.55,Math.min(1.65,scale+v));render()}function resetGrid(){scale=1;render()}function pick(b){document.querySelectorAll('.answer').forEach(x=>x.classList.remove('selected'));b.classList.add('selected')}render();
</script></body></html>"""

    def public_assessment_html(self) -> str:
        """一般與專業使用者共用的完整雙眼流程：右散光→左散光→右黃斑→左黃斑→結果；點選答案後自動進入下一項。"""
        return r"""<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>視覺功能自我測驗｜Cloud Vision</title>
<style>
*{box-sizing:border-box}body{margin:0;background:#f4f7fb;color:#172033;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif}.wrap{max-width:1040px;margin:auto;padding:22px 14px 70px}.card{background:#fff;border:1px solid #dbe5f0;border-radius:20px;padding:24px;box-shadow:0 7px 24px rgba(29,51,84,.08)}.hidden{display:none!important}h1,h2{margin-top:0}.step{color:#1769e0;font-weight:900;margin-bottom:8px}.eyeTag{display:inline-block;background:#1769e0;color:#fff;border-radius:999px;padding:8px 16px;font-size:18px;font-weight:900;margin:0 0 12px}.notice{background:#fff7df;border:1px solid #efd58c;border-radius:13px;padding:13px;line-height:1.6;margin:14px 0}.answers,.nav,.scaleControls{display:flex;gap:10px;flex-wrap:wrap}.answer,.btn,.scaleBtn{border:0;border-radius:13px;padding:13px 18px;font-size:17px;font-weight:900;cursor:pointer}.answer{flex:1;min-width:220px;background:#f4f7fb;border:2px solid #cfdaea;color:#172033}.answer.selected{background:#e8f6ed;border-color:#159447;color:#0e6c34}.btn{background:#1769e0;color:#fff}.btn.green{background:#14823b}.btn.gray{background:#64748b}.nav{justify-content:center;margin-top:22px}.dialWrap{display:flex;justify-content:center;align-items:center;min-height:510px;overflow:auto}.dialStage{position:relative;width:760px;height:500px;transform-origin:center bottom;overflow:visible}.dialSvg{width:100%;height:100%;display:block;overflow:visible}.dialControls{display:flex;justify-content:center;align-items:center;gap:9px;flex-wrap:wrap;margin:-4px 0 14px}.dialScaleBtn{border:0;border-radius:12px;padding:10px 14px;font-size:16px;font-weight:900;background:#e8eef7;color:#244f91;cursor:pointer}.dialScaleBtn.active{background:#1769e0;color:#fff}.dialReadout{font-weight:900;color:#244f91;min-width:145px;text-align:center}.amslerWrap{display:flex;justify-content:center;padding:16px;overflow:auto}.amsler{width:378px;height:378px;flex:0 0 auto;background-color:#fff;background-image:linear-gradient(#111 1px,transparent 1px),linear-gradient(90deg,#111 1px,transparent 1px);background-size:5% 5%;border:2px solid #111;position:relative}.amsler:after{content:'';position:absolute;left:50%;top:50%;width:11px;height:11px;border-radius:50%;background:#111;transform:translate(-50%,-50%)}.scaleControls{justify-content:center;align-items:center;margin:6px 0 12px}.scaleBtn{background:#e8eef7;color:#244f91;padding:10px 15px}.scaleReadout{font-weight:900;min-width:190px;text-align:center}.resultGrid{display:grid;grid-template-columns:1fr 1fr;gap:12px}.resultBox{background:#f7f9fc;border:1px solid #dbe5f0;border-radius:14px;padding:16px;line-height:1.7}.warn{background:#fff0f0;border-color:#efb2b2}.ok{background:#ecf8f0;border-color:#b8e3c5}.small{font-size:14px;color:#64748b;line-height:1.6}.survey{margin-top:20px;border-top:1px solid #dbe5f0;padding-top:20px}.question{margin:18px 0}.questionTitle{font-size:18px;font-weight:900;margin-bottom:10px}.surveyChoices{display:flex;gap:10px;flex-wrap:wrap}.surveyChoice{border:2px solid #cfdaea;background:#f4f7fb;color:#172033;border-radius:12px;padding:12px 17px;font-size:16px;font-weight:850;cursor:pointer}.surveyChoice.selected{background:#e8f6ed;border-color:#159447;color:#0e6c34}.survey textarea{width:100%;min-height:105px;border:1px solid #aebed1;border-radius:12px;padding:12px;font-size:16px;resize:vertical}.submitStatus{text-align:center;margin-top:12px;font-weight:850}.home{display:block;text-align:center;margin-top:20px;color:#244f91;text-decoration:none;font-weight:900}@media(max-width:760px){.dialStage{width:620px;height:400px}.dialWrap{min-height:430px}}@media(max-width:650px){.resultGrid{grid-template-columns:1fr}.dialStage{width:520px;height:340px}.dialWrap{min-height:365px}.card{padding:18px}.answer{min-width:100%}}
</style></head><body><main class='wrap'>
<section id='astig' class='card'><div id='astigStep' class='step'></div><div id='astigEye' class='eyeTag'></div><h1>散光鐘自我觀察</h1><div class='notice'>請戴平常使用的眼鏡，在約 30 cm 距離，遮住另一眼並注視中央黑點。點選答案後會自動進入下一項：先完成右眼與左眼散光鐘，再進行右眼與左眼黃斑部。</div><div class='dialControls'><button class='dialScaleBtn' onclick='setDialScale(.75)'>75%</button><button class='dialScaleBtn' onclick='setDialScale(1)'>100%</button><button class='dialScaleBtn' onclick='setDialScale(1.25)'>125%</button><button class='dialScaleBtn' onclick='setDialScale(1.5)'>150%</button><button class='dialScaleBtn' onclick='adjustDialScale(-.025)'>－微縮</button><div id='dialReadout' class='dialReadout'>倍率 100%</div><button class='dialScaleBtn' onclick='adjustDialScale(.025)'>＋微放</button></div><div class='dialWrap'><div id='dialStage' class='dialStage'><svg id='dialSvg' class='dialSvg' viewBox='0 0 720 460'></svg></div></div><div class='answers'><button class='answer' data-value='所有線條看起來一樣' onclick='chooseAstig(this)'>所有線條看起來一樣</button><button class='answer' data-value='有些方向較深或較清楚' onclick='chooseAstig(this)'>有些方向較深或較清楚</button></div></section>
<section id='amsler' class='card hidden'><div id='amslerStep' class='step'></div><div id='amslerEye' class='eyeTag'></div><h1>黃斑部 Amsler 方格自我觀察</h1><div class='notice'><b>測驗距離：30 公分；方格標準尺寸：10 × 10 公分。</b><br>請遮住另一眼，持續注視中央黑點。點選答案後會自動進入下一眼；完成左眼後直接顯示測驗結果，不需要再按確定。</div><div class='scaleControls'><button class='scaleBtn' onclick='adjustAmsler(-0.02)'>－ 縮小</button><div id='scaleReadout' class='scaleReadout'></div><button class='scaleBtn' onclick='adjustAmsler(0.02)'>＋ 放大</button><button class='scaleBtn' onclick='resetAmsler()'>恢復 10 cm</button></div><div class='amslerWrap'><div id='amslerGrid' class='amsler'></div></div><div class='answers'><button class='answer' data-value='格線筆直且完整' onclick='chooseAmsler(this)'>格線筆直且完整</button><button class='answer' data-value='格線彎曲、模糊或缺損' onclick='chooseAmsler(this)'>格線彎曲、模糊或缺損</button></div><div class='nav'><button class='btn gray' onclick='backOneStep()'>返回上一項</button></div></section>
<section id='result' class='card hidden'><div class='step'>完成</div><h1>個人測驗結果</h1><p class='small'>以下為本次右眼、左眼、雙眼視力及左右眼完整紀錄，僅供參考。</p><div id='resultGrid' class='resultGrid'></div><div id='recommendation' class='notice'></div><section class='survey'><h2>使用問卷</h2><p class='small'>問卷不需要填寫姓名、電話或 Email。</p><div class='question'><div class='questionTitle'>1. 您對本平台的整體使用感受？</div><div class='surveyChoices' data-group='satisfaction'><button class='surveyChoice' data-value='滿意' onclick='chooseSurvey(this)'>滿意</button><button class='surveyChoice' data-value='普通' onclick='chooseSurvey(this)'>普通</button><button class='surveyChoice' data-value='不滿意' onclick='chooseSurvey(this)'>不滿意</button></div></div><div class='question'><div class='questionTitle'>2. 您覺得本平台需要改善嗎？</div><div class='surveyChoices' data-group='needs_improvement'><button class='surveyChoice' data-value='不需要改善' onclick='chooseSurvey(this)'>不需要改善</button><button class='surveyChoice' data-value='需要改善' onclick='chooseSurvey(this)'>需要改善</button></div></div><div class='question'><div class='questionTitle'>3. 改善建議（選填）</div><textarea id='improvementSuggestion' maxlength='500' placeholder='請寫下您的建議'></textarea></div><div class='nav'><button id='submitSurveyBtn' class='btn green' onclick='submitSurvey()'>送出問卷</button></div><div id='submitStatus' class='submitStatus'></div></section><div class='nav'><button class='btn' onclick='restart()'>重新測驗</button><button class='btn green' onclick="location.href='/cloud'">返回首頁</button></div></section></main>
<script>
const storeKey='cloudVisionPublicAssessmentV35',scaleKey='cloudVisionPublicAmslerScaleV19',dialScaleKey='cloudVisionPublicDialScaleV19',calKey='cloudVisionCalibrationV10';
let eyeVA={right_eye_va:'',left_eye_va:'',both_eyes_va:''};
try{eyeVA=JSON.parse(sessionStorage.getItem('cloudVisionEyeVA')||localStorage.getItem('cloudVisionEyeVA')||'{}')||eyeVA}catch(e){}
// 以網址參數為最高優先，避免 Safari 私密瀏覽或頁面切換時 storage 遺失。
try{
  const q=new URLSearchParams(location.search);
  if(q.get('right_eye_va'))eyeVA.right_eye_va=q.get('right_eye_va');
  if(q.get('left_eye_va'))eyeVA.left_eye_va=q.get('left_eye_va');
  if(q.get('both_eyes_va'))eyeVA.both_eyes_va=q.get('both_eyes_va');
}catch(e){}
let state={stage:'right_astig',astig:{right:'',left:''},amsler:{right:'',left:''},vision:{right:String(eyeVA.right_eye_va||''),left:String(eyeVA.left_eye_va||''),both:String(eyeVA.both_eyes_va||'')}};
// 每次由一般使用者或專業人員進入完整測驗，都從右眼散光鐘重新開始。
// 不讀取舊版流程狀態，避免上一輪答案造成跳步或混用左右眼。
try{localStorage.removeItem(storeKey)}catch(e){}
let amslerScale=parseFloat(localStorage.getItem(scaleKey)||'1')||1,dialScale=parseFloat(localStorage.getItem(dialScaleKey)||'1')||1;dialScale=Math.max(.60,Math.min(1.80,dialScale));
function save(){localStorage.setItem(storeKey,JSON.stringify(state))}
function currentEye(){return state.stage.startsWith('left')?'left':'right'}
function buildDial(){const svg=document.getElementById('dialSvg'),NS='http://www.w3.org/2000/svg';svg.setAttribute('viewBox','0 0 760 500');svg.innerHTML='';const cx=380,cy=445,r=300,labelR=342;for(let theta=180;theta>=0;theta-=10){const rad=theta*Math.PI/180,x=cx+r*Math.cos(rad),y=cy-r*Math.sin(rad),line=document.createElementNS(NS,'line');line.setAttribute('x1',cx);line.setAttribute('y1',cy);line.setAttribute('x2',x);line.setAttribute('y2',y);line.setAttribute('stroke','#111');line.setAttribute('stroke-width','2.4');svg.appendChild(line);let shown=theta===90?180:(theta>90?270-theta:90-theta);const tx=cx+labelR*Math.cos(rad),ty=cy-labelR*Math.sin(rad)+7,t=document.createElementNS(NS,'text');t.setAttribute('x',tx);t.setAttribute('y',ty);t.setAttribute('text-anchor','middle');t.setAttribute('font-size','21');t.setAttribute('font-weight','900');t.textContent=String(shown);svg.appendChild(t)}const c=document.createElementNS(NS,'circle');c.setAttribute('cx',cx);c.setAttribute('cy',cy);c.setAttribute('r','10');c.setAttribute('fill','#111');svg.appendChild(c);renderDialScale()}
function renderDialScale(){const d=document.getElementById('dialStage');d.style.transform='scale('+dialScale+')';d.parentElement.style.minHeight=Math.round(510*dialScale)+'px';document.getElementById('dialReadout').textContent='倍率 '+Math.round(dialScale*100)+'%'}function setDialScale(v){dialScale=Math.max(.60,Math.min(1.80,Number(v)||1));localStorage.setItem(dialScaleKey,String(dialScale));renderDialScale()}function adjustDialScale(v){setDialScale(dialScale+v)}
function calibrationFactor(){try{const d=JSON.parse(localStorage.getItem(calKey)||'null');return d&&d.factor?Number(d.factor):1}catch(e){return 1}}function renderAmsler(){const px=10*37.7952755906*calibrationFactor()*amslerScale,g=document.getElementById('amslerGrid');g.style.width=px+'px';g.style.height=px+'px';document.getElementById('scaleReadout').textContent='目標顯示：約 '+(10*amslerScale).toFixed(1)+' cm｜倍率 '+amslerScale.toFixed(2)}function adjustAmsler(d){amslerScale=Math.max(.60,Math.min(1.40,amslerScale+d));localStorage.setItem(scaleKey,String(amslerScale));renderAmsler()}function resetAmsler(){amslerScale=1;localStorage.setItem(scaleKey,'1');renderAmsler()}
function renderStage(){const eye=currentEye(),label=eye==='right'?'右眼':'左眼',isAstig=state.stage.endsWith('astig');document.getElementById('astig').classList.toggle('hidden',!isAstig);document.getElementById('amsler').classList.toggle('hidden',isAstig);document.getElementById('result').classList.add('hidden');if(isAstig){document.getElementById('astigEye').textContent=label;document.getElementById('astigStep').textContent=eye==='right'?'步驟 1／4':'步驟 2／4';document.querySelectorAll('#astig .answer').forEach(b=>b.classList.toggle('selected',b.dataset.value===state.astig[eye]))}else{document.getElementById('amslerEye').textContent=label;document.getElementById('amslerStep').textContent=eye==='right'?'步驟 3／4':'步驟 4／4';document.querySelectorAll('#amsler .answer').forEach(b=>b.classList.toggle('selected',b.dataset.value===state.amsler[eye]));renderAmsler()}scrollTo(0,0)}
const NEXT_STAGE={right_astig:'left_astig',left_astig:'right_amsler',right_amsler:'left_amsler'};
function clearAnswerButtons(sectionId){document.querySelectorAll('#'+sectionId+' .answer').forEach(x=>x.classList.remove('selected'))}
function chooseAstig(btn){
  const stage=state.stage,eye=stage==='right_astig'?'right':'left';
  if(stage!=='right_astig'&&stage!=='left_astig')return;
  state.astig[eye]=btn.dataset.value;
  document.querySelectorAll('#astig .answer').forEach(x=>x.classList.toggle('selected',x===btn));
  const next=NEXT_STAGE[stage];
  setTimeout(()=>{clearAnswerButtons('astig');state.stage=next;save();renderStage()},160);
}
function chooseAmsler(btn){
  const stage=state.stage,eye=stage==='right_amsler'?'right':'left';
  if(stage!=='right_amsler'&&stage!=='left_amsler')return;
  state.amsler[eye]=btn.dataset.value;
  document.querySelectorAll('#amsler .answer').forEach(x=>x.classList.toggle('selected',x===btn));
  if(stage==='right_amsler'){
    setTimeout(()=>{clearAnswerButtons('amsler');state.stage='left_amsler';save();renderStage()},160);
  }else{
    setTimeout(()=>{clearAnswerButtons('amsler');save();showResult()},160);
  }
}
function backOneStep(){const prev={left_astig:'right_astig',right_amsler:'left_astig',left_amsler:'right_amsler'};state.stage=prev[state.stage]||'right_astig';save();renderStage()}
let surveyState={satisfaction:'',needs_improvement:''};
function showResult(){
  if(!(state.astig.right&&state.astig.left&&state.amsler.right&&state.amsler.left))return;
  document.getElementById('astig').classList.add('hidden');document.getElementById('amsler').classList.add('hidden');document.getElementById('result').classList.remove('hidden');
  const rows=[['右眼視力',state.vision.right?Number(state.vision.right).toFixed(2):'未記錄',false],['左眼視力',state.vision.left?Number(state.vision.left).toFixed(2):'未記錄',false],['雙眼視力',state.vision.both?Number(state.vision.both).toFixed(2):'未記錄',false],['右眼散光鐘',state.astig.right,state.astig.right.includes('有些')],['左眼散光鐘',state.astig.left,state.astig.left.includes('有些')],['右眼黃斑部',state.amsler.right,state.amsler.right.includes('彎曲')],['左眼黃斑部',state.amsler.left,state.amsler.left.includes('彎曲')]],bad=rows.slice(3).some(r=>r[2]);
  document.getElementById('resultGrid').innerHTML=rows.map(r=>`<div class="resultBox ${r[2]?'warn':'ok'}"><strong>${r[0]}</strong><br>${r[1]}</div>`).join('');
  document.getElementById('recommendation').textContent=bad?'本次有項目呈現不一致、彎曲、模糊或缺損。結果僅供參考。':'本次左右眼自我觀察未發現明顯異常，結果僅供參考。';scrollTo(0,0);saveCompletedResult()
}
function resultPayload(includeSurvey=false){let visitor='';try{visitor=localStorage.getItem('cloudVisionVisitorV1019')||''}catch(e){}let pro=null;try{pro=JSON.parse(localStorage.getItem('cloudVisionProfessionalV2')||'null')}catch(e){}const payload={consent:'yes',user_type:(sessionStorage.getItem('cloudVisionCurrentUserType')||((pro&&pro.name)?'專業人員':'一般使用者')),professional_name:(pro&&pro.name)||'',professional_role:(pro&&pro.role)||'',professional_purpose:(pro&&pro.purpose)||'',visual_acuity_right:state.vision.right,visual_acuity_left:state.vision.left,visual_acuity_both:state.vision.both,astigmatism_right:state.astig.right,astigmatism_left:state.astig.left,amsler_right:state.amsler.right,amsler_left:state.amsler.left,visitor_id:visitor,session_id:sessionStorage.getItem('cloudVisionCurrentSession')||((pro&&pro.session_id)||'')};if(includeSurvey){payload.satisfaction=surveyState.satisfaction;payload.needs_improvement=surveyState.needs_improvement;payload.improvement_suggestion=document.getElementById('improvementSuggestion').value.trim()}return payload}
let completionSaved=false;async function saveCompletedResult(){if(completionSaved)return;try{const r=await fetch('/cloud/result',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(resultPayload(false)),keepalive:true}),data=await r.json();if(!r.ok||!data.ok)throw new Error(data.error||'完成紀錄送出失敗');completionSaved=true;sessionStorage.setItem('cloudVisionCompletionSaved','1')}catch(e){setTimeout(saveCompletedResult,1500)}}
function chooseSurvey(btn){const group=btn.parentElement.dataset.group;surveyState[group]=btn.dataset.value;btn.parentElement.querySelectorAll('.surveyChoice').forEach(x=>x.classList.toggle('selected',x===btn))}
async function submitSurvey(){if(!surveyState.satisfaction){alert('請選擇整體使用感受。');return}if(!surveyState.needs_improvement){alert('請選擇是否需要改善。');return}const btn=document.getElementById('submitSurveyBtn'),status=document.getElementById('submitStatus');btn.disabled=true;status.textContent='正在送出問卷…';try{const r=await fetch('/cloud/result',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(resultPayload(true))}),data=await r.json();if(!r.ok||!data.ok)throw new Error(data.error||'送出失敗');completionSaved=true;status.textContent='✅ 感謝您的回饋，問卷已送出。';btn.textContent='問卷已送出';document.querySelectorAll('.surveyChoice,#improvementSuggestion').forEach(x=>x.disabled=true)}catch(e){btn.disabled=false;status.textContent='⚠️ 問卷未送出，請再試一次。'}}
function restart(){localStorage.removeItem(storeKey);location.reload()}buildDial();renderAmsler();renderStage();
</script></body></html>"""

    def public_geometry_center_html(self) -> str:
        """完整瀏覽器版視力幾何教學中心，將 A/B/C、結果卡、圖解、進度與互動表格放在同一頁。"""
        return r"""<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>視力幾何教學中心（完整教學模式）｜Cloud Vision</title><style>
*{box-sizing:border-box}body{margin:0;background:#f2f5f8;color:#172033;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif}.wrap{max-width:1540px;margin:auto;padding:14px}.header{display:flex;justify-content:space-between;gap:16px;align-items:flex-start;background:#fff;border:1px solid #d9e1ea;padding:18px 22px}.title h1{margin:0;font-size:32px}.title small{font-size:14px;color:#64748b}.subtitle{margin-top:9px;font-weight:700}.toolbar{display:flex;gap:9px;flex-wrap:wrap}.toolbtn,.btn{border:1px solid #aeb9c7;background:#fff;border-radius:5px;padding:10px 14px;font-weight:800;cursor:pointer}.layout{display:grid;grid-template-columns:1.48fr 1fr;gap:14px;margin-top:12px}.left,.right{display:grid;gap:12px}.abc{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}.panel{background:#fff;border:1px solid #d6dfe9;padding:15px}.panel h2{margin:0 0 12px;font-size:19px}.row{display:grid;grid-template-columns:1fr 1fr;gap:8px}.field label{display:block;font-weight:800;font-size:14px;margin-bottom:5px}.field input,.field select{width:100%;padding:9px;border:1px solid #aebaca;border-radius:3px;font-size:16px}.btn{background:#f4f5f7}.btn.primary{background:#1769e0;color:#fff;border-color:#1769e0}.result{min-height:43px;padding:9px;margin-top:9px;border:1px solid #dbe3ec;background:#fafcfe;font-weight:800}.status{background:#517d16;color:#fff;text-align:center;font-size:25px;font-weight:900;padding:11px}.cards{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-top:10px}.infoCard{border:1px solid #dbe3ec;padding:13px;min-height:190px}.infoCard h3{margin:0 0 8px}.infoGrid{display:grid;grid-template-columns:1fr 1fr;gap:8px;font-size:14px}.teach{line-height:1.55;font-size:14px}.progressTitle{font-weight:900;margin-top:12px}.progress{height:25px;background:#dbe4ef;border:1px solid #a9b7c7;margin:7px 0}.progress i{display:block;height:100%;background:#6f9f25}.flow{padding:9px 0;font-weight:800}.tableWrap{overflow:auto;border:1px solid #d6dfe9}.data{width:100%;border-collapse:collapse}.data th,.data td{border-bottom:1px solid #dde4ec;padding:7px;text-align:center;white-space:nowrap}.data th{background:#eef2f6}.diagramPanel{background:#fff;border:1px solid #d6dfe9;padding:12px}.diagramTitle{font-weight:900;font-size:18px;margin-bottom:6px}.diagram{position:relative;height:340px;border:1px solid #dbe3ec;background:#fff;overflow:hidden}.eye{position:absolute;left:40px;top:145px;width:44px;height:44px;border:5px solid #111;border-radius:50%}.eye:after{content:'';position:absolute;width:10px;height:10px;background:#111;border-radius:50%;left:12px;top:12px}.letter{position:absolute;right:32px;top:104px;font-size:105px;font-weight:900}.ray{position:absolute;left:84px;top:166px;width:calc(100% - 190px);height:2px;background:#5f7082;transform-origin:left}.distanceLabel{position:absolute;left:43%;top:132px;font-weight:800}.heightLabel{position:absolute;right:30px;top:221px;font-weight:800}.angleLabel{position:absolute;left:95px;top:128px;color:#1769e0;font-weight:900}.theory{margin-top:10px;border:1px solid #d6dfe9;padding:10px}.theory h3{margin:0 0 8px}.theoryStatus{background:#517d16;color:white;text-align:center;font-size:22px;font-weight:900;padding:10px}.sliders{display:grid;gap:8px;margin-top:12px}.sliderRow{display:grid;grid-template-columns:95px 1fr 86px;gap:8px;align-items:center}.quick{display:flex;gap:8px;flex-wrap:wrap;margin-top:9px}.quick button{padding:7px 12px;border:1px solid #aebaca;background:#fff;border-radius:4px;font-weight:800}.footerBtns{display:flex;gap:10px;flex-wrap:wrap;margin-top:12px}.back{display:inline-block;color:#244f91;font-weight:900;text-decoration:none;padding:12px}.muted{color:#64748b}.greenText{color:#397114;font-weight:900}@media(max-width:1100px){.layout{grid-template-columns:1fr}.abc{grid-template-columns:1fr}.cards{grid-template-columns:1fr}}@media(max-width:650px){.row,.infoGrid{grid-template-columns:1fr}.title h1{font-size:25px}.diagram{height:280px}}
</style></head><body><main class='wrap'><header class='header'><div class='title'><h1>視力幾何教學中心　Visual Geometry Teaching Center</h1><small>Teaching & Research Edition</small><div class='subtitle'>由字高、觀看距離與視力，連結完整字母視角、MAR、CPD、cycles/letter 與理論辨識結果。</div></div><div class='toolbar'><button class='toolbtn' onclick='loadExample()'>載入教學範例</button><button class='toolbtn' onclick='window.print()'>輸出報告</button><button class='toolbtn' onclick='toggleFull()'>最大化／還原</button></div></header>
<div class='layout'><section class='left'><div class='abc'><div class='panel'><h2>A｜字高＋視力 → 最遠辨識距離</h2><div class='row'><div class='field'><label>字母高度（cm）</label><input id='aH' type='number' step='.01' value='1.0'></div><div class='field'><label>十進位視力（decimal VA）</label><input id='aVA' type='number' step='.01' value='.05'></div></div><button class='btn' onclick='calcA()'>計算最遠距離</button><div id='aR' class='result'>輸入字高與十進位視力後，按「計算最遠距離」。</div></div><div class='panel'><h2>B｜字高＋距離 → 所需最低視力</h2><div class='row'><div class='field'><label>字母高度（cm）</label><input id='bH' type='number' step='.01' value='1.0'></div><div class='field'><label>觀看距離（cm）</label><input id='bD' type='number' step='1' value='55'></div></div><button class='btn' onclick='calcB()'>計算所需視力</button><div id='bR' class='result'>輸入字高與觀看距離後，按「計算所需視力」。</div></div><div class='panel'><h2>C｜字高＋距離＋受試者視力 → 整合判斷</h2><div class='row'><div class='field'><label>字高（cm）</label><input id='cH' type='number' step='.01' value='.5'></div><div class='field'><label>距離（cm）</label><input id='cD' type='number' step='1' value='57'></div><div class='field'><label>受試者視力（decimal VA）</label><input id='cVA' type='number' step='.01' value='1.0'></div></div><button class='btn primary' onclick='calcC()'>整合判斷</button><div id='cR' class='result'></div></div></div>
<div class='panel' style='border:2px solid #2f6b5f'><h2 style='margin:0 0 8px'>以下為完整專業分析功能</h2><p class='muted' style='margin-top:0'>C 整合判斷、教學資訊卡、達成度、判斷流程、字高連動表、兩張理論圖與互動滑桿完整保留。</p><div id='status' class='status'>YES｜清楚可辨識</div><div class='cards'><div class='infoCard'><h3>△ 幾何資訊</h3><div id='geoInfo' class='infoGrid'></div></div><div class='infoCard'><h3>◉ 視力資訊</h3><div id='visionInfo' class='infoGrid'></div></div><div class='infoCard'><h3>▣ 空間頻率資訊</h3><div id='freqInfo' class='teach'></div></div><div class='infoCard'><h3>⌖ 教學重點</h3><div id='teachInfo' class='teach'></div></div></div><div class='progressTitle'>2.5 cycles/letter 達成度</div><div class='progress'><i id='bar1'></i></div><div id='barText1' class='flow'></div><div class='progressTitle'>受試者視力需求達成度</div><div class='progress'><i id='bar2'></i></div><div id='barText2' class='flow'></div><div class='flow' id='flow'></div><h3>cutoff × 字體大小連動表</h3><div class='tableWrap'><table class='data'><thead><tr><th>字高 cm</th><th>視角 °</th><th>完整字母視角 arcmin</th><th>通過上限 cycles/letter</th><th>最低視力</th><th>視力條件</th><th>最終判斷</th></tr></thead><tbody id='tbody'></tbody></table></div><div id='visibleSummary' class='flow'></div><div class='footerBtns'><button class='btn' onclick='window.print()'>輸出視力幾何文字報告</button><a class='back' href='/cloud/professional/teaching'>返回專業教學</a></div></div></section>
<section class='right'><div class='diagramPanel'><div class='diagramTitle'>圖一｜觀看幾何完整字母視角</div><div class='diagram'><div class='eye'></div><div class='ray' id='ray1' style='transform:rotate(-2deg)'></div><div class='ray' id='ray2' style='transform:rotate(2deg)'></div><div class='letter'>E</div><div class='angleLabel'>視角 θ</div><div id='distanceLabel' class='distanceLabel'>觀看距離：57 cm</div><div id='heightLabel' class='heightLabel'>字高：0.50 cm</div></div><div class='theory'><h3>圖二｜理論辨識結果</h3><div id='theoryStatus' class='theoryStatus'>理論可辨識</div><div id='theoryText' style='padding:10px;text-align:center;font-weight:800'></div></div><h3>互動示範｜拖動滑桿，上方 C 區與兩張圖即時改變</h3><div class='sliders'><div class='sliderRow'><b>字高</b><input id='sH' type='range' min='.17' max='15' step='.01' value='.5' oninput='syncSlider()'><span id='sHv'>0.50 cm</span></div><div class='sliderRow'><b>觀看距離</b><input id='sD' type='range' min='30' max='688' step='1' value='57' oninput='syncSlider()'><span id='sDv'>57 cm</span></div><div class='sliderRow'><b>受試者視力</b><input id='sVA' type='range' min='.05' max='1.5' step='.01' value='1' oninput='syncSlider()'><span id='sVAv'>1.00 VA</span></div></div><div class='quick'><button onclick='setDistance(57)'>57 cm</button><button onclick='setDistance(114)'>114 cm</button><button onclick='setDistance(300)'>300 cm</button><button onclick='setDistance(688)'>688 cm</button></div><p class='muted'>右圖資料來源：C 區。推薦流程：字高（cm）＋距離（cm）→ 完整字母視角（arcmin）→ MAR／最低視力 → CPD／cycles per letter → 理論辨識判斷。</p></div></section></div></main>
<script>
const heights=[.17,.25,.5,1,2.5,5,10,15],cutoff=10,sloanCPL=2.5;function n(id){return Number(document.getElementById(id).value)||0}function geom(h,d){const deg=2*Math.atan((h/2)/d)*180/Math.PI,arc=deg*60,cpl=arc/6,minVA=5/arc,mar=arc/5;return{deg,arc,cpl,minVA,mar}}function calcA(){const h=n('aH'),va=Math.max(.001,n('aVA'));const arc=5/va,d=h/(2*Math.tan((arc/60)*Math.PI/180/2));document.getElementById('aR').innerHTML='<b>最遠辨識距離：約 '+d.toFixed(1)+' cm（'+(d/100).toFixed(3)+' m）</b>'}function calcB(){const g=geom(n('bH'),n('bD'));document.getElementById('bR').innerHTML='<b>最低所需視力：約 '+g.minVA.toFixed(3)+' VA</b><br>完整字母視角 '+g.arc.toFixed(3)+' arcmin｜MAR '+g.mar.toFixed(3)}function calcC(){const h=n('cH'),d=n('cD'),va=n('cVA'),g=geom(h,d),freqOK=g.cpl>=sloanCPL,visionOK=va>=g.minVA,ok=freqOK&&visionOK;document.getElementById('cR').innerHTML=(ok?'<b class="greenText">理論可辨識</b>':'<b style="color:#a93232">理論可能不足</b>')+'｜最低需求 '+g.minVA.toFixed(3)+' VA｜視力裕度 '+(va/g.minVA).toFixed(2)+'×';const status=document.getElementById('status');status.textContent=(ok?'YES｜清楚可辨識':'NO｜可能無法辨識');status.style.background=ok?'#517d16':'#a93232';document.getElementById('theoryStatus').textContent=ok?'理論可辨識':'理論可能不足';document.getElementById('theoryStatus').style.background=ok?'#517d16':'#a93232';document.getElementById('theoryText').textContent='受試者視力 '+va.toFixed(2)+' VA ｜ 最低需求 '+g.minVA.toFixed(3)+' VA ｜ 視力裕度 '+(va/g.minVA).toFixed(2)+'×';document.getElementById('geoInfo').innerHTML=`<div>字母高度</div><b>${h.toFixed(2)} cm</b><div>觀看距離</div><b>${d.toFixed(2)} cm</b><div>完整字母視角</div><b>${g.arc.toFixed(3)} arcmin</b><div>理論最遠辨識距離</div><b>${(h/(2*Math.tan((5/va/60)*Math.PI/180/2))/100).toFixed(3)} m</b>`;document.getElementById('visionInfo').innerHTML=`<div>受試者視力</div><b>${va.toFixed(3)} VA</b><div>最低所需視力</div><b>${g.minVA.toFixed(3)} VA</b><div>受試者 MAR</div><b>${(1/va).toFixed(3)} arcmin</b><div>所需 MAR</div><b>${g.mar.toFixed(3)} arcmin</b><div>視力裕度</div><b>${(va/g.minVA).toFixed(2)} 倍</b>`;document.getElementById('freqInfo').innerHTML=`Sloan 主要辨識尺度：${sloanCPL.toFixed(1)} cycles/letter<br>目前完整字母可提供：${g.cpl.toFixed(3)} cycles/letter<br>系統高頻上限：30.00 cpd<br>目前 cutoff：${cutoff.toFixed(1)} cpd` ;document.getElementById('teachInfo').innerHTML=`<b>${freqOK?'已通過':'未通過'}：</b>目前 ${g.cpl.toFixed(3)} cycles/letter ${freqOK?'≥':'<'} ${sloanCPL.toFixed(1)} cycles/letter。<br><br>字高增加 → 視角增加 → 所需視力降低。<br>距離增加 → 視角減少 → 所需視力提高。<br>視力愈好 → 視力裕度愈高。`;const p1=Math.min(100,g.cpl/sloanCPL*100),p2=Math.min(100,va/g.minVA*100);document.getElementById('bar1').style.width=p1+'%';document.getElementById('bar2').style.width=p2+'%';document.getElementById('barText1').textContent='影像資訊：'+g.cpl.toFixed(3)+' ÷ '+sloanCPL.toFixed(1)+' = '+(g.cpl/sloanCPL*100).toFixed(1)+'%｜'+(freqOK?'已通過主要辨識尺度':'尚未通過主要辨識尺度');document.getElementById('barText2').textContent='視力條件：'+va.toFixed(3)+' ÷ '+g.minVA.toFixed(3)+' = '+(va/g.minVA*100).toFixed(1)+'%｜'+(visionOK?'視力已足夠':'視力可能不足');document.getElementById('flow').textContent='cutoff '+cutoff+' cpd × 視角 '+g.deg.toFixed(4)+'° = '+(cutoff*g.deg).toFixed(3)+' cycles/letter → '+(ok?'YES｜主要辨識尺度已通過，可辨識':'NO｜主要辨識尺度或視力條件不足');document.getElementById('distanceLabel').textContent='觀看距離：'+d.toFixed(0)+' cm';document.getElementById('heightLabel').textContent='字高：'+h.toFixed(2)+' cm';const a=Math.min(10,Math.max(.5,g.deg*4));document.getElementById('ray1').style.transform='rotate(-'+a+'deg)';document.getElementById('ray2').style.transform='rotate('+a+'deg)';renderTable(d,va);}
function renderTable(d,va){let visible=[];document.getElementById('tbody').innerHTML=heights.map(h=>{const g=geom(h,d),freqOK=g.cpl>=sloanCPL,visionOK=va>=g.minVA,ok=freqOK&&visionOK;if(ok)visible.push(h);return `<tr><td>${h}</td><td>${g.deg.toFixed(4)}</td><td>${g.arc.toFixed(3)}</td><td>${g.cpl.toFixed(3)}</td><td>${g.minVA.toFixed(3)}</td><td>${visionOK?'☑ 足夠':'△ 不足'}</td><td>${ok?'☑ 可辨識':(freqOK?'△ 視力不足':'△ 模糊可用')}</td></tr>`}).join('');document.getElementById('visibleSummary').textContent='目前條件下：'+(visible.length?visible.join(' cm、')+' cm 看得到。':'沒有字高通過完整條件。')}
function syncSlider(){const h=n('sH'),d=n('sD'),va=n('sVA');document.getElementById('sHv').textContent=h.toFixed(2)+' cm';document.getElementById('sDv').textContent=d.toFixed(0)+' cm';document.getElementById('sVAv').textContent=va.toFixed(2)+' VA';document.getElementById('cH').value=h;document.getElementById('cD').value=d;document.getElementById('cVA').value=va;calcC()}function setDistance(v){document.getElementById('sD').value=v;syncSlider()}function loadExample(){document.getElementById('aH').value=1;document.getElementById('aVA').value=.05;document.getElementById('bH').value=1;document.getElementById('bD').value=55;document.getElementById('sH').value=.5;document.getElementById('sD').value=57;document.getElementById('sVA').value=1;calcA();calcB();syncSlider()}function toggleFull(){if(!document.fullscreenElement)document.documentElement.requestFullscreen?.();else document.exitFullscreen?.()}loadExample();
</script></body></html>"""


    def public_calibration_html(self) -> str:
        """獨立且可保存於瀏覽器的 5 cm 校正頁。"""
        return r"""<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
<title>5 cm 螢幕校正｜Cloud Vision</title>
<style>
*{box-sizing:border-box} body{margin:0;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","Noto Sans TC",sans-serif;background:#f6f8fb;color:#172033}.wrap{max-width:760px;margin:auto;padding:20px 16px 42px}h1{text-align:center;margin:8px 0}.tip{background:#eaf3ff;border:1px solid #bad4ff;border-radius:14px;padding:14px;line-height:1.65}.stage{margin:20px 0;background:white;border:1px solid #dce5f0;border-radius:18px;padding:30px 10px;text-align:center;overflow:hidden}.line{height:6px;background:#111;margin:25px auto;border-radius:3px;width:188.976px;max-width:95%}.label{font-weight:800;font-size:20px}.controls{display:grid;grid-template-columns:repeat(2,1fr);gap:10px}.btn{border:0;border-radius:12px;padding:14px;font-size:17px;font-weight:750;background:#e8eef7}.primary{background:#1769e0;color:white}.inputbox{margin-top:18px;background:white;border:1px solid #dce5f0;border-radius:16px;padding:16px}.inputbox input{width:100%;font-size:20px;padding:12px;border:1px solid #b8c6d8;border-radius:10px;margin:8px 0 12px}.status{text-align:center;margin:14px 0;font-weight:800;color:#087a3d}.back{display:block;text-align:center;margin-top:16px;color:#1769e0;text-decoration:none;font-weight:700}@media(max-width:500px){.controls{grid-template-columns:1fr 1fr}}
@media(min-width:701px) and (max-width:1180px){.wrap{max-width:900px;padding:34px 34px 80px}h1{font-size:40px}.tip{font-size:19px;padding:18px}.stage{padding:44px 20px}.label{font-size:25px}.controls{max-width:650px;margin:0 auto;gap:16px}.btn{font-size:21px;padding:18px}.inputbox{font-size:18px;padding:22px}.inputbox input{font-size:23px;padding:15px}.status{font-size:19px}}

.compactVerifyScreen{padding:max(8px,env(safe-area-inset-top)) 10px max(10px,env(safe-area-inset-bottom))!important;justify-content:flex-start!important;overflow:hidden!important}
.compactVerifyCard{width:min(560px,98vw)!important;max-height:calc(100dvh - max(18px,env(safe-area-inset-top)) - max(18px,env(safe-area-inset-bottom)))!important;padding:12px 14px!important;display:flex;flex-direction:column;align-items:stretch;gap:6px;overflow:hidden}
.compactVerifyCard .brand{margin:0;font-size:16px}.compactVerifyTitle{font-size:clamp(25px,7vw,34px)!important;margin:0 0 2px!important;line-height:1.1}.compactDistanceLine{display:flex;align-items:center;justify-content:center;gap:7px;font-size:19px;font-weight:800}.compactDistanceInput{width:92px!important;height:42px!important;font-size:21px!important;text-align:center}.compactPresets{display:grid;grid-template-columns:repeat(3,1fr);gap:7px}.compactPresets .presetBtn{min-height:38px!important;padding:4px 8px!important;font-size:18px!important}.compactLetterBox{margin:2px auto!important;padding:10px!important;min-height:116px!important;width:100%;display:flex;align-items:center;justify-content:center}.compactVerifyValue{font-size:17px;font-weight:900;text-align:center;line-height:1.15}.compactApplyBtn{width:100%;min-height:44px!important;padding:7px!important;font-size:20px!important}.compactStatus{font-size:15px!important;font-weight:800!important;min-height:20px!important;margin:0!important;line-height:1.2}.compactNextBtn{width:100%;min-height:52px!important;padding:8px!important;font-size:21px!important;margin:0!important}.compactRestartBtn{border:0;background:transparent;color:#315b9f;text-decoration:underline;font-size:15px;font-weight:700;padding:4px;min-height:28px}.compactVerifyScreen .formulaBox,.compactVerifyScreen .calNote{display:none!important}
@media(max-height:700px){.compactVerifyCard{gap:3px;padding:8px 11px!important}.compactVerifyTitle{font-size:25px!important}.compactLetterBox{min-height:90px!important;padding:5px!important}.compactNextBtn{min-height:46px!important;font-size:19px!important}.compactApplyBtn{min-height:40px!important;font-size:18px!important}.compactPresets .presetBtn{min-height:34px!important}}
</style></head><body><main class="wrap"><h1>5 cm 螢幕校正</h1><div class="tip">請拿實體尺測量下方黑線。可使用按鈕調整，或直接輸入目前量到的實際長度。目標是讓黑線實際長度正好為 <b>5.00 cm</b>。</div>
<div class="stage"><div class="label">5.00 cm</div><div id="line" class="line"></div><div id="pxText"></div></div>
<div class="controls"><button class="btn" onclick="step(.90)">大幅縮短</button><button class="btn" onclick="step(1.10)">大幅加長</button><button class="btn" onclick="step(.99)">細微縮短</button><button class="btn" onclick="step(1.01)">細微加長</button></div>
<div class="inputbox"><label for="measured"><b>黑線目前實際量到幾公分？</b></label><input id="measured" inputmode="decimal" placeholder="例如：4.20"><button class="btn primary" style="width:100%" onclick="autoCorrect()">依實測長度自動修正</button></div>
<div id="status" class="status"></div><button class="btn primary" style="width:100%" onclick="saveCalibration()">儲存本裝置校正</button><button class="btn" style="width:100%;margin-top:10px" onclick="resetCalibration()">恢復預設值</button><a class="back" href="/cloud">← 回 Cloud Vision 首頁</a></main>
<script>
const KEY='cloudVisionCalibrationV10'; const BASE=188.976; let factor=1;
try{const d=JSON.parse(localStorage.getItem(KEY)||'null');if(d&&d.factor)factor=Number(d.factor)||1}catch(e){}
function render(){const px=BASE*factor;document.getElementById('line').style.width=px+'px';document.getElementById('pxText').textContent='目前顯示寬度：'+px.toFixed(1)+' px｜倍率：'+factor.toFixed(4)}
function step(v){factor*=v;render();document.getElementById('status').textContent='請再用尺確認黑線長度'}
function autoCorrect(){const m=parseFloat(document.getElementById('measured').value);if(!m||m<=0){alert('請輸入尺量到的實際長度');return}factor*=5/m;render();document.getElementById('status').textContent='已依 '+m.toFixed(2)+' cm 自動修正，請再次用尺確認'}
function saveCalibration(){localStorage.setItem(KEY,JSON.stringify({factor:factor,date:new Date().toISOString()}));document.getElementById('status').textContent='✅ 校正已儲存在這個瀏覽器';const next=new URLSearchParams(location.search).get('next')||'';let target='/cloud';if(next==='general')target='/cloud/general';else if(next==='professional-test')target='/cloud/professional/test';else if(next==='teaching')target='/cloud/professional/teaching';setTimeout(()=>location.href=target,650)}
function resetCalibration(){factor=1;localStorage.removeItem(KEY);render();document.getElementById('status').textContent='已恢復預設值'}render();
</script></body></html>"""

    def start_remote_server(self) -> None:
        if self.remote_server is not None:
            return
        app = self
        # V7.3：HTTP 執行緒只使用這一份權威初估視力狀態。
        # POST /estimate、GET /estimate、GET /state 全部讀寫同一個 closure dict，
        # 再同步回 Tkinter app 欄位，避免不同路由看見不同資料。
        estimate_lock = threading.RLock()
        estimate_store = {
            "value": app.remote_estimated_va,
            "version": int(app.remote_estimated_va_version),
            "seen_version": int(app.remote_estimated_va_seen_version),
            "last_submit": app.remote_estimated_va_last_submit,
            "last_control_read": app.remote_estimated_va_last_control_read,
            "last_submit_ip": app.remote_estimated_va_last_submit_ip,
            "last_control_ip": app.remote_estimated_va_last_control_ip,
        }

        def sync_estimate_to_app():
            app.remote_estimated_va = estimate_store["value"]
            app.remote_estimated_va_version = estimate_store["version"]
            app.remote_estimated_va_seen_version = estimate_store["seen_version"]
            app.remote_estimated_va_last_submit = estimate_store["last_submit"]
            app.remote_estimated_va_last_control_read = estimate_store["last_control_read"]
            app.remote_estimated_va_last_submit_ip = estimate_store["last_submit_ip"]
            app.remote_estimated_va_last_control_ip = estimate_store["last_control_ip"]

        def estimate_payload():
            with estimate_lock:
                value = estimate_store["value"]
                version = estimate_store["version"]
                seen = estimate_store["seen_version"]
                return {
                    "ok": True,
                    "estimated_va": value,
                    "estimated_va_version": version,
                    "estimated_va_seen_version": seen,
                    "estimated_va_last_submit": estimate_store["last_submit"],
                    "estimated_va_last_control_read": estimate_store["last_control_read"],
                    "estimated_va_last_submit_ip": estimate_store["last_submit_ip"],
                    "estimated_va_last_control_ip": estimate_store["last_control_ip"],
                    "estimated_va_ack": (value is not None and seen >= version),
                }

        class RemoteHandler(BaseHTTPRequestHandler):
            def log_message(self, _format, *_args):
                return
            def _send(self, body: str, status: int = 200, content_type: str = "text/html; charset=utf-8"):
                data = body.encode("utf-8")
                self.send_response(status)
                self.send_header("Content-Type", content_type)
                self.send_header("Content-Length", str(len(data)))
                self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
                self.send_header("Pragma", "no-cache")
                self.send_header("Expires", "0")
                self.send_header("Connection", "close")
                self.end_headers()
                self.wfile.write(data)
            def _send_bytes(self, data: bytes, status: int = 200, content_type: str = "application/octet-stream", filename: str = ""):
                self.send_response(status)
                self.send_header("Content-Type", content_type)
                self.send_header("Content-Length", str(len(data)))
                if filename:
                    self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Cache-Control", "no-store")
                self.send_header("Connection", "close")
                self.end_headers()
                self.wfile.write(data)

            def _examiner_cookie_token(self) -> str:
                cookie = self.headers.get("Cookie", "")
                for part in cookie.split(";"):
                    name, sep, value = part.strip().partition("=")
                    if sep and name == "cloudvision_examiner":
                        return value.strip()
                return ""

            def _examiner_authenticated(self) -> bool:
                token = self._examiner_cookie_token()
                if not token:
                    return False
                with app.examiner_sessions_lock:
                    return token in app.examiner_sessions

            def _redirect(self, location: str, cookie: str = "") -> None:
                self.send_response(303)
                self.send_header("Location", location)
                if cookie:
                    self.send_header("Set-Cookie", cookie)
                self.send_header("Cache-Control", "no-store")
                self.end_headers()

            def _examiner_login_html(self, error: str = "") -> str:
                error_html = f"<p style='color:#b42318;font-weight:800'>{html.escape(error)}</p>" if error else ""
                return f"""<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>測驗者登入｜Cloud Vision</title><style>body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif;background:#f4f7fb;margin:0;color:#172033}}.wrap{{min-height:100vh;display:grid;place-items:center;padding:22px;box-sizing:border-box}}.card{{width:min(440px,100%);background:#fff;border:1px solid #dbe5f1;border-radius:20px;padding:28px;box-sizing:border-box;box-shadow:0 14px 40px rgba(20,45,80,.10)}}h1{{margin:0 0 8px}}p{{line-height:1.6}}input{{width:100%;box-sizing:border-box;padding:15px;font-size:20px;border:1px solid #aebed1;border-radius:12px;margin:8px 0 16px}}button{{width:100%;padding:15px;border:0;border-radius:12px;background:#1769e0;color:white;font-size:20px;font-weight:800}}a{{color:#1769e0}}</style></head><body><main class='wrap'><section class='card'><h1>🔒 測驗者登入</h1><p>此頁僅供測驗者查看回傳資料與操作測驗端。</p>{error_html}<form method='post' action='/examiner/login'><label for='password'>管理密碼</label><input id='password' name='password' type='password' minlength='6' maxlength='6' autocomplete='current-password' autofocus required><button type='submit'>登入測驗者後台</button></form><p style='text-align:center'><a href='/cloud'>返回受試者公開頁面</a></p></section></main></body></html>"""

            def _examiner_dashboard_html(self, message: str = "") -> str:
                token = app.connection_session_token
                stats = app._today_dashboard_stats()
                msg = f"<p class='ok'>{html.escape(message)}</p>" if message else ""
                stat_html = (
                    f"<div class='stat'><b>{stats['general_users']}</b><span>今日一般使用者</span></div>"
                    f"<div class='stat'><b>{stats['professional_users']}</b><span>今日專業人員</span></div>"
                    f"<div class='stat'><b>{stats['starts']}</b><span>今日登錄／測驗</span></div>"
                    f"<div class='stat'><b>{stats['completes']}</b><span>完成測驗</span></div>"
                    f"<div class='stat'><b>{stats['incomplete']}</b><span>未完成</span></div>"
                    f"<div class='stat'><b>{app._format_duration(stats['general_avg_seconds'])}</b><span>一般平均時間</span></div>"
                    f"<div class='stat'><b>{app._format_duration(stats['professional_avg_seconds'])}</b><span>專業平均時間</span></div>"
                )
                try:
                    app._save_today_excel_to_disk()
                except Exception:
                    pass
                return f"""<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>測驗者後台｜Cloud Vision</title><style>body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif;background:#f4f7fb;margin:0;color:#172033}}.wrap{{max-width:980px;margin:auto;padding:24px}}.stats{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin:18px 0}}.stat{{background:white;border:1px solid #dbe5f1;border-radius:16px;padding:16px;text-align:center}}.stat b{{display:block;font-size:30px;color:#1769e0}}.stat span{{font-size:14px;font-weight:750;color:#5a687a}}.cards{{display:grid;grid-template-columns:1fr 1fr;gap:14px}}.card{{display:block;background:white;border:1px solid #dbe5f1;border-radius:18px;padding:22px;text-decoration:none;color:#172033;font-size:21px;font-weight:800}}.sub{{font-size:15px;font-weight:500;color:#5a687a;margin-top:6px;line-height:1.5}}.ok{{background:#eaf8ef;border:1px solid #b7e4c7;padding:12px;border-radius:12px;font-weight:800}}a.logout{{color:#1769e0;font-weight:700}}@media(max-width:760px){{.stats{{grid-template-columns:repeat(2,1fr)}}.cards{{grid-template-columns:1fr}}}}@media(max-width:480px){{.stats{{grid-template-columns:1fr 1fr}}}}</style></head><body><main class='wrap'><h1>Cloud Vision 測驗者後台</h1><p><strong>📅 今天：{datetime.now().strftime('%Y/%m/%d')}</strong></p><p>此固定後台可直接用手機重新整理，查看今天每次登錄、完成狀態與平均測驗時間。</p>{msg}<section class='stats'>{stat_html}</section><section class='cards'><a class='card' href='/examiner/today'>📋 查看今日資料<div class='sub'>查看今天每次登錄、完成狀態、使用時間、左右眼結果及問卷。</div></a><a class='card' href='/control?session={token}'>📱 測驗者手機／iPad<div class='sub'>進入原本的專業雙端連線與遙控出題畫面；此功能完整保留。</div></a></section><p><a class='logout' href='/examiner/logout'>登出測驗者後台</a>　｜　<a class='logout' href='/cloud'>前往公開首頁</a></p></main><script>setTimeout(()=>location.replace('/examiner?refresh='+Date.now()),15000)</script></body></html>"""

            def _examiner_today_html(self) -> str:
                rows = app._today_test_records()
                def esc(value):
                    return html.escape(str(value or ""))
                cards = []
                for index, r in enumerate(rows, 1):
                    user_type = r.get("user_type", "一般使用者")
                    identity = r.get("professional_name", "") or r.get("anonymous_id", "") or r.get("visitor_id", "") or f"匿名使用者 {index}"
                    survey = "已填" if (r.get("satisfaction") or r.get("needs_improvement") or r.get("improvement_suggestion")) else "未填"
                    status = r.get("completion_status", "未完成")
                    status_class = "done" if status == "已完成" else "pending"
                    test_id = r.get("test_id", "") or r.get("session_id", "") or "—"
                    login_text = app._display_datetime(r.get("login_at", ""))
                    complete_text = app._display_datetime(r.get("completed_at", "")) if r.get("completed_at") else "未完成"
                    cards.append(f"""<details class='record'><summary><span>{esc(r.get('time'))}　{esc(user_type)}　{esc(identity)}</span><b class='{status_class}'>{esc(status)}</b></summary><div class='detail'><div><strong>基本資料</strong><span>Test ID：{esc(test_id)}<br>使用者類型：{esc(user_type)}<br>姓名／編號：{esc(identity)}</span></div><div><strong>使用時間</strong><span>登入：{esc(login_text)}<br>完成：{esc(complete_text)}<br>測驗耗時：{esc(r.get('duration_text')) or '—'}</span></div><div><strong>初估視力</strong><span>右眼：{esc(r.get('visual_acuity_right')) or '未測'}<br>左眼：{esc(r.get('visual_acuity_left')) or '未測'}<br>雙眼：{esc(r.get('visual_acuity_both')) or '未測'}</span></div><div><strong>散光鐘</strong><span>右眼：{esc(r.get('astigmatism_right')) or '未測'}<br>左眼：{esc(r.get('astigmatism_left')) or '未測'}</span></div><div><strong>黃斑部</strong><span>右眼：{esc(r.get('amsler_right')) or '未測'}<br>左眼：{esc(r.get('amsler_left')) or '未測'}</span></div><div><strong>使用問卷</strong><span>填寫狀態：{survey}<br>滿意度：{esc(r.get('satisfaction')) or '—'}<br>是否需要改善：{esc(r.get('needs_improvement')) or '—'}<br>改善建議：{esc(r.get('improvement_suggestion')) or '—'}</span></div><div><strong>專業人員資料</strong><span>身分：{esc(r.get('professional_role')) or '—'}<br>用途：{esc(r.get('professional_purpose')) or '—'}</span></div><div><strong>裝置資料</strong><span>{esc(r.get('device')) or '—'}</span></div></div></details>""")
                body = "".join(cards) if cards else "<div class='empty'>今天尚無登錄紀錄。</div>"
                return f"""<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>今日測驗紀錄｜Cloud Vision</title><style>*{{box-sizing:border-box}}body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans TC',sans-serif;background:#f4f7fb;margin:0;color:#172033}}.wrap{{max-width:980px;margin:auto;padding:24px}}h1{{margin-bottom:6px}}.muted{{color:#64748b;margin-top:0}}.record{{background:#fff;border:1px solid #dbe5f1;border-radius:16px;margin:12px 0;overflow:hidden}}summary{{cursor:pointer;padding:18px;display:flex;justify-content:space-between;gap:12px;align-items:center;font-weight:850}}summary b{{font-size:14px;padding:5px 9px;border-radius:99px}}.done{{color:#137333;background:#e8f5e9}}.pending{{color:#a15c00;background:#fff4d6}}.detail{{border-top:1px solid #e3eaf2;padding:16px;display:grid;grid-template-columns:1fr 1fr;gap:12px}}.detail>div{{background:#f8fafc;border-radius:12px;padding:13px;overflow-wrap:anywhere}}strong{{display:block;margin-bottom:6px}}span{{line-height:1.65}}.empty{{background:#fff;border:1px solid #dbe5f1;border-radius:16px;padding:26px;text-align:center;color:#64748b}}.back{{display:inline-block;margin-top:18px;color:#1769e0;font-weight:800;text-decoration:none}}@media(max-width:650px){{summary{{align-items:flex-start;flex-direction:column}}.detail{{grid-template-columns:1fr}}}}</style></head><body><main class='wrap'><h1>📋 今日測驗紀錄（{len(rows)} 筆）</h1><p class='muted'>包含已完成與未完成紀錄；點選即可查看 Test ID、時間、結果與問卷。</p>{body}<a class='back' href='/examiner'>← 返回測驗者後台</a></main></body></html>"""

            def do_POST(self):
                parsed = urlparse(self.path)
                query = parse_qs(parsed.query)
                if parsed.path == "/examiner/login":
                    try:
                        length = int(self.headers.get("Content-Length", "0") or "0")
                        raw = self.rfile.read(length).decode("utf-8") if length > 0 else ""
                        form = parse_qs(raw)
                        supplied = form.get("password", [""])[0]
                    except Exception:
                        supplied = ""
                    expected = app.examiner_password_var.get().strip() or "123456"
                    if secrets.compare_digest(str(supplied), str(expected)):
                        login_token = secrets.token_urlsafe(32)
                        with app.examiner_sessions_lock:
                            app.examiner_sessions.add(login_token)
                        self._redirect("/examiner?login=ok", f"cloudvision_examiner={login_token}; Path=/; Max-Age=43200; HttpOnly; SameSite=Lax")
                    else:
                        self._send(self._examiner_login_html("密碼錯誤，請重新輸入。"), 403)
                    return
                if parsed.path == "/examiner/password":
                    if not self._examiner_authenticated():
                        self._redirect("/examiner")
                        return
                    try:
                        length = int(self.headers.get("Content-Length", "0") or "0")
                        form = parse_qs(self.rfile.read(length).decode("utf-8") if length > 0 else "")
                        current = form.get("current_password", [""])[0]
                        new = form.get("new_password", [""])[0]
                        confirm = form.get("confirm_password", [""])[0]
                        expected = app.examiner_password_var.get().strip() or "123456"
                        if not secrets.compare_digest(current, expected):
                            raise ValueError("目前密碼不正確")
                        if len(new) != 6:
                            raise ValueError("新密碼必須剛好 6 個字元")
                        if new != confirm:
                            raise ValueError("兩次輸入的新密碼不一致")
                        if new == current:
                            raise ValueError("新密碼不可與目前密碼相同")
                        app.examiner_password_var.set(new)
                        app._save_hotspot_settings()
                        self._redirect("/examiner?notice=password_saved")
                    except ValueError as exc:
                        self._send(self._examiner_dashboard_html(str(exc)), 400)
                    return
                if parsed.path == "/examiner/email/settings":
                    if not self._examiner_authenticated():
                        self._redirect("/examiner")
                        return
                    try:
                        length = int(self.headers.get("Content-Length", "0") or "0")
                        form = parse_qs(self.rfile.read(length).decode("utf-8") if length > 0 else "")
                        settings = app._load_email_settings()
                        recipient = form.get("recipient", [""])[0].strip()
                        sender = form.get("sender", [""])[0].strip()
                        new_password = form.get("app_password", [""])[0].replace(" ", "").strip()
                        if "@" not in recipient or "@" not in sender:
                            raise ValueError("Email 格式不正確")
                        settings.update({
                            "enabled": form.get("enabled", [""])[0] == "yes",
                            "recipient": recipient,
                            "sender": sender,
                            "smtp_server": "smtp.mail.yahoo.com",
                            "smtp_port": 465,
                        })
                        if new_password:
                            settings["app_password"] = new_password
                        app._save_email_settings(settings)
                        self._redirect("/examiner?notice=email_saved")
                    except ValueError as exc:
                        self._send(self._examiner_dashboard_html(str(exc)), 400)
                    return
                if parsed.path == "/examiner/email/test":
                    if not self._examiner_authenticated():
                        self._redirect("/examiner")
                        return
                    try:
                        app._send_test_email()
                        self._redirect("/examiner?notice=email_test_ok")
                    except Exception as exc:
                        app._write_email_log(False, str(exc))
                        self._send(self._examiner_dashboard_html("❌ 測試信寄送失敗：" + html.escape(str(exc))), 400)
                    return
                if parsed.path == "/examiner/appointments/read":
                    if not self._examiner_authenticated():
                        self._redirect("/examiner")
                        return
                    count = app._mark_today_appointments_read()
                    self._redirect(f"/examiner?notice=read&count={count}")
                    return
                if parsed.path == "/examiner/appointment/status":
                    if not self._examiner_authenticated():
                        self._redirect("/examiner")
                        return
                    try:
                        length = int(self.headers.get("Content-Length", "0") or "0")
                        form = parse_qs(self.rfile.read(length).decode("utf-8") if length > 0 else "")
                        record_id = form.get("id", [""])[0].strip()
                        status = form.get("status", [""])[0].strip()
                        if not record_id:
                            raise ValueError("缺少預約資料編號")
                        app._set_appointment_status(record_id, status)
                        self._redirect("/examiner?notice=status_saved")
                    except ValueError as exc:
                        self._send(str(exc), 400, "text/plain; charset=utf-8")
                    return
                if parsed.path == "/cloud/event":
                    try:
                        length = int(self.headers.get("Content-Length", "0") or "0")
                        raw = self.rfile.read(length).decode("utf-8") if length > 0 else "{}"
                        payload = json.loads(raw)
                        saved = app._save_public_event(payload, self.client_address[0], self.headers.get("User-Agent", ""))
                        self._send(json.dumps({"ok": True, "id": saved["id"]}, ensure_ascii=False), content_type="application/json; charset=utf-8")
                    except Exception as exc:
                        self._send(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False), 400, "application/json; charset=utf-8")
                    return
                if parsed.path == "/cloud/result":
                    try:
                        length = int(self.headers.get("Content-Length", "0") or "0")
                        if length <= 0 or length > 65536:
                            raise ValueError("invalid_length")
                        raw = self.rfile.read(length).decode("utf-8")
                        payload = json.loads(raw)
                        if not isinstance(payload, dict):
                            raise ValueError("invalid_payload")
                        consent = str(payload.get("consent", "")).lower() in ("yes", "true", "1", "on")
                        if not consent:
                            payload["name"] = ""
                            payload["phone"] = ""
                        saved = app._save_public_result(payload, self.client_address[0], self.headers.get("User-Agent", ""))
                        self._send(json.dumps({"ok": True, "id": saved["id"]}, ensure_ascii=False), content_type="application/json; charset=utf-8")
                    except Exception as exc:
                        self._send(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False), 400, "application/json; charset=utf-8")
                    return
                if parsed.path != "/estimate":
                    self._send(json.dumps({"ok": False, "error": "not_found"}, ensure_ascii=False), 404, "application/json; charset=utf-8")
                    return
                if not app._valid_connection_session(query):
                    self._send(json.dumps({"ok": False, "error": "invalid_session"}, ensure_ascii=False), 403, "application/json; charset=utf-8")
                    return
                role = query.get("role", [""])[0]
                if role != "participant":
                    self._send(json.dumps({"ok": False, "error": "invalid_role"}, ensure_ascii=False), 403, "application/json; charset=utf-8")
                    return
                try:
                    length = int(self.headers.get("Content-Length", "0") or "0")
                    raw = self.rfile.read(length).decode("utf-8") if length > 0 else ""
                    content_type = self.headers.get("Content-Type", "")
                    if "application/json" in content_type:
                        payload = json.loads(raw or "{}")
                        raw_value = payload.get("estimated_va", payload.get("value"))
                    else:
                        form = parse_qs(raw)
                        raw_value = form.get("estimated_va", form.get("value", [None]))[0]
                    estimate = float(raw_value)
                    if not 0.10 <= estimate <= 1.50:
                        raise ValueError
                except Exception as exc:
                    print(f"[Cloud Vision V7.3][POST /estimate] 無效資料 raw={raw!r} error={exc}", flush=True)
                    self._send(json.dumps({"ok": False, "error": "invalid_estimated_va"}, ensure_ascii=False), 400, "application/json; charset=utf-8")
                    return
                app._register_device_connection("participant", self.client_address[0], notify=False, user_agent=self.headers.get("User-Agent", ""))
                with estimate_lock:
                    estimate_store["value"] = estimate
                    estimate_store["version"] += 1
                    estimate_store["last_submit"] = time.strftime("%H:%M:%S")
                    estimate_store["last_submit_ip"] = self.client_address[0]
                    version = estimate_store["version"]
                    sync_estimate_to_app()
                print(f"[Cloud Vision V7.3][participant POST→server] 初估視力={estimate:.2f} version={version} ip={self.client_address[0]}", flush=True)
                app.remote_command_queue.put(("estimated_va_received", str(estimate)))
                self._send(json.dumps({
                    "ok": True,
                    "estimated_va": estimate,
                    "estimated_va_version": version,
                    "received_by_server": True,
                }, ensure_ascii=False), content_type="application/json; charset=utf-8")

            def do_GET(self):
                parsed = urlparse(self.path)
                query = parse_qs(parsed.query)
                if parsed.path == "/examiner/logout":
                    token = self._examiner_cookie_token()
                    if token:
                        with app.examiner_sessions_lock:
                            app.examiner_sessions.discard(token)
                    self._redirect("/cloud", "cloudvision_examiner=; Path=/; Max-Age=0; HttpOnly; SameSite=Lax")
                    return
                if parsed.path == "/examiner/today":
                    if not self._examiner_authenticated():
                        self._redirect("/examiner")
                        return
                    self._send(self._examiner_today_html())
                    return
                if parsed.path == "/examiner":
                    if self._examiner_authenticated():
                        try:
                            message = "✅ 已成功登入測驗者後台" if query.get("login", [""])[0] == "ok" else ""
                            notice = query.get("notice", [""])[0]
                            if notice == "email_saved":
                                message = "✅ Email 設定已儲存"
                            elif notice == "email_test_ok":
                                message = "✅ 測試信已寄出，請查看收件匣或垃圾郵件"
                            elif notice == "password_saved":
                                message = "✅ 管理密碼已更新並儲存"
                            elif notice == "excel_opened":
                                message = "✅ 已在這台電腦開啟今日 Excel"
                            elif notice == "folder_opened":
                                message = "✅ 已在這台電腦開啟 CloudVision\\Excel 資料夾"
                            elif notice == "read":
                                message = f"✅ 已將 {query.get('count', ['0'])[0]} 筆新預約標示為已讀"
                            elif notice == "status_saved":
                                message = "✅ 預約狀態已更新，Excel 也已同步"
                            self._send(self._examiner_dashboard_html(message))
                        except Exception as exc:
                            try:
                                app._write_email_log(False, "後台頁面產生失敗：" + str(exc))
                            except Exception:
                                pass
                            safe_error = html.escape(str(exc))
                            self._send(f"<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>後台讀取失敗</title></head><body style='font-family:sans-serif;padding:24px'><h1>後台暫時無法顯示</h1><p>伺服器仍在運作，請返回後台重新整理。</p><pre style='white-space:pre-wrap'>{safe_error}</pre><p><a href='/examiner'>重新整理後台</a>｜<a href='/cloud'>返回受試者頁面</a></p></body></html>", 500)
                    else:
                        self._send(self._examiner_login_html())
                    return
                # 今日資料、CSV 與測驗者控制頁都必須先通過管理密碼。
                if parsed.path in ("/cloud/today", "/cloud/detail", "/cloud/today.csv", "/cloud/today.xlsx", "/examiner/open-excel", "/examiner/open-folder", "/control") and not self._examiner_authenticated():
                    self._redirect("/examiner")
                    return
                role_for_auth = query.get("role", [""])[0]
                if role_for_auth == "control" and parsed.path in ("/command", "/state", "/estimate", "/control_estimate") and not self._examiner_authenticated():
                    self._send(json.dumps({"ok": False, "error": "examiner_login_required"}, ensure_ascii=False), 403, "application/json; charset=utf-8")
                    return
                if parsed.path in ("/cloud", "/cloud/"):
                    self._send(app.public_home_html())
                elif parsed.path == "/cloud/general":
                    self._send(app.public_general_entry_html())
                elif parsed.path == "/cloud/calibration":
                    self._send(app.public_calibration_html())
                elif parsed.path == "/cloud/professional":
                    self._send(app.public_professional_html())
                elif parsed.path == "/cloud/professional/disclaimer":
                    self._send(app.public_professional_disclaimer_html())
                elif parsed.path == "/cloud/professional/hub":
                    self._send(app.public_professional_hub_html())
                elif parsed.path == "/cloud/professional/data":
                    self._send(app.public_professional_data_html(query.get("q", [""])[0]))
                elif parsed.path == "/cloud/professional/today.xlsx":
                    try:
                        data = app._build_today_excel()
                        self._send_bytes(data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="CloudVision_Today.xlsx")
                    except Exception as exc:
                        self._send(f"Excel 匯出失敗：{html.escape(str(exc))}", 500, "text/plain; charset=utf-8")
                elif parsed.path == "/cloud/professional/test":
                    self._send(app.public_professional_test_html())
                elif parsed.path == "/cloud/professional/teaching":
                    self._send(app.public_professional_teaching_html())
                elif parsed.path == "/cloud/test/astigmatism":
                    self._send(app.public_astigmatism_test_html())
                elif parsed.path == "/cloud/test/amsler":
                    self._send(app.public_amsler_test_html())
                elif parsed.path == "/cloud/tools/size":
                    self._send(app.public_size_tool_html())
                elif parsed.path == "/cloud/tools/geometry":
                    self._send(app.public_geometry_center_html())
                elif parsed.path.startswith("/cloud/teaching/"):
                    self._send(app.public_teaching_topic_html(parsed.path.rsplit("/", 1)[-1]))
                elif parsed.path == "/cloud/disclaimer":
                    self._send(app.public_disclaimer_html())
                elif parsed.path == "/cloud/assessment":
                    self._send(app.public_assessment_html())
                elif parsed.path == "/examiner/open-excel":
                    try:
                        excel_path = app._sync_daily_excel()
                        app._open_on_server(excel_path)
                        self._redirect("/examiner?notice=excel_opened")
                    except Exception as exc:
                        self._send(f"無法開啟今日 Excel：{html.escape(str(exc))}", 500, "text/plain; charset=utf-8")
                elif parsed.path == "/examiner/open-folder":
                    try:
                        os.makedirs(app.public_excel_dir, exist_ok=True)
                        app._open_on_server(app.public_excel_dir)
                        self._redirect("/examiner?notice=folder_opened")
                    except Exception as exc:
                        self._send(f"無法開啟 Excel 資料夾：{html.escape(str(exc))}", 500, "text/plain; charset=utf-8")
                elif parsed.path == "/cloud/today.xlsx":
                    try:
                        data = app._build_today_excel()
                        filename = "CloudVision_" + datetime.now().strftime("%Y-%m-%d") + ".xlsx"
                        self._send_bytes(data, 200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename)
                    except Exception as exc:
                        self._send(f"Excel 匯出失敗：{html.escape(str(exc))}", 500, "text/plain; charset=utf-8")
                elif parsed.path == "/cloud/today.csv":
                    rows = app._today_public_results()
                    fields = [
                        "time", "name", "phone", "visual_acuity_right", "visual_acuity_left",
                        "astigmatism_right", "astigmatism_left", "amsler_right", "amsler_left",
                        "near_reading", "appointment", "appointment_date", "appointment_time", "note"
                    ]
                    output = io.StringIO()
                    writer = csv.DictWriter(output, fieldnames=fields)
                    writer.writeheader()
                    for row in rows:
                        writer.writerow({key: row.get(key, "") for key in fields})
                    self._send("\ufeff" + output.getvalue(), 200, "text/csv; charset=utf-8")
                elif parsed.path == "/cloud/today":
                    self._send(app.public_today_html(query.get("filter", ["all"])[0], query.get("q", [""])[0]))
                elif parsed.path == "/cloud/detail":
                    self._send(app.public_detail_html(query.get("id", [""])[0]))
                elif parsed.path in ("/", "/remote"):
                    # 正式雲端首頁直接顯示原本已完成的雙入口頁面。
                    # 不再使用「已連到 Cloud Vision」的臨時單按鈕轉接頁。
                    self._send(app.public_home_html())
                elif parsed.path in ("/control", "/participant", "/command", "/answer", "/state", "/estimate", "/control_estimate") and not app._valid_connection_session(query):
                    self._send("此連線頁面已失效，請重新掃描電腦畫面上的最新 QR Code。", 403, "text/plain; charset=utf-8")
                elif parsed.path == "/control":
                    app._register_device_connection("control", self.client_address[0], user_agent=self.headers.get("User-Agent", ""))
                    self._send(app.control_html())
                elif parsed.path == "/participant":
                    app._register_device_connection("participant", self.client_address[0], user_agent=self.headers.get("User-Agent", ""))
                    self._send(app.answer_html())
                elif parsed.path == "/estimate":
                    role = query.get("role", [""])[0]
                    if role not in ("participant", "control"):
                        self._send(json.dumps({"ok": False, "error": "invalid_role"}, ensure_ascii=False), 403, "application/json; charset=utf-8")
                        return
                    app._register_device_connection(role, self.client_address[0], notify=False, user_agent=self.headers.get("User-Agent", ""))
                    raw_value = query.get("value", [None])[0]
                    if role == "participant" and raw_value not in (None, ""):
                        try:
                            estimate = float(raw_value)
                            if not 0.10 <= estimate <= 1.50:
                                raise ValueError
                        except (TypeError, ValueError):
                            self._send(json.dumps({"ok": False, "error": "invalid_estimated_va"}, ensure_ascii=False), 400, "application/json; charset=utf-8")
                            return
                        with estimate_lock:
                            estimate_store["value"] = estimate
                            estimate_store["version"] += 1
                            estimate_store["last_submit"] = time.strftime("%H:%M:%S")
                            estimate_store["last_submit_ip"] = self.client_address[0]
                            version = estimate_store["version"]
                            sync_estimate_to_app()
                        print(f"[Cloud Vision V7.3][participant→server] 初估視力={estimate:.2f} version={version} ip={self.client_address[0]}", flush=True)
                        app.remote_command_queue.put(("estimated_va_received", str(estimate)))
                        self._send(json.dumps({"ok": True, "estimated_va": estimate, "estimated_va_version": version}, ensure_ascii=False), content_type="application/json; charset=utf-8")
                        return
                    # 測驗者與受試者都讀取同一筆伺服器狀態。
                    # 只有測驗者端真正讀到此筆資料時，才更新 seen_version；
                    # 受試者必須等 seen_version 追上資料版本後，才可進入等待畫面。
                    with estimate_lock:
                        if role == "control" and estimate_store["value"] is not None:
                            estimate_store["seen_version"] = max(estimate_store["seen_version"], estimate_store["version"])
                            estimate_store["last_control_read"] = time.strftime("%H:%M:%S")
                            estimate_store["last_control_ip"] = self.client_address[0]
                            sync_estimate_to_app()
                            print(
                                f"[Cloud Vision V7.3][control /estimate read] 初估視力={estimate_store['value']:.2f} "
                                f"seen_version={estimate_store['seen_version']} ip={self.client_address[0]}",
                                flush=True,
                            )
                        payload = estimate_payload()
                    self._send(json.dumps(payload, ensure_ascii=False), content_type="application/json; charset=utf-8")
                elif parsed.path == "/control_estimate":
                    # V7.9：測驗者直接讀取 app 的最新初估視力，避免 /state 與 closure 狀態不同步。
                    role = query.get("role", [""])[0]
                    if role != "control":
                        self._send(json.dumps({"ok": False, "error": "invalid_role"}, ensure_ascii=False), 403, "application/json; charset=utf-8")
                        return
                    app._register_device_connection("control", self.client_address[0], notify=False, user_agent=self.headers.get("User-Agent", ""))
                    with estimate_lock:
                        # 先將 closure 同步回 app，再由 app 欄位建立最小、穩定的回傳資料。
                        sync_estimate_to_app()
                        value = app.remote_estimated_va
                        version = int(app.remote_estimated_va_version)
                        if value is not None:
                            estimate_store["seen_version"] = max(int(estimate_store["seen_version"]), version)
                            estimate_store["last_control_read"] = time.strftime("%H:%M:%S")
                            estimate_store["last_control_ip"] = self.client_address[0]
                            sync_estimate_to_app()
                    self._send(json.dumps({
                        "ok": True,
                        "estimated_va": value,
                        "estimated_va_version": version,
                    }, ensure_ascii=False), content_type="application/json; charset=utf-8")
                elif parsed.path == "/command":
                    role = query.get("role", [""])[0]
                    cmd = query.get("cmd", [""])[0]
                    value = query.get("value", [""])[0]
                    if role == "control":
                        app._register_device_connection("control", self.client_address[0], notify=False, user_agent=self.headers.get("User-Agent", ""))
                    elif role == "participant" and cmd in ("undo_answer", "reset_answers", "set_estimated_va", "reset_estimated_va"):
                        app._register_device_connection("participant", self.client_address[0], notify=False, user_agent=self.headers.get("User-Agent", ""))
                    else:
                        self._send("Invalid role", 403, "text/plain; charset=utf-8")
                        return
                    if cmd == "reset_estimated_va":
                        with estimate_lock:
                            estimate_store["value"] = None
                            estimate_store["version"] += 1
                            estimate_store["seen_version"] = estimate_store["version"]
                            estimate_store["last_submit"] = "已重設"
                            estimate_store["last_control_read"] = "已重設"
                            estimate_store["last_submit_ip"] = ""
                            estimate_store["last_control_ip"] = ""
                            version = estimate_store["version"]
                            sync_estimate_to_app()
                        print(f"[Cloud Vision V7.3] 初估視力流程重設｜version={version}", flush=True)
                        self._send(json.dumps({"ok": True, "estimated_va": None, "estimated_va_version": version}, ensure_ascii=False), content_type="application/json; charset=utf-8")
                        return
                    if cmd == "set_estimated_va":
                        # 舊路徑保留相容；V6.8 受試者改走獨立 /estimate 通道。
                        # 避免只排入 Tkinter 佇列後先回 OK，造成受試者端等待 ACK 超時。
                        try:
                            estimate = float(value)
                            if not 0.05 <= estimate <= 2.0:
                                raise ValueError
                        except (TypeError, ValueError):
                            self._send(
                                json.dumps({"ok": False, "error": "invalid_estimated_va"}, ensure_ascii=False),
                                status=400,
                                content_type="application/json; charset=utf-8",
                            )
                            return
                        with estimate_lock:
                            estimate_store["value"] = estimate
                            estimate_store["version"] += 1
                            estimate_store["last_submit"] = time.strftime("%H:%M:%S")
                            estimate_store["last_submit_ip"] = self.client_address[0]
                            version = estimate_store["version"]
                            sync_estimate_to_app()
                        print(f"[Cloud Vision] 收到初估視力：{estimate:.2f}｜version={version}", flush=True)
                        # Tkinter 相關 UI 更新仍交由主執行緒處理。
                        app.remote_command_queue.put(("estimated_va_received", str(estimate)))
                        self._send(
                            json.dumps({
                                "ok": True,
                                "estimated_va": estimate,
                                "estimated_va_version": version,
                            }, ensure_ascii=False),
                            content_type="application/json; charset=utf-8",
                        )
                        return
                    app.remote_command_queue.put((cmd, value))
                    self._send("OK", content_type="text/plain; charset=utf-8")
                elif parsed.path == "/answer":
                    if query.get("role", [""])[0] != "participant":
                        self._send("Invalid role", 403, "text/plain; charset=utf-8")
                        return
                    app._register_device_connection("participant", self.client_address[0], notify=False, user_agent=self.headers.get("User-Agent", ""))
                    answer = query.get("value", [""])[0].upper()
                    app.remote_command_queue.put(("answer", answer))
                    self._send("OK", content_type="text/plain; charset=utf-8")
                elif parsed.path == "/ping":
                    self._send("JOHNVA_OK", content_type="text/plain; charset=utf-8")
                elif parsed.path == "/favicon.ico":
                    self._send("", status=204, content_type="image/x-icon")
                elif parsed.path == "/state":
                    role = query.get("role", [""])[0]
                    if role in ("control", "participant"):
                        app._register_device_connection(role, self.client_address[0], notify=False, user_agent=self.headers.get("User-Agent", ""))
                    # V7.0：測驗者頁面本來就會固定讀取 /state。
                    # 因此只要 control 成功讀到目前狀態，就視為測驗者已實際收到初估視力，
                    # 同步更新 seen_version，避免另外的 /estimate 輪詢被 Safari 暫停時無法 ACK。
                    with estimate_lock:
                        if role == "control" and estimate_store["value"] is not None:
                            estimate_store["last_control_read"] = time.strftime("%H:%M:%S")
                            estimate_store["last_control_ip"] = self.client_address[0]
                            if estimate_store["seen_version"] < estimate_store["version"]:
                                estimate_store["seen_version"] = estimate_store["version"]
                                print(
                                    f"[Cloud Vision V7.3][control /state ACK] 初估視力={estimate_store['value']:.2f} "
                                    f"seen_version={estimate_store['seen_version']} ip={self.client_address[0]}",
                                    flush=True,
                                )
                        sync_estimate_to_app()
                        state_data = app.remote_state()
                        state_data.update(estimate_payload())
                    payload = json.dumps(state_data, ensure_ascii=False)
                    self._send(payload, content_type="application/json; charset=utf-8")
                else:
                    self._send("Not found", 404, "text/plain; charset=utf-8")

        try:
            self.remote_server = ThreadingHTTPServer(("0.0.0.0", self.remote_port), RemoteHandler)
        except OSError as exc:
            self.status_var.set(f"遙控伺服器無法啟動：{exc}")
            return
        self.remote_thread = threading.Thread(target=self.remote_server.serve_forever, daemon=True)
        self.remote_thread.start()
        self._update_remote_urls()

    def _poll_remote_commands(self) -> None:
        """只在 Tkinter 主執行緒中執行手機送來的命令。"""
        try:
            while True:
                cmd, value = self.remote_command_queue.get_nowait()
                try:
                    self.execute_remote_command(cmd, value)
                    if cmd != "answer":
                        self.status_var.set(f"手機遙控指令已執行：{cmd}{' ' + value if value else ''}")
                except Exception as exc:
                    self.status_var.set(f"手機遙控執行失敗：{cmd}｜{exc}")
        except queue.Empty:
            pass

        try:
            while True:
                role, ip = self.device_event_queue.get_nowait()
                if not self._device_notified_online.get(role, False):
                    self._device_notified_online[role] = True
                    role_name = "測驗者手機／iPad" if role == "control" else "受試者手機"
                    self.status_var.set(f"✅ {role_name}已連線｜IP：{ip}")
                    # 讓提示一定出現在最前面，避免被全螢幕或 QR Code 視窗遮住。
                    try:
                        self.root.bell()
                        if self.qr_window is not None and self.qr_window.winfo_exists():
                            self.qr_window.lift()
                            self.qr_window.attributes("-topmost", True)
                            self.qr_window.after(1200, lambda: self.qr_window.attributes("-topmost", False)
                                                 if self.qr_window is not None and self.qr_window.winfo_exists() else None)
                    except tk.TclError:
                        pass
                    messagebox.showinfo(
                        "連線成功",
                        f"✅ {role_name}已成功連線！\n\n裝置 IP：{ip}\n\n現在可以開始操作。",
                        parent=self.qr_window if self.qr_window is not None and self.qr_window.winfo_exists() else self.root,
                    )
        except queue.Empty:
            pass
        self._update_device_connection_vars()
        self.root.after(50, self._poll_remote_commands)

    def show_remote_url(self) -> None:
        self.show_remote_qr()

    def show_remote_qr(self) -> None:
        if not self.remote_url:
            self.start_remote_server()
        if not self.remote_url:
            return
        try:
            import qrcode
        except ImportError:
            # 嘗試安裝到目前實際執行本程式的 Python（例如 PsychoPy 內建 Python）。
            try:
                import subprocess, sys
                subprocess.check_call(
                    [sys.executable, "-m", "pip", "install", "qrcode[pil]"],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )
                import qrcode
            except Exception:
                messagebox.showinfo(
                    "雙手機連線網址",
                    "QR Code 元件目前無法載入，但兩個網址仍可正常使用。\n\n"
                    f"測驗者遙控：\n{self.control_url}\n\n"
                    f"受試者答題：\n{self.answer_url}",
                    parent=self.root,
                )
                return

        if self.qr_window is not None and self.qr_window.winfo_exists():
            self.qr_window.lift()
            return

        control_img = qrcode.make(self.control_url).convert("RGB").resize((230, 230))
        public_scan_url = self._public_cloud_url()
        answer_img = qrcode.make(public_scan_url).convert("RGB").resize((360, 360))
        self.qr_photo_control = ImageTk.PhotoImage(control_img)
        self.qr_photo_answer = ImageTk.PhotoImage(answer_img)
        self.qr_photo_hotspot = None
        if self.connection_mode == "hotspot" and self.hotspot_ssid_var.get().strip():
            hotspot_img = qrcode.make(self._hotspot_wifi_payload()).convert("RGB").resize((250, 250))
            self.qr_photo_hotspot = ImageTk.PhotoImage(hotspot_img)

        win = tk.Toplevel(self.root)
        self.qr_window = win
        win.title("Cloud Vision V10.2｜掃描開始測驗")
        win.configure(background="white")
        win.attributes("-topmost", True)

        mode_title = (
            "電腦行動熱點模式：手機請連接電腦開出的熱點"
            if self.connection_mode == "hotspot"
            else "共用 Wi-Fi 模式：手機與電腦請連接同一個 Wi-Fi"
        )
        tk.Label(
            win, text=mode_title, bg="white",
            font=("Microsoft JhengHei", 17, "bold")
        ).pack(padx=24, pady=(18, 4))
        tk.Label(
            win, text=f"目前連線 IP：{self.connection_ip}", bg="white", fg="#555555",
            font=("Microsoft JhengHei", 11, "bold")
        ).pack(padx=24, pady=(0, 10))

        qr_row = tk.Frame(win, bg="white")
        qr_row.pack(padx=18, pady=4)

        if self.qr_photo_hotspot is not None:
            hotspot = tk.Frame(qr_row, bg="#eaf8ea", bd=2, relief="groove")
            hotspot.pack(side=tk.LEFT, padx=8, pady=4)
            tk.Label(hotspot, text="① 先連電腦行動熱點", bg="#eaf8ea", fg="#176b2c",
                     font=("Microsoft JhengHei", 15, "bold")).pack(padx=12, pady=(14, 4))
            tk.Label(hotspot, text="所有手機先掃這一張", bg="#eaf8ea", font=("Microsoft JhengHei", 12)).pack()
            tk.Label(hotspot, image=self.qr_photo_hotspot, bg="#eaf8ea").pack(padx=10, pady=8)
            tk.Label(hotspot, text=f"熱點：{self.hotspot_ssid_var.get().strip()}", bg="#eaf8ea", fg="#176b2c",
                     font=("Microsoft JhengHei", 10, "bold"), wraplength=275).pack(padx=8, pady=(0, 12))

        left = tk.Frame(qr_row, bg="#eef5ff", bd=2, relief="groove")
        left.pack(side=tk.LEFT, padx=10, pady=4)
        tk.Label(left, text="專業人員遙控", bg="#eef5ff", fg="#174ea6",
                 font=("Microsoft JhengHei", 15, "bold")).pack(padx=16, pady=(14, 4))
        tk.Label(left, text="遙控出題", bg="#eef5ff", font=("Microsoft JhengHei", 13)).pack()
        tk.Label(left, image=self.qr_photo_control, bg="#eef5ff").pack(padx=14, pady=8)
        tk.Label(left, text=self.control_url, bg="#eef5ff", fg="#174ea6",
                 font=("Arial", 10, "bold"), wraplength=300).pack(padx=10, pady=(0, 14))

        right = tk.Frame(qr_row, bg="#fff5eb", bd=2, relief="groove")
        right.pack(side=tk.LEFT, padx=10, pady=4)
        tk.Label(right, text="📱 公開自我測驗 QR Code", bg="#fff5eb", fg="#9a4b00",
                 font=("Microsoft JhengHei", 19, "bold")).pack(padx=16, pady=(14, 4))
        tk.Label(right, text="掃描後先進入公開首頁，不會跳到專業確認視標頁", bg="#fff5eb", font=("Microsoft JhengHei", 14, "bold")).pack()
        tk.Label(right, image=self.qr_photo_answer, bg="#fff5eb").pack(padx=14, pady=8)
        tk.Label(right, text=public_scan_url, bg="#fff5eb", fg="#9a4b00",
                 font=("Arial", 10, "bold"), wraplength=300).pack(padx=10, pady=(0, 14))

        tk.Label(
            win, text=("先掃綠色 QR Code 加入電腦熱點，再依身分掃描藍色或橘色 QR Code。"
                       if self.qr_photo_hotspot is not None else
                       "一般使用者直接掃描右側大型 QR Code，先進入公開自我測驗首頁；左側仍為專業人員遙控。"),
            bg="white", font=("Microsoft JhengHei", 12, "bold")
        ).pack(padx=20, pady=(8, 12))
        def close_qr() -> None:
            self.qr_window = None
            win.destroy()
        ttk.Button(win, text="關閉", command=close_qr).pack(pady=(0, 16))
        win.protocol("WM_DELETE_WINDOW", close_qr)

    def execute_remote_command(self, cmd: str, value: str = "") -> None:
        if cmd == "answer":
            self.receive_answer(value)
            return
        if cmd == "mode_adult":
            self.set_test_mode("adult")
            return
        if cmd == "mode_child":
            self.set_test_mode("child")
            return
        if cmd == "send_question":
            self.send_current_question()
            return
        if cmd == "clear_question":
            self.clear_current_question()
            return
        if cmd == "undo_answer":
            self.undo_last_answer()
            return
        if cmd == "reset_answers":
            self.reset_current_answers()
            return
        if cmd == "estimated_va_received":
            try:
                estimate = float(value)
                self.status_var.set(f"✅ 已收到受試者初估視力：{estimate:.2f}；可由此附近開始正式測驗。")
                self._update_result_panel()
                try:
                    self.root.bell()
                except tk.TclError:
                    pass
            except (TypeError, ValueError):
                self.status_var.set("受試者初估視力輸入無效。")
            return
        if cmd == "set_estimated_va":
            # 保留舊指令相容性；新版本由 HTTP handler 先寫入狀態。
            try:
                estimate = float(value)
                if not 0.05 <= estimate <= 2.0:
                    raise ValueError
                self.remote_estimated_va = estimate
                self.remote_estimated_va_version += 1
                self.status_var.set(f"受試者回報初估視力：{estimate:g}；可由此附近開始正式測驗。")
                self._update_result_panel()
            except ValueError:
                self.status_var.set("受試者初估視力輸入無效。")
            return
        if cmd == "sticker":
            self.select_reward_sticker(value)
            return
        if cmd == "send_stickers":
            self.send_selected_stickers()
            return
        if cmd == "clear_pending_stickers":
            self.clear_pending_stickers()
            return
        if cmd == "clear_stickers":
            self.clear_reward_stickers()
            return
        question_changing = {"random", "optotype", "mode_full", "mode_single", "stim_prev", "stim_next", "stim_random", "va_prev", "va_next", "row_full", "stim_count", "stimulus", "va", "distance", "chart", "dial", "amsler", "worth", "bagolini", "thorington", "worth_focus", "focus"}
        # 兒童模式完成一題並送完貼圖後，測驗者可直接切換視力值、距離、
        # 上下列或其他功能，不必先返回主畫面。若本題已完成，系統在執行
        # 新指令前自動清除本題作答狀態；只有仍在作答中時才阻止換題。
        if self.test_mode == "child" and cmd in question_changing:
            # 任何會更換視標、視力列或功能畫面的指令，都先同步結束舊題。
            # 避免 C／E 尚在作答時切換畫面後，舊的 active_targets 與新畫面互相衝突。
            if self.question_status in ("awaiting_answer", "completed"):
                self.clear_current_question(silent=True)
        actions = {
            "up": lambda: self.canvas.yview_scroll(-3, "units"),
            "down": lambda: self.canvas.yview_scroll(3, "units"),
            "home": lambda: self.canvas.yview_moveto(0.0),
            "random": self.randomize_letters,
            "chart": self.show_full_acuity_chart,
            "mode_full": self.remote_enter_full_mode,
            "mode_single": self.remote_enter_single_mode,
            "dial": self.show_astigmatic_dial,
            "amsler": self.show_amsler_grid,
            "worth": self.show_worth_four_dot,
            "bagolini": self.show_bagolini_test,
            "thorington": self.show_thorington_test,
            "worth_focus": self.toggle_worth_focus_mode,
            "focus": self.toggle_test_focus_mode,
            "fullscreen": self.toggle_fullscreen,
            "hide_controls": self.toggle_controls,
            "stim_prev": self.remote_previous_stimulus,
            "stim_next": self.remote_next_stimulus,
            "stim_random": self.remote_random_stimulus,
            "va_prev": self.remote_previous_va_row,
            "va_next": self.remote_next_va_row,
            "row_full": lambda: self.remote_set_stimulus_count(5),
        }
        if cmd == "optotype":
            self.set_optotype_mode(value)
            return
        if cmd == "distance":
            try:
                self.remote_set_distance(float(value))
            except ValueError:
                self.status_var.set("手機距離設定失敗：請輸入數字。")
            return
        if cmd == "stim_count":
            try:
                self.remote_set_stimulus_count(int(value))
            except ValueError:
                self.status_var.set("手機刺激數量設定失敗。")
            return
        if cmd == "stimulus":
            try:
                self.remote_select_stimulus(int(value))
            except ValueError:
                self.status_var.set("手機視標選擇失敗。")
            return
        if cmd == "va":
            try:
                if self.remote_mode != "single":
                    self.status_var.set("請先在手機按『單一刺激模式』，再選擇視力值。")
                else:
                    self.remote_jump_to_va(float(value))
            except ValueError:
                self.status_var.set("手機視力列設定失敗。")
            return
        action = actions.get(cmd)
        if action:
            action()

    def remote_enter_full_mode(self) -> None:
        """保留並恢復完整 Bailey-Lovie 視力表。"""
        self.remote_mode = "full"
        self.remote_single_row_index = None
        self.remote_stimulus_index = None
        self.remote_stimulus_count = 1
        self.show_acuity_chart()
        self.status_var.set("目前為完整視力表模式；可先做整體視力測驗。")

    def remote_enter_single_mode(self) -> None:
        """進入單一刺激準備狀態；仍保留完整表，等待選擇視力值。"""
        self.remote_mode = "single"
        self.remote_single_row_index = None
        self.remote_stimulus_index = None
        self.remote_stimulus_count = 1
        self.show_acuity_chart()
        self.status_var.set("已進入刺激模式；請選擇視力值，再選擇顯示 1～5 個刺激。")

    def remote_set_distance(self, distance_cm: float) -> None:
        if not 20 <= distance_cm <= 2000:
            self.status_var.set("手機距離設定失敗：距離必須為 20～2000 cm。")
            return
        self.distance_var.set(f"{distance_cm:g}")
        # 設定距離後回到完整視力表，先做整體測驗。
        self.remote_mode = "full"
        self.remote_single_row_index = None
        self.remote_stimulus_index = None
        self.remote_stimulus_count = 1
        self.show_acuity_chart()
        self.root.update_idletasks()
        self.canvas.yview_moveto(0.0)
        self.status_var.set(f"手機已設定觀看距離：{distance_cm:g} cm；請再選擇視力列。")

    def remote_jump_to_va(self, target: float) -> None:
        """手機選擇視力值後，只保留該列，避免靠捲動定位。"""
        nearest = min(range(len(DECIMAL_LEVELS)), key=lambda i: abs(DECIMAL_LEVELS[i] - target))
        self.remote_mode = "single"
        self.remote_single_row_index = nearest
        # 選擇視力值後，預設顯示該列第一個刺激。
        self.remote_stimulus_index = 0
        self.remote_stimulus_count = 1
        self.current_view = "chart"
        self.refresh_chart()
        self.mark_new_question()
        distance_text = self.distance_var.get().strip()
        self.status_var.set(
            f"距離 {distance_text} cm；目前只顯示視力 {DECIMAL_LEVELS[nearest]:.2f} 這一列。"
        )



    def remote_set_stimulus_count(self, count: int) -> None:
        """設定同一視力列同時顯示 1～5 個刺激；5 即完整一排。"""
        if self.remote_single_row_index is None:
            self.status_var.set("請先選擇視力值，再設定刺激數量。")
            return
        if count not in (1, 2, 3, 4, 5):
            self.status_var.set("刺激數量必須為 1～5 個。")
            return
        self.remote_stimulus_count = count
        if self.remote_stimulus_index is None:
            self.remote_stimulus_index = 0
        self.refresh_chart()
        self.mark_new_question()
        self.status_var.set(
            f"視力 {DECIMAL_LEVELS[self.remote_single_row_index]:.2f}：目前顯示 {count} 個刺激。"
        )

    def remote_select_stimulus(self, index: int) -> None:
        """顯示目前視力列中的第 1～5 個單一刺激視標。"""
        if self.remote_single_row_index is None:
            self.status_var.set("請先在手機選擇視力值，再選擇刺激視標。")
            return
        if not 0 <= index <= 4:
            self.status_var.set("刺激視標必須是第 1～5 個。")
            return
        self.remote_stimulus_index = index
        self.refresh_chart()
        self.mark_new_question()

    def remote_previous_stimulus(self) -> None:
        if self.remote_single_row_index is None:
            self.status_var.set("請先選擇視力值。")
            return
        current = 0 if self.remote_stimulus_index is None else self.remote_stimulus_index
        self.remote_stimulus_index = (current - 1) % 5
        self.refresh_chart()
        self.mark_new_question()

    def remote_next_stimulus(self) -> None:
        if self.remote_single_row_index is None:
            self.status_var.set("請先選擇視力值。")
            return
        current = -1 if self.remote_stimulus_index is None else self.remote_stimulus_index
        self.remote_stimulus_index = (current + 1) % 5
        self.refresh_chart()
        self.mark_new_question()

    def remote_random_stimulus(self) -> None:
        if self.remote_single_row_index is None:
            self.status_var.set("請先選擇視力值。")
            return
        self.remote_stimulus_index = random.randrange(5)
        self.refresh_chart()
        self.mark_new_question()

    def remote_previous_va_row(self) -> None:
        """切換到上一個視力值（較大的字），保留目前刺激數量。"""
        if self.remote_single_row_index is None:
            self.status_var.set("請先選擇視力值。")
            return
        if self.remote_single_row_index <= 0:
            self.status_var.set("目前已經是最前面的視力值 0.10。")
            return
        self.remote_single_row_index -= 1
        self.remote_stimulus_index = 0
        self.refresh_chart()
        self.mark_new_question()
        self.status_var.set(
            f"已切換到上一個視力值：{DECIMAL_LEVELS[self.remote_single_row_index]:.2f}；等待受試者作答。"
        )

    def remote_next_va_row(self) -> None:
        """切換到下一個視力值（較小的字），保留目前刺激數量。"""
        if self.remote_single_row_index is None:
            self.status_var.set("請先選擇視力值。")
            return
        if self.remote_single_row_index >= len(DECIMAL_LEVELS) - 1:
            self.status_var.set("目前已經是最後一個視力值 1.58。")
            return
        self.remote_single_row_index += 1
        self.remote_stimulus_index = 0
        self.refresh_chart()
        self.mark_new_question()
        self.status_var.set(
            f"已切換到下一個視力值：{DECIMAL_LEVELS[self.remote_single_row_index]:.2f}；等待受試者作答。"
        )

    def control_html(self) -> str:
        va_buttons = "".join(
            f'<button class="va" onclick="chooseVA(\'{v}\')">{v:.2f}</button>'
            for v in DECIMAL_LEVELS
        )
        va_values_json = json.dumps([float(v) for v in DECIMAL_LEVELS])
        return """<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1,maximum-scale=1,user-scalable=no'><title>視力表遙控器</title><style>
body{font-family:-apple-system,BlinkMacSystemFont,'Microsoft JhengHei',sans-serif;background:#f3f4f6;margin:0;padding:10px;text-align:center;color:#222}h1{font-size:24px;margin:6px}.step,.resultBox,.navPanel{max-width:680px;margin:10px auto;background:white;border-radius:16px;padding:13px;box-shadow:0 1px 6px #c8c8c8;box-sizing:border-box}.grid{display:grid;grid-template-columns:repeat(3,1fr);gap:9px}.modeGrid,.actionGrid{display:grid;grid-template-columns:1fr 1fr;gap:10px}.functionGrid{display:grid;grid-template-columns:repeat(2,1fr);gap:9px}.functionBtn{min-height:62px;background:#f7f9ff;color:#173b76;border:2px solid #8aa7d8}.stimGrid{display:grid;grid-template-columns:repeat(4,1fr);gap:7px}button{font-size:18px;font-weight:800;min-height:58px;border:1px solid #999;border-radius:12px;background:white;padding:8px}.active{background:#174ea6!important;color:white!important}.homeBtn{width:100%;background:#edf8ed;color:#176b27;border:2px solid #58a966}.randomBtn{width:100%;background:#fff4e5;color:#9a4b00;border:2px solid #e0a044;margin-bottom:10px}.sendBtn{background:#174ea6;color:white;border:2px solid #174ea6;font-size:21px}.clearBtn{background:#fff4e5;color:#9a4b00;border:2px solid #e0a044;font-size:21px}.resultBox{background:#eef5ff;border:2px solid #8ab0ee}.resultLine{font-size:18px;font-weight:700;margin:5px}.correct{color:#14823b}.wrong{color:#c62828}.waiting{color:#555}.navGrid{display:grid;grid-template-columns:1fr 1.15fr 1fr;grid-template-rows:auto auto auto;gap:9px}.navBtn{border:2px solid #4f7edb;color:#174ea6}.upBtn{grid-column:2;grid-row:1}.downBtn{grid-column:2;grid-row:3}.leftBtn{grid-column:1;grid-row:2}.rightBtn{grid-column:3;grid-row:2}.vaCenter{grid-column:2;grid-row:2;background:#174ea6;color:white;border-radius:14px;display:flex;flex-direction:column;justify-content:center;min-height:78px}.vaCenter strong{font-size:32px}.hint{font-size:15px;color:#555;margin-top:8px}.mainTitle{font-size:27px;margin:5px}.mainHint{font-size:17px;color:#444;line-height:1.5}.distanceRow{display:grid;grid-template-columns:1fr auto auto;gap:8px;align-items:center}.distanceInput{font-size:24px;font-weight:800;text-align:center;min-width:0;height:58px;border:2px solid #4f7edb;border-radius:12px;padding:0 8px;box-sizing:border-box}.distanceUnit{font-size:21px;font-weight:800}.applyBtn{background:#174ea6;color:white;border:2px solid #174ea6}.letterBtn{width:100%;background:#edf8ed;color:#176b27;border:2px solid #58a966;margin-top:10px}.distanceNow{font-size:17px;font-weight:800;color:#174ea6;margin-top:8px}.stickerGrid{display:grid;grid-template-columns:repeat(4,1fr);gap:8px}.stickerBtn{font-size:34px;min-height:66px;background:#fffdf5;border:2px solid #f0b84f}.collectionLine{font-size:25px;line-height:1.5;word-break:break-all}.pendingLine{font-size:25px;font-weight:900;min-height:42px;margin:10px 0;color:#9a4b00}.sendStickerBtn{width:100%;background:#14823b;color:white;border:2px solid #14823b;font-size:22px;margin-top:8px}.hidden{display:none!important}
</style></head><body><h1>📱 測驗者模式｜視力表遙控器</h1><div id='status'>正在連線…</div>
<div id='mainScreen'>
  <div class='step'>
    <div class='mainTitle'><b>主畫面／完整視力表</b></div>
    <div class='mainHint'>先設定觀看距離，再選擇要開始的視力值。</div>
  </div>
  <div id='estimateNotice' style='display:none;position:sticky;top:8px;z-index:50;margin:8px auto;padding:14px 16px;max-width:680px;border:2px solid #16863b;border-radius:14px;background:#eaf8ef;color:#0d6b2c;font-size:22px;font-weight:900;text-align:center;box-shadow:0 4px 16px rgba(0,0,0,.16)'>已收到受試者初估視力</div>
  <div class='step' id='mainEstimateCard'>
    <h2>受試者初估視力回報</h2>
    <div id='mainEstimatedVA' class='resultLine waiting'>尚未收到初估視力</div>
  </div>
  <div class='step'>
    <h2>測驗功能主選單</h2>
    <div class='functionGrid'>
      <button class='functionBtn' onclick="sendCmd('chart')">完整視力表</button>
      <button class='functionBtn' onclick="sendCmd('dial')">散光鐘</button>
      <button class='functionBtn' onclick="sendCmd('amsler')">阿姆斯勒方格</button>
      <button class='functionBtn' onclick="sendCmd('worth')">Worth 四點</button>
      <button class='functionBtn' onclick="sendCmd('worth_focus')">純四點</button>
      <button class='functionBtn' onclick="sendCmd('bagolini')">Bagolini</button>
      <button class='functionBtn' onclick="sendCmd('thorington')">Thorington</button>
      <button class='functionBtn' onclick="sendCmd('focus')">單點全黑</button>
      <button class='functionBtn' onclick="sendCmd('hide_controls')">顯示／隱藏設定</button>
    </div>
  </div>
  <div class='step'>
    <h2>① 選擇視標</h2>
    <div class='grid'><button id='optLetter' onclick="sendCmd('optotype','letter')">Sloan Letter</button><button id='optC' onclick="sendCmd('optotype','landolt_c')">Landolt C</button><button id='optE' onclick="sendCmd('optotype','tumbling_e')">Tumbling E</button></div>
    <h2>② 設定觀看距離</h2>
    <div class='distanceRow'><input id='distanceInput' class='distanceInput' type='number' inputmode='decimal' min='1' step='1' value='300'><span class='distanceUnit'>cm</span><button class='applyBtn' onclick='applyDistance()'>套用</button></div>
    <div id='distanceNow' class='distanceNow'>目前距離：300 cm</div>
    <button class='letterBtn' onclick="sendCmd('random')">重新排列視標</button>
  </div>
  <div class='step'>
    <button class='randomBtn' onclick='chooseRandomVA()'>隨機選擇一個視力值</button>
    <h2>③ 選擇起始視力值</h2>
    <div class='grid'>""" + va_buttons + """</div>
  </div>
</div>
<div id='operationScreen' class='hidden'>
  <div class='step'><button class='homeBtn' onclick='goHome()'>⌂ 回主畫面／完整視力表</button><button class='letterBtn' onclick="sendCmd('random')">重新排列視標</button><div id='operationDistance' class='distanceNow'>目前距離：300 cm</div></div>
  <div class='step'><h2>測試模式</h2><div class='modeGrid'><button id='adultBtn' onclick="setMode('adult')">成人模式<br><small>只有測驗者端</small></button><button id='childBtn' onclick="setMode('child')">兒童模式<br><small>手機作答與自動判分</small></button></div><div id='modeState' class='resultLine'>目前模式：兒童模式</div></div>
  <div class='navPanel'><h2>目前測驗操作</h2><div class='navGrid'><button class='navBtn upBtn' onclick="sendCmd('va_prev')">↑ 上一列<span id='upLabel'></span></button><button class='navBtn leftBtn' onclick="sendCmd('stim_prev')">← 左邊</button><div class='vaCenter'><small>目前視力值</small><strong id='currentVA'>0.50</strong></div><button class='navBtn rightBtn' onclick="sendCmd('stim_next')">右邊 →</button><button class='navBtn downBtn' onclick="sendCmd('va_next')">↓ 下一列<span id='downLabel'></span></button></div></div>
  <div class='step'><h2>顯示刺激數量</h2><div class='stimGrid'><button id='stim5' onclick='setStim(5)'>5 個</button><button id='stim4' onclick='setStim(4)'>4 個</button><button id='stim3' onclick='setStim(3)'>3 個</button><button id='stim2' onclick='setStim(2)'>2 個</button><button id='stim1' onclick='setStim(1)'>1 個</button></div></div>
  <div class='resultBox'><h2>本題狀態與結果</h2><div id='questionStatus' class='resultLine waiting'>目前狀態：等待發送</div><div id='resultQuestion' class='resultLine'>目前預覽：尚未選題</div><div id='resultAnswer' class='resultLine'>受試者答案：尚未作答</div><div id='estimatedVA' class='resultLine'>受試者初估視力：尚未回報</div><div id='resultJudge' class='resultLine waiting'></div><div id='resultTime' class='resultLine'></div></div>
  <div class='step'><h2>答對後，由測驗者選擇貼圖</h2><div class='stickerGrid'><button class="stickerBtn" onclick="sendCmd('sticker','🍎')">🍎</button><button class="stickerBtn" onclick="sendCmd('sticker','⭐')">⭐</button><button class="stickerBtn" onclick="sendCmd('sticker','🐶')">🐶</button><button class="stickerBtn" onclick="sendCmd('sticker','🚗')">🚗</button><button class="stickerBtn" onclick="sendCmd('sticker','🦖')">🦖</button><button class="stickerBtn" onclick="sendCmd('sticker','🐱')">🐱</button><button class="stickerBtn" onclick="sendCmd('sticker','🧸')">🧸</button><button class="stickerBtn" onclick="sendCmd('sticker','🐼')">🐼</button><button class="stickerBtn" onclick="sendCmd('sticker','🚀')">🚀</button><button class="stickerBtn" onclick="sendCmd('sticker','👑')">👑</button></div><div id='pendingStickers' class='pendingLine'>已選：尚未選擇貼圖</div><div class='actionGrid'><button class='clearBtn' onclick="sendCmd('clear_pending_stickers')">清除選擇</button><button class='sendStickerBtn' onclick="sendCmd('send_stickers')">📤 發送貼圖</button></div><div id='stickerCollection' class='collectionLine'>目前收藏：0 個</div><button class='clearBtn' onclick="sendCmd('clear_stickers')">清空貼圖收藏</button><div class='hint'>可以先按兩次蘋果，再按「發送貼圖」一次送出兩顆蘋果。</div></div>
  <div class='step'><div class='actionGrid'><button class='sendBtn' onclick="sendCmd('send_question')">▶ 開始作答</button><button class='clearBtn' onclick="sendCmd('clear_question')">確認結果並清除</button></div><div class='hint'>兒童模式：發送 → 作答判分 → 測驗者送貼圖 → 清除 → 下一題。</div></div>
</div>
<script>
const SESSION='""" + self.connection_session_token + """';
let connectionFailures=0,lastEstimatedVA=null,lastEstimatedVersion=0;
function markOnline(){connectionFailures=0;const e=document.getElementById('status');if(e){e.textContent='✅ 已連線到電腦';e.style.color='#14823b'}}
function markOffline(){connectionFailures++;if(connectionFailures>=2){const e=document.getElementById('status');if(e){e.textContent='❌ 與電腦斷線，請確認仍連著 JohnVA／同一 Wi-Fi';e.style.color='#c62828'}}}
const VA_VALUES=""" + va_values_json + """;
async function sendCmd(cmd,value=''){
  try{
    const r=await fetch('/command?role=control&session='+encodeURIComponent(SESSION)+'&cmd='+encodeURIComponent(cmd)+'&value='+encodeURIComponent(value)+'&t='+Date.now(),{cache:'no-store'});
    if(!r.ok)throw new Error('HTTP '+r.status);
    markOnline();
    // 指令已送達後，立即重讀一次最新狀態，不等待下一輪計時器。
    await refresh(true);
    await refreshEstimatedVADirect();
  }catch(e){markOffline();}
}
function setMode(m){sendCmd(m==='adult'?'mode_adult':'mode_child')}
function showMain(){document.getElementById('mainScreen').classList.remove('hidden');document.getElementById('operationScreen').classList.add('hidden')}
function showOperation(){document.getElementById('mainScreen').classList.add('hidden');document.getElementById('operationScreen').classList.remove('hidden')}
function goHome(){sendCmd('mode_full');showMain()}
function chooseVA(v){sendCmd('mode_single');setTimeout(()=>sendCmd('va',v),120);showOperation()}
function chooseRandomVA(){const v=VA_VALUES[Math.floor(Math.random()*VA_VALUES.length)];chooseVA(v)}
function applyDistance(){const el=document.getElementById('distanceInput');const v=Number(el.value);if(!Number.isFinite(v)||v<=0){alert('請輸入正確的觀看距離（cm）');return}sendCmd('distance',String(v))}
function setStim(n){sendCmd('stim_count',String(n))}
function updateEstimatedVA(value,version){
  const text=value==null?'尚未收到初估視力':'✅ 已收到初估視力：'+Number(value).toFixed(2);
  const main=document.getElementById('mainEstimatedVA');
  if(main){main.textContent=text;main.className='resultLine '+(value==null?'waiting':'correct');}
  const operation=document.getElementById('estimatedVA');
  if(operation)operation.textContent='受試者初估視力：'+(value==null?'尚未回報':Number(value).toFixed(2));
  const notice=document.getElementById('estimateNotice');
  if(value!=null&&notice){notice.style.display='block';notice.textContent='✅ 已收到受試者初估視力：'+Number(value).toFixed(2);}
  const v=Number(version||0);
  if(value!=null&&v>lastEstimatedVersion){
    lastEstimatedVersion=v;
    lastEstimatedVA=Number(value);
    if(navigator.vibrate)navigator.vibrate([120,80,120]);
  }
}

async function refreshEstimatedVADirect(){
  try{
    const r=await fetch('/control_estimate?role=control&session='+encodeURIComponent(SESSION)+'&t='+Date.now(),{cache:'no-store'});
    if(!r.ok)throw new Error('state '+r.status);
    const d=await r.json();
    updateEstimatedVA(d.estimated_va,d.estimated_va_version);
  }catch(e){}
}
let controlRefreshBusy=false;
let controlPollTimer=null;
function reconnectNow(){connectionFailures=0;startControlPolling(true);refreshEstimatedVADirect()}
window.addEventListener('online',reconnectNow);
window.addEventListener('focus',reconnectNow);
window.addEventListener('pageshow',reconnectNow);
document.addEventListener('visibilitychange',()=>{if(!document.hidden)reconnectNow()});
function setParticipantDialScale(v){participantDialScale=Math.max(.60,Math.min(1.80,Number(v)||1));participantSafeStore(DIAL_SCALE_KEY,participantDialScale);renderParticipantDialScale();if(typeof drawDial==='function')drawDial()}function adjustParticipantDialScale(v){setParticipantDialScale(participantDialScale+v)}function renderParticipantDialScale(){const r=document.getElementById('dialSizeReadout');if(r)r.textContent='倍率 '+Math.round(participantDialScale*100)+'%';const c=document.getElementById('dialCanvas');if(c){c.style.transform='scale('+participantDialScale+')';c.style.transformOrigin='center center'}}
function renderParticipantAmsler(){
  const grid=document.getElementById('amslerGrid');
  const svg=document.getElementById('amslerSvg');
  if(!grid||!svg||!(pxPerCm>0))return;
  // 這台 iPad 實測舊版標示 10 cm 時只有約 5.7 cm，先套用固定補償；
  // 再讓使用者用下方按鈕微調，直到實體尺量到真正 10 cm。
  const AMSLER_DEVICE_CORRECTION=10/5.7;
  const size=10*pxPerCm*AMSLER_DEVICE_CORRECTION*amslerManualScale;
  const fixed=size.toFixed(1)+'px';
  grid.style.setProperty('width',fixed,'important');
  grid.style.setProperty('height',fixed,'important');
  grid.style.setProperty('min-width',fixed,'important');
  grid.style.setProperty('min-height',fixed,'important');
  grid.style.setProperty('max-width','none','important');
  grid.style.setProperty('max-height','none','important');
  grid.style.setProperty('flex','0 0 '+fixed,'important');
  grid.style.setProperty('transform','none','important');
  svg.setAttribute('viewBox','0 0 1000 1000');
  svg.setAttribute('width',size.toFixed(1));
  svg.setAttribute('height',size.toFixed(1));
  svg.style.width='100%';svg.style.height='100%';svg.style.display='block';
  // 每次顯示都重畫，避免 iPad Safari 從隱藏畫面切換後漏掉格線。
  const ns='http://www.w3.org/2000/svg';
  while(svg.firstChild)svg.removeChild(svg.firstChild);
  const bg=document.createElementNS(ns,'rect');
  bg.setAttribute('x','0');bg.setAttribute('y','0');bg.setAttribute('width','1000');bg.setAttribute('height','1000');bg.setAttribute('fill','#fff');svg.appendChild(bg);
  for(let i=0;i<=20;i++){
    const pos=i*50;
    const v=document.createElementNS(ns,'line');v.setAttribute('x1',String(pos));v.setAttribute('y1','0');v.setAttribute('x2',String(pos));v.setAttribute('y2','1000');v.setAttribute('stroke','#000');v.setAttribute('stroke-width',i===0||i===20?'3':'2');svg.appendChild(v);
    const h=document.createElementNS(ns,'line');h.setAttribute('x1','0');h.setAttribute('y1',String(pos));h.setAttribute('x2','1000');h.setAttribute('y2',String(pos));h.setAttribute('stroke','#000');h.setAttribute('stroke-width',i===0||i===20?'3':'2');svg.appendChild(h);
  }
  const dot=document.createElementNS(ns,'circle');dot.setAttribute('cx','500');dot.setAttribute('cy','500');dot.setAttribute('r','10');dot.setAttribute('fill','#000');svg.appendChild(dot);
  const estimatedCm=10*AMSLER_DEVICE_CORRECTION*amslerManualScale;
  const readout=document.getElementById('amslerSizeReadout');
  if(readout)readout.textContent='目前倍率 '+amslerManualScale.toFixed(2)+'（請用尺量）';
  const info=document.getElementById('amslerInfo');
  if(info)info.textContent='標準測試距離 30 cm｜20 × 20 格｜目標實體大小 10 × 10 cm';
}
function safeStore(key,value){try{localStorage.setItem(key,String(value));}catch(e){}}
function adjustAmslerScale(delta){
  amslerManualScale=Math.max(0.35,Math.min(3.5,amslerManualScale+Number(delta||0)));
  renderParticipantAmsler();renderParticipantDialScale();
  safeStore(AMSLER_SCALE_KEY,amslerManualScale);
  const status=document.getElementById('amslerStatus');
  if(status)status.textContent='已調整阿姆斯勒倍率：'+amslerManualScale.toFixed(2)+'，請用尺確認實際寬度';
}
function resetAmslerScale(){
  amslerManualScale=1;
  renderParticipantAmsler();
  safeStore(AMSLER_SCALE_KEY,1);
  const status=document.getElementById('amslerStatus');
  if(status)status.textContent='已恢復阿姆斯勒預設倍率，請再用尺確認 10 公分';
}
function renderParticipantBagolini(mode,dotPx){
  const dot=document.getElementById('bagoliniDot');
  if(dot){
    const d=Math.max(2,Math.min(5,Number(dotPx)||3));
    dot.style.width=d+'px';dot.style.height=d+'px';
  }
  const info=document.getElementById('bagoliniInfo');
  const near=(mode||'near')==='near';
  if(info)info.textContent=(near?'近距模式：33 cm':'遠距模式：6 m')+'｜中央純白光點 '+Math.max(2,Math.min(5,Number(dotPx)||3))+' px';
}
function renderParticipantWorth(mode){
  // 受測者端固定為純刺激畫面：全黑背景，只顯示四個光點。
  // 不顯示近距／遠距文字、標題或按鈕。
  const stage=document.getElementById('worthStage');
  if(!stage)return;
  const shortSide=Math.min(window.innerWidth,window.innerHeight);
  const dot=Math.max(18,Math.min(34,shortSide*0.035));
  const dx=Math.max(80,Math.min(165,shortSide*0.18));
  const dy=Math.max(72,Math.min(145,shortSide*0.16));
  stage.style.setProperty('--worth-dot',dot.toFixed(1)+'px');
  stage.style.setProperty('--worth-dx',dx.toFixed(1)+'px');
  stage.style.setProperty('--worth-dy',dy.toFixed(1)+'px');
}
function renderParticipantThorington(distanceCm,dotPx){
  const svg=document.getElementById('thoringtonSvg');
  if(!svg)return;
  const ns='http://www.w3.org/2000/svg';
  const W=1400,H=440,cx=700,cy=260,startX=100,endX=1300,step=50;
  const baseWidth=Math.max(340,window.innerWidth*0.94);
  const widthPx=baseWidth*thoringtonManualScale;
  const heightPx=widthPx*(H/W);

  // iPad Safari 在隱藏畫面切換或縮放後，可能只保留黑底而漏掉 SVG 線條。
  // 每次顯示與每次縮放都重新建立完整刺激，確保白線、刻度與文字一定出現。
  while(svg.firstChild)svg.removeChild(svg.firstChild);
  svg.setAttribute('viewBox','0 0 '+W+' '+H);
  svg.setAttribute('preserveAspectRatio','xMidYMid meet');
  svg.setAttribute('width',widthPx.toFixed(1));
  svg.setAttribute('height',heightPx.toFixed(1));
  svg.style.setProperty('width',widthPx.toFixed(1)+'px','important');
  svg.style.setProperty('height',heightPx.toFixed(1)+'px','important');
  svg.style.setProperty('min-width',widthPx.toFixed(1)+'px','important');
  svg.style.setProperty('min-height',heightPx.toFixed(1)+'px','important');
  svg.style.setProperty('max-width','none','important');
  svg.style.setProperty('max-height','none','important');
  svg.style.setProperty('display','block','important');
  svg.style.setProperty('visibility','visible','important');
  svg.style.setProperty('opacity','1','important');

  function el(name,attrs,text){
    const node=document.createElementNS(ns,name);
    Object.keys(attrs||{}).forEach(k=>node.setAttribute(k,String(attrs[k])));
    if(text!==undefined)node.textContent=text;
    svg.appendChild(node);
    return node;
  }
  el('rect',{x:0,y:0,width:W,height:H,fill:'#000'});
  el('text',{x:cx,y:58,fill:'#f4f4f4','font-size':34,'font-family':'Arial','font-weight':700,'text-anchor':'middle'},'Modified Thorington 水平眼位測量');
  const mode=Number(distanceCm)===40?'近距 40 cm':(Number(distanceCm)===300?'遠距 3 m':'中距 1 m');
  el('text',{id:'thoringtonModeText',x:cx,y:100,fill:'#d7dce3','font-size':25,'font-family':'Arial','text-anchor':'middle'},mode+'｜每格 1Δ');

  // 主白線先畫，使用明顯純白與較粗線寬。
  el('line',{x1:startX,y1:cy,x2:endX,y2:cy,stroke:'#ffffff','stroke-width':5,'stroke-linecap':'round'});
  const letters='LKJIHGFEDCBA';
  for(let i=-12;i<=12;i++){
    const x=cx+i*step;
    const isCenter=i===0;
    el('line',{x1:x,y1:cy-14,x2:x,y2:cy+18,stroke:isCenter?'#ffffff':'#e7edf5','stroke-width':isCenter?4:3});
    if(i<0){
      el('text',{x:x,y:220,fill:'#ffffff','font-size':30,'font-family':'Arial','font-weight':600,'text-anchor':'middle'},letters[i+12]);
    }else if(i>0){
      el('text',{x:x,y:220,fill:'#ffffff','font-size':30,'font-family':'Arial','font-weight':600,'text-anchor':'middle'},String(i));
    }
  }
  const r=Math.max(5,Math.min(10,Number(dotPx)||5));
  el('circle',{id:'thoringtonCenterDot',cx:cx,cy:cy,r:r,fill:'#ffffff',stroke:'#ffffff','stroke-width':2});

  const readout=document.getElementById('thoringtonSizeReadout');
  if(readout)readout.textContent='目前倍率 '+thoringtonManualScale.toFixed(2);
  // Safari 需要下一個繪圖週期再確認尺寸，避免全螢幕切換後只剩黑底。
  requestAnimationFrame(()=>{
    svg.setAttribute('width',widthPx.toFixed(1));
    svg.setAttribute('height',heightPx.toFixed(1));
    svg.style.width=widthPx.toFixed(1)+'px';
    svg.style.height=heightPx.toFixed(1)+'px';
  });
}
function adjustThoringtonScale(delta){
  thoringtonManualScale=Math.max(0.35,Math.min(3.5,thoringtonManualScale+Number(delta||0)));
  renderParticipantThorington(lastThoringtonDistance,lastThoringtonDotPx);
  safeStore(THORINGTON_SCALE_KEY,thoringtonManualScale);
}
function resetThoringtonScale(){
  thoringtonManualScale=1;
  renderParticipantThorington(lastThoringtonDistance,lastThoringtonDotPx);
  safeStore(THORINGTON_SCALE_KEY,1);
}
async function refresh(force=false){
  if(controlRefreshBusy&&!force)return;
  controlRefreshBusy=true;
  try{
    const controller=new AbortController();
    const timeout=setTimeout(()=>controller.abort(),2500);
    const r=await fetch('/state?role=control&session='+encodeURIComponent(SESSION)+'&t='+Date.now(),{cache:'no-store',signal:controller.signal});
    clearTimeout(timeout);
    if(!r.ok)throw new Error('HTTP '+r.status);
    const d=await r.json();
    markOnline();
    updateEstimatedVA(d.estimated_va,d.estimated_va_version);
    if(d.va==null){showMain();return}else{showOperation()}
    const map={waiting_send:'等待發送',awaiting_answer:'等待受試者作答',completed:'判分完成'};
    document.getElementById('questionStatus').textContent='目前狀態：'+(map[d.question_status]||d.question_status||'等待發送');
    const shown=d.question_status==='waiting_send'?(d.preview_targets||[]):(d.active_targets||[]);
    document.getElementById('resultQuestion').textContent=(d.question_status==='waiting_send'?'目前預覽：':'本題題目：')+(shown.length?shown.join(' '):'尚未選題');
    const answers=d.answers||[];
    document.getElementById('resultAnswer').textContent='受試者答案：'+(answers.length?answers.join(' '):'尚未作答');
    const j=document.getElementById('resultJudge');
    if(d.question_status==='completed'){
      const active=d.active_targets||[];
      j.textContent='結果：'+(d.result_marks||[]).join(' ')+'　正確 '+Number(d.correct_count||0)+'/'+active.length;
      j.className='resultLine '+(Number(d.correct_count||0)===active.length?'correct':'wrong');
    }else{
      j.textContent=d.question_status==='awaiting_answer'?'作答進度：'+answers.length+'/'+(d.active_targets||[]).length:'';
      j.className='resultLine waiting';
    }
    document.getElementById('resultTime').textContent=d.elapsed==null?'':'反應時間：'+Number(d.elapsed).toFixed(2)+' 秒';
    document.getElementById('currentVA').textContent=Number(d.va).toFixed(2);
    document.getElementById('upLabel').textContent=d.prev_va==null?'':' '+Number(d.prev_va).toFixed(2);
    document.getElementById('downLabel').textContent=d.next_va==null?'':' '+Number(d.next_va).toFixed(2);
    for(let i=1;i<=5;i++)document.getElementById('stim'+i).classList.toggle('active',i===Number(d.stimulus_count));
    document.getElementById('adultBtn').classList.toggle('active',d.test_mode==='adult');
    document.getElementById('childBtn').classList.toggle('active',d.test_mode==='child');
    document.getElementById('modeState').textContent='目前模式：'+(d.test_mode==='adult'?'成人模式':'兒童模式');
    const om=d.optotype_mode||'letter';
    const ol=document.getElementById('optLetter'),oc=document.getElementById('optC'),oe=document.getElementById('optE');
    if(ol){ol.classList.toggle('active',om==='letter');oc.classList.toggle('active',om==='landolt_c');oe.classList.toggle('active',om==='tumbling_e');}
    const sc=document.getElementById('stickerCollection');if(sc)sc.textContent='目前收藏：'+(d.sticker_count||0)+' 個　'+((d.stickers||[]).join(' '));
    const ps=document.getElementById('pendingStickers');if(ps)ps.textContent='已選：'+((d.pending_stickers||[]).length?(d.pending_stickers||[]).join(' '):'尚未選擇貼圖');
    if(d.distance_cm!=null){
      const txt='目前距離：'+Number(d.distance_cm).toFixed(Number(d.distance_cm)%1?1:0)+' cm';
      document.getElementById('distanceNow').textContent=txt;
      document.getElementById('operationDistance').textContent=txt;
      const input=document.getElementById('distanceInput');if(document.activeElement!==input)input.value=Number(d.distance_cm).toFixed(Number(d.distance_cm)%1?1:0);
    }
  }catch(e){markOffline();}
  finally{controlRefreshBusy=false;}
}
async function controlPollLoop(){
  await refresh();
  await refreshEstimatedVADirect();
  controlPollTimer=setTimeout(controlPollLoop,400);
}
function startControlPolling(immediate=false){
  if(controlPollTimer){clearTimeout(controlPollTimer);controlPollTimer=null;}
  if(immediate){refresh(true);refreshEstimatedVADirect();}
  controlPollTimer=setTimeout(controlPollLoop,150);
}
ensureCalibration();startControlPolling(true);
</script></body></html>"""

    def answer_html(self) -> str:
        buttons = "".join(
            f'<button class="answer letterAnswer" onclick="sendAnswer(\'{letter}\')">{letter}</button>'
            for letter in SLOAN_LETTERS
        ) + "".join(
            f'<button class="answer directionAnswer cardinalAnswer hidden" onclick="sendAnswer(\'{symbol}\')">{symbol}</button>'
            for symbol in ("↑", "→", "↓", "←")
        ) + "".join(
            f'<button class="answer directionAnswer diagonalAnswer hidden" onclick="sendAnswer(\'{symbol}\')">{symbol}</button>'
            for symbol in ("↗", "↘", "↙", "↖")
        )
        return """<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'>
<meta name='viewport' content='width=device-width,initial-scale=1,maximum-scale=1,user-scalable=no,viewport-fit=cover'>
<meta name='apple-mobile-web-app-capable' content='yes'><meta name='mobile-web-app-capable' content='yes'>
<title>Cloud Vision V10.7 受試者端</title><style>
*{box-sizing:border-box;-webkit-tap-highlight-color:transparent}html,body{width:100%;height:100%;margin:0;background:#fff;overscroll-behavior:none}body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;color:#111}.hidden{display:none!important}button{font:inherit;touch-action:manipulation}.screen{min-height:100vh;width:100%;padding:max(18px,env(safe-area-inset-top)) max(18px,env(safe-area-inset-right)) max(20px,env(safe-area-inset-bottom)) max(18px,env(safe-area-inset-left));background:#fff}.center{display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center}.brand{font-size:18px;font-weight:800;letter-spacing:.04em;color:#254b87;margin-bottom:8px}.calCard{width:min(760px,96vw);border:2px solid #234f98;border-radius:22px;padding:28px 20px;box-shadow:0 10px 30px rgba(0,0,0,.08)}.calCard h1{font-size:clamp(28px,5vw,44px);margin:0 0 12px}.calCard p{font-size:clamp(18px,2.5vw,25px);line-height:1.55;margin:8px 0}.calLineBox{width:100%;overflow:hidden;padding:44px 0 30px}.calLine{height:10px;background:#111;margin:auto;position:relative}.calLine:before,.calLine:after{content:"";position:absolute;top:-13px;width:4px;height:36px;background:#111}.calLine:before{left:0}.calLine:after{right:0}.calValue{font-size:20px;font-weight:800;margin:4px 0 18px}.calAdjust{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-top:8px}.calBtn{min-height:72px;border:2px solid #315b9f;border-radius:16px;background:#fff;color:#21477f;font-size:24px;font-weight:800}.calDone{width:100%;min-height:76px;border:0;border-radius:16px;background:#174f9b;color:#fff;font-size:24px;font-weight:900;margin-top:16px}.calNote{font-size:16px;color:#666;margin-top:14px}.waitWrap{min-height:100vh}.waitIcon{font-size:58px}.waitTitle{font-size:clamp(34px,6vw,64px);font-weight:900;margin:14px 0}.waitSub{font-size:clamp(19px,3vw,28px);color:#555}.chartWrap{min-height:auto;padding-top:70px;padding-bottom:140px}#chartScreen{height:100dvh;min-height:100dvh;overflow-y:auto;-webkit-overflow-scrolling:touch;overscroll-behavior-y:contain;padding-bottom:max(140px,calc(env(safe-area-inset-bottom) + 120px))}.chartTitle{font-size:clamp(26px,4vw,40px);font-weight:900;text-align:center;margin:0 0 8px}.chartHint{text-align:center;font-size:18px;color:#555;margin-bottom:14px}.acuityChart{width:min(900px,100%);margin:0 auto 18px;display:flex;flex-direction:column;align-items:center;gap:5px}.acuityRow{display:grid;grid-template-columns:62px 1fr 62px;align-items:center;width:100%}.acuityLabel{font-size:15px;font-weight:800;color:#333}.acuityLetters{text-align:center;font-family:Arial,sans-serif;font-weight:900;line-height:1;white-space:nowrap}.estimateCard{width:min(760px,100%);margin:12px auto;border:2px solid #315b9f;border-radius:18px;padding:18px}.estimateGrid{display:grid;grid-template-columns:repeat(4,1fr);gap:9px;margin:12px 0}.estimateBtn{min-height:58px;border:2px solid #315b9f;border-radius:12px;background:#fff;color:#21477f;font-size:20px;font-weight:900}.estimateRow{display:grid;grid-template-columns:1fr auto;gap:10px}.eyeAcuityForm{display:grid;gap:14px;margin-top:18px}.eyeAcuityRow{display:grid;grid-template-columns:140px 1fr;align-items:center;gap:12px;font-size:22px;font-weight:900}.eyeAcuityRow span{text-align:right}.eyeSubmitBtn{width:100%;margin-top:4px;min-height:58px}.estimateCard .estimateInput{min-height:54px;padding:8px 12px}@media(max-width:650px){.eyeAcuityRow{grid-template-columns:1fr}.eyeAcuityRow span{text-align:left}}.estimateInput{min-width:0;border:2px solid #315b9f;border-radius:12px;font-size:22px;text-align:center}.estimateSend{border:0;border-radius:12px;background:#174f9b;color:#fff;font-weight:900;padding:12px 18px;font-size:20px}.topActions{position:fixed;right:14px;top:max(12px,env(safe-area-inset-top));z-index:5;display:flex;gap:8px}.estimateLauncher{position:fixed;left:50%;bottom:max(14px,env(safe-area-inset-bottom));transform:translateX(-50%);z-index:20;border:0;border-radius:16px;background:#174f9b;color:#fff;font-size:20px;font-weight:900;padding:15px 28px;box-shadow:0 6px 20px rgba(0,0,0,.25)}.estimateModal{position:fixed;inset:0;z-index:30;background:rgba(0,0,0,.42);display:flex;align-items:center;justify-content:center;padding:20px}.estimateModalCard{width:min(720px,96vw);max-height:90vh;overflow-y:auto;background:#fff;border:2px solid #315b9f;border-radius:20px;padding:22px;box-shadow:0 12px 40px rgba(0,0,0,.3)}.estimateModalHeader{display:flex;justify-content:space-between;align-items:center;gap:12px;margin-bottom:8px}.estimateModalTitle{font-size:26px;font-weight:900}.estimateClose{border:1px solid #777;border-radius:12px;background:#fff;font-size:22px;padding:8px 14px}.estimateModal .estimateGrid{grid-template-columns:repeat(4,1fr)}.smallBtn{border:1px solid #777;border-radius:12px;background:#fff;padding:10px 14px;font-weight:700}.testScreen{display:flex;flex-direction:column;min-height:100vh}.testHeader{display:flex;justify-content:space-between;align-items:center;gap:12px;margin-bottom:10px}.progress{font-size:clamp(18px,2.5vw,28px);font-weight:900;color:#8b251d}.sizeInfo{font-size:15px;color:#315b9f}.questionPanel{flex:0 0 auto;min-height:130px;border:2px solid #2d4f96;border-radius:20px;display:flex;align-items:center;justify-content:center;padding:16px 18px;margin:6px 0 14px;overflow:hidden}.questionDisplay{font-family:Arial,sans-serif;font-weight:900;white-space:nowrap;line-height:1;text-align:center;display:flex;align-items:center;justify-content:center;gap:var(--opt-gap,12px)}.optotypeSvg{display:block;flex:0 0 auto;width:var(--opt-size,24px);height:var(--opt-size,24px)}.correctionBar{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:14px}.correctionBtn{min-height:54px;border-radius:14px;border:2px solid #b46a22;background:#fff7ec;color:#7e4210;font-size:18px;font-weight:800}.resetBtn{border-color:#b44253;background:#fff3f5;color:#8b2332}.correctionBtn:disabled{opacity:.3}.answerArea{margin-top:auto}.answerGrid{display:grid;grid-template-columns:repeat(4,1fr);gap:10px}.answer{min-height:88px;border:2px solid #222;border-radius:16px;background:#fff;font-size:clamp(32px,5vw,56px);font-weight:900}.answer:active{transform:scale(.97);background:#eef4ff}.answer:disabled{opacity:.28}.directionAnswer{grid-column:auto}.answerGrid.directionMode{grid-template-columns:repeat(4,1fr)}.status{font-size:16px;color:#14823b;margin-top:10px;min-height:24px}.collection{border:1px solid #d9a62c;border-radius:16px;padding:10px 14px;margin-bottom:12px}.collectionTitle{font-weight:800;color:#8a3c1c}.stickers{font-size:34px;min-height:38px}.newSticker{animation:pop .55s ease}@keyframes pop{0%{transform:scale(.75)}60%{transform:scale(1.15)}100%{transform:scale(1)}}.dialScreen{display:flex;flex-direction:column;min-height:100vh}.dialHeader{display:flex;justify-content:space-between;align-items:center;gap:12px}.dialTitle{font-size:clamp(26px,4vw,42px);font-weight:900}.dialInfo{font-size:16px;color:#315b9f;margin-top:4px}.dialWrap{flex:1;display:flex;align-items:center;justify-content:center;min-height:420px}.dialCanvas{width:100%;height:100%;min-height:420px;display:block}.dialControls{display:flex;justify-content:center;align-items:center;gap:9px;flex-wrap:wrap;padding:8px 10px}.dialAdjustBtn{min-width:92px;min-height:48px;border:2px solid #315b9f;border-radius:13px;background:#fff;color:#21477f;font-size:18px;font-weight:900}.dialSizeReadout{font-size:18px;font-weight:900;color:#21477f;min-width:135px;text-align:center}.dialStatus{text-align:center;font-size:18px;color:#14823b;font-weight:800;padding:8px 0}.amslerScreen{display:flex;flex-direction:column;min-height:100vh;overflow:auto}.amslerHeader{display:flex;justify-content:space-between;align-items:center;gap:12px}.amslerTitle{font-size:clamp(26px,4vw,42px);font-weight:900}.amslerInfo{font-size:16px;color:#315b9f;margin-top:4px}.amslerWrap{flex:1;display:flex;align-items:center;justify-content:center;min-height:420px;padding:10px;overflow:auto}.amslerGrid{position:relative;background:#fff;flex:0 0 auto;overflow:hidden}.amslerSvg{display:block;width:100%;height:100%;background:#fff}.amslerSvg line,.amslerSvg rect{vector-effect:non-scaling-stroke;shape-rendering:crispEdges}.amslerSvg circle{vector-effect:non-scaling-stroke}.amslerStatus{text-align:center;font-size:18px;color:#14823b;font-weight:800;padding:8px 0}.amslerControls{display:flex;justify-content:center;align-items:center;gap:10px;flex-wrap:wrap;padding:8px 10px 2px}.amslerAdjustBtn{min-width:120px;min-height:52px;border:2px solid #315b9f;border-radius:14px;background:#fff;color:#21477f;font-size:20px;font-weight:900}.amslerSizeReadout{font-size:18px;font-weight:900;color:#21477f;min-width:165px;text-align:center}.worthScreen{display:flex;flex-direction:column;min-height:100vh;background:#000;color:#fff}.worthHeader{display:flex;justify-content:space-between;align-items:center;gap:12px}.worthTitle{font-size:clamp(26px,4vw,42px);font-weight:900}.worthInfo{font-size:16px;color:#d0d0d0;margin-top:4px}.worthWrap{flex:1;display:flex;align-items:center;justify-content:center;min-height:420px}.worthStage{position:relative;width:min(88vw,760px);height:min(68vh,560px);--worth-dot:26px;--worth-dx:130px;--worth-dy:110px}.worthDot{position:absolute;width:var(--worth-dot);height:var(--worth-dot);border-radius:50%;transform:translate(-50%,-50%)}.worthRed{left:50%;top:calc(50% - var(--worth-dy));background:#ff351f}.worthGreenL{left:calc(50% - var(--worth-dx));top:50%;background:#39e600}.worthGreenR{left:calc(50% + var(--worth-dx));top:50%;background:#39e600}.worthWhite{left:50%;top:calc(50% + var(--worth-dy));background:#fff}.worthStatus{text-align:center;font-size:18px;color:#9df0b3;font-weight:800;padding:8px 0}.worthScreen .worthHeader,.worthScreen .worthStatus{display:none!important}.worthScreen{padding:0!important;overflow:hidden}.worthWrap{min-height:100vh!important}.thoringtonScreen{display:flex;flex-direction:column;min-height:100vh;background:#000;color:#fff;padding:0!important;overflow:auto}.thoringtonControls{position:sticky;top:0;z-index:8;display:flex;justify-content:center;align-items:center;gap:10px;flex-wrap:wrap;padding:max(10px,env(safe-area-inset-top)) 12px 10px;background:rgba(0,0,0,.88)}.thoringtonAdjustBtn{min-width:112px;min-height:48px;border:2px solid #d8e4f5;border-radius:13px;background:#111;color:#fff;font-size:19px;font-weight:900}.thoringtonSizeReadout{min-width:150px;text-align:center;color:#fff;font-size:18px;font-weight:900}.thoringtonWrap{flex:1;display:flex;align-items:center;justify-content:center;min-height:calc(100vh - 78px);overflow:auto;padding:16px}.thoringtonSvg{display:block;height:auto;max-width:none;max-height:none;flex:0 0 auto}.bagoliniScreen{display:flex;flex-direction:column;min-height:100vh;background:#000;color:#fff}.bagoliniHeader{display:flex;justify-content:space-between;align-items:center;gap:12px}.bagoliniTitle{font-size:clamp(26px,4vw,42px);font-weight:900}.bagoliniInfo{font-size:16px;color:#d0d0d0;margin-top:4px}.bagoliniWrap{flex:1;display:flex;align-items:center;justify-content:center;min-height:420px}.bagoliniDot{width:3px;height:3px;min-width:2px;min-height:2px;max-width:5px;max-height:5px;background:#fff;border-radius:50%;box-shadow:0 0 1px rgba(255,255,255,.8)}.bagoliniStatus{text-align:center;font-size:18px;color:#9df0b3;font-weight:800;padding:8px 0}@media(max-width:650px){.answerGrid{gap:7px}.answer{min-height:72px}.questionPanel{min-height:105px}.correctionBar{grid-template-columns:1fr}.calAdjust{grid-template-columns:1fr 1fr}}@media(orientation:landscape) and (max-height:650px){#chartScreen{padding-top:8px}.chartWrap{padding-top:52px;padding-bottom:160px}.acuityChart{gap:2px}.estimateCard{margin-top:20px}.testScreen{padding-top:10px}.questionPanel{min-height:80px;margin:2px 0 8px}.answer{min-height:58px}.collection{display:none}.correctionBtn{min-height:44px}.testHeader{margin-bottom:4px}}

/* V10.7 只修第一個「確認視標大小」頁：允許捲動，底部保留 Safari 安全空間。
   第二個 0.5 驗證頁及後續功能完全不變。 */
.compactValidationScreen{
  min-height:100dvh!important;
  height:auto!important;
  padding:max(10px,env(safe-area-inset-top)) 10px max(170px,calc(env(safe-area-inset-bottom) + 150px))!important;
  justify-content:flex-start!important;
  overflow-y:auto!important;
  overflow-x:hidden!important;
  -webkit-overflow-scrolling:touch!important;
  overscroll-behavior-y:contain!important;
}
.compactValidationCard{
  width:min(520px,96vw)!important;
  height:auto!important;
  min-height:0!important;
  max-height:none!important;
  padding:14px 14px 18px!important;
  border-radius:16px!important;
  display:flex!important;
  flex-direction:column!important;
  align-items:stretch!important;
  gap:8px!important;
  overflow:visible!important;
}
.compactValidationCard .brand{font-size:16px!important;line-height:1.1!important;margin:0!important}
.compactValidationTitle{font-size:25px!important;line-height:1.1!important;margin:0!important}
.compactValidationHint{font-size:16px!important;line-height:1.3!important;margin:0!important}
.compactValidationDistance{font-size:17px!important;line-height:1.2!important;margin:0!important}
.compactValidationLetterBox{
  flex:0 0 auto!important;
  min-height:145px!important;
  max-height:none!important;
  margin:3px auto!important;
  padding:10px!important;
  width:100%!important;
  border:2px solid #cfd8e6!important;
  border-radius:14px!important;
  background:#fff!important;
  display:flex!important;
  align-items:center!important;
  justify-content:center!important;
  overflow:hidden!important;
}
.compactValidationValue{font-size:17px!important;line-height:1.15!important;margin:0!important;text-align:center!important;font-weight:900!important}
.compactValidationActions{
  margin-top:4px!important;
  display:grid!important;
  gap:8px!important;
  width:100%!important;
  flex:0 0 auto!important;
}
.compactValidationPrimary{width:100%!important;min-height:52px!important;margin:0!important;padding:8px 10px!important;font-size:19px!important;border-radius:13px!important}
.compactValidationRestart{width:100%!important;min-height:40px!important;margin:0!important;padding:6px 10px!important;border:1px solid #315b9f!important;border-radius:11px!important;background:#fff!important;color:#315b9f!important;font-size:15px!important;font-weight:800!important}
@media(max-height:680px){
  .compactValidationScreen{padding-bottom:max(190px,calc(env(safe-area-inset-bottom) + 170px))!important}
  .compactValidationCard{gap:6px!important;padding:10px 12px 16px!important}
  .compactValidationTitle{font-size:22px!important}
  .compactValidationHint{font-size:14px!important}
  .compactValidationLetterBox{min-height:125px!important}
}

/* V10.4 手機 0.5 驗證：真正放在受試者頁面的單畫面樣式 */
.compactVerifyScreen{padding:max(6px,env(safe-area-inset-top)) 8px max(8px,env(safe-area-inset-bottom))!important;justify-content:flex-start!important;overflow:hidden!important}
.compactVerifyCard{width:min(520px,98vw)!important;height:auto!important;max-height:calc(100dvh - 14px)!important;padding:9px 11px!important;display:flex!important;flex-direction:column!important;align-items:stretch!important;gap:5px!important;overflow:hidden!important;border-radius:16px!important}
.compactVerifyCard .brand{margin:0!important;font-size:14px!important;line-height:1.1!important}
.compactVerifyTitle{font-size:24px!important;margin:0!important;line-height:1.05!important}
.compactDistanceLine{display:flex!important;align-items:center!important;justify-content:center!important;gap:5px!important;font-size:17px!important;font-weight:800!important;line-height:1!important}
.compactDistanceInput{width:78px!important;height:35px!important;font-size:19px!important;text-align:center!important;padding:2px 5px!important}
.compactPresets{display:grid!important;grid-template-columns:repeat(3,1fr)!important;gap:5px!important}
.compactPresets .presetBtn{min-height:31px!important;padding:2px 6px!important;font-size:16px!important;border-radius:14px!important}
.compactApplyBtn{width:100%!important;min-height:39px!important;padding:5px!important;font-size:18px!important;margin:0!important}
.compactLetterBox{margin:0 auto!important;padding:5px!important;min-height:92px!important;max-height:100px!important;width:100%!important;display:flex!important;align-items:center!important;justify-content:center!important;border-radius:14px!important}
.compactVerifyValue{font-size:15px!important;font-weight:900!important;text-align:center!important;line-height:1.05!important;margin:0!important}
.compactNextBtn{width:100%!important;min-height:46px!important;padding:6px!important;font-size:19px!important;margin:0!important}
.compactRestartBtn{border:0!important;background:transparent!important;color:#315b9f!important;text-decoration:underline!important;font-size:13px!important;font-weight:700!important;padding:1px!important;min-height:22px!important}
.compactStatus{display:none!important}
.compactVerifyScreen .formulaBox,.compactVerifyScreen .calNote{display:none!important}
@media(max-height:680px){.compactVerifyCard{gap:3px!important;padding:6px 9px!important}.compactVerifyTitle{font-size:21px!important}.compactLetterBox{min-height:76px!important;max-height:82px!important}.compactNextBtn{min-height:41px!important;font-size:17px!important}.compactApplyBtn{min-height:35px!important;font-size:16px!important}.compactPresets .presetBtn{min-height:28px!important;font-size:15px!important}}
</style></head><body>
<div id='calibrationScreen' class='screen center'>
  <div class='calCard'>
    <div class='brand'>Cloud Vision</div>
    <h1>📏 裝置校正</h1>
    <p>請拿實體尺量下面這條黑線。</p>
    <p>按「縮短」或「加長」，直到黑線實際長度正好是 <b>4.00 cm</b>。</p>
    <div class='calLineBox'><div id='calLine' class='calLine'></div></div>
    <div id='calValue' class='calValue'>請量到 4.00 cm</div>
    <div class='calAdjust'><button class='calBtn' onclick='adjustCalibration(-5)'>－ 縮短</button><button class='calBtn' onclick='adjustCalibration(5)'>＋ 加長</button></div>
    <button class='calDone' onclick='finishCalibration()'>完成校正並開始</button>
    <div class='calNote'>校正值只儲存在這台裝置的瀏覽器中；更換裝置時需重新校正。</div>
  </div>
</div>
<div id='validationScreen' class='screen center hidden compactValidationScreen'>
  <div class='calCard compactValidationCard'>
    <div class='brand'>Cloud Vision</div>
    <h1 class='compactValidationTitle'>🔎 確認視標大小</h1>
    <p class='compactValidationHint'>請用尺確認下面 Z 的高度。</p>
    <p class='compactValidationDistance'><b>觀看距離：57 cm</b></p>
    <div class='compactValidationLetterBox'>
      <div id='validationLetter' style='font-family:Arial Black,Arial,sans-serif;font-weight:900;line-height:1;display:inline-block'>Z</div>
    </div>
    <div class='compactValidationValue'>目標外框高度：1.00 cm</div>
    <div class='compactValidationActions'>
      <button class='calDone compactValidationPrimary' onclick='confirmValidation()'>✓ 大小正確，進入測驗</button>
      <button class='compactValidationRestart' onclick='restartCalibrationFromValidation()'>尺寸不正確，重新校正</button>
    </div>
  </div>
</div>
<div id='acuityVerifyScreen' class='screen center hidden compactVerifyScreen'>
  <div class='calCard compactVerifyCard'>
    <div class='brand'>Cloud Vision</div>
    <h1 class='compactVerifyTitle'>視力 0.5 驗證</h1>
    <div class='compactDistanceLine'>觀看距離
      <input id='verifyDistance' class='verifyInput compactDistanceInput' type='number' min='1' step='1' value='57'>
      <span>cm</span>
    </div>
    <div class='compactPresets'>
      <button class='presetBtn' onclick='setVerifyDistance(40)'>40</button>
      <button class='presetBtn' onclick='setVerifyDistance(57)'>57</button>
      <button class='presetBtn' onclick='setVerifyDistance(100)'>100</button>
    </div>
    <button id='applyVerifyBtn' class='calBtn compactApplyBtn' onclick='applyVerifyDistance()'>套用距離</button>
    <div class='verifyLetterBox compactLetterBox'><div id='acuityVerifyLetter' style='font-family:Arial Black,Arial,sans-serif;font-weight:900;line-height:1;display:inline-block'>Z</div></div>
    <div id='acuityVerifyValue' class='compactVerifyValue'>57 cm｜視標高度 0.170 cm</div>
    <div id='verifyUpdateStatus' class='status compactStatus'>目前使用 57 cm</div>
    <button class='calDone compactNextBtn' onclick='confirmAcuityVerification()'>✓ 正確，開始測驗</button>
    <button class='compactRestartBtn' onclick='restartCalibrationFromValidation()'>尺寸不正確，重新校正</button>
  </div>
</div>
<div id='chartScreen' class='screen hidden'>
  <div class='chartWrap'>
    <div class='brand' style='text-align:center'>Cloud Vision</div>
    <div class='chartTitle'>完整 Sloan 視力表</div>
    <div id='chartDistanceInfo' class='chartHint'>觀看距離 57 cm｜請先找出大約能看清楚的最小一列</div>
    <div id='acuityChart' class='acuityChart'></div>
    <div class='estimateCard'>
      <div style='font-size:26px;font-weight:900;text-align:center'>請分別測量三種視力</div>
      <div style='font-size:16px;color:#666;text-align:center;margin-top:5px'>依序遮住另一眼測量右眼、左眼，最後雙眼一起看。</div>
      <div class='eyeAcuityForm'>
        <label class='eyeAcuityRow'><span>右眼視力</span><select id='rightEyeVA' class='estimateInput' aria-label='右眼視力'><option value=''>請選擇右眼視力</option><option value='0.10'>0.10</option><option value='0.13'>0.13</option><option value='0.16'>0.16</option><option value='0.20'>0.20</option><option value='0.25'>0.25</option><option value='0.32'>0.32</option><option value='0.40'>0.40</option><option value='0.50'>0.50</option><option value='0.63'>0.63</option><option value='0.79'>0.79</option><option value='1.00'>1.00</option><option value='1.26'>1.26</option><option value='1.50'>1.50</option></select></label>
        <label class='eyeAcuityRow'><span>左眼視力</span><select id='leftEyeVA' class='estimateInput' aria-label='左眼視力'><option value=''>請選擇左眼視力</option><option value='0.10'>0.10</option><option value='0.13'>0.13</option><option value='0.16'>0.16</option><option value='0.20'>0.20</option><option value='0.25'>0.25</option><option value='0.32'>0.32</option><option value='0.40'>0.40</option><option value='0.50'>0.50</option><option value='0.63'>0.63</option><option value='0.79'>0.79</option><option value='1.00'>1.00</option><option value='1.26'>1.26</option><option value='1.50'>1.50</option></select></label>
        <label class='eyeAcuityRow'><span>雙眼視力</span><select id='bothEyesVA' class='estimateInput' aria-label='雙眼視力'><option value=''>請選擇雙眼視力</option><option value='0.10'>0.10</option><option value='0.13'>0.13</option><option value='0.16'>0.16</option><option value='0.20'>0.20</option><option value='0.25'>0.25</option><option value='0.32'>0.32</option><option value='0.40'>0.40</option><option value='0.50'>0.50</option><option value='0.63'>0.63</option><option value='0.79'>0.79</option><option value='1.00'>1.00</option><option value='1.26'>1.26</option><option value='1.50'>1.50</option></select></label>
        <button class='estimateSend eyeSubmitBtn' onclick='submitThreeEyeVA()'>送出右眼、左眼與雙眼視力</button>
      </div>
      <div id='estimateStatus' class='status' style='text-align:center'></div>
    </div>
  </div>
  
  
  <div id='estimateModal' class='estimateModal hidden' onclick='closeEstimateModal(event)'>
    <div class='estimateModalCard' onclick='event.stopPropagation()'>
      <div class='estimateModalHeader'><div class='estimateModalTitle'>你大約看到哪一個視力？</div><button class='estimateClose' onclick='closeEstimateModal()'>✕</button></div>
      <div style='font-size:16px;color:#666;text-align:center;margin-bottom:10px'>只作為正式測驗的起始參考。</div>
      <div class='estimateRow' style='margin-top:14px'>
        <select id='estimateSelectModal' class='estimateInput' aria-label='初估視力'>
          <option value=''>請選擇視力</option>
          <option value='0.10'>0.10</option><option value='0.13'>0.13</option><option value='0.16'>0.16</option>
          <option value='0.20'>0.20</option><option value='0.25'>0.25</option><option value='0.32'>0.32</option>
          <option value='0.40'>0.40</option><option value='0.50'>0.50</option><option value='0.63'>0.63</option>
          <option value='0.79'>0.79</option><option value='1.00'>1.00</option><option value='1.26'>1.26</option>
          <option value='1.50'>1.50</option>
        </select>
        <button class='estimateSend' onclick='submitModalEstimate()'>確定送出</button>
      </div>
    </div>
  </div>
</div>
<div id='dialScreen' class='screen dialScreen hidden'>
  <div class='dialHeader'><div><div class='dialTitle'>散光鐘</div><div id='dialInfo' class='dialInfo'>等待測驗者設定</div></div></div>
  <div class='dialControls'><button type='button' class='dialAdjustBtn' onclick='setParticipantDialScale(.75)'>75%</button><button type='button' class='dialAdjustBtn' onclick='setParticipantDialScale(1)'>100%</button><button type='button' class='dialAdjustBtn' onclick='setParticipantDialScale(1.25)'>125%</button><button type='button' class='dialAdjustBtn' onclick='setParticipantDialScale(1.5)'>150%</button><button type='button' class='dialAdjustBtn' onclick='adjustParticipantDialScale(-.025)'>－微縮</button><div id='dialSizeReadout' class='dialSizeReadout'>倍率 100%</div><button type='button' class='dialAdjustBtn' onclick='adjustParticipantDialScale(.025)'>＋微放</button></div>
  <div class='dialWrap'><canvas id='dialCanvas' class='dialCanvas'></canvas></div>
  <div id='dialStatus' class='dialStatus'>已連線，請注視散光鐘中心</div>
</div>
<div id='amslerScreen' class='screen amslerScreen hidden'>
  <div class='amslerHeader'><div><div class='amslerTitle'>阿姆斯勒方格（Amsler Grid）</div><div id='amslerInfo' class='amslerInfo'>測試距離 30 公分｜方格寬度 10 公分｜方格高度 10 公分｜20 × 20 格</div></div></div>
  <div class='amslerWrap'><div id='amslerGrid' class='amslerGrid'><svg id='amslerSvg' class='amslerSvg' viewBox='0 0 1000 1000' preserveAspectRatio='none' xmlns='http://www.w3.org/2000/svg' aria-label='20乘20阿姆斯勒方格'><line x1='0' y1='0' x2='0' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='0' x2='1000' y2='0' stroke='#000' stroke-width='1'/><line x1='50' y1='0' x2='50' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='50' x2='1000' y2='50' stroke='#000' stroke-width='1'/><line x1='100' y1='0' x2='100' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='100' x2='1000' y2='100' stroke='#000' stroke-width='1'/><line x1='150' y1='0' x2='150' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='150' x2='1000' y2='150' stroke='#000' stroke-width='1'/><line x1='200' y1='0' x2='200' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='200' x2='1000' y2='200' stroke='#000' stroke-width='1'/><line x1='250' y1='0' x2='250' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='250' x2='1000' y2='250' stroke='#000' stroke-width='1'/><line x1='300' y1='0' x2='300' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='300' x2='1000' y2='300' stroke='#000' stroke-width='1'/><line x1='350' y1='0' x2='350' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='350' x2='1000' y2='350' stroke='#000' stroke-width='1'/><line x1='400' y1='0' x2='400' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='400' x2='1000' y2='400' stroke='#000' stroke-width='1'/><line x1='450' y1='0' x2='450' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='450' x2='1000' y2='450' stroke='#000' stroke-width='1'/><line x1='500' y1='0' x2='500' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='500' x2='1000' y2='500' stroke='#000' stroke-width='1'/><line x1='550' y1='0' x2='550' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='550' x2='1000' y2='550' stroke='#000' stroke-width='1'/><line x1='600' y1='0' x2='600' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='600' x2='1000' y2='600' stroke='#000' stroke-width='1'/><line x1='650' y1='0' x2='650' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='650' x2='1000' y2='650' stroke='#000' stroke-width='1'/><line x1='700' y1='0' x2='700' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='700' x2='1000' y2='700' stroke='#000' stroke-width='1'/><line x1='750' y1='0' x2='750' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='750' x2='1000' y2='750' stroke='#000' stroke-width='1'/><line x1='800' y1='0' x2='800' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='800' x2='1000' y2='800' stroke='#000' stroke-width='1'/><line x1='850' y1='0' x2='850' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='850' x2='1000' y2='850' stroke='#000' stroke-width='1'/><line x1='900' y1='0' x2='900' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='900' x2='1000' y2='900' stroke='#000' stroke-width='1'/><line x1='950' y1='0' x2='950' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='950' x2='1000' y2='950' stroke='#000' stroke-width='1'/><line x1='1000' y1='0' x2='1000' y2='1000' stroke='#000' stroke-width='1'/><line x1='0' y1='1000' x2='1000' y2='1000' stroke='#000' stroke-width='1'/><rect x='1' y='1' width='998' height='998' fill='none' stroke='#000' stroke-width='2'/><circle cx='500' cy='500' r='10' fill='#000'/></svg></div></div>
  <div class='amslerControls'><button type='button' class='amslerAdjustBtn' onclick='adjustAmslerScale(-0.10)'>－ 縮小</button><div id='amslerSizeReadout' class='amslerSizeReadout'>目前倍率 0.50</div><button type='button' class='amslerAdjustBtn' onclick='adjustAmslerScale(0.10)'>＋ 放大</button><button type='button' class='amslerAdjustBtn' onclick='resetAmslerScale()'>恢復</button></div>
  <div id='amslerStatus' class='amslerStatus'>請用尺確認外框寬度與高度都為 10 公分；若不正確，可按放大或縮小微調</div>
</div>
<div id='worthScreen' class='screen worthScreen hidden'>
  <div class='worthHeader'><div><div class='worthTitle'>Worth 四點測試（Worth Four Dot Test）</div><div id='worthInfo' class='worthInfo'>近距模式：33 cm</div></div></div>
  <div class='worthWrap'><div id='worthStage' class='worthStage'><div class='worthDot worthRed'></div><div class='worthDot worthGreenL'></div><div class='worthDot worthGreenR'></div><div class='worthDot worthWhite'></div></div></div>
  <div id='worthStatus' class='worthStatus'>請配戴紅綠眼鏡並注視四點</div>
</div>
<div id='bagoliniScreen' class='screen bagoliniScreen hidden'>
  <div class='bagoliniHeader'><div><div class='bagoliniTitle'>Bagolini 條紋鏡測試</div><div id='bagoliniInfo' class='bagoliniInfo'>近距模式：33 cm</div></div></div>
  <div class='bagoliniWrap'><div id='bagoliniDot' class='bagoliniDot'></div></div>
  <div id='bagoliniStatus' class='bagoliniStatus'>請配戴 Bagolini 條紋鏡並注視中央白色光點</div>
</div>
<div id='thoringtonScreen' class='screen thoringtonScreen hidden'>
  <div class='thoringtonControls'><button type='button' class='thoringtonAdjustBtn' onclick='adjustThoringtonScale(-0.10)'>－ 縮小</button><div id='thoringtonSizeReadout' class='thoringtonSizeReadout'>目前倍率 1.00</div><button type='button' class='thoringtonAdjustBtn' onclick='adjustThoringtonScale(0.10)'>＋ 放大</button><button type='button' class='thoringtonAdjustBtn' onclick='resetThoringtonScale()'>恢復</button></div>
  <div class='thoringtonWrap'>
    <svg id='thoringtonSvg' class='thoringtonSvg' viewBox='0 0 1400 440' preserveAspectRatio='xMidYMid meet' xmlns='http://www.w3.org/2000/svg' aria-label='Thorington card'><rect x='0' y='0' width='1400' height='440' fill='#000'/><text x='700' y='58' fill='#f4f4f4' font-size='34' font-family='Arial' font-weight='700' text-anchor='middle'>Modified Thorington 水平眼位測量</text><text id='thoringtonModeText' x='700' y='100' fill='#d7dce3' font-size='25' font-family='Arial' text-anchor='middle'>中距 1 m｜每格 1Δ</text><line x1='100' y1='260' x2='1300' y2='260' stroke='#f2f4f7' stroke-width='4'/><line x1='100.00' x2='100.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='100.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>L</text><line x1='150.00' x2='150.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='150.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>K</text><line x1='200.00' x2='200.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='200.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>J</text><line x1='250.00' x2='250.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='250.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>I</text><line x1='300.00' x2='300.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='300.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>H</text><line x1='350.00' x2='350.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='350.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>G</text><line x1='400.00' x2='400.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='400.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>F</text><line x1='450.00' x2='450.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='450.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>E</text><line x1='500.00' x2='500.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='500.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>D</text><line x1='550.00' x2='550.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='550.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>C</text><line x1='600.00' x2='600.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='600.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>B</text><line x1='650.00' x2='650.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='650.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>A</text><line x1='700.00' x2='700.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='4'/><line x1='750.00' x2='750.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='750.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>1</text><line x1='800.00' x2='800.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='800.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>2</text><line x1='850.00' x2='850.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='850.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>3</text><line x1='900.00' x2='900.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='900.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>4</text><line x1='950.00' x2='950.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='950.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>5</text><line x1='1000.00' x2='1000.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='1000.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>6</text><line x1='1050.00' x2='1050.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='1050.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>7</text><line x1='1100.00' x2='1100.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='1100.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>8</text><line x1='1150.00' x2='1150.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='1150.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>9</text><line x1='1200.00' x2='1200.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='1200.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>10</text><line x1='1250.00' x2='1250.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='1250.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>11</text><line x1='1300.00' x2='1300.00' y1='247' y2='277' stroke='#cfd5dc' stroke-width='2'/><text x='1300.00' y='220' fill='#e5e9ee' font-size='30' font-family='Arial' text-anchor='middle'>12</text><circle id='thoringtonCenterDot' cx='700' cy='260' r='6' fill='#fff' stroke='#fff' stroke-width='2'/></svg>
  </div>
</div>
<div id='waitingScreen' class='screen center hidden'>
  <div class='topActions'><button class='smallBtn' onclick='startCalibration()'>重新校正</button></div>
  <div class='waitWrap center'><div class='waitIcon'>✓</div><div class='waitTitle'>尺寸確認完成</div><div class='waitSub'>等待測驗者按「開始作答」…</div><div id='waitStatus' class='status'>正在連線…</div></div>
</div>
<div id='testScreen' class='screen testScreen hidden'>
  <div class='testHeader'><div><div id='progress' class='progress'>作答進度：0 / 0</div><div id='sizeInfo' class='sizeInfo'></div></div></div>
  <div id='questionPanel' class='questionPanel'><div id='questionDisplay' class='questionDisplay'></div></div>
  <div class='collection'><div id='collectionTitle' class='collectionTitle'>我的貼圖收藏（0）</div><div id='stickers' class='stickers'>等待測驗者送貼圖</div></div>
  <div class='correctionBar'><button id='undoBtn' class='correctionBtn' onclick="sendCmd('undo_answer')">↶ 退回上一個答案</button><button id='resetBtn' class='correctionBtn resetBtn' onclick="sendCmd('reset_answers')">全部清除，重新作答</button></div>
  <section class='answerArea'><div class='answerGrid'>""" + buttons + """</div></section>
  <div id='status' class='status'>正在連線…</div>
</div>
<script>
const SESSION='""" + self.connection_session_token + """';
const CAL_KEY='cloudVisionPxPerCmV65';
const VERIFIED_KEY='cloudVisionSizeVerifiedV65';
let busy=false,lastStickerVersion=-1,connectionFailures=0;
let formalTestStarted=false,lastQuestionTargets=[],lastOptotypeMode='letter';
// V10.1 每次新開受試者頁面，都從 4.00 cm 校正線開始，避免舊版紀錄跳過驗證流程。
localStorage.removeItem(CAL_KEY);
localStorage.removeItem(VERIFIED_KEY);
let pxPerCm=0;
const AMSLER_SCALE_KEY='cloudVisionAmslerScaleV98';
const DIAL_SCALE_KEY='cloudVisionDialScaleV18';
const THORINGTON_SCALE_KEY='cloudVisionThoringtonScaleV98';
function safeLoadNumber(key,fallback){try{const v=Number(localStorage.getItem(key));return Number.isFinite(v)&&v>0?v:fallback;}catch(e){return fallback;}}
let amslerManualScale=safeLoadNumber(AMSLER_SCALE_KEY,0.50);
amslerManualScale=Math.max(0.35,Math.min(3.5,amslerManualScale));
let participantDialScale=safeLoadNumber(DIAL_SCALE_KEY,1);
participantDialScale=Math.max(.60,Math.min(1.80,participantDialScale));
let thoringtonManualScale=safeLoadNumber(THORINGTON_SCALE_KEY,1);
thoringtonManualScale=Math.max(0.35,Math.min(3.5,thoringtonManualScale));
let lastThoringtonDistance=100;
let lastThoringtonDotPx=5;

// V9.7：縮放函式必須定義在受試者頁面。上一版誤放在測驗者頁面，
// 因此 iPad 看得到按鈕，但 onclick 找不到函式，倍率永遠停在 1.00。
function participantSafeStore(key,value){try{localStorage.setItem(key,String(value));}catch(e){}}
function participantSvgElement(svg,name,attrs,text){
  const node=document.createElementNS('http://www.w3.org/2000/svg',name);
  Object.keys(attrs||{}).forEach(k=>node.setAttribute(k,String(attrs[k])));
  if(text!==undefined)node.textContent=text;
  svg.appendChild(node);
  return node;
}
function renderParticipantAmsler(){
  const grid=document.getElementById('amslerGrid');
  const svg=document.getElementById('amslerSvg');
  const readout=document.getElementById('amslerSizeReadout');
  if(readout)readout.textContent='目前倍率 '+amslerManualScale.toFixed(2);
  if(!grid||!svg)return;
  const calibrated=(Number(pxPerCm)>0?Number(pxPerCm):37.8);
  const correction=10/5.7;
  const size=Math.max(120,10*calibrated*correction*amslerManualScale);
  const fixed=size.toFixed(1)+'px';
  grid.style.setProperty('width',fixed,'important');
  grid.style.setProperty('height',fixed,'important');
  grid.style.setProperty('min-width',fixed,'important');
  grid.style.setProperty('min-height',fixed,'important');
  grid.style.setProperty('max-width','none','important');
  grid.style.setProperty('max-height','none','important');
  grid.style.setProperty('flex','0 0 '+fixed,'important');
  svg.setAttribute('viewBox','0 0 1000 1000');
  svg.setAttribute('width',size.toFixed(1));
  svg.setAttribute('height',size.toFixed(1));
  svg.style.setProperty('width','100%','important');
  svg.style.setProperty('height','100%','important');
  while(svg.firstChild)svg.removeChild(svg.firstChild);
  participantSvgElement(svg,'rect',{x:0,y:0,width:1000,height:1000,fill:'#fff'});
  for(let i=0;i<=20;i++){
    const p=i*50;
    participantSvgElement(svg,'line',{x1:p,y1:0,x2:p,y2:1000,stroke:'#000','stroke-width':(i===0||i===20)?3:2});
    participantSvgElement(svg,'line',{x1:0,y1:p,x2:1000,y2:p,stroke:'#000','stroke-width':(i===0||i===20)?3:2});
  }
  participantSvgElement(svg,'circle',{cx:500,cy:500,r:10,fill:'#000'});
}
function adjustAmslerScale(delta){
  const d=Number(delta)||0;
  amslerManualScale=Math.round(Math.max(0.35,Math.min(3.5,amslerManualScale+d))*100)/100;
  participantSafeStore(AMSLER_SCALE_KEY,amslerManualScale);
  renderParticipantAmsler();
  const status=document.getElementById('amslerStatus');
  if(status)status.textContent='已調整為 '+amslerManualScale.toFixed(2)+' 倍，請用尺確認寬度';
}
function resetAmslerScale(){
  amslerManualScale=0.50;
  participantSafeStore(AMSLER_SCALE_KEY,0.50);
  renderParticipantAmsler();
}
function renderParticipantThorington(distanceCm,dotPx){
  const svg=document.getElementById('thoringtonSvg');
  const readout=document.getElementById('thoringtonSizeReadout');
  if(readout)readout.textContent='目前倍率 '+thoringtonManualScale.toFixed(2);
  if(!svg)return;
  const W=1400,H=440,cx=700,cy=260,startX=100,endX=1300,step=50;
  const widthPx=Math.max(340,window.innerWidth*0.94)*thoringtonManualScale;
  const heightPx=widthPx*(H/W);
  svg.setAttribute('viewBox','0 0 '+W+' '+H);
  svg.setAttribute('width',widthPx.toFixed(1));
  svg.setAttribute('height',heightPx.toFixed(1));
  svg.style.setProperty('width',widthPx.toFixed(1)+'px','important');
  svg.style.setProperty('height',heightPx.toFixed(1)+'px','important');
  svg.style.setProperty('min-width',widthPx.toFixed(1)+'px','important');
  svg.style.setProperty('max-width','none','important');
  while(svg.firstChild)svg.removeChild(svg.firstChild);
  participantSvgElement(svg,'rect',{x:0,y:0,width:W,height:H,fill:'#000'});
  participantSvgElement(svg,'text',{x:cx,y:58,fill:'#fff','font-size':34,'font-family':'Arial','font-weight':700,'text-anchor':'middle'},'Modified Thorington 水平眼位測量');
  const mode=Number(distanceCm)===40?'近距 40 cm':(Number(distanceCm)===300?'遠距 3 m':'中距 1 m');
  participantSvgElement(svg,'text',{x:cx,y:100,fill:'#ddd','font-size':25,'font-family':'Arial','text-anchor':'middle'},mode+'｜每格 1Δ');
  participantSvgElement(svg,'line',{x1:startX,y1:cy,x2:endX,y2:cy,stroke:'#fff','stroke-width':2,'stroke-linecap':'round'});
  const letters='LKJIHGFEDCBA';
  for(let i=-12;i<=12;i++){
    const x=cx+i*step;
    participantSvgElement(svg,'line',{x1:x,y1:cy-14,x2:x,y2:cy+18,stroke:'#fff','stroke-width':i===0?2:1.5});
    if(i<0)participantSvgElement(svg,'text',{x:x,y:220,fill:'#fff','font-size':30,'font-family':'Arial','font-weight':600,'text-anchor':'middle'},letters[i+12]);
    else if(i>0)participantSvgElement(svg,'text',{x:x,y:220,fill:'#fff','font-size':30,'font-family':'Arial','font-weight':600,'text-anchor':'middle'},String(i));
  }
  participantSvgElement(svg,'circle',{cx:cx,cy:cy,r:12,fill:'#000'});
  participantSvgElement(svg,'circle',{cx:cx,cy:cy,r:Math.max(6,Math.min(9,Number(dotPx)||6)),fill:'#fff'});
}
function adjustThoringtonScale(delta){
  thoringtonManualScale=Math.round(Math.max(0.35,Math.min(3.5,thoringtonManualScale+(Number(delta)||0)))*100)/100;
  participantSafeStore(THORINGTON_SCALE_KEY,thoringtonManualScale);
  renderParticipantThorington(lastThoringtonDistance,lastThoringtonDotPx);
}
function resetThoringtonScale(){
  thoringtonManualScale=1;
  participantSafeStore(THORINGTON_SCALE_KEY,1);
  renderParticipantThorington(lastThoringtonDistance,lastThoringtonDotPx);
}
// 明確掛到 window，確保 iPad Safari 的 inline onclick 一定找得到。
window.adjustAmslerScale=adjustAmslerScale;
window.resetAmslerScale=resetAmslerScale;
window.renderParticipantAmsler=renderParticipantAmsler;
window.adjustThoringtonScale=adjustThoringtonScale;
window.resetThoringtonScale=resetThoringtonScale;
window.renderParticipantThorington=renderParticipantThorington;
let sizeVerified=false;
let estimateSubmitted=false;
let validationPassed=false;
// V9.4：受試者端 5 cm 校正線的預設長度縮短為原本約 46%，避免首次校正必須連按很多次。
let calibrationLinePx=Math.min(window.innerWidth*0.65,300);
function showOnly(id){
  const chart=document.getElementById('chartScreen');
  if(chart&&id!=='chartScreen')chart.dataset.justOpened='0';
  const ids=['calibrationScreen','validationScreen','acuityVerifyScreen','chartScreen','dialScreen','amslerScreen','worthScreen','bagoliniScreen','thoringtonScreen','waitingScreen','testScreen'];
  ids.forEach(x=>{
    const el=document.getElementById(x);
    if(el)el.classList.toggle('hidden',x!==id);
  });
  // Safari 會保留上一個畫面的捲動位置；切換畫面後強制回到最上方。
  document.documentElement.scrollTop=0;
  document.body.scrollTop=0;
  requestAnimationFrame(()=>{
    window.scrollTo({top:0,left:0,behavior:'auto'});
    requestAnimationFrame(()=>window.scrollTo(0,0));
  });
}
function renderCalibration(){const line=document.getElementById('calLine');line.style.width=calibrationLinePx+'px';document.getElementById('calValue').textContent='請用尺確認：4.00 cm';}
function startCalibration(){resetEstimateOnServer();sizeVerified=false;validationPassed=false;localStorage.removeItem(VERIFIED_KEY);showOnly('calibrationScreen');renderCalibration();}
function adjustCalibration(delta){calibrationLinePx=Math.max(80,Math.min(window.innerWidth*0.92,calibrationLinePx+delta));renderCalibration();}
function renderValidation(){const el=document.getElementById('validationLetter');if(!el||!(pxPerCm>0))return;const targetPx=1.00*pxPerCm;el.style.fontSize=(targetPx/.72).toFixed(1)+'px';}
let verifyDistanceCm=57;
function targetHeightForVA05(distanceCm){return 0.17*Number(distanceCm)/57;}
function renderAcuityVerification(){if(!(pxPerCm>0))return;const input=document.getElementById('verifyDistance');if(input&&document.activeElement!==input)input.value=String(verifyDistanceCm);const h=targetHeightForVA05(verifyDistanceCm);const el=document.getElementById('acuityVerifyLetter');if(el)el.style.fontSize=((h*pxPerCm)/.72).toFixed(2)+'px';const info=document.getElementById('acuityVerifyValue');if(info)info.textContent='視標高度 '+h.toFixed(3)+' cm';}
function showVerifyUpdateFeedback(message){
  const status=document.getElementById('verifyUpdateStatus');
  if(status){status.textContent=message;status.style.color='#14823b';}
  const btn=document.getElementById('applyVerifyBtn');
  if(btn){
    const original='套用距離';
    btn.textContent='✓ 已更新';
    btn.style.background='#eaf7ee';
    setTimeout(()=>{btn.textContent=original;btn.style.background='#fff';},900);
  }
}
function setVerifyDistance(v){
  verifyDistanceCm=Number(v);
  const input=document.getElementById('verifyDistance');
  if(input)input.value=String(verifyDistanceCm);
  renderAcuityVerification();
  showVerifyUpdateFeedback('已切換為 '+verifyDistanceCm+' cm，視標已更新');
}
function applyVerifyDistance(){
  const input=document.getElementById('verifyDistance');
  const v=Number(input&&input.value);
  if(!Number.isFinite(v)||v<=0){alert('請輸入正確的觀看距離');if(input)input.focus();return;}
  verifyDistanceCm=v;
  if(input)input.blur();
  renderAcuityVerification();
  showVerifyUpdateFeedback('已套用 '+v+' cm，視標已更新');
  const letter=document.getElementById('acuityVerifyLetter');
  if(letter){letter.animate([{transform:'scale(1)'},{transform:'scale(1.35)'},{transform:'scale(1)'}],{duration:420,easing:'ease-out'});}
}
function confirmAcuityVerification(){
  sizeVerified=true;
  localStorage.setItem(VERIFIED_KEY,'1');
  showOnly('chartScreen');
  renderFullChart();
}
const FULL_CHART_ROWS=[
  [0.10,'OZNVK'],[0.13,'VRSOD'],[0.16,'NKZRS'],[0.20,'DKSHV'],[0.25,'RKVHN'],
  [0.32,'SKROC'],[0.40,'VRHSN'],[0.50,'OVCRH'],[0.63,'SZVKN'],[0.79,'KODRH'],
  [1.00,'NSOKD'],[1.26,'CVHRS'],[1.50,'ZKOND']
];
function chartLetterHeightCm(va){return (0.085/Number(va))*(verifyDistanceCm/57);}
function renderFullChart(){
  if(!(pxPerCm>0))return;
  const chartScreen=document.getElementById('chartScreen');
  if(chartScreen&&chartScreen.dataset.justOpened!=='1'){chartScreen.scrollTop=0;chartScreen.dataset.justOpened='1';}
  const di=document.getElementById('chartDistanceInfo');
  if(di)di.textContent='觀看距離 '+Number(verifyDistanceCm).toFixed(Number(verifyDistanceCm)%1?1:0)+' cm｜請先找出大約能看清楚的最小一列';
  const box=document.getElementById('acuityChart');
  if(!box)return;
  box.innerHTML='';
  FULL_CHART_ROWS.forEach(([va,letters])=>{
    const row=document.createElement('div');row.className='acuityRow';
    const left=document.createElement('div');left.className='acuityLabel';left.textContent=Number(va).toFixed(2);
    const mid=document.createElement('div');mid.className='acuityLetters';mid.textContent=letters.split('').join('  ');
    const h=chartLetterHeightCm(va);mid.style.fontSize=Math.max(3,(h*pxPerCm/.72)).toFixed(2)+'px';
    const right=document.createElement('div');right.className='acuityLabel';right.style.textAlign='right';right.textContent=h.toFixed(3)+' cm';
    row.append(left,mid,right);box.appendChild(row);
  });
}
function openEstimateModal(){const m=document.getElementById('estimateModal');if(m)m.classList.remove('hidden');}
function closeEstimateModal(event){if(event&&event.target!==document.getElementById('estimateModal'))return;const m=document.getElementById('estimateModal');if(m)m.classList.add('hidden');}
function submitModalEstimate(){const input=document.getElementById('estimateSelectModal');submitEstimate(input?input.value:'');}
function submitSelectedEstimate(){const input=document.getElementById('estimateSelect');submitEstimate(input?input.value:'');}
async function readRemoteState(){
  const r=await fetch('/estimate?role=participant&session='+encodeURIComponent(SESSION)+'&t='+Date.now(),{cache:'no-store'});
  if(!r.ok)throw new Error('state');
  return await r.json();
}
async function waitEstimateAck(value,submittedVersion){
  const deadline=Date.now()+10000;
  while(Date.now()<deadline){
    await new Promise(resolve=>setTimeout(resolve,220));
    const d=await readRemoteState();
    const same=d.estimated_va!=null&&Math.abs(Number(d.estimated_va)-value)<0.0001;
    const stored=Number(d.estimated_va_version||0)>=Number(submittedVersion||0);
    const controlSawIt=Number(d.estimated_va_seen_version||0)>=Number(submittedVersion||0);
    if(same&&stored&&controlSawIt)return true;
  }
  return false;
}
async function submitEstimate(v){
  const value=Number(v);
  if(!Number.isFinite(value)||value<0.10||value>1.50){alert('請先從下拉選單選擇 0.10～1.50 的視力值');return}
  const status=document.getElementById('estimateStatus');
  try{
    if(status){status.textContent='正在送出，等待測驗者端實際收到…';status.style.color='#8a5a00';}
    console.log('[Cloud Vision V8.9][participant] 透過下午已成功通道送出初估視力',value);
    if(status){status.textContent='正在送出 '+value.toFixed(2)+' 到測驗者…';}
    const url='/command?role=participant&session='+encodeURIComponent(SESSION)+'&cmd=set_estimated_va&value='+encodeURIComponent(value.toFixed(2))+'&t='+Date.now();
    const r=await fetch(url,{cache:'no-store'});
    const rawText=await r.text();
    if(!r.ok)throw new Error('傳送失敗 HTTP '+r.status+' '+rawText);
    let ack;try{ack=JSON.parse(rawText)}catch(parseError){throw new Error('回傳格式錯誤')}
    const ackValue=Number(ack.estimated_va);
    if(!ack.ok||!Number.isFinite(ackValue)||Math.abs(ackValue-value)>=0.0001)throw new Error('電腦未確認資料');
    // 下午已成功驗證的流程：電腦伺服器回覆 OK 即完成送出。
    // 不再要求受試者端等待測驗者輪詢後的 seen_version，避免 server_state_not_confirmed。
    estimateSubmitted=true;
    const modal=document.getElementById('estimateModal');if(modal)modal.classList.add('hidden');
    if(status){status.textContent='✓ 電腦已收到初估視力 '+value.toFixed(2)+'。請按「下一步：散光鐘」繼續。';status.style.color='#14823b';}
    // 公開自我測驗流程：初估視力完成後，提供明確入口依序進入散光鐘與黃斑部。
    showOnly('chartScreen');
    renderFullChart();
    const launcher=document.getElementById('estimateLauncher');
    if(launcher){launcher.classList.add('hidden');}
    const nextBtn=document.getElementById('nextScreeningBtn');
    if(nextBtn){nextBtn.classList.remove('hidden');}
    if(status){status.textContent='✓ 已送出初估視力 '+value.toFixed(2)+'。下一步請測驗散光鐘與黃斑部。';status.style.color='#14823b';}
  }catch(e){
    console.error('[Cloud Vision V8.9][participant] 初估視力送出失敗',e);
    const msg=String(e&&e.message?e.message:e);
    if(status){status.textContent='❌ 送出失敗：'+msg;status.style.color='#c62828';}
    alert('初估視力送出失敗：'+msg);
  }
}
async function submitThreeEyeVA(){
  const right=document.getElementById('rightEyeVA')?.value||'';
  const left=document.getElementById('leftEyeVA')?.value||'';
  const both=document.getElementById('bothEyesVA')?.value||'';
  if(!right){alert('請先選擇右眼視力');document.getElementById('rightEyeVA')?.focus();return;}
  if(!left){alert('請再選擇左眼視力');document.getElementById('leftEyeVA')?.focus();return;}
  if(!both){alert('最後請選擇雙眼視力');document.getElementById('bothEyesVA')?.focus();return;}
  const values={right_eye_va:Number(right),left_eye_va:Number(left),both_eyes_va:Number(both)};
  try{
    sessionStorage.setItem('cloudVisionEyeVA',JSON.stringify(values));
    localStorage.setItem('cloudVisionEyeVA',JSON.stringify(values));
  }catch(e){}
  const status=document.getElementById('estimateStatus');
  if(status){status.textContent='正在送出右眼、左眼與雙眼視力…';status.style.color='#8a5a00';}
  try{
    const url='/command?role=participant&session='+encodeURIComponent(SESSION)+'&cmd=set_estimated_va&value='+encodeURIComponent(Number(both).toFixed(2))+'&right_eye_va='+encodeURIComponent(Number(right).toFixed(2))+'&left_eye_va='+encodeURIComponent(Number(left).toFixed(2))+'&both_eyes_va='+encodeURIComponent(Number(both).toFixed(2))+'&t='+Date.now();
    const r=await fetch(url,{cache:'no-store'}); const data=await r.json();
    if(!r.ok||!data.ok)throw new Error(data.error||'送出失敗');
    estimateSubmitted=true;
    document.getElementById('estimateLauncher')?.classList.add('hidden');
    document.getElementById('nextScreeningBtn')?.classList.add('hidden');
    if(status){status.textContent='✓ 已記錄：右眼 '+Number(right).toFixed(2)+'、左眼 '+Number(left).toFixed(2)+'、雙眼 '+Number(both).toFixed(2)+'，正在進入散光鐘…';status.style.color='#14823b';}
    // 不再多按一次「下一步」；直接把三項視力帶入後續測驗與最終結果頁。
    const next='/cloud/assessment?right_eye_va='+encodeURIComponent(Number(right).toFixed(2))+'&left_eye_va='+encodeURIComponent(Number(left).toFixed(2))+'&both_eyes_va='+encodeURIComponent(Number(both).toFixed(2))+'&v=4.1&t='+Date.now();
    setTimeout(()=>{window.location.href=next;},220);
  }catch(e){
    if(status){status.textContent='❌ 送出失敗：'+String(e.message||e);status.style.color='#c62828';}
    alert('三項視力送出失敗，請再試一次');
  }
}
window.submitThreeEyeVA=submitThreeEyeVA;
function submitCustomEstimate(){submitThreeEyeVA();}
async function resetEstimateOnServer(){
  try{
    await fetch('/command?role=participant&session='+encodeURIComponent(SESSION)+'&cmd=reset_estimated_va&t='+Date.now(),{cache:'no-store'});
  }catch(e){}
  estimateSubmitted=false;
}
function finishCalibration(){pxPerCm=calibrationLinePx/4;localStorage.setItem(CAL_KEY,String(pxPerCm));sizeVerified=false;validationPassed=false;localStorage.removeItem(VERIFIED_KEY);showOnly('validationScreen');renderValidation();}
function confirmValidation(){
  // 只切換狀態，不刪除或重建任何畫面。
  // validationPassed 用來避免定時 refresh 又把畫面切回 1.00 cm 驗證頁。
  sizeVerified=false;
  validationPassed=true;
  localStorage.removeItem(VERIFIED_KEY);
  showOnly('acuityVerifyScreen');
  renderAcuityVerification();
}
function restartCalibrationFromValidation(){pxPerCm=0;sizeVerified=false;validationPassed=false;localStorage.removeItem(CAL_KEY);localStorage.removeItem(VERIFIED_KEY);calibrationLinePx=Math.min(window.innerWidth*.65,300);startCalibration();}
function ensureCalibration(){if(!(pxPerCm>0)){startCalibration();return false}if(!sizeVerified){if(validationPassed){showOnly('acuityVerifyScreen');renderAcuityVerification()}else{showOnly('validationScreen');renderValidation()}return false}return true}
function participantOnline(){connectionFailures=0;for(const id of ['status','waitStatus']){const s=document.getElementById(id);if(s){s.style.color='#14823b'}}}
function participantOffline(){connectionFailures++;if(connectionFailures>=2){for(const id of ['status','waitStatus']){const s=document.getElementById(id);if(s){s.textContent='❌ 與電腦斷線，請確認仍連著同一個 Wi‑Fi';s.style.color='#c62828'}}enable(false)}}
function enable(v){document.querySelectorAll('.answer').forEach(b=>b.disabled=!v)}
async function sendAnswer(x){
  if(busy)return;busy=true;
  try{
    const r=await fetch('/answer?role=participant&session='+encodeURIComponent(SESSION)+'&value='+encodeURIComponent(x)+'&t='+Date.now(),{cache:'no-store'});
    if(!r.ok)throw new Error();
    participantOnline();
    await refresh(true);
  }catch(e){participantOffline()}
  finally{setTimeout(()=>busy=false,220)}
}
async function sendCmd(cmd){
  try{
    const r=await fetch('/command?role=participant&session='+encodeURIComponent(SESSION)+'&cmd='+encodeURIComponent(cmd)+'&t='+Date.now(),{cache:'no-store'});
    if(!r.ok)throw new Error();
    participantOnline();
    await refresh(true);
  }catch(e){participantOffline()}
}
function updateStickers(d){const el=document.getElementById('stickers'),title=document.getElementById('collectionTitle'),items=d.stickers||[];title.textContent='我的貼圖收藏（'+items.length+'）';el.innerHTML=items.length?items.map(x=>'<span>'+x+'</span>').join(''):'<span style="font-size:17px;color:#777">等待測驗者送貼圖</span>';if(lastStickerVersion>=0&&d.sticker_version!==lastStickerVersion){el.classList.remove('newSticker');void el.offsetWidth;el.classList.add('newSticker')}lastStickerVersion=d.sticker_version}
function arrowAngle(symbol){return {'→':0,'↘':45,'↓':90,'↙':135,'←':180,'↖':225,'↑':270,'↗':315}[symbol]??0;}
function landoltSvg(symbol){const a=arrowAngle(symbol);return `<svg class="optotypeSvg" viewBox="0 0 100 100" aria-label="Landolt C ${symbol}"><circle cx="50" cy="50" r="40" fill="none" stroke="#000" stroke-width="20"/><line x1="50" y1="50" x2="104" y2="50" stroke="#fff" stroke-width="22" transform="rotate(${a} 50 50)"/></svg>`;}
function tumblingESvg(symbol){const a=arrowAngle(symbol);return `<svg class="optotypeSvg" viewBox="0 0 100 100" aria-label="Tumbling E ${symbol}"><g transform="rotate(${a} 50 50)"><rect x="10" y="10" width="20" height="80" fill="#000"/><rect x="10" y="10" width="80" height="20" fill="#000"/><rect x="10" y="40" width="80" height="20" fill="#000"/><rect x="10" y="70" width="80" height="20" fill="#000"/></g></svg>`;}
function renderQuestionTargets(mode,targets){const qd=document.getElementById('questionDisplay');const list=targets||[];if(mode==='landolt_c')qd.innerHTML=list.map(landoltSvg).join('');else if(mode==='tumbling_e')qd.innerHTML=list.map(tumblingESvg).join('');else qd.textContent=list.join('　');}
function applyPhysicalSize(d){if(!(pxPerCm>0)||!(d.target_height_cm>0))return;const qd=document.getElementById('questionDisplay');const n=Math.max(1,(d.active_targets||[]).length);const targetPx=d.target_height_cm*pxPerCm;const available=(window.innerWidth-70)/Math.max(1,n*1.18);const actualPx=Math.max(6,Math.min(targetPx,available));if((d.optotype_mode||'letter')==='letter'){const fontPx=actualPx/.72;qd.style.fontSize=fontPx.toFixed(1)+'px';qd.style.letterSpacing=n>1?Math.min(fontPx*.45,20)+'px':'0px';}else{qd.style.fontSize='';qd.style.letterSpacing='0';qd.style.setProperty('--opt-size',actualPx.toFixed(2)+'px');qd.style.setProperty('--opt-gap',Math.max(8,Math.min(actualPx*.45,24)).toFixed(1)+'px');}document.getElementById('sizeInfo').textContent='目標外框高 '+Number(d.target_height_cm).toFixed(3)+' cm｜距離 '+Number(d.distance_cm).toFixed(0)+' cm';}
function toggleFullscreen(){}
let participantRefreshBusy=false;
let participantPollTimer=null;
function reconnectNow(){connectionFailures=0;startParticipantPolling(true)}
window.addEventListener('online',reconnectNow);
window.addEventListener('focus',reconnectNow);
window.addEventListener('pageshow',reconnectNow);
document.addEventListener('visibilitychange',()=>{if(!document.hidden)reconnectNow()});
function renderParticipantDial(distanceCm){
  const canvas=document.getElementById('dialCanvas');
  if(!canvas)return;
  const wrap=canvas.parentElement;
  const cssW=Math.max(320,wrap.clientWidth||window.innerWidth-36);
  const cssH=Math.max(420,wrap.clientHeight||window.innerHeight-150);
  const dpr=Math.max(1,window.devicePixelRatio||1);
  canvas.width=Math.round(cssW*dpr);canvas.height=Math.round(cssH*dpr);
  canvas.style.width=cssW+'px';canvas.style.height=cssH+'px';
  const ctx=canvas.getContext('2d');ctx.setTransform(dpr,0,0,dpr,0,0);
  ctx.clearRect(0,0,cssW,cssH);ctx.fillStyle='#fff';ctx.fillRect(0,0,cssW,cssH);
  const cx=cssW/2,cy=cssH*0.78;
  const pxcm=Math.max(1,Number(pxPerCm)||37.8);
  const requested=12*(Number(distanceCm||300)/300)*pxcm;
  const maxR=Math.min(cssW*0.40,cy-55,cssH*0.62);
  const r=Math.max(90,Math.min(requested,maxR));
  ctx.strokeStyle='#000';ctx.lineWidth=Math.max(1.5,cssW/550);ctx.lineCap='butt';
  for(let a=0;a<=180;a+=10){const t=a*Math.PI/180;ctx.beginPath();ctx.moveTo(cx,cy);ctx.lineTo(cx+r*Math.cos(t),cy-r*Math.sin(t));ctx.stroke();}
  const labelR=r+Math.max(20,r*.10);ctx.fillStyle='#000';ctx.font='700 '+Math.max(12,Math.min(22,r/12))+'px Arial';ctx.textAlign='center';ctx.textBaseline='middle';
  for(let a=0;a<=180;a+=10){const t=a*Math.PI/180;let label=a===90?'180':(a<90?String(90-a):String(270-a));ctx.fillText(label,cx+labelR*Math.cos(t),cy-labelR*Math.sin(t));}
  document.getElementById('dialInfo').textContent='觀看距離 '+Number(distanceCm||300).toFixed(Number(distanceCm||300)%1?1:0)+' cm｜已套用本機 5 公分校正';
}
async function refresh(force=false){
  if(!ensureCalibration())return;
  if(participantRefreshBusy&&!force)return;
  participantRefreshBusy=true;
  try{
    const controller=new AbortController();
    const timeout=setTimeout(()=>controller.abort(),2500);
    const r=await fetch('/state?role=participant&session='+encodeURIComponent(SESSION)+'&t='+Date.now(),{cache:'no-store',signal:controller.signal});
    clearTimeout(timeout);
    if(!r.ok)throw new Error();
    const d=await r.json();
    participantOnline();
    updateStickers(d);
    if((d.current_view||'chart')==='dial'){
      showOnly('dialScreen');
      renderParticipantDial(d.distance_cm||300);
      document.getElementById('dialStatus').textContent='散光鐘已收到，請比較各方向線條的清晰度';
      return;
    }
    if((d.current_view||'chart')==='amsler'){
      showOnly('amslerScreen');
      renderParticipantAmsler();
      document.getElementById('amslerStatus').textContent='阿姆斯勒方格已收到：請在 30 cm 單眼注視中央黑點，觀察格線是否彎曲、缺損或模糊';
      return;
    }
    if((d.current_view||'chart')==='worth'){
      showOnly('worthScreen');
      renderParticipantWorth(d.worth_mode);
      document.getElementById('worthStatus').textContent='Worth 四點已收到：請配戴紅綠眼鏡並回報看到的點數與顏色';
      return;
    }
    if((d.current_view||'chart')==='bagolini'){
      showOnly('bagoliniScreen');
      renderParticipantBagolini(d.bagolini_mode,d.bagolini_dot_px);
      document.getElementById('bagoliniStatus').textContent='Bagolini 已收到：請配戴條紋鏡並持續注視中央白色光點';
      return;
    }
    if((d.current_view||'chart')==='thorington'){
      showOnly('thoringtonScreen');
      lastThoringtonDistance=Number(d.thorington_distance_cm)||100;
      lastThoringtonDotPx=Number(d.thorington_dot_px)||5;
      renderParticipantThorington(lastThoringtonDistance,lastThoringtonDotPx);
      return;
    }
    if(d.question_status==='waiting_send'){
      // 正式測驗尚未開始前，維持完整視力表供初估視力使用。
      // 一旦已進入正式作答，清除上一題後仍停留在作答頁，等待測驗者發送下一題，
      // 不再自動跳回完整視力表。
      if(!formalTestStarted){
        showOnly('chartScreen');
        renderFullChart();
      }else{
        showOnly('testScreen');
        enable(false);
        const qd=document.getElementById('questionDisplay');
        if(qd&&lastQuestionTargets.length)renderQuestionTargets(lastOptotypeMode,lastQuestionTargets);
        document.getElementById('progress').textContent='等待下一題';
        document.getElementById('status').textContent='本題已完成，等待測驗者發送下一題';
        document.getElementById('undoBtn').disabled=true;
        document.getElementById('resetBtn').disabled=true;
      }
      return;
    }
    const om=d.optotype_mode||'letter';
    lastOptotypeMode=om;
    document.querySelectorAll('.letterAnswer').forEach(b=>b.classList.toggle('hidden',om!=='letter'));
    document.querySelectorAll('.cardinalAnswer').forEach(b=>b.classList.toggle('hidden',om==='letter'));
    document.querySelectorAll('.diagonalAnswer').forEach(b=>b.classList.toggle('hidden',om!=='landolt_c'));
    document.querySelector('.answerGrid').classList.toggle('directionMode',om!=='letter');
    const answers=d.answers||[];
    const hasAnswers=answers.length>0;
    document.getElementById('undoBtn').disabled=!hasAnswers;
    document.getElementById('resetBtn').disabled=!hasAnswers;
    if(d.test_mode!=='child'){
      enable(false);showOnly('waitingScreen');
      document.getElementById('waitStatus').textContent='目前為成人模式，請等待測驗者';
      return;
    }
    formalTestStarted=true;
    showOnly('testScreen');
    const targets=d.active_targets||[];
    if(targets.length)lastQuestionTargets=targets.slice();
    renderQuestionTargets(om,(targets.length?targets:lastQuestionTargets));
    applyPhysicalSize(d);
    if(d.question_status==='awaiting_answer'){
      enable(true);
      document.getElementById('progress').textContent='作答進度：'+answers.length+' / '+targets.length;
      document.getElementById('status').textContent=answers.length?'已送出：'+answers.join(' '):'題目已收到，請開始作答';
    }else{
      enable(false);
      document.getElementById('progress').textContent='本題已完成';
      document.getElementById('status').textContent='答案已送出，等待下一題';
    }
  }catch(e){participantOffline()}
  finally{participantRefreshBusy=false;}
}
async function participantPollLoop(){
  await refresh();
  participantPollTimer=setTimeout(participantPollLoop,400);
}
function startParticipantPolling(immediate=false){
  if(participantPollTimer){clearTimeout(participantPollTimer);participantPollTimer=null;}
  if(immediate)refresh(true);
  participantPollTimer=setTimeout(participantPollLoop,150);
}
window.addEventListener('resize',()=>{if(!document.getElementById('calibrationScreen').classList.contains('hidden'))renderCalibration();if(!document.getElementById('validationScreen').classList.contains('hidden'))renderValidation();if(!document.getElementById('acuityVerifyScreen').classList.contains('hidden'))renderAcuityVerification();if(!document.getElementById('dialScreen').classList.contains('hidden'))renderParticipantDial(300);if(!document.getElementById('amslerScreen').classList.contains('hidden'))renderParticipantAmsler();if(!document.getElementById('thoringtonScreen').classList.contains('hidden'))renderParticipantThorington(lastThoringtonDistance,lastThoringtonDotPx);if(!document.getElementById('worthScreen').classList.contains('hidden'))renderParticipantWorth('near')});
document.addEventListener('DOMContentLoaded',()=>{
  const mainBtn=document.querySelector('#chartScreen .estimateSend');
  if(mainBtn&&!mainBtn.dataset.bound){mainBtn.dataset.bound='1';mainBtn.addEventListener('click',event=>{event.preventDefault();submitSelectedEstimate();});}
  const modalBtn=document.querySelector('#estimateModal .estimateSend');
  if(modalBtn&&!modalBtn.dataset.bound){modalBtn.dataset.bound='1';modalBtn.addEventListener('click',event=>{event.preventDefault();submitModalEstimate();});}
});
startCalibration();startParticipantPolling(true);
</script></body></html>"""

    def current_displayed_letters(self) -> list[str]:
        if self.remote_single_row_index is None or not self.row_letters:
            return []
        index = self.remote_single_row_index
        count = max(1, min(5, int(self.remote_stimulus_count)))
        start = 0 if self.remote_stimulus_index is None else self.remote_stimulus_index % 5
        positions = list(range(5)) if count == 5 else [(start + offset) % 5 for offset in range(count)]
        if self.optotype_mode_var.get() == "letter":
            return [self.row_letters[index][p] for p in positions]
        return [DIRECTION_SYMBOLS[self.row_directions[index][p]] for p in positions]

    def _safe_remote_distance(self) -> float:
        try:
            return float(self.distance_var.get()) if self.distance_var.get().strip() else 300.0
        except (TypeError, ValueError, tk.TclError):
            return 300.0

    def remote_state(self) -> dict:
        preview = self.current_displayed_letters()
        return {
            "question_id": self.question_id,
            "current_view": self.current_view,
            "worth_mode": self.worth_mode,
            "bagolini_mode": self.bagolini_mode,
            "bagolini_dot_px": self.bagolini_dot_px,
            "thorington_distance_cm": self.thorington_distance_cm,
            "thorington_dot_px": self.thorington_dot_px,
            "test_mode": self.test_mode,
            "optotype_mode": self.optotype_mode_var.get(),
            "question_status": self.question_status,
            "preview_targets": preview,
            "active_targets": list(self.active_targets),
            "answers": list(self.answer_sequence),
            "result_marks": ["✓" if ok else "✗" for ok in self.answer_results],
            "correct_count": sum(self.answer_results),
            "elapsed": self.last_elapsed,
            "va": None if self.remote_single_row_index is None else DECIMAL_LEVELS[self.remote_single_row_index],
            "prev_va": None if self.remote_single_row_index is None or self.remote_single_row_index <= 0 else DECIMAL_LEVELS[self.remote_single_row_index - 1],
            "next_va": None if self.remote_single_row_index is None or self.remote_single_row_index >= len(DECIMAL_LEVELS)-1 else DECIMAL_LEVELS[self.remote_single_row_index + 1],
            "stimulus_count": self.remote_stimulus_count,
            "mode": self.remote_mode,
            "distance_cm": self._safe_remote_distance(),
            "target_height_cm": (None if self.remote_single_row_index is None else self.letter_height_cm(DECIMAL_LEVELS[self.remote_single_row_index], self._safe_remote_distance())),
            "estimated_va": self.remote_estimated_va,
            "estimated_va_version": self.remote_estimated_va_version,
            "estimated_va_seen_version": self.remote_estimated_va_seen_version,
            "estimated_va_last_submit": self.remote_estimated_va_last_submit,
            "estimated_va_last_control_read": self.remote_estimated_va_last_control_read,
            "estimated_va_last_submit_ip": self.remote_estimated_va_last_submit_ip,
            "estimated_va_last_control_ip": self.remote_estimated_va_last_control_ip,
            "estimated_va_ack": (self.remote_estimated_va is not None and self.remote_estimated_va_seen_version >= self.remote_estimated_va_version),
            "stickers": list(self.sticker_collection),
            "pending_stickers": list(self.pending_stickers),
            "sticker_count": len(self.sticker_collection),
            "sticker_version": self.sticker_version,
        }

    def set_test_mode(self, mode: str) -> None:
        if mode not in ("adult", "child"):
            return
        self.clear_current_question(silent=True)
        self.test_mode = mode
        self.status_var.set("已切換為成人模式：只使用測驗者端。" if mode == "adult" else "已切換為兒童模式：受試者手機可作答。")
        self._update_result_panel()

    def send_current_question(self) -> None:
        targets = self.current_displayed_letters()
        if not targets:
            self.status_var.set("請先選擇視力值與刺激數量，再發送題目。")
            return
        if self.test_mode != "child":
            self.status_var.set("成人模式不需要發送到受試者手機。")
            return
        if self.question_status != "waiting_send":
            self.status_var.set("上一題尚未清除，請先按『確認並清除』。")
            return
        self.active_targets = list(targets)
        self.answer_sequence = []
        self.answer_results = []
        self.last_answer = ""
        self.last_answer_correct = None
        self.last_elapsed = None
        self.question_started_at = time.monotonic()
        self.question_id += 1
        self.question_status = "awaiting_answer"
        self._update_result_panel()
        self.status_var.set(f"題目已發送：{' '.join(self.active_targets)}；等待 {len(self.active_targets)} 個答案。")

    def clear_current_question(self, silent: bool = False) -> None:
        self.active_targets = []
        self.answer_sequence = []
        self.answer_results = []
        self.last_answer = ""
        self.last_answer_correct = None
        self.last_elapsed = None
        self.question_status = "waiting_send"
        self.question_id += 1
        self.question_started_at = time.monotonic()
        self._update_result_panel()
        if not silent:
            self.status_var.set("上一題已同步清除；可以設定並發送下一題。")

    def _update_result_panel(self) -> None:
        preview = self.current_displayed_letters()
        targets = self.active_targets if self.question_status != "waiting_send" else preview
        prefix = "目前預覽：" if self.question_status == "waiting_send" else "本題題目："
        self.result_question_var.set(prefix + (" ".join(targets) if targets else "尚未選題"))
        self.result_answer_var.set("受試者答案：" + (" ".join(self.answer_sequence) if self.answer_sequence else "尚未作答"))
        if self.question_status == "completed":
            marks = " ".join("✓" if ok else "✗" for ok in self.answer_results)
            count = sum(self.answer_results)
            self.result_judgement_var.set(f"{marks}　{count}/{len(self.active_targets)}")
            self.result_judgement_label.configure(foreground="#14823b" if count == len(self.active_targets) else "#c62828")
        elif self.question_status == "awaiting_answer":
            self.result_judgement_var.set(f"等待作答：{len(self.answer_sequence)}/{len(self.active_targets)}")
            self.result_judgement_label.configure(foreground="#555555")
        else:
            self.result_judgement_var.set("等待發送")
            self.result_judgement_label.configure(foreground="#555555")
        self.result_time_var.set("" if self.last_elapsed is None else f"反應時間：{self.last_elapsed:.2f} 秒")

    def receive_answer(self, answer: str) -> None:
        answer = answer.strip().upper()
        valid_answers = set(SLOAN_LETTERS) | set(DIRECTION_SYMBOLS.values())
        if answer not in valid_answers:
            self.status_var.set("手機收到無效答案。")
            return
        if self.test_mode != "child":
            self.status_var.set("目前是成人模式，受試者手機答案未採用。")
            return
        if self.question_status != "awaiting_answer" or not self.active_targets:
            self.status_var.set("目前沒有已發送且等待作答的題目。")
            return
        index = len(self.answer_sequence)
        if index >= len(self.active_targets):
            return
        correct = answer == self.active_targets[index]
        self.answer_sequence.append(answer)
        self.answer_results.append(correct)
        self.last_answer = answer
        self.last_answer_correct = correct
        if len(self.answer_sequence) >= len(self.active_targets):
            self.last_elapsed = max(0.0, time.monotonic() - self.question_started_at)
            self.question_status = "completed"
            self.status_var.set(f"本題完成：{' '.join(self.active_targets)}｜回答：{' '.join(self.answer_sequence)}｜正確 {sum(self.answer_results)}/{len(self.active_targets)}；請確認並清除。")
        else:
            self.status_var.set(f"已收到第 {len(self.answer_sequence)}/{len(self.active_targets)} 個答案：{answer}")
        self._update_result_panel()

    def undo_last_answer(self) -> None:
        """Remove the most recently entered participant answer and continue the same question."""
        if self.test_mode != "child" or not self.active_targets:
            self.status_var.set("目前沒有可以退回的受試者答案。")
            return
        if not self.answer_sequence:
            self.status_var.set("目前尚未作答，沒有答案可以退回。")
            return
        removed = self.answer_sequence.pop()
        if self.answer_results:
            self.answer_results.pop()
        self.last_answer = self.answer_sequence[-1] if self.answer_sequence else ""
        self.last_answer_correct = self.answer_results[-1] if self.answer_results else None
        self.last_elapsed = None
        self.question_status = "awaiting_answer"
        self.status_var.set(f"已退回上一個答案：{removed}；目前作答 {len(self.answer_sequence)}/{len(self.active_targets)}。")
        self._update_result_panel()

    def reset_current_answers(self) -> None:
        """Clear all answers while keeping the currently displayed question active."""
        if self.test_mode != "child" or not self.active_targets:
            self.status_var.set("目前沒有可以重新作答的題目。")
            return
        self.answer_sequence = []
        self.answer_results = []
        self.last_answer = ""
        self.last_answer_correct = None
        self.last_elapsed = None
        self.question_status = "awaiting_answer"
        self.question_started_at = time.monotonic()
        self.status_var.set("本題答案已全部清除，受試者可以重新作答。")
        self._update_result_panel()

    def mark_new_question(self) -> None:
        if self.question_status == "waiting_send":
            self.question_started_at = time.monotonic()
            self._update_result_panel()

    def _sticker_text(self, limit: int = 30) -> str:
        shown = self.sticker_collection[-limit:]
        lines = ["  ".join(shown[i:i + 5]) for i in range(0, len(shown), 5)]
        return "\n".join(lines) if lines else "尚未收到貼圖"

    def _refresh_sticker_display(self) -> None:
        label = self.sticker_collection_label
        if label is None:
            return
        try:
            if label.winfo_exists():
                label.configure(
                    text=f"我的貼圖收藏（{len(self.sticker_collection)}）\n{self._sticker_text(limit=40)}"
                )
            else:
                self.sticker_collection_label = None
        except tk.TclError:
            self.sticker_collection_label = None

    def _refresh_pending_sticker_display(self) -> None:
        text = " ".join(self.pending_stickers) if self.pending_stickers else "尚未選擇貼圖"
        self.pending_sticker_var.set(f"已選：{text}")

    def select_reward_sticker(self, sticker: str) -> None:
        if sticker not in REWARD_STICKERS:
            self.status_var.set("無效的獎勵貼圖。")
            return
        self.pending_stickers.append(sticker)
        self._refresh_pending_sticker_display()
        self.status_var.set(f"已選擇 {sticker}；按『發送貼圖』後才會送給受試者。")

    def clear_pending_stickers(self) -> None:
        self.pending_stickers.clear()
        self._refresh_pending_sticker_display()
        self.status_var.set("已清除尚未發送的貼圖。")

    def send_selected_stickers(self) -> None:
        if self.test_mode != "child":
            self.status_var.set("目前為成人模式，未送出兒童獎勵貼圖。")
            return
        if not self.pending_stickers:
            self.status_var.set("請先選擇至少一張貼圖，再按『發送貼圖』。")
            return
        sent = list(self.pending_stickers)
        self.sticker_collection.extend(sent)
        self.pending_stickers.clear()
        self.sticker_version += 1
        self._refresh_pending_sticker_display()
        self._refresh_sticker_display()
        self.status_var.set(
            f"測驗者已送出 {' '.join(sent)}；目前共收集 {len(self.sticker_collection)} 個貼圖。"
            "現在可直接切換視力值、距離、上一列、下一列或其他功能。"
        )

    def send_reward_sticker(self, sticker: str) -> None:
        # 相容舊呼叫：改成先選擇，再由發送按鈕送出。
        self.select_reward_sticker(sticker)

    def clear_reward_stickers(self) -> None:
        self.sticker_collection.clear()
        self.sticker_version += 1
        self._refresh_sticker_display()
        self.status_var.set("兒童貼圖收藏已清空。")

    # ---------- geometry ----------
    @staticmethod
    def decimal_va_to_logmar(decimal_va: float) -> float:
        return -math.log10(decimal_va)

    @staticmethod
    def letter_height_cm(decimal_va: float, distance_cm: float) -> float:
        """完整字母視角 = 5 / VA arcmin。"""
        angle_deg = (5.0 / decimal_va) / 60.0
        return 2.0 * distance_cm * math.tan(math.radians(angle_deg) / 2.0)

    def screen_pixels_per_cm(self) -> float:
        try:
            return float(self.root.winfo_fpixels("1i")) / 2.54
        except Exception:
            return 96.0 / 2.54

    @staticmethod
    def reading_font(size: int):
        for name in (
            "DejaVuSans-Bold.ttf",
            "arialbd.ttf",
            "Arial Bold.ttf",
            "LiberationSans-Bold.ttf",
        ):
            try:
                return ImageFont.truetype(name, size)
            except OSError:
                continue
        return ImageFont.load_default()

    def make_exact_height_text_image(self, text: str, target_height_px: int) -> Image.Image:
        """將黑色字形本身縮放到指定高度，而非只設定字型 point size。"""
        target_height_px = max(1, int(round(target_height_px)))
        font = self.reading_font(260)
        probe = Image.new("L", (5000, 600), 255)
        draw = ImageDraw.Draw(probe)
        bbox = draw.textbbox((0, 0), text, font=font, spacing=0)
        width = max(1, bbox[2] - bbox[0])
        height = max(1, bbox[3] - bbox[1])

        image = Image.new("L", (width + 20, height + 20), 255)
        ImageDraw.Draw(image).text(
            (10 - bbox[0], 10 - bbox[1]), text, fill=0, font=font
        )
        ink_bbox = ImageOps.invert(image).getbbox()
        if ink_bbox:
            image = image.crop(ink_bbox)

        new_width = max(
            1, int(round(image.width * target_height_px / max(1, image.height)))
        )
        resampling = getattr(Image, "Resampling", Image).LANCZOS
        return image.resize((new_width, target_height_px), resampling)

    # ---------- actions ----------
    def randomize_letters(self, refresh: bool = True) -> None:
        """重新排列 Sloan 字母以及 Landolt C／Tumbling E 的方向。"""
        self.row_letters = []
        self.row_directions = []
        previous_letters = None
        previous_directions = None
        for _ in DECIMAL_LEVELS:
            letters = random.sample(list(SLOAN_LETTERS), 5)
            while previous_letters is not None and letters == previous_letters:
                letters = random.sample(list(SLOAN_LETTERS), 5)
            self.row_letters.append(letters)
            previous_letters = letters

            direction_pool = LANDOLT_DIRECTIONS if self.optotype_mode_var.get() == "landolt_c" else CARDINAL_DIRECTIONS
            directions = [random.choice(direction_pool) for _ in range(5)]
            while previous_directions is not None and directions == previous_directions:
                directions = [random.choice(direction_pool) for _ in range(5)]
            self.row_directions.append(directions)
            previous_directions = directions
        if refresh:
            self.refresh_chart()

    def set_optotype_mode(self, mode: str) -> None:
        if mode not in ("letter", "landolt_c", "tumbling_e"):
            return
        self.optotype_mode_var.set(mode)
        self.change_optotype_mode()

    def change_optotype_mode(self) -> None:
        if self.question_status == "awaiting_answer":
            self.status_var.set("受試者仍在作答中；請先完成或清除本題，再切換視標。")
            return
        if self.question_status == "completed":
            self.clear_current_question()
        self.randomize_letters(refresh=False)
        self.refresh_chart()

    @staticmethod
    def optotype_mode_name(mode: str) -> str:
        return {"letter": "Sloan Letter", "landolt_c": "Landolt C", "tumbling_e": "Tumbling E"}.get(mode, "Sloan Letter")

    @staticmethod
    def _v63_rotated_rectangle(center, angle_deg, radial_start, radial_end, half_width):
        angle = math.radians(angle_deg)
        ux, uy = math.cos(angle), math.sin(angle)
        vx, vy = -uy, ux
        cx, cy = center
        return [
            (cx + ux * radial_start + vx * half_width, cy + uy * radial_start + vy * half_width),
            (cx + ux * radial_end + vx * half_width, cy + uy * radial_end + vy * half_width),
            (cx + ux * radial_end - vx * half_width, cy + uy * radial_end - vy * half_width),
            (cx + ux * radial_start - vx * half_width, cy + uy * radial_start - vy * half_width),
        ]

    def make_landolt_c_image(self, target_px: int, direction: str) -> Image.Image:
        """沿用今日 V6.3 正式 Landolt C：外徑、環寬、缺口＝5:1:1。"""
        target_px = max(6, int(round(target_px)))
        supersample = 10
        hi = target_px * supersample
        pad = max(supersample * 2, int(round(hi * 0.10)))
        canvas = hi + pad * 2
        center = (canvas / 2.0, canvas / 2.0)

        alpha = Image.new("L", (canvas, canvas), 0)
        draw = ImageDraw.Draw(alpha)
        outer_r = hi / 2.0
        stroke = hi / 5.0
        inner_r = outer_r - stroke
        draw.ellipse(
            [center[0]-outer_r, center[1]-outer_r, center[0]+outer_r, center[1]+outer_r],
            fill=255,
        )
        draw.ellipse(
            [center[0]-inner_r, center[1]-inner_r, center[0]+inner_r, center[1]+inner_r],
            fill=0,
        )

        angle_map = {"right": 0.0, "down_right": 45.0, "down": 90.0, "down_left": 135.0, "left": 180.0, "up_left": 225.0, "up": 270.0, "up_right": 315.0}
        gap_width = hi / 5.0
        draw.polygon(
            self._v63_rotated_rectangle(
                center, angle_map.get(direction, 0.0), 0.0,
                outer_r + pad * 2, gap_width / 2.0
            ),
            fill=0,
        )

        rgba = Image.new("RGBA", (canvas, canvas), (0, 0, 0, 0))
        rgba.putalpha(alpha)
        resampling = getattr(Image, "Resampling", Image).LANCZOS
        rgba = rgba.resize((max(target_px + 6, round(canvas / supersample)),) * 2, resampling)
        bbox = rgba.getbbox()
        if bbox:
            rgba = rgba.crop(bbox)
        rgba.thumbnail((target_px, target_px), resampling)
        result = Image.new("L", (target_px, target_px), 255)
        mask = rgba.getchannel("A")
        black = Image.new("L", rgba.size, 0)
        result.paste(black, ((target_px-rgba.width)//2, (target_px-rgba.height)//2), mask)
        return result

    def make_tumbling_e_image(self, target_px: int, direction: str) -> Image.Image:
        """沿用今日 V6.3 正式 Tumbling E：完整 5×5 標準網格。"""
        target_px = max(6, int(round(target_px)))
        supersample = 10
        hi = target_px * supersample
        pad = max(supersample * 2, int(round(hi * 0.10)))
        canvas = hi + pad * 2
        alpha = Image.new("L", (canvas, canvas), 0)
        draw = ImageDraw.Draw(alpha)
        unit = hi / 5.0
        x0 = y0 = pad
        x1 = y1 = pad + hi
        draw.rectangle([x0, y0, x0 + unit, y1], fill=255)
        draw.rectangle([x0, y0, x1, y0 + unit], fill=255)
        draw.rectangle([x0, y0 + 2*unit, x1, y0 + 3*unit], fill=255)
        draw.rectangle([x0, y0 + 4*unit, x1, y1], fill=255)

        rgba = Image.new("RGBA", (canvas, canvas), (0, 0, 0, 0))
        rgba.putalpha(alpha)
        rotate_map = {"right": 0, "down": -90, "left": 180, "up": 90}
        angle = rotate_map.get(direction, 0)
        if angle:
            bicubic = getattr(Image, "Resampling", Image).BICUBIC
            rgba = rgba.rotate(angle, resample=bicubic, expand=False)

        resampling = getattr(Image, "Resampling", Image).LANCZOS
        rgba = rgba.resize((max(target_px + 6, round(canvas / supersample)),) * 2, resampling)
        bbox = rgba.getbbox()
        if bbox:
            rgba = rgba.crop(bbox)
        rgba.thumbnail((target_px, target_px), resampling)
        result = Image.new("L", (target_px, target_px), 255)
        mask = rgba.getchannel("A")
        black = Image.new("L", rgba.size, 0)
        result.paste(black, ((target_px-rgba.width)//2, (target_px-rgba.height)//2), mask)
        return result

    def make_optotype_row_image(self, index: int, target_px: int, selected_positions: list[int]) -> Image.Image:
        mode = self.optotype_mode_var.get()
        gap = max(4, int(round(target_px * 0.65)))
        row_width = len(selected_positions) * target_px + max(0, len(selected_positions) - 1) * gap
        row = Image.new("L", (max(1, row_width), target_px), 255)
        for out_pos, source_pos in enumerate(selected_positions):
            if mode == "landolt_c":
                optotype = self.make_landolt_c_image(target_px, self.row_directions[index][source_pos])
            elif mode == "tumbling_e":
                optotype = self.make_tumbling_e_image(target_px, self.row_directions[index][source_pos])
            else:
                glyph = self.make_exact_height_text_image(self.row_letters[index][source_pos], target_px)
                optotype = Image.new("L", (target_px, target_px), 255)
                if glyph.width > target_px:
                    glyph = glyph.resize((target_px, target_px), getattr(Image, "Resampling", Image).LANCZOS)
                optotype.paste(glyph, ((target_px - glyph.width)//2, (target_px - glyph.height)//2))
            row.paste(optotype, (out_pos * (target_px + gap), 0))
        return row

    def _read_settings(self) -> tuple[float, float]:
        try:
            distance_cm = float(self.distance_var.get().strip())
            calibration = float(self.calibration_var.get().strip())
        except ValueError as exc:
            raise ValueError("觀看距離與校正倍率必須輸入數字。") from exc

        if not 20 <= distance_cm <= 2000:
            raise ValueError("觀看距離請設定在 20～2000 cm。")
        if not 0.2 <= calibration <= 5.0:
            raise ValueError("校正倍率請設定在 0.2～5.0。")
        return distance_cm, calibration


    def calibrate_from_ruler(self) -> None:
        """依使用者用實體尺量到的 5 cm 校正線長度，自動更新倍率。

        新倍率 = 目前倍率 × 5.00 / 實際量得長度。
        例如目前倍率 1.000、實測 4.2 cm，則新倍率為 1.190476。
        """
        try:
            current = float(self.calibration_var.get().strip())
        except ValueError:
            current = 1.0

        measured = simpledialog.askfloat(
            "5 cm 實尺校正",
            "請用尺量上方『5.00 cm 校正線』，輸入實際量到的公分數：",
            initialvalue=4.2,
            minvalue=1.0,
            maxvalue=10.0,
            parent=self.root,
        )
        if measured is None:
            return

        new_calibration = current * 5.0 / measured
        self.calibration_var.set(f"{new_calibration:.6f}")
        self.apply_current_view()
        self.status_var.set(
            f"校正完成：原校正線實測 {measured:.3f} cm；新倍率 {new_calibration:.6f}。"
            "請再用尺確認校正線為 5.00 cm。"
        )
        messagebox.showinfo(
            "校正完成",
            f"實測長度：{measured:.3f} cm\n"
            f"新校正倍率：{new_calibration:.6f}\n\n"
            "請再次用尺確認畫面上的校正線為 5.00 cm。",
            parent=self.root,
        )

    def _prepare_view_change(self) -> None:
        """切換功能畫面前，清除舊題與失效的畫面元件參照。"""
        if self.question_status != "waiting_send":
            self.clear_current_question(silent=True)
        # chart_frame 內容即將被銷毀；不可再保留舊貼圖 Label 的 Tk 參照。
        self.sticker_collection_label = None

    def apply_current_view(self) -> None:
        """依目前畫面套用觀看距離與校正倍率。"""
        if self.current_view == "dial":
            self.show_astigmatic_dial()
        elif self.current_view == "amsler":
            self.show_amsler_grid()
        elif self.current_view == "worth":
            self.show_worth_four_dot(self.worth_mode)
        elif self.current_view == "bagolini":
            self.show_bagolini_test(self.bagolini_mode)
        elif self.current_view == "thorington":
            self.show_thorington_test()
        else:
            self.refresh_chart()

    def show_full_acuity_chart(self) -> None:
        """由手機「完整視力表」按鈕恢復全部 13 列。"""
        self._prepare_view_change()
        self.remote_mode = "full"
        self.remote_single_row_index = None
        self.remote_stimulus_index = None
        self.show_acuity_chart()

    def show_acuity_chart(self) -> None:
        if self.bagolini_focus_mode:
            self.toggle_bagolini_focus_mode()
        self.current_view = "chart"
        self.canvas.configure(background="white")
        self.chart_frame.configure(background="white")
        self.chart_frame.pack_propagate(True)
        self.info_frame.pack(fill=tk.X, before=self.canvas.master) if self.controls_visible and not self.info_frame.winfo_ismapped() else None
        self.v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.configure(yscrollcommand=self.v_scroll.set)
        self.refresh_chart()

    def show_astigmatic_dial(self) -> None:
        self._prepare_view_change()
        self.current_view = "dial"
        self.canvas.configure(background="white")
        self.chart_frame.configure(background="white")
        try:
            distance_cm, calibration = self._read_settings()
        except ValueError as exc:
            messagebox.showerror("設定錯誤", str(exc))
            return

        self.info_frame.pack_forget()
        self.v_scroll.pack_forget()
        self.canvas.configure(yscrollcommand=None)

        for child in self.chart_frame.winfo_children():
            child.destroy()
        self.photo_images.clear()

        dial = tk.Canvas(
            self.chart_frame,
            background="white",
            highlightthickness=0,
            borderwidth=0,
        )
        dial.pack(fill=tk.BOTH, expand=True)
        self.dial_canvas = dial

        # 散光鐘畫面必須與外層 Canvas 同高。原版本只設定寬度，
        # chart_frame 的高度會維持在很小的 requested height，造成圖片下半部
        # （也就是所有線條的共同中心）被外層 Canvas 裁掉。
        self.root.update_idletasks()
        viewport_w = max(700, self.canvas.winfo_width())
        viewport_h = max(500, self.canvas.winfo_height())
        self.canvas.itemconfigure(
            self.chart_window_id, width=viewport_w, height=viewport_h
        )
        self.chart_frame.configure(width=viewport_w, height=viewport_h)
        self.chart_frame.pack_propagate(False)

        def draw_dial(event=None):
            dial.delete("all")
            w = max(700, dial.winfo_width())
            h = max(500, dial.winfo_height())

            # 半圓的中心下移，使散光鐘完整落在畫面中央偏下，
            # 上方數字不會再貼近或超出視窗頂端。
            cx = w / 2
            cy = h * 0.70

            px_per_cm = self.screen_pixels_per_cm() * calibration
            reference_distance_cm = 300.0
            reference_radius_cm = 12.0
            requested_radius_px = (
                reference_radius_cm
                * (distance_cm / reference_distance_cm)
                * px_per_cm
            )

            # 依中心位置限制半徑，保留上方數字與下方狀態資訊空間。
            max_radius_px = min(w * 0.40, cy - 72, h * 0.56)
            radius = max(220.0, min(requested_radius_px, max_radius_px))

            # Tkinter Canvas 的斜線沒有反鋸齒，容易出現階梯與扭曲感。
            # 改用 Pillow 在 4 倍解析度繪製，再以 LANCZOS 縮回螢幕尺寸。
            scale = 4
            img_w = max(1, int(round(w * scale)))
            img_h = max(1, int(round(h * scale)))
            image = Image.new("RGB", (img_w, img_h), "white")
            draw = ImageDraw.Draw(image)

            # 所有放射線必須使用完全相同的整數中心座標，
            # 避免縮放後在中心附近形成缺口或看起來沒有接到底部。
            center_x_hi = int(round(cx * scale))
            center_y_hi = int(round(cy * scale))
            sradius = radius * scale

            # 縮小後約 2 px 的純黑線，確保每條線從共同中心一路連到外端。
            line_width = max(8, int(round(2.0 * scale)))

            for angle in range(0, 181, 10):
                theta = math.radians(angle)
                end_x_hi = int(round(center_x_hi + sradius * math.cos(theta)))
                end_y_hi = int(round(center_y_hi - sradius * math.sin(theta)))
                draw.line(
                    [(center_x_hi, center_y_hi), (end_x_hi, end_y_hi)],
                    fill=(0, 0, 0),
                    width=line_width,
                )

            label_r = radius + max(22, radius * 0.10)
            font_size = max(11, min(20, int(round(radius / 17))))
            try:
                label_font = ImageFont.truetype("arialbd.ttf", font_size * scale)
            except OSError:
                try:
                    label_font = ImageFont.truetype("DejaVuSans-Bold.ttf", font_size * scale)
                except OSError:
                    label_font = ImageFont.load_default()

            # 數字依使用者指定重新排列：
            # 中央最上方為 180；右側由上往下為 10～90；
            # 左側由上往下為 170～100。
            # 圖形的 19 條放射線完整保留，不刪除任何線條。
            for angle in range(0, 181, 10):
                theta = math.radians(angle)
                x = (cx + label_r * math.cos(theta)) * scale
                y = (cy - label_r * math.sin(theta)) * scale

                if angle == 90:
                    label_text = "180"
                elif angle < 90:
                    # 右側：靠近中央為 10，最右側為 90。
                    label_text = str(90 - angle)
                else:
                    # 左側：靠近中央為 170，最左側為 100。
                    label_text = str(270 - angle)

                bbox = draw.textbbox((0, 0), label_text, font=label_font)
                tw = bbox[2] - bbox[0]
                th = bbox[3] - bbox[1]
                draw.text(
                    (x - tw / 2, y - th / 2), label_text,
                    fill=(0, 0, 0), font=label_font
                )

            resampling = getattr(Image, "Resampling", Image).LANCZOS
            image = image.resize((int(w), int(h)), resampling)
            photo = ImageTk.PhotoImage(image)
            dial._dial_photo = photo
            dial.create_image(0, 0, image=photo, anchor="nw")

            actual_radius_cm = radius / px_per_cm
            dial.create_text(
                16, h - 16, anchor="sw",
                text=(
                    f"觀看距離 {distance_cm:g} cm｜散光鐘半徑 {actual_radius_cm:.2f} cm"
                    + ("｜已達螢幕可容納上限" if requested_radius_px > max_radius_px else "")
                ),
                fill="#555555", font=("Microsoft JhengHei", 9)
            )

        dial.bind("<Configure>", draw_dial)
        self.root.after(80, draw_dial)
        self.root.update_idletasks()
        self._center_chart_frame()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self.canvas.yview_moveto(0.0)
        self.status_var.set(
            f"散光鐘已依 {distance_cm:g} cm 距離等比例調整；"
            "共 19 條抗鋸齒半圓直線；所有線條完整顯示並連到同一個下方中心點；中央標示 180，右側 10～90，左側 170～100；中央不畫圓圈或黑點。按 F4 回視力表。"
        )


    def show_amsler_grid(self) -> None:
        self._prepare_view_change()
        """顯示臨床標準固定尺寸的阿姆斯勒方格表。

        固定規格：30 cm 觀看距離、10 cm × 10 cm、20 × 20 格，
        每格 5 mm。方格大小只受螢幕校正倍率影響，不受上方觀看距離影響。
        """
        self.current_view = "amsler"
        self.canvas.configure(background="white")
        self.chart_frame.configure(background="white")
        try:
            _distance_cm, calibration = self._read_settings()
        except ValueError as exc:
            messagebox.showerror("設定錯誤", str(exc))
            return

        self.info_frame.pack_forget()
        self.v_scroll.pack_forget()
        self.canvas.configure(yscrollcommand=None)

        for child in self.chart_frame.winfo_children():
            child.destroy()
        self.photo_images.clear()

        amsler = tk.Canvas(
            self.chart_frame, background="white", highlightthickness=0, borderwidth=0
        )
        amsler.pack(fill=tk.BOTH, expand=True)
        self.amsler_canvas = amsler

        self.root.update_idletasks()
        viewport_w = max(700, self.canvas.winfo_width())
        viewport_h = max(500, self.canvas.winfo_height())
        self.canvas.itemconfigure(self.chart_window_id, width=viewport_w, height=viewport_h)
        self.chart_frame.configure(width=viewport_w, height=viewport_h)
        self.chart_frame.pack_propagate(False)

        def draw_grid(event=None):
            amsler.delete("all")
            w = max(700, amsler.winfo_width())
            h = max(500, amsler.winfo_height())

            # 臨床標準：固定 10 cm × 10 cm，不隨上方觀看距離改變。
            px_per_cm = self.screen_pixels_per_cm() * calibration
            requested_size_px = 10.0 * px_per_cm

            # 正常校正時使用真實 10 cm；若螢幕太小，才做安全上限。
            max_size_px = min(w * 0.82, h * 0.68)
            grid_size = min(requested_size_px, max_size_px)
            cell = grid_size / 20.0

            cx = w / 2.0
            cy = h / 2.0 + 35
            left = cx - grid_size / 2.0
            top = cy - grid_size / 2.0
            right = cx + grid_size / 2.0
            bottom = cy + grid_size / 2.0

            # 標題與固定測試距離提示。
            amsler.create_text(
                cx, 38,
                text="阿姆斯勒方格（Amsler Grid）",
                fill="black",
                font=("Microsoft JhengHei", 21, "bold"),
                anchor="center",
            )
            amsler.create_text(
                cx, 72,
                text="標準測試距離：30 cm",
                fill="#333333",
                font=("Microsoft JhengHei", 13, "bold"),
                anchor="center",
            )

            # 21 條垂直線與 21 條水平線，形成 20 × 20 格。
            line_width = max(1, int(round(px_per_cm * 0.018)))
            for i in range(21):
                x = left + i * cell
                y = top + i * cell
                amsler.create_line(x, top, x, bottom, fill="black", width=line_width)
                amsler.create_line(left, y, right, y, fill="black", width=line_width)

            amsler.create_rectangle(
                left, top, right, bottom, outline="black", width=max(2, line_width + 1)
            )

            # 中央固定注視點。
            dot_r = max(3.0, cell * 0.14)
            amsler.create_oval(
                cx - dot_r, cy - dot_r, cx + dot_r, cy + dot_r,
                fill="black", outline="black"
            )

            actual_size_cm = grid_size / px_per_cm
            footer = (
                f"固定規格：20 × 20 格｜每格 5 mm｜方格 {actual_size_cm:.2f} × {actual_size_cm:.2f} cm"
            )
            if requested_size_px > max_size_px:
                footer += "｜目前螢幕空間不足，已自動縮小顯示"
            amsler.create_text(
                cx, min(h - 24, bottom + 34),
                text=footer, fill="#555555",
                font=("Microsoft JhengHei", 10), anchor="center"
            )

        amsler.bind("<Configure>", draw_grid)
        self.root.after(80, draw_grid)
        self.root.update_idletasks()
        self._center_chart_frame()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self.canvas.yview_moveto(0.0)
        self.status_var.set(
            "阿姆斯勒方格（Amsler Grid）：固定 30 cm 測試距離、10 × 10 cm、20 × 20 格；"
            "不受上方觀看距離影響。按 F4 回視力表，F3 顯示散光鐘。"
        )


    def toggle_worth_focus_mode(self) -> None:
        """F9 切換 Worth 純四點測試畫面。

        純四點模式會隱藏工具列、標題、距離文字、操作按鈕與狀態列，
        全螢幕只保留紅、綠、綠、白四個刺激點，避免其他畫面元素干擾。
        """
        entering = not self.worth_focus_mode
        if entering:
            self._worth_previous_fullscreen = bool(self.fullscreen_var.get())
            self.worth_focus_mode = True
            self.set_fullscreen(True)
        else:
            self.worth_focus_mode = False
            if not self._worth_previous_fullscreen:
                self.set_fullscreen(False)
        self.show_worth_four_dot(self.worth_mode)

    def show_worth_four_dot(self, mode: str | None = None) -> None:
        self._prepare_view_change()
        """顯示 Worth 四點測試的標準四燈排列。

        近距模式：33 cm，四點中心間距約 6°。
        遠距模式：6 m，四點中心間距約 1.25°。
        配戴紅綠眼鏡時，通常紅片置於右眼、綠片置於左眼。
        本畫面只負責顯示刺激，不自動記錄或判讀結果。
        """
        if mode in ("near", "far"):
            self.worth_mode = mode
        self.current_view = "worth"

        # 純四點模式：隱藏所有介面，只保留測試刺激。
        if self.worth_focus_mode:
            self.control_frame.pack_forget()
            self.info_frame.pack_forget()
            self.status_label.pack_forget()
        else:
            if self.controls_visible and not self.control_frame.winfo_ismapped():
                self.control_frame.pack(fill=tk.X, before=self.canvas.master)
            if self.controls_visible and not self.info_frame.winfo_ismapped():
                self.info_frame.pack(fill=tk.X, before=self.canvas.master)
            if not self.status_label.winfo_ismapped():
                self.status_label.pack(fill=tk.X)

        try:
            _distance_cm, calibration = self._read_settings()
        except ValueError as exc:
            messagebox.showerror("設定錯誤", str(exc))
            return

        self.info_frame.pack_forget()
        self.v_scroll.pack_forget()
        self.canvas.configure(yscrollcommand=None, background="black")

        for child in self.chart_frame.winfo_children():
            child.destroy()
        self.photo_images.clear()

        worth = tk.Canvas(
            self.chart_frame,
            background="black",
            highlightthickness=0,
            borderwidth=0,
        )
        worth.pack(fill=tk.BOTH, expand=True)
        self.worth_canvas = worth

        self.root.update_idletasks()
        viewport_w = max(700, self.canvas.winfo_width())
        viewport_h = max(500, self.canvas.winfo_height())
        self.canvas.itemconfigure(self.chart_window_id, width=viewport_w, height=viewport_h)
        self.chart_frame.configure(width=viewport_w, height=viewport_h, background="black")
        self.chart_frame.pack_propagate(False)

        def draw_worth(event=None):
            worth.delete("all")
            w = max(700, worth.winfo_width())
            h = max(500, worth.winfo_height())
            px_per_cm = self.screen_pixels_per_cm() * calibration

            if self.worth_mode == "near":
                viewing_distance_cm = 33.0
                separation_deg = 6.0
                mode_text = "近距模式：33 cm"
            else:
                viewing_distance_cm = 600.0
                separation_deg = 1.25
                mode_text = "遠距模式：6 m"

            # 相鄰光點中心距離依指定視角換算；若螢幕不足，才等比例縮小。
            separation_cm = 2.0 * viewing_distance_cm * math.tan(
                math.radians(separation_deg) / 2.0
            )
            requested_sep_px = separation_cm * px_per_cm
            max_sep_px = min(w * 0.22, h * 0.24)
            sep = min(requested_sep_px, max_sep_px)

            # 光點直徑採 5 mm 實體尺寸，並設最小顯示值以保持清楚。
            requested_dot_px = 0.5 * px_per_cm
            dot_diameter = max(12.0, min(requested_dot_px, sep * 0.42))
            r = dot_diameter / 2.0

            cx = w / 2.0
            cy = h / 2.0 if self.worth_focus_mode else h / 2.0 + 25

            # 菱形排列：上紅、左右綠、下白。
            positions = [
                (cx, cy - sep, "#ff0000"),
                (cx - sep, cy, "#00ff00"),
                (cx + sep, cy, "#00ff00"),
                (cx, cy + sep, "#ffffff"),
            ]
            for x, y, color in positions:
                worth.create_oval(
                    x - r, y - r, x + r, y + r,
                    fill=color, outline=color, width=0,
                )

            if not self.worth_focus_mode:
                worth.create_text(
                    cx, 36,
                    text="Worth 四點測試（Worth Four Dot Test）",
                    fill="white",
                    font=("Microsoft JhengHei", 20, "bold"),
                    anchor="center",
                )
                worth.create_text(
                    cx, 70,
                    text=mode_text,
                    fill="#d0d0d0",
                    font=("Microsoft JhengHei", 13, "bold"),
                    anchor="center",
                )
                worth.create_text(
                    cx, h - 48,
                    text="請配戴紅綠眼鏡：右眼紅片、左眼綠片",
                    fill="#d0d0d0",
                    font=("Microsoft JhengHei", 11),
                    anchor="center",
                )

                # 畫面內切換按鈕，不加入結果記錄。
                button_y = h - 22
                near_fill = "#eeeeee" if self.worth_mode == "near" else "#777777"
                far_fill = "#eeeeee" if self.worth_mode == "far" else "#777777"
                near_text = "black" if self.worth_mode == "near" else "white"
                far_text = "black" if self.worth_mode == "far" else "white"

                near_id = worth.create_text(
                    cx - 90, button_y, text="近距 33 cm",
                    fill=near_text, font=("Microsoft JhengHei", 11, "bold"),
                    tags=("near_mode",),
                )
                far_id = worth.create_text(
                    cx + 90, button_y, text="遠距 6 m",
                    fill=far_text, font=("Microsoft JhengHei", 11, "bold"),
                    tags=("far_mode",),
                )
                for item_id, fill, tag in (
                    (near_id, near_fill, "near_mode"),
                    (far_id, far_fill, "far_mode"),
                ):
                    bbox = worth.bbox(item_id)
                    if bbox:
                        rect = worth.create_rectangle(
                            bbox[0] - 12, bbox[1] - 6, bbox[2] + 12, bbox[3] + 6,
                            fill=fill, outline="#aaaaaa", width=1,
                            tags=(tag,),
                        )
                        worth.tag_lower(rect, item_id)

                if requested_sep_px > max_sep_px:
                    worth.create_text(
                        w - 14, h - 14,
                        text="螢幕空間不足，已等比例縮小",
                        fill="#999999", font=("Microsoft JhengHei", 8),
                        anchor="se",
                    )
        def set_near(_event=None):
            self.worth_mode = "near"
            draw_worth()
            self.status_var.set("Worth 四點：近距 33 cm；上紅、左右綠、下白。")

        def set_far(_event=None):
            self.worth_mode = "far"
            draw_worth()
            self.status_var.set("Worth 四點：遠距 6 m；上紅、左右綠、下白。")

        worth.tag_bind("near_mode", "<Button-1>", set_near)
        worth.tag_bind("far_mode", "<Button-1>", set_far)

        def canvas_click(_event=None):
            if self.worth_focus_mode:
                self.toggle_worth_focus_mode()

        worth.bind("<Button-1>", canvas_click, add="+")
        worth.bind("<Configure>", draw_worth)
        self.root.after(80, draw_worth)
        self.root.update_idletasks()
        self._center_chart_frame()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self.canvas.yview_moveto(0.0)
        if self.worth_focus_mode:
            self.status_var.set("Worth 純四點模式：全螢幕只保留四個刺激點；再按 F9 或點擊畫面返回。")
        else:
            self.status_var.set(
                "Worth 四點測試：近距 33 cm／遠距 6 m；上方紅點、左右綠點、下方白點；"
                "按 F9 進入純四點模式，只保留四個刺激點。按 F4 回視力表。"
            )


    def toggle_test_focus_mode(self) -> None:
        """F10 依目前測試切換純測試畫面。"""
        if self.current_view == "thorington":
            self.toggle_thorington_focus_mode()
        else:
            self.toggle_bagolini_focus_mode()

    def toggle_thorington_focus_mode(self) -> None:
        """切換 Modified Thorington 暗室測試模式。

        隱藏所有操作介面，但保留中央亮點、水平刻度線、左側字母與右側數字。
        再按 F10 返回。
        """
        if self.current_view != "thorington":
            self.show_thorington_test()
            self.root.after(120, self.toggle_thorington_focus_mode)
            return

        self.thorington_focus_mode = not self.thorington_focus_mode
        if self.thorington_focus_mode:
            self._thorington_previous_fullscreen = bool(self.fullscreen_var.get())
            self.control_frame.pack_forget()
            self.info_frame.pack_forget()
            self.status_label.pack_forget()
            self.root.configure(cursor="none")
            self.set_fullscreen(True)
        else:
            self.root.configure(cursor="")
            self.control_frame.pack(fill=tk.X, before=self.canvas.master)
            self.status_label.pack(fill=tk.X, after=self.canvas.master)
            if not self._thorington_previous_fullscreen:
                self.set_fullscreen(False)
        self.show_thorington_test()

    def set_thorington_distance(self, distance_cm: float) -> None:
        """設定 Modified Thorington 的觀看距離並重新繪圖。

        1：40 cm（近距）
        2：100 cm（中距／研究）
        3：300 cm（遠距）
        """
        if self.current_view != "thorington":
            return
        if distance_cm not in (40.0, 100.0, 300.0):
            return
        self.thorington_distance_cm = distance_cm
        self.show_thorington_test()

    def _thorington_mode_name(self) -> str:
        if self.thorington_distance_cm == 40.0:
            return "近距 40 cm"
        if self.thorington_distance_cm == 100.0:
            return "中距 1 m"
        return "遠距 3 m"

    def show_thorington_test(self) -> None:
        self._prepare_view_change()
        """顯示電腦化 Modified Thorington 水平隱斜測試。

        右眼持水平條紋的 Maddox rod，右眼會看到垂直光線；
        左眼看到中央亮點與刻度。受試者回報垂直光線穿過哪個字母或數字。
        左側字母代表外偏，右側數字代表內偏。每格均依目前觀看距離校正為 1Δ。
        """
        self.current_view = "thorington"
        try:
            _distance_cm, calibration = self._read_settings()
        except ValueError as exc:
            messagebox.showerror("設定錯誤", str(exc))
            return

        self.info_frame.pack_forget()
        self.v_scroll.pack_forget()
        self.canvas.configure(yscrollcommand=None, background="black")
        for child in self.chart_frame.winfo_children():
            child.destroy()
        self.photo_images.clear()

        th = tk.Canvas(self.chart_frame, background="black", highlightthickness=0, borderwidth=0)
        th.pack(fill=tk.BOTH, expand=True)
        self.thorington_canvas = th

        self.root.update_idletasks()
        viewport_w = max(700, self.canvas.winfo_width())
        viewport_h = max(500, self.canvas.winfo_height())
        self.canvas.itemconfigure(self.chart_window_id, width=viewport_w, height=viewport_h)
        self.chart_frame.configure(width=viewport_w, height=viewport_h, background="black")
        self.chart_frame.pack_propagate(False)

        def draw_thorington(event=None):
            th.delete("all")
            w = max(700, th.winfo_width())
            h = max(500, th.winfo_height())
            px_per_cm = self.screen_pixels_per_cm() * calibration
            cx, cy = w / 2.0, h / 2.0

            # 1 prism diopter = 1 cm displacement at 1 m。
            # 因此螢幕上的每格實體距離 = 觀看距離（m）× 1 cm。
            step_px = px_per_cm * (self.thorington_distance_cm / 100.0)
            usable_half_width = w * 0.42
            max_steps = max(2, min(12, int(usable_half_width // max(step_px, 1))))

            line_fill = "#777777" if self.thorington_focus_mode else "#999999"
            text_fill = "#c8c8c8" if self.thorington_focus_mode else "#e0e0e0"
            line_half = max_steps * step_px
            th.create_line(cx-line_half, cy, cx+line_half, cy, fill=line_fill, width=2)

            # 近距刻度較密，字體需避免重疊；中、遠距則可放大。
            font_size = max(12, min(30, int(step_px * 0.48)))
            font = ("Microsoft JhengHei", font_size, "bold")
            letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
            label_y = cy - max(30, font_size + 18)
            for i in range(1, max_steps + 1):
                x_left = cx - i * step_px
                x_right = cx + i * step_px
                tick = 8 if i % 2 else 13
                th.create_line(x_left, cy-tick, x_left, cy+tick, fill=line_fill, width=2)
                th.create_line(x_right, cy-tick, x_right, cy+tick, fill=line_fill, width=2)
                th.create_text(x_left, label_y, text=letters[i-1], fill=text_fill, font=font)
                th.create_text(x_right, label_y, text=str(i), fill=text_fill, font=font)

            # 中央亮點模擬傳統卡片中央 LED／筆燈孔。
            r = self.thorington_dot_px / 2.0
            for extra, shade in ((7, "#202020"), (4, "#505050"), (2, "#909090")):
                rr = r + extra
                th.create_oval(cx-rr, cy-rr, cx+rr, cy+rr, fill=shade, outline=shade, width=0)
            th.create_oval(cx-r, cy-r, cx+r, cy+r, fill="#ffffff", outline="#ffffff", width=0)

            if self.thorington_focus_mode:
                return

            mode_name = self._thorington_mode_name()
            th.create_text(cx, 34, text="Modified Thorington 水平眼位測量",
                           fill="white", font=("Microsoft JhengHei", 19, "bold"))
            th.create_text(cx, 66,
                           text=f"右眼 Maddox rod：條紋水平 → 看見垂直光線｜{mode_name}｜每格 1Δ",
                           fill="#d0d0d0", font=("Microsoft JhengHei", 12, "bold"))

            # 三個距離模式按鈕。
            modes = [(40.0, "近距 40 cm"), (100.0, "中距 1 m"), (300.0, "遠距 3 m")]
            button_y = 105
            button_w = 145
            gap = 16
            total_w = len(modes)*button_w + (len(modes)-1)*gap
            start_x = cx-total_w/2
            for index, (distance_cm, label) in enumerate(modes):
                x1 = start_x + index*(button_w+gap)
                x2 = x1 + button_w
                active = self.thorington_distance_cm == distance_cm
                fill = "#e8eef7" if active else "#20242a"
                outline = "#ffffff" if active else "#777777"
                text_color = "#111111" if active else "#eeeeee"
                tag = f"thorington_mode_{int(distance_cm)}"
                th.create_rectangle(x1, button_y-17, x2, button_y+17,
                                    fill=fill, outline=outline, width=2, tags=(tag,))
                th.create_text((x1+x2)/2, button_y, text=label, fill=text_color,
                               font=("Microsoft JhengHei", 10, "bold"), tags=(tag,))

            th.create_text(cx, h-70,
                           text="請注視中央白光，回答垂直光線穿過哪一個字母或數字",
                           fill="#d0d0d0", font=("Microsoft JhengHei", 11))
            th.create_text(cx, h-42,
                           text="字母側＝外偏（XP）｜數字側＝內偏（EP）｜穿過白光＝正位（0Δ）",
                           fill="#d0d0d0", font=("Microsoft JhengHei", 11, "bold"))
            th.create_text(cx, h-16,
                           text="按 1／2／3 切換 40 cm／1 m／3 m｜F10 暗室模式｜F12 進入本測試",
                           fill="#999999", font=("Microsoft JhengHei", 9))

        th.tag_bind("thorington_mode_40", "<Button-1>", lambda _e: self.set_thorington_distance(40.0))
        th.tag_bind("thorington_mode_100", "<Button-1>", lambda _e: self.set_thorington_distance(100.0))
        th.tag_bind("thorington_mode_300", "<Button-1>", lambda _e: self.set_thorington_distance(300.0))
        th.bind("<Configure>", draw_thorington)
        self.root.after(80, draw_thorington)
        self.root.update_idletasks()
        self._center_chart_frame()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self.canvas.yview_moveto(0.0)
        mode_name = self._thorington_mode_name()
        if self.thorington_focus_mode:
            self.status_var.set(f"Modified Thorington 暗室模式：{mode_name}、每格 1Δ；按 F10 返回。")
        else:
            self.status_var.set(
                f"Modified Thorington：{mode_name}、每格 1Δ；字母側為 XP，數字側為 EP；"
                "按 1／2／3 切換距離，F10 進入暗室測試模式。"
            )


    def toggle_bagolini_focus_mode(self) -> None:
        """切換 Bagolini 純黑單點測試模式。

        進入後隱藏所有介面、狀態列與滑鼠游標，只保留黑底中央白點。
        再按 F10 或點擊黑色畫面即可返回原本操作畫面。
        """
        if self.current_view != "bagolini":
            self.show_bagolini_test(self.bagolini_mode)
            self.root.after(120, self.toggle_bagolini_focus_mode)
            return

        self.bagolini_focus_mode = not self.bagolini_focus_mode

        if self.bagolini_focus_mode:
            self._bagolini_previous_fullscreen = bool(self.fullscreen_var.get())
            self.control_frame.pack_forget()
            self.info_frame.pack_forget()
            self.status_label.pack_forget()
            self.root.configure(cursor="none")
            self.set_fullscreen(True)
        else:
            self.root.configure(cursor="")
            # 將設定列放回主內容區之前，狀態列放回最下方。
            self.control_frame.pack(fill=tk.X, before=self.canvas.master)
            self.status_label.pack(fill=tk.X, after=self.canvas.master)
            if not self._bagolini_previous_fullscreen:
                self.set_fullscreen(False)

        self.show_bagolini_test(self.bagolini_mode)



    def adjust_bagolini_dot(self, delta: int) -> None:
        """調整 Bagolini 中央白色光點直徑，限制在 2～5 px。"""
        if self.current_view != "bagolini":
            return
        new_size = max(2, min(5, self.bagolini_dot_px + delta))
        if new_size == self.bagolini_dot_px:
            return
        self.bagolini_dot_px = new_size
        self.show_bagolini_test(self.bagolini_mode)

    def show_bagolini_test(self, mode: str | None = None) -> None:
        self._prepare_view_change()
        """顯示 Bagolini 條紋鏡測試用的單一白色注視光點。

        受試者需配戴真正的 Bagolini Striated Glasses。
        螢幕只顯示單一白色光點；交叉光帶由條紋鏡產生，
        程式不在畫面上繪製任何斜線或 X。
        """
        if mode in ("near", "far"):
            self.bagolini_mode = mode
        self.current_view = "bagolini"

        try:
            _distance_cm, calibration = self._read_settings()
        except ValueError as exc:
            messagebox.showerror("設定錯誤", str(exc))
            return

        self.info_frame.pack_forget()
        self.v_scroll.pack_forget()
        self.canvas.configure(yscrollcommand=None, background="black")

        for child in self.chart_frame.winfo_children():
            child.destroy()
        self.photo_images.clear()

        bagolini = tk.Canvas(
            self.chart_frame,
            background="black",
            highlightthickness=0,
            borderwidth=0,
        )
        bagolini.pack(fill=tk.BOTH, expand=True)
        self.bagolini_canvas = bagolini

        self.root.update_idletasks()
        viewport_w = max(700, self.canvas.winfo_width())
        viewport_h = max(500, self.canvas.winfo_height())
        self.canvas.itemconfigure(self.chart_window_id, width=viewport_w, height=viewport_h)
        self.chart_frame.configure(width=viewport_w, height=viewport_h, background="black")
        self.chart_frame.pack_propagate(False)

        def draw_bagolini(event=None):
            bagolini.delete("all")
            w = max(700, bagolini.winfo_width())
            h = max(500, bagolini.winfo_height())
            px_per_cm = self.screen_pixels_per_cm() * calibration

            if self.bagolini_mode == "near":
                mode_text = "近距模式：33 cm"
            else:
                mode_text = "遠距模式：6 m"

            # 中央光點改為可調整的純白像素點，直徑 2～5 px。
            # 這比 5 mm 大圓點更接近 Bagolini 測試所需的小型點光源。
            dot_diameter = float(self.bagolini_dot_px)
            r = dot_diameter / 2.0
            cx = w / 2.0
            cy = h / 2.0 + 15

            # 畫中央白色注視點。前一版殘留了未定義的 rr／shade，
            # 導致繪圖函式在執行時中斷，因此 Bagolini 畫面看不到白點。
            # 這裡只保留清楚、純白、可調直徑的實心點。
            bagolini.create_oval(
                cx - r, cy - r, cx + r, cy + r,
                fill="#ffffff", outline="#ffffff", width=0,
                tags=("bagolini_fixation_dot",),
            )

            # 純黑測試模式：畫面只保留中央白點。
            if self.bagolini_focus_mode:
                return

            bagolini.create_text(
                cx, 38,
                text="Bagolini 條紋鏡測試（Bagolini Striated Glasses Test）",
                fill="white",
                font=("Microsoft JhengHei", 19, "bold"),
                anchor="center",
            )
            bagolini.create_text(
                cx, 72,
                text=mode_text,
                fill="#d0d0d0",
                font=("Microsoft JhengHei", 13, "bold"),
                anchor="center",
            )
            bagolini.create_text(
                cx, h - 54,
                text="請配戴 Bagolini 條紋鏡並注視中央白色光點",
                fill="#d0d0d0",
                font=("Microsoft JhengHei", 11),
                anchor="center",
            )

            bagolini.create_text(
                cx, h - 82,
                text=f"中央純白光點：{self.bagolini_dot_px} px（＋／－調整）",
                fill="#d0d0d0",
                font=("Microsoft JhengHei", 10, "bold"),
                anchor="center",
            )

            button_y = h - 24
            near_fill = "#eeeeee" if self.bagolini_mode == "near" else "#777777"
            far_fill = "#eeeeee" if self.bagolini_mode == "far" else "#777777"
            near_text = "black" if self.bagolini_mode == "near" else "white"
            far_text = "black" if self.bagolini_mode == "far" else "white"

            near_id = bagolini.create_text(
                cx - 90, button_y, text="近距 33 cm",
                fill=near_text, font=("Microsoft JhengHei", 11, "bold"),
                tags=("bagolini_near",),
            )
            far_id = bagolini.create_text(
                cx + 90, button_y, text="遠距 6 m",
                fill=far_text, font=("Microsoft JhengHei", 11, "bold"),
                tags=("bagolini_far",),
            )
            minus_id = bagolini.create_text(
                cx - 200, button_y, text="－ 光點",
                fill="black", font=("Microsoft JhengHei", 11, "bold"),
                tags=("bagolini_minus",),
            )
            plus_id = bagolini.create_text(
                cx + 200, button_y, text="＋ 光點",
                fill="black", font=("Microsoft JhengHei", 11, "bold"),
                tags=("bagolini_plus",),
            )
            for item_id, fill, tag in (
                (near_id, near_fill, "bagolini_near"),
                (far_id, far_fill, "bagolini_far"),
                (minus_id, "#eeeeee", "bagolini_minus"),
                (plus_id, "#eeeeee", "bagolini_plus"),
            ):
                bbox = bagolini.bbox(item_id)
                if bbox:
                    rect = bagolini.create_rectangle(
                        bbox[0] - 12, bbox[1] - 6, bbox[2] + 12, bbox[3] + 6,
                        fill=fill, outline="#aaaaaa", width=1, tags=(tag,),
                    )
                    bagolini.tag_lower(rect, item_id)

        def set_near(_event=None):
            self.bagolini_mode = "near"
            draw_bagolini()
            self.status_var.set("Bagolini 條紋鏡測試：近距 33 cm，注視中央白色光點。")

        def set_far(_event=None):
            self.bagolini_mode = "far"
            draw_bagolini()
            self.status_var.set("Bagolini 條紋鏡測試：遠距 6 m，注視中央白色光點。")

        def make_smaller(_event=None):
            self.adjust_bagolini_dot(-1)

        def make_larger(_event=None):
            self.adjust_bagolini_dot(1)

        bagolini.tag_bind("bagolini_near", "<Button-1>", set_near)
        bagolini.tag_bind("bagolini_far", "<Button-1>", set_far)
        bagolini.tag_bind("bagolini_minus", "<Button-1>", make_smaller)
        bagolini.tag_bind("bagolini_plus", "<Button-1>", make_larger)

        def canvas_click(_event=None):
            if self.bagolini_focus_mode:
                self.toggle_bagolini_focus_mode()

        bagolini.bind("<Button-1>", canvas_click, add="+")
        bagolini.bind("<Configure>", draw_bagolini)
        self.root.after(80, draw_bagolini)
        self.root.update_idletasks()
        self._center_chart_frame()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self.canvas.yview_moveto(0.0)
        if self.bagolini_focus_mode:
            self.status_var.set("Bagolini 單點全黑模式：再按 F10 或點擊畫面即可返回。")
        else:
            self.status_var.set(
                f"Bagolini 條紋鏡測試：中央純白光點 {self.bagolini_dot_px} px；近距 33 cm／遠距 6 m；"
                "＋／－調整光點；F10 隱藏全部介面，只保留中央白點，再按 F10 返回。按 F4 回視力表。"
            )


    def refresh_chart(self) -> None:
        if self.worth_focus_mode:
            self.worth_focus_mode = False
            if self.controls_visible and not self.control_frame.winfo_ismapped():
                self.control_frame.pack(fill=tk.X, before=self.canvas.master)
            if self.controls_visible and not self.info_frame.winfo_ismapped():
                self.info_frame.pack(fill=tk.X, before=self.canvas.master)
            if not self.status_label.winfo_ismapped():
                self.status_label.pack(fill=tk.X)
        self.current_view = "chart"
        self.canvas.configure(background="white")
        self.chart_frame.configure(background="white")
        try:
            distance_cm, calibration = self._read_settings()
        except ValueError as exc:
            messagebox.showerror("設定錯誤", str(exc))
            return

        if len(self.row_letters) != len(DECIMAL_LEVELS) or len(self.row_directions) != len(DECIMAL_LEVELS):
            self.randomize_letters(refresh=False)

        px_per_cm = self.screen_pixels_per_cm() * calibration
        self.photo_images.clear()
        self.row_widgets = []

        for child in self.chart_frame.winfo_children():
            child.destroy()

        single_index = self.remote_single_row_index
        if single_index is None:
            if not self.v_scroll.winfo_ismapped():
                self.v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
            self.canvas.configure(yscrollcommand=self.v_scroll.set)
            display_indices = list(range(len(DECIMAL_LEVELS)))
        else:
            self.v_scroll.pack_forget()
            self.canvas.configure(yscrollcommand=None)
            display_indices = [single_index]

        self._draw_calibration_line(px_per_cm)

        header = tk.Label(
            self.chart_frame,
            text=f"{self.optotype_mode_name(self.optotype_mode_var.get()).upper()} VISUAL ACUITY CHART   VIEWING DISTANCE {distance_cm:g} cm",
            background="white",
            foreground="black",
            font=("Arial", 11, "bold"),
        )
        header.pack(pady=(3, 10))

        # 兒童單列模式：大貼圖收藏固定在作答字母上方，從左到右排列。
        if single_index is not None and self.test_mode == "child":
            sticker_box = tk.Frame(self.chart_frame, background="#fff8df", bd=2, relief="solid")
            sticker_box.pack(fill=tk.X, padx=70, pady=(0, 14))
            self.sticker_collection_label = tk.Label(
                sticker_box,
                text=f"我的貼圖收藏（{len(self.sticker_collection)}）\n{self._sticker_text(limit=40)}",
                background="#fff8df", foreground="#9a4b00", justify="left", anchor="w",
                font=("Segoe UI Emoji", 34), wraplength=max(700, self.canvas.winfo_width() - 180),
                padx=18, pady=10
            )
            self.sticker_collection_label.pack(fill=tk.X)

        viewport_h = max(500, self.canvas.winfo_height())
        for index in display_indices:
            decimal_va = DECIMAL_LEVELS[index]
            logmar = self.decimal_va_to_logmar(decimal_va)
            size_cm = self.letter_height_cm(decimal_va, distance_cm)
            target_px = max(2, int(round(size_cm * px_per_cm)))
            if single_index is not None:
                count = max(1, min(5, int(self.remote_stimulus_count)))
                start = 0 if self.remote_stimulus_index is None else self.remote_stimulus_index % 5
                selected_positions = list(range(5)) if count == 5 else [(start + offset) % 5 for offset in range(count)]
            else:
                selected_positions = list(range(5))

            image = self.make_optotype_row_image(index, target_px, selected_positions)
            photo = ImageTk.PhotoImage(image.convert("RGB"))
            self.photo_images.append(photo)

            # 字母列距約等於字母高度；以 1.45 倍保留清楚空間。
            row_height = max(24, int(round(target_px * 1.45)))
            row = tk.Frame(self.chart_frame, background="white", height=row_height)
            if single_index is None:
                row.pack(fill=tk.X)
            else:
                vertical_pad = max(20, int((viewport_h - row_height - 190) / 2))
                row.pack(fill=tk.X, pady=(vertical_pad, vertical_pad))
            self.row_widgets.append(row)
            row.pack_propagate(False)

            label_width = 105
            if self.show_labels_var.get():
                tk.Label(
                    row,
                    text=f"{decimal_va:.2f}",
                    width=8,
                    anchor="e",
                    background="white",
                    foreground="black",
                    font=("Arial", 10, "bold"),
                ).pack(side=tk.LEFT, padx=(0, 12))
            else:
                tk.Frame(row, width=label_width, background="white").pack(side=tk.LEFT)

            center = tk.Canvas(
                row,
                height=row_height,
                background="white",
                highlightthickness=0,
                borderwidth=0,
            )
            center.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            center.create_image(
                0, row_height // 2, image=photo, anchor="center", tags="optotype"
            )
            center.bind(
                "<Configure>",
                lambda event, canvas=center: canvas.coords(
                    "optotype", event.width // 2, event.height // 2
                ),
            )

            if self.show_labels_var.get():
                tk.Label(
                    row,
                    text=f"{logmar:+.1f}",
                    width=8,
                    anchor="w",
                    background="white",
                    foreground="black",
                    font=("Arial", 10, "bold"),
                ).pack(side=tk.RIGHT, padx=(12, 0))
            else:
                tk.Frame(row, width=label_width, background="white").pack(side=tk.RIGHT)

        footer_text = (
            "左：十進位視力　右：logMAR　｜　"
            f"距離 {distance_cm:g} cm　｜　螢幕換算 {px_per_cm:.2f} px/cm　｜　"
            "請用尺確認上方 5 cm 校正線"
        )
        if single_index is not None:
            footer_text = (
                f"刺激模式：視力 {DECIMAL_LEVELS[single_index]:.2f}　｜　"
                f"距離 {distance_cm:g} cm　｜　顯示 {self.remote_stimulus_count} 個刺激"
            )
        footer = tk.Label(
            self.chart_frame,
            text=footer_text,
            background="white",
            foreground="#555555",
            font=("Microsoft JhengHei", 9),
        )
        footer.pack(pady=(12, 18))

        self.root.update_idletasks()
        self._center_chart_frame()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self.canvas.yview_moveto(0.0)
        if single_index is None:
            self.status_var.set(
                f"已依 {distance_cm:g} cm 觀看距離重建 13 列視力表；"
                f"最大列視標高度約 {self.letter_height_cm(0.10, distance_cm):.2f} cm。"
            )
        else:
            selected_va = DECIMAL_LEVELS[single_index]
            self.status_var.set(
                f"刺激模式：距離 {distance_cm:g} cm，視力 {selected_va:.2f}，"
                f"目前同時顯示 {self.remote_stimulus_count} 個刺激；"
                f"視標外框約 {self.letter_height_cm(selected_va, distance_cm):.2f} cm。"
            )

    def _draw_calibration_line(self, px_per_cm: float) -> None:
        c = self.calibration_canvas
        c.delete("all")
        line_px = max(1, int(round(5.0 * px_per_cm)))
        x0, y = 18, 23
        c.create_line(x0, y, x0 + line_px, y, width=4, fill="black")
        c.create_line(x0, y - 9, x0, y + 9, width=2, fill="black")
        c.create_line(x0 + line_px, y - 9, x0 + line_px, y + 9, width=2, fill="black")
        c.create_text(
            x0 + line_px + 12,
            y,
            text="5.00 cm 校正線",
            anchor="w",
            fill="black",
            font=("Microsoft JhengHei", 10, "bold"),
        )

    def set_fullscreen(self, enabled: bool) -> None:
        self.fullscreen_var.set(enabled)
        try:
            self.root.attributes("-fullscreen", enabled)
        except tk.TclError:
            if enabled:
                self.root.state("zoomed")
        if not enabled:
            try:
                self.root.state("zoomed")
            except tk.TclError:
                pass

    def toggle_fullscreen(self) -> None:
        self.set_fullscreen(not bool(self.fullscreen_var.get()))

    def toggle_controls(self) -> None:
        self.controls_visible = not self.controls_visible
        if self.controls_visible:
            self.control_frame.pack(fill=tk.X, before=self.info_frame)
            self.info_frame.pack(fill=tk.X, before=self.canvas.master)
        else:
            self.control_frame.pack_forget()
            self.info_frame.pack_forget()
        self.root.after(50, self._center_chart_frame)

    def _on_canvas_resize(self, event: tk.Event) -> None:
        width = max(700, event.width)
        if self.current_view in ("dial", "amsler", "worth"):
            height = max(500, event.height)
            self.canvas.itemconfigure(
                self.chart_window_id, width=width, height=height
            )
            self.chart_frame.configure(width=width, height=height)
            if self.current_view == "dial" and self.dial_canvas is not None and self.dial_canvas.winfo_exists():
                self.dial_canvas.configure(width=width, height=height)
            if self.current_view == "amsler" and self.amsler_canvas is not None and self.amsler_canvas.winfo_exists():
                self.amsler_canvas.configure(width=width, height=height)
            if self.current_view == "worth" and self.worth_canvas is not None and self.worth_canvas.winfo_exists():
                self.worth_canvas.configure(width=width, height=height)
        else:
            self.chart_frame.pack_propagate(True)
            self.canvas.itemconfigure(self.chart_window_id, width=width, height="")
        self.canvas.coords(self.chart_window_id, event.width // 2, 0)
        if self._resize_after_id is not None:
            self.root.after_cancel(self._resize_after_id)
        self._resize_after_id = self.root.after(80, self._center_chart_frame)

    def _center_chart_frame(self) -> None:
        if not self.canvas.winfo_exists():
            return
        width = max(700, self.canvas.winfo_width())
        if self.current_view in ("dial", "amsler", "worth"):
            height = max(500, self.canvas.winfo_height())
            self.canvas.itemconfigure(
                self.chart_window_id, width=width, height=height
            )
            self.chart_frame.configure(width=width, height=height)
        else:
            self.chart_frame.pack_propagate(True)
            self.canvas.itemconfigure(self.chart_window_id, width=width, height="")
        self.canvas.coords(self.chart_window_id, width // 2, 0)

    def _on_mousewheel(self, event: tk.Event) -> None:
        if self.current_view != "chart":
            return
        delta = getattr(event, "delta", 0)
        if delta:
            self.canvas.yview_scroll(-3 if delta > 0 else 3, "units")


def main() -> None:
    cloud_mode = os.environ.get("CLOUDVISION_BACKEND", "").strip().lower() in {"1", "true", "yes", "on"}
    root = tk.Tk()
    FullscreenAcuityChart(root, cloud_mode=cloud_mode)
    root.mainloop()


# -----------------------------------------------------------------------------
# Railway / Gunicorn WSGI compatibility layer
# -----------------------------------------------------------------------------
_backend_process = None
_xvfb_process = None
_backend_lock = threading.Lock()
_backend_port = 8765


def _ensure_virtual_display(env: dict | None = None) -> str:
    """Ensure a working Xvfb display exists and return its DISPLAY value."""
    global _xvfb_process
    target_env = env if env is not None else os.environ
    display = str(target_env.get("DISPLAY", "")).strip()
    if display:
        # Railway/Nixpacks 有時會留下 DISPLAY=:0，但實際沒有可連線的 X server。
        # 必須先用 Tk 實測，成功才沿用；失敗則清除並啟動自己的 Xvfb。
        probe = None
        try:
            probe = tk.Tk()
            probe.withdraw()
            probe.update_idletasks()
            return display
        except tk.TclError:
            target_env.pop("DISPLAY", None)
            os.environ.pop("DISPLAY", None)
            display = ""
        finally:
            if probe is not None:
                try:
                    probe.destroy()
                except Exception:
                    pass

    xvfb = shutil.which("Xvfb")
    if not xvfb:
        raise RuntimeError("找不到 Xvfb；請確認 nixpacks.toml 已安裝 xvfb。")

    display = ":99"
    socket_path = "/tmp/.X11-unix/X99"
    lock_path = "/tmp/.X99-lock"
    # Remove stale files only when no Xvfb process owned by this app is alive.
    if _xvfb_process is None or _xvfb_process.poll() is not None:
        for stale in (socket_path, lock_path):
            try:
                os.remove(stale)
            except FileNotFoundError:
                pass
            except OSError:
                pass
        _xvfb_process = subprocess.Popen(
            [xvfb, display, "-screen", "0", "1280x1024x24", "-ac", "-nolisten", "tcp"],
            stdout=sys.stdout,
            stderr=sys.stderr,
        )

    target_env["DISPLAY"] = display
    os.environ["DISPLAY"] = display

    deadline = time.time() + 10
    while time.time() < deadline:
        if _xvfb_process is not None and _xvfb_process.poll() is not None:
            raise RuntimeError("Xvfb 啟動後立即結束，無法提供虛擬螢幕。")
        if os.path.exists(socket_path):
            # Verify that Tk can actually connect before continuing.
            probe = None
            try:
                probe = tk.Tk()
                probe.withdraw()
                probe.update_idletasks()
                return display
            except tk.TclError:
                pass
            finally:
                if probe is not None:
                    try:
                        probe.destroy()
                    except Exception:
                        pass
        time.sleep(0.2)
    raise RuntimeError("Xvfb 已啟動，但 Tkinter 在 10 秒內仍無法連上 DISPLAY=:99。")


def _backend_is_ready(timeout: float = 0.5) -> bool:
    try:
        with socket.create_connection(("127.0.0.1", _backend_port), timeout=timeout):
            return True
    except OSError:
        return False


def _ensure_backend_started() -> None:
    global _backend_process
    if _backend_is_ready():
        return
    with _backend_lock:
        if _backend_is_ready():
            return
        if _backend_process is not None and _backend_process.poll() is None:
            return
        env = os.environ.copy()
        env["CLOUDVISION_BACKEND"] = "1"
        env["CLOUDVISION_BACKEND_PORT"] = str(_backend_port)
        display = _ensure_virtual_display(env)
        print(f"Cloud Vision parent prepared DISPLAY={display}", flush=True)
        script_path = os.path.abspath(__file__)
        command = [sys.executable, script_path]
        _backend_process = subprocess.Popen(command, env=env, stdout=sys.stdout, stderr=sys.stderr)


def _status_response(start_response, status: str, message: str):
    body = message.encode("utf-8")
    start_response(status, [("Content-Type", "text/plain; charset=utf-8"), ("Content-Length", str(len(body))), ("Cache-Control", "no-store")])
    return [body]


def app(environ, start_response):
    """Railway 使用的 WSGI 入口；把請求轉送到原本 V4.4 網頁伺服器。"""
    method = environ.get("REQUEST_METHOD", "GET")
    path = environ.get("PATH_INFO", "/") or "/"

    # Railway 的啟動健康檢查等待時間很短。先立即回覆 200，
    # 同時在背景啟動原本的 Tk / HTTP 後端，避免部署被誤判失敗。
    if path in {"/health", "/healthz", "/__health"}:
        return _status_response(start_response, "200 OK", "ok")

    _ensure_backend_started()
    if not _backend_is_ready():
        if path == "/":
            body = ("<!doctype html><html lang='zh-Hant'><head><meta charset='utf-8'>"
                    "<meta http-equiv='refresh' content='3'>"
                    "<meta name='viewport' content='width=device-width,initial-scale=1'>"
                    "<title>Cloud Vision 啟動中</title></head>"
                    "<body style='font-family:sans-serif;text-align:center;padding:48px'>"
                    "<h2>Cloud Vision 正在啟動</h2><p>系統準備完成後會自動重新整理。</p>"
                    "</body></html>").encode("utf-8")
            start_response("200 OK", [
                ("Content-Type", "text/html; charset=utf-8"),
                ("Content-Length", str(len(body))),
                ("Cache-Control", "no-store"),
            ])
            return [body]
        return _status_response(start_response, "503 Service Unavailable", "Cloud Vision 後端正在啟動，請稍後重試。")

    from http.client import HTTPConnection
    query = environ.get("QUERY_STRING", "")
    target = path + (("?" + query) if query else "")
    try:
        length = int(environ.get("CONTENT_LENGTH", "") or 0)
    except ValueError:
        length = 0
    body = environ["wsgi.input"].read(length) if length > 0 else None
    headers = {}
    for key, value in environ.items():
        if key.startswith("HTTP_"):
            name = key[5:].replace("_", "-").title()
            if name not in {"Host", "Connection", "Content-Length"}:
                headers[name] = value
    if environ.get("CONTENT_TYPE"):
        headers["Content-Type"] = environ["CONTENT_TYPE"]
    if body is not None:
        headers["Content-Length"] = str(len(body))
    headers["Host"] = f"127.0.0.1:{_backend_port}"
    headers["X-Forwarded-Proto"] = environ.get("HTTP_X_FORWARDED_PROTO", "https")
    headers["X-Forwarded-For"] = environ.get("HTTP_X_FORWARDED_FOR", environ.get("REMOTE_ADDR", ""))

    connection = HTTPConnection("127.0.0.1", _backend_port, timeout=120)
    try:
        connection.request(method, target, body=body, headers=headers)
        response = connection.getresponse()
        response_body = response.read()
        excluded = {"connection", "keep-alive", "proxy-authenticate", "proxy-authorization", "te", "trailers", "transfer-encoding", "upgrade", "content-length"}
        response_headers = [(name, value) for name, value in response.getheaders() if name.lower() not in excluded]
        response_headers.append(("Content-Length", str(len(response_body))))
        start_response(f"{response.status} {response.reason}", response_headers)
        return [response_body]
    except Exception as exc:
        return _status_response(start_response, "502 Bad Gateway", f"Cloud Vision 連線失敗：{exc}")
    finally:
        connection.close()


def _run_railway_entrypoint() -> None:
    """Railway 即使以 `python app.py` 啟動，也先切換到 Web 伺服器。"""
    port = os.environ.get("PORT", "").strip()
    is_railway = bool(port) or bool(os.environ.get("RAILWAY_ENVIRONMENT")) or bool(os.environ.get("RAILWAY_PROJECT_ID"))
    is_backend = os.environ.get("CLOUDVISION_BACKEND", "").strip().lower() in {"1", "true", "yes", "on"}

    if is_backend:
        # The parent normally prepares DISPLAY before spawning this process.
        # Keep this fallback so direct backend starts are safe too.
        display = _ensure_virtual_display()
        print(f"Cloud Vision backend verified DISPLAY={display}", flush=True)
        main()
        return

    if is_railway:
        # 避免 Railway 的自訂啟動命令仍是 `python app.py` 時直接呼叫 tk.Tk()。
        bind_port = port or "8080"
        gunicorn = shutil.which("gunicorn")
        if gunicorn:
            os.execv(gunicorn, [
                gunicorn, "app:app",
                "--bind", f"0.0.0.0:{bind_port}",
                "--workers", "1",
                "--threads", "4",
                "--timeout", "180",
                "--access-logfile", "-",
                "--error-logfile", "-",
            ])
        # 極端情況：gunicorn 指令不存在時，使用 Python 內建 WSGI 伺服器。
        from wsgiref.simple_server import make_server
        with make_server("0.0.0.0", int(bind_port), app) as server:
            print(f"Cloud Vision Web listening on 0.0.0.0:{bind_port}", flush=True)
            server.serve_forever()
        return

    # 本機直接執行時仍保留原本桌面版。
    main()


if __name__ == "__main__":
    _run_railway_entrypoint()
